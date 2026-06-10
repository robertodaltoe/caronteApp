"""
modules/export_excel_sett.py

Aggiorna il file Banca_Ore_Docenti_v3.xlsm con i dati del DB per una settimana specifica:
- Colonna 9 (supplenze svolte R)
- Colonna 6 (permessi orari)
- Colonna 7 (Ed. Civica libero)
- Colonna 8 (delta = saldo settimana)
Aggiorna anche il foglio Riepilogo:
- Colonna 3 (totale supplenze R)
- Colonna 4 (ore a pagamento)
- Colonna 6 (saldo lordo)
- Colonna 7 (saldo effettivo)
- Colonna 8 (CREDITO/DEBITO/OK)
"""
import os
from openpyxl import load_workbook
from modules.import_banca_ore import _parse_data_settimana


def aggiorna_sett_excel(sett_n, movimenti_sett, file_path):
    """
    Aggiorna il foglio sett.N del file Excel con i movimenti del DB.
    movimenti_sett: dict {cognome_upper: {'sup': int, 'perm': int, 'civ': int}}
    """
    wb = load_workbook(file_path, keep_vba=True)
    nome_foglio = f'sett.{sett_n}'

    if nome_foglio not in wb.sheetnames:
        return False, f'Foglio {nome_foglio} non trovato'

    ws = wb[nome_foglio]
    aggiornati = 0

    for r in range(3, ws.max_row + 1):
        nome = ws.cell(r, 1).value
        if not nome:
            continue
        cognome = str(nome).strip().upper().replace('\u2019', "'")

        dati = movimenti_sett.get(cognome)
        if not dati:
            continue

        sup  = dati.get('sup', 0)
        perm = dati.get('perm', 0)
        civ  = dati.get('civ', 0)
        delta = sup - perm - civ

        if sup  > 0: ws.cell(r, 9).value = sup
        if perm > 0: ws.cell(r, 6).value = perm
        if civ  > 0: ws.cell(r, 7).value = civ
        ws.cell(r, 8).value = delta if (sup or perm or civ) else None
        aggiornati += 1

    wb.save(file_path)
    return True, f'Aggiornati {aggiornati} docenti nel foglio {nome_foglio}'


def aggiorna_riepilogo_excel(saldi_tutti, pagamenti, file_path):
    """
    Aggiorna il foglio Riepilogo con i saldi completi del DB.
    saldi_tutti: dict {cognome_upper: {'sup': int, 'perm': int, 'civ': int}}
    pagamenti:   dict {cognome_upper: int (ore)}
    """
    wb = load_workbook(file_path, keep_vba=True)
    ws = wb['Riepilogo']
    aggiornati = 0

    for r in range(2, ws.max_row + 1):
        nome = ws.cell(r, 1).value
        if not nome:
            continue
        cognome = str(nome).strip().upper().replace('\u2019', "'")

        s = saldi_tutti.get(cognome)
        if not s:
            continue

        sup  = s.get('sup', 0)
        perm = s.get('perm', 0)
        civ  = s.get('civ', 0)
        pag  = pagamenti.get(cognome, 0)

        lordo    = sup - perm - civ
        effettivo = lordo - pag

        ws.cell(r, 3).value = sup  if sup  else None
        ws.cell(r, 4).value = pag  if pag  else None
        ws.cell(r, 6).value = abs(lordo)
        ws.cell(r, 7).value = abs(effettivo)
        ws.cell(r, 8).value = 'CREDITO' if effettivo > 0 else ('DEBITO' if effettivo < 0 else 'OK')
        aggiornati += 1

    wb.save(file_path)
    return True, f'Riepilogo aggiornato per {aggiornati} docenti'
