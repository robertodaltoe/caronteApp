"""
modules/export_piano_xlsx.py
Export Excel del Piano Annuale delle Attività, nello stile del modello
originale fornito da Roberto (BOZZA_PIANO_ATTIVITA_2026_27_CORRETTO.xlsx):
un foglio per mese con titolo/intestazione colorati, colonna mese
verticale a sinistra, banner di giorno/sospensione/termine lezioni, e
un foglio "Riepilogo ore" per classe.

Non replica il testo libero del foglio originale (agenda dei Collegio,
"ORDINE DEL GIORNO" dei Consigli) — non è un dato strutturato in
CaronteApp, quelle celle restano vuote da compilare a mano dopo
l'export (concordato con Roberto). Non replica nemmeno il
raggruppamento editoriale multi-giorno del foglio originale (es. la
scritta "CONSIGLI DI CLASSE — Classi prime" che copre più giorni
consecutivi): qui ogni giorno ha il proprio banner e, sotto, un
sotto-titolo per ogni gruppo di eventi dello stesso tipo in quel
giorno — stessi colori e stessa gerarchia visiva del modello, ma
derivato meccanicamente dai dati invece che scritto a mano.
"""
from datetime import time as dtime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from modules.prospetto_supplenze import GIORNI_IT

FONT_NAME = 'Open Sans'

FILL_TITOLO      = PatternFill('solid', fgColor='1F3864')
FILL_INTESTAZIONE = PatternFill('solid', fgColor='2E5395')
FILL_GIORNO      = PatternFill('solid', fgColor='4472C4')
FILL_GRUPPO      = PatternFill('solid', fgColor='6C7A96')
FILL_COLLEGIO    = PatternFill('solid', fgColor='EAD1DC')
FILL_FESTIVITA   = PatternFill('solid', fgColor='C00000')
FILL_SOSPENSIONE = PatternFill('solid', fgColor='375623')

_THIN = Side(style='thin', color='BFBFBF')
BORDER_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

COLONNE = ['Attività', 'Indirizzo', 'Classe', 'Inizio', 'Fine', 'Ore', 'Categoria']
LARGHEZZE = {'A': 3.5, 'B': 32, 'C': 14, 'D': 10, 'E': 9, 'F': 9, 'G': 8, 'H': 24}


def _to_time(s):
    if not s:
        return None
    h, m = s.split(':')
    return dtime(int(h), int(m))


def _bianco_grassetto(size=10):
    return Font(name=FONT_NAME, size=size, bold=True, color='FFFFFFFF')


def _riga_banner(ws, r, testo, fill):
    ws.merge_cells(f'B{r}:H{r}')
    c = ws[f'B{r}']
    c.value = testo
    c.font = _bianco_grassetto()
    c.fill = fill
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    for col in 'BCDEFGH':
        ws[f'{col}{r}'].border = BORDER_THIN
    ws.row_dimensions[r].height = 16.5
    return r + 1


def _riga_evento(ws, r, ev, con_titolo):
    if con_titolo:
        c = ws[f'B{r}']
        c.value = ev.titolo
        c.font = Font(name=FONT_NAME, size=9, bold=True)
        if ev.tipo == 'collegio':
            c.fill = FILL_COLLEGIO
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws[f'C{r}'].value = ev.col_indirizzo or ''
    ws[f'D{r}'].value = ev.col_classe or ''

    t_ini, t_fine = _to_time(ev.ora_inizio), _to_time(ev.ora_fine)
    if t_ini and t_fine:
        ws[f'E{r}'].value, ws[f'E{r}'].number_format = t_ini, 'h:mm'
        ws[f'F{r}'].value, ws[f'F{r}'].number_format = t_fine, 'h:mm'
        ws[f'G{r}'].value = f'=(F{r}-E{r})*24'
    else:
        ws[f'G{r}'].value = round(ev.durata_ore, 2)
    ws[f'G{r}'].number_format = '0.0'
    ws[f'H{r}'].value = ev.col_categoria

    for col in 'CDEFGH':
        cc = ws[f'{col}{r}']
        cc.font = Font(name=FONT_NAME, size=10)
        cc.alignment = Alignment(horizontal='center', vertical='center')
    for col in 'BCDEFGH':
        ws[f'{col}{r}'].border = BORDER_THIN
    ws.row_dimensions[r].height = 15.75
    return r + 1


def _scrivi_eventi_giorno(ws, r, eventi):
    i, n = 0, len(eventi)
    while i < n:
        tipo = eventi[i].tipo
        gruppo = [eventi[i]]
        j = i + 1
        while j < n and eventi[j].tipo == tipo:
            gruppo.append(eventi[j])
            j += 1
        if len(gruppo) == 1:
            r = _riga_evento(ws, r, gruppo[0], con_titolo=True)
        else:
            r = _riga_banner(ws, r, gruppo[0].col_categoria, FILL_GRUPPO)
            for ev in gruppo:
                r = _riga_evento(ws, r, ev, con_titolo=False)
        i = j
    return r


def _crea_foglio_mese(wb, etichetta_mese, righe, anno):
    ws = wb.create_sheet(title=etichetta_mese[:31])
    ws.sheet_view.showGridLines = False
    for col, larg in LARGHEZZE.items():
        ws.column_dimensions[col].width = larg

    ws.merge_cells('A1:H1')
    c = ws['A1']
    c.value = f'PIANO ANNUALE DELLE ATTIVITÀ DEI DOCENTI — {etichetta_mese.upper()}'
    c.font = Font(name=FONT_NAME, size=12, bold=True, color='FFFFFFFF')
    c.fill = FILL_TITOLO
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 25.5

    ws.merge_cells('A2:H2')
    c = ws['A2']
    c.value = (f'PROPOSTA — a.s. {anno}. Indirizzo e Classe da assegnare. '
               'Date/orari da confermare in Collegio Docenti.')
    c.font = Font(name=FONT_NAME, size=8, italic=True, color='FF595959')
    ws.row_dimensions[2].height = 12.75

    riga_intestazione = 3
    for i, testo in enumerate(COLONNE):
        col = get_column_letter(2 + i)
        c = ws[f'{col}{riga_intestazione}']
        c.value = testo
        c.font = _bianco_grassetto()
        c.fill = FILL_INTESTAZIONE
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER_THIN
    ws.row_dimensions[riga_intestazione].height = 18

    r = riga_intestazione + 1
    for data_r, tipo_r, contenuto in righe:
        if tipo_r == 'sospensione':
            multi = contenuto.data_fine != contenuto.data_inizio
            testo = data_r.strftime('%d/%m')
            if multi:
                testo += f"–{contenuto.data_fine.strftime('%d/%m/%Y')}"
            testo += f' — {contenuto.descrizione} ({contenuto.tipo_label})'
            r = _riga_banner(ws, r, testo, FILL_SOSPENSIONE if multi else FILL_FESTIVITA)
        elif tipo_r == 'termine_lezioni':
            r = _riga_banner(ws, r, f"{data_r.strftime('%d/%m/%Y')} — Termine lezioni",
                              FILL_SOSPENSIONE)
        else:
            r = _riga_banner(ws, r, f'{GIORNI_IT[data_r.weekday()].upper()} {data_r.day}',
                              FILL_GIORNO)
            r = _scrivi_eventi_giorno(ws, r, contenuto)

    ultima_riga = max(r - 1, riga_intestazione + 1)
    ws.merge_cells(f'A{riga_intestazione}:A{ultima_riga}')
    c = ws[f'A{riga_intestazione}']
    c.value = etichetta_mese.upper()
    c.font = Font(name=FONT_NAME, size=11, bold=True, color='FFFFFFFF')
    c.fill = FILL_TITOLO
    c.alignment = Alignment(horizontal='center', vertical='center',
                             text_rotation=90, wrap_text=True)

    ws.freeze_panes = f'B{riga_intestazione + 1}'
    return ws


def _crea_foglio_riepilogo(wb, classi_ore, anno):
    ws = wb.create_sheet(title='Riepilogo ore')
    for col, larg in {'A': 12, 'B': 10, 'C': 22, 'D': 14, 'E': 10}.items():
        ws.column_dimensions[col].width = larg

    ws.merge_cells('A1:E1')
    c = ws['A1']
    c.value = 'RIEPILOGO ORE PER CLASSE — Consigli di classe e Scrutini'
    c.font = Font(name=FONT_NAME, size=12, bold=True, color='FFFFFFFF')
    c.fill = FILL_TITOLO
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 25.5

    ws.merge_cells('A2:E2')
    c = ws['A2']
    c.value = (f'a.s. {anno} — generato automaticamente da CaronteApp a ogni export, '
               'elenco classi da Assegnazioni.')
    c.font = Font(name=FONT_NAME, size=8, italic=True, color='FF595959')
    ws.row_dimensions[2].height = 12.75

    intest = ['Indirizzo', 'Classe', 'Ore Consigli di classe', 'Ore Scrutini', 'Totale']
    for i, testo in enumerate(intest):
        col = get_column_letter(1 + i)
        c = ws[f'{col}4']
        c.value = testo
        c.font = _bianco_grassetto()
        c.fill = FILL_INTESTAZIONE
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER_THIN
    ws.row_dimensions[4].height = 18

    r = 5
    for indirizzo, classe, ore_cdc, ore_scrutinio in classi_ore:
        ws[f'A{r}'].value = indirizzo
        ws[f'B{r}'].value = classe
        ws[f'C{r}'].value = round(ore_cdc, 2)
        ws[f'D{r}'].value = round(ore_scrutinio, 2)
        ws[f'E{r}'].value = f'=C{r}+D{r}'
        for col in 'ABCDE':
            cc = ws[f'{col}{r}']
            cc.font = Font(name=FONT_NAME, size=10)
            cc.alignment = Alignment(horizontal='center', vertical='center')
            cc.border = BORDER_THIN
        for col in 'CDE':
            ws[f'{col}{r}'].number_format = '0.0'
        ws.row_dimensions[r].height = 15.75
        r += 1
    return ws


def genera_xlsx_piano_annuale(mesi, classi_ore, anno):
    """
    mesi: [(etichetta_mese, [(data, tipo_riga, contenuto), ...])] —
    stessa struttura di routes.attivita_ist._righe_piano_annuale().
    classi_ore: [(indirizzo, classe, ore_consigli, ore_scrutini), ...].
    Ritorna un openpyxl.Workbook pronto per il salvataggio.
    """
    wb = Workbook()
    wb.remove(wb.active)

    for etichetta_mese, righe in mesi:
        _crea_foglio_mese(wb, etichetta_mese, righe, anno)

    _crea_foglio_riepilogo(wb, classi_ore, anno)

    return wb
