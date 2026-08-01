"""
modules/import_banca_ore.py

Legge tutte le settimane dal file Banca_Ore_Docenti_v3.xlsm e restituisce
una lista di movimenti da importare nel DB, con gestione doppioni.
Le date vengono estratte automaticamente dai titoli dei fogli.
"""
import re
from openpyxl import load_workbook
from datetime import date

MESI = {
    'OTTOBRE': 10, 'NOVEMBRE': 11, 'DICEMBRE': 12,
    'GENNAIO':  1, 'FEBBRAIO':  2, 'MARZO':     3,
    'APRILE':   4, 'MAGGIO':    5, 'GIUGNO':    6,
    'LUGLIO':   7, 'AGOSTO':    8, 'SETTEMBRE': 9,
}


def _parse_data_settimana(titolo, sn_precedente_data=None):
    """
    Estrae la data di inizio settimana dal titolo del foglio.
    Es: "Settimana 2 — 27-01 NOVEMBRE" -> il 27 è ottobre (mese precedente)
    Usa sn_precedente_data per inferire il mese corretto nei casi limite.
    """
    if not titolo:
        return None

    # Pattern: "GG-GG MESE" o "GG-GG MESE"
    m = re.search(r'(\d{1,2})[-–](\d{1,2})\s+([A-ZÀÈÌ]+)', str(titolo))
    if not m:
        return None

    giorno_inizio = int(m.group(1))
    giorno_fine   = int(m.group(2))
    mese_nome     = m.group(3)
    mese          = MESI.get(mese_nome)
    if not mese:
        return None

    # Caso limite: "27-01 NOVEMBRE" — il giorno 27 è nel mese precedente
    # Se giorno_inizio > giorno_fine, il mese è quello PRECEDENTE al mese citato
    if giorno_inizio > giorno_fine:
        mese_inizio = mese - 1 if mese > 1 else 12
    else:
        mese_inizio = mese

    anno = 2025 if mese_inizio >= 9 else 2026

    try:
        return date(anno, mese_inizio, giorno_inizio)
    except ValueError:
        return None


def leggi_movimenti_file(path):
    """
    Legge tutte le settimane dal file XLSM.
    Restituisce lista di dict:
    {sett_n, cognome, data, tipo, ore, descrizione}
    """
    wb = load_workbook(path, data_only=True)
    movimenti = []

    for sn in range(1, 35):
        nome_foglio = f'sett.{sn}'
        if nome_foglio not in wb.sheetnames:
            continue

        ws = wb[nome_foglio]
        titolo = ws.cell(1, 1).value
        data_sett = _parse_data_settimana(titolo)

        if not data_sett:
            # Foglio senza data leggibile (es. sett.30 "DA DEFINIRE") — salta
            continue

        for r in range(3, ws.max_row + 1):
            cognome = ws.cell(r, 1).value
            if not cognome or not str(cognome).strip():
                continue
            # Salta righe non-dati (navigazione, ecc.)
            if str(cognome).startswith('▤︎') or str(cognome).startswith('▲︎'):
                continue
            cognome = str(cognome).strip()

            sup  = _num(ws.cell(r, 9).value)  # supplenze svolte
            perm = _num(ws.cell(r, 6).value)  # permessi orari
            civ  = _num(ws.cell(r, 7).value)  # civica libera

            if sup > 0:
                movimenti.append({
                    'sett_n':      sn,
                    'cognome':     cognome,
                    'data':        data_sett,
                    'tipo':        'supplenza_recupero',
                    'ore':         sup,
                    'descrizione': f'Supplenza svolta — sett.{sn}',
                })
            if perm > 0:
                movimenti.append({
                    'sett_n':      sn,
                    'cognome':     cognome,
                    'data':        data_sett,
                    'tipo':        'permesso_orario',
                    'ore':         perm,
                    'descrizione': f'Permesso orario — sett.{sn}',
                })
            if civ > 0:
                movimenti.append({
                    'sett_n':      sn,
                    'cognome':     cognome,
                    'data':        data_sett,
                    'tipo':        'civica',
                    'ore':         civ,
                    'descrizione': f'Libero Ed. Civica — sett.{sn}',
                })

    return movimenti


def _num(v):
    """Converte valore cella in numero intero, 0 se None o non numerico."""
    if v is None:
        return 0
    try:
        n = int(float(str(v)))
        return max(0, n)  # mai negativo
    except (ValueError, TypeError):
        return 0
