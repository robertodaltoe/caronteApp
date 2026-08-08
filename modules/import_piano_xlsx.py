"""
Import del Piano Annuale delle Attività dei Docenti da file .xlsx.

Legge un file .xlsx costruito secondo lo "standard" adottato dalla scuola
(banner-giorno a piena larghezza, righe-slot per Consigli/GLO, righe-evento
per Collegio/Formazione/Incontri, righe-scadenza) e lo trasforma in una
lista di eventi pronti per essere salvati come AttivitaIst.

Questo modulo NON tocca il database: `parse_piano_xlsx()` restituisce solo
dati. Il salvataggio è responsabilità del chiamante (vedi routes/attivita_ist.py).

Riconoscimento righe — basato sul colore di riempimento della cella in
colonna B (merge a piena larghezza) e sulla struttura circostante:

  · titolo foglio (riga 1)         → estrazione mese/i e anno
  · intestazione tabella           → individua colonna "Attività" (rif. colonne successive)
  · banner blu (giorno)            → "LUNEDÌ 7" ecc. — aggiorna la data corrente
  · banner grigio-blu (sezione)    → titolo Consigli/GLO SE seguito da un banner-giorno,
                                      altrimenti è solo informativo (festività, sospensioni…) e viene ignorato
  · banner rosso/verde (festività) → informativo, ignorato (non genera eventi)
  · riga grigia "ORDINE DEL GIORNO"+ testo → nota, allegata retroattivamente
                                      agli slot della sezione corrente
  · riga verde "⚑ scadenza"        → informativa, ignorata (nessun orario associato)
  · riga con orario e colonna B vuota   → SLOT del Consiglio/GLO corrente
  · riga con orario e colonna B piena   → EVENTO (Collegio, Formazione, Incontro, ...)

Limiti noti: la data si ricava contando i giorni banner in ordine (i mesi
combinati in un unico foglio, es. "NOVEMBRE-DICEMBRE", vengono distinti
assumendo che il numero del giorno non possa mai tornare indietro senza
essere passati al mese successivo del foglio).
"""
import re
import io
import datetime
import openpyxl

MESI_MAP = {
    'GENNAIO': 1, 'FEBBRAIO': 2, 'MARZO': 3, 'APRILE': 4, 'MAGGIO': 5,
    'GIUGNO': 6, 'LUGLIO': 7, 'AGOSTO': 8, 'SETTEMBRE': 9, 'OTTOBRE': 10,
    'NOVEMBRE': 11, 'DICEMBRE': 12,
}

DAY_FILL   = '4472C4'
TITLE_FILL = '6C7A96'
FEST_FILL  = 'C00000'
LEZ_FILL   = '375623'
ODG_FILL   = '808080'
SCAD_FILL  = 'C6E0B4'
INFO_FILLS = {FEST_FILL, LEZ_FILL}


def _fill_rgb(cell):
    try:
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == 'rgb':
            rgb = cell.fill.fgColor.rgb
            if rgb and len(rgb) == 8:
                return rgb[2:]  # strip alpha 'FF'
            return rgb
    except Exception:
        pass
    return None


def _is_wide_merge(ws, row, col, min_span=3):
    for mc in ws.merged_cells.ranges:
        if mc.min_row == row and mc.min_col == col and (mc.max_col - mc.min_col) >= min_span:
            return True
    return False


def _extract_months_year(title_text):
    if not title_text:
        return [], None
    ym = re.search(r'(20\d{2})', title_text)
    year = int(ym.group(1)) if ym else None
    hits = []
    for nome, num in MESI_MAP.items():
        idx = title_text.upper().find(nome)
        if idx >= 0:
            hits.append((idx, num))
    hits.sort()
    return [num for _, num in hits], year


def _classify_sezione(titolo):
    t = titolo.upper()
    if t.strip() == 'GLO' or ' GLO' in f' {t}':
        return 'glo'
    if 'SCRUTIN' in t:
        return 'scrutinio'
    if 'CONSIGL' in t:
        return 'consiglio_classe'
    return 'altro'


def _classify_evento(label):
    l = label.upper()
    if 'COLLEGIO' in l:
        return 'collegio'
    if 'FORMAZIONE' in l:
        return 'formazione'
    if 'GLO' in l:
        return 'glo'
    if 'REFERENT' in l:
        return 'riunione_referenti'
    if 'DIPARTIMENT' in l:
        return 'dipartimento'
    if 'MATERIA' in l or 'MATERIE' in l:
        return 'riunione_materia'
    if ('INCONTRO' in l or 'COLLOQ' in l) and ('FAMIGL' in l or 'GENITOR' in l or 'SCUOLA' in l):
        return 'incontro_famiglie'
    if 'SCRUTIN' in l:
        return 'scrutinio'
    return 'altro'


def _to_hhmm(value):
    if isinstance(value, datetime.time):
        return value.strftime('%H:%M')
    if isinstance(value, datetime.datetime):
        return value.strftime('%H:%M')
    return None


def parse_piano_xlsx(file_bytes_or_path):
    """
    Restituisce {'fogli': [...], 'eventi': [...], 'avvisi': [...]}.
    Ogni evento: dict con tipo, titolo, data (iso str), ora_ini, ora_fin,
    classe, note, foglio (nome foglio sorgente), riga (numero riga excel,
    per debug).
    """
    if isinstance(file_bytes_or_path, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes_or_path), data_only=True)
    else:
        wb = openpyxl.load_workbook(file_bytes_or_path, data_only=True)

    eventi = []
    avvisi = []
    fogli_letti = []

    for ws in wb.worksheets:
        title_text = ws.cell(1, 1).value or ''
        mesi, anno = _extract_months_year(str(title_text))
        if not mesi or not anno:
            continue  # non è un foglio mensile del piano (es. Introduzione)

        # individua riga di intestazione ("Attività" in colonna B o C)
        header_row = None
        col_attivita = None
        for r in range(1, 8):
            for c in range(1, 6):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip().lower() == 'attività':
                    header_row = r
                    col_attivita = c
                    break
            if header_row:
                break
        if not header_row:
            avvisi.append(f'Foglio "{ws.title}": intestazione non trovata, saltato.')
            continue

        col_indirizzo = col_attivita + 1
        col_classe    = col_attivita + 2
        col_inizio    = col_attivita + 3
        col_fine      = col_attivita + 4

        # limite di lettura: rispetta l'area di stampa se impostata (esclude
        # eventuali blocchi "modello da copiare" fuori stampa)
        max_row = ws.max_row
        if ws.print_area:
            m = re.search(r'\$?[A-Z]+\$?(\d+)\s*$', str(ws.print_area).split(',')[0])
            if m:
                max_row = min(max_row, int(m.group(1)))

        fogli_letti.append(ws.title)

        month_idx = 0
        last_day = 0
        data_corrente = None
        sezione_titolo = None
        sezione_tipo = None
        sezione_indices = []  # indici in `eventi` degli slot della sezione aperta

        r = header_row + 1
        while r <= max_row:
            cell_b = ws.cell(r, col_attivita)
            fill = _fill_rgb(cell_b)
            wide = _is_wide_merge(ws, r, col_attivita)
            testo_b = (cell_b.value or '').strip() if isinstance(cell_b.value, str) else ''

            if wide and fill == DAY_FILL:
                m = re.search(r'(\d{1,2})', testo_b)
                if m:
                    day = int(m.group(1))
                    if day < last_day:
                        month_idx = min(month_idx + 1, len(mesi) - 1)
                    last_day = day
                    try:
                        data_corrente = datetime.date(anno, mesi[month_idx], day)
                    except ValueError:
                        avvisi.append(f'Foglio "{ws.title}" riga {r}: data non valida ({mesi[month_idx]}/{day}/{anno}).')
                        data_corrente = None
                r += 1
                continue

            if wide and fill == TITLE_FILL:
                # sezione (Consigli/GLO) SE seguita da un banner-giorno entro 2 righe,
                # altrimenti è un banner puramente informativo (es. "ESAMI INTEGRATIVI")
                prossimo_fill = _fill_rgb(ws.cell(r + 1, col_attivita)) if r + 1 <= max_row else None
                if prossimo_fill == DAY_FILL:
                    sezione_titolo = testo_b
                    sezione_tipo = _classify_sezione(testo_b)
                    sezione_indices = []
                r += 1
                continue

            if wide and fill in INFO_FILLS:
                r += 1
                continue

            if wide and fill == SCAD_FILL:
                r += 1
                continue

            if fill == ODG_FILL and testo_b.upper().startswith('ORDINE'):
                nota = ws.cell(r + 1, col_attivita).value if r + 1 <= max_row else None
                if nota and sezione_indices:
                    for idx in sezione_indices:
                        eventi[idx]['note'] = nota
                r += 2
                continue

            ora_ini_val = ws.cell(r, col_inizio).value
            ora_fin_val = ws.cell(r, col_fine).value
            has_orario = isinstance(ora_ini_val, (datetime.time, datetime.datetime))

            if has_orario and not testo_b:
                # riga-slot di un Consiglio/GLO
                if not sezione_titolo or not data_corrente:
                    avvisi.append(f'Foglio "{ws.title}" riga {r}: slot orario senza sezione/data riconosciuta, saltato.')
                    r += 1
                    continue
                indirizzo = ws.cell(r, col_indirizzo).value
                classe_v  = ws.cell(r, col_classe).value
                classe = None
                parts = [str(x).strip() for x in (classe_v, indirizzo) if x]
                if parts:
                    classe = ' '.join(parts)
                eventi.append({
                    'tipo': sezione_tipo, 'titolo': sezione_titolo,
                    'data': data_corrente.isoformat(),
                    'ora_ini': _to_hhmm(ora_ini_val), 'ora_fin': _to_hhmm(ora_fin_val),
                    'classe': classe, 'note': None,
                    'foglio': ws.title, 'riga': r,
                })
                sezione_indices.append(len(eventi) - 1)
                r += 1
                continue

            if testo_b:
                # riga-evento (Collegio, Formazione, Incontro, ...)
                desc = ws.cell(r, col_indirizzo).value
                desc = str(desc).strip() if desc else ''
                titolo = f'{testo_b} — {desc}' if desc else testo_b
                oi = _to_hhmm(ora_ini_val)
                of = _to_hhmm(ora_fin_val)
                if oi == of:
                    oi = of = None
                if data_corrente:
                    eventi.append({
                        'tipo': _classify_evento(testo_b), 'titolo': titolo,
                        'data': data_corrente.isoformat(),
                        'ora_ini': oi, 'ora_fin': of,
                        'classe': None, 'note': None,
                        'foglio': ws.title, 'riga': r,
                    })
                r += 1
                continue

            r += 1

    return {'fogli': fogli_letti, 'eventi': eventi, 'avvisi': avvisi}
