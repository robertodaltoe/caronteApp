"""
modules/parser_orario.py
Logica di parsing e importazione orario.
Usata sia da import_orario.py che dalla route /sincronizzazione.
"""
import re, datetime, os, json
from openpyxl import load_workbook

SHEET_ORARIO  = '7_ORARIO DEFINITIVO_teachers_ti'
SHEET_DOCENTI = 'Docenti'
GIORNI = ['Lunedì','Martedì','Mercoledì','Giovedì','Venerdì','Sabato']
LIBERO = {'---', '-x-', '', 'none'}

def clean(v):
    return str(v).strip() if v is not None else ''

def is_libero(v):
    return clean(v).lower() in LIBERO

def is_classe(s):
    return bool(re.match(r'^\d[A-Z]', s.strip()))

def build_col_map(ws):
    giorno_map = {}
    for c in range(1, ws.max_column + 1):
        v = clean(ws.cell(2, c).value)
        for i, g in enumerate(GIORNI):
            if g.lower() in v.lower():
                giorno_map[c] = i
                break
    col_map = {}
    ora_counter = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(3, c).value
        if not isinstance(v, datetime.time):
            continue
        g_num = None
        for gc in sorted(giorno_map.keys(), reverse=True):
            if gc <= c:
                g_num = giorno_map[gc]
                break
        if g_num is None:
            continue
        ora_counter[g_num] = ora_counter.get(g_num, 0) + 1
        col_map[c] = (g_num, ora_counter[g_num])
    return col_map


def parse_file(excel_path):
    wb = load_workbook(excel_path, data_only=True)
    ws_or = wb[SHEET_ORARIO]

    merged = {}
    for merge in ws_or.merged_cells.ranges:
        mr, mc = merge.min_row, merge.min_col
        for r in range(merge.min_row, merge.max_row + 1):
            for c in range(merge.min_col, merge.max_col + 1):
                merged[(r, c)] = (mr, mc)

    def val(r, c):
        mr, mc = merged.get((r, c), (r, c))
        return ws_or.cell(mr, mc).value

    col_map = build_col_map(ws_or)

    anagrafica = []
    if SHEET_DOCENTI in wb.sheetnames:
        ws_doc = wb[SHEET_DOCENTI]
        for r in range(3, ws_doc.max_row + 1):
            cv = ws_doc.cell(r, 1).value
            if not cv:
                continue
            attivo_s = clean(ws_doc.cell(r, 6).value).upper()
            anagrafica.append({
                'cognome':       clean(cv).upper(),
                'nome':          clean(ws_doc.cell(r, 2).value),
                'materia':       clean(ws_doc.cell(r, 3).value),
                'ore_contratto': int(ws_doc.cell(r, 4).value or 0),
                'attivo':        attivo_s in ('SÌ','SI','S','1','TRUE'),
            })

    slots = []
    docente_corrente = None

    for r in range(4, ws_or.max_row + 1):
        tipo_riga = clean(val(r, 1)).upper()
        if tipo_riga not in ('CLASSE', 'MATERIE', 'COMPRESENZA'):
            continue

        if tipo_riga == 'CLASSE':
            nd = clean(val(r, 2)).upper()
            if nd:
                docente_corrente = nd
            if not docente_corrente:
                continue
            riga_mat = r + 1
            for c, (giorno, ora) in col_map.items():
                cs = clean(val(r, c))
                ms = clean(val(riga_mat, c))
                if is_libero(cs):
                    continue
                if is_classe(cs):
                    tipo = 'lezione'
                elif 'POTENZ' in cs.upper():
                    tipo = 'potenziamento'
                elif 'DISPOS' in cs.upper():
                    tipo = 'disposizione'
                else:
                    tipo = 'altro'
                slots.append({'cognome_file': docente_corrente,
                               'giorno': giorno, 'ora': ora,
                               'classe': cs, 'materia': ms, 'tipo_ora': tipo})

        elif tipo_riga == 'COMPRESENZA' and docente_corrente:
            riga_ref = None
            for rr in range(r - 1, 3, -1):
                if clean(val(rr, 1)).upper() == 'CLASSE':
                    riga_ref = rr
                    break
            for c, (giorno, ora) in col_map.items():
                cs = clean(val(r, c))
                if is_libero(cs) or '|' not in cs:
                    continue
                cognomi = [x.strip().upper() for x in cs.split('|')]
                cr = clean(val(riga_ref, c)) if riga_ref else ''
                mc = clean(val(riga_ref + 1, c)) if riga_ref else ''
                for cog in cognomi:
                    slots.append({'cognome_file': cog, 'giorno': giorno,
                                  'ora': ora, 'classe': cr, 'materia': mc,
                                  'tipo_ora': 'compresenza'})

    return {'docenti_anagrafica': anagrafica, 'slots': slots}


def applica_importazione(excel_path, db_session):
    from models.docente import Docente
    from models.orario_docente import OrarioDocente
    from models.sync_orario import AliasDocente, LogImportazione

    parsed = parse_file(excel_path)

    alias_map  = {a.nome_file.upper(): a.id_docente for a in AliasDocente.query.all()}
    docenti_db = {d.cognome.upper(): d for d in Docente.query.all()}

    def risolvi(cognome_file):
        cog = cognome_file.upper().strip()
        if cog in alias_map:
            from models import db
            return db.session.get(Docente, alias_map[cog])
        return docenti_db.get(cog)

    stats = {'slot_totali': 0, 'docenti_nuovi': 0,
             'aggiornati': 0, 'non_riconosciuti': set()}

    # Aggiorna/crea docenti
    for ana in parsed['docenti_anagrafica']:
        doc = risolvi(ana['cognome'])
        if doc is None:
            doc = Docente(
                cognome=ana['cognome'], nome=ana['nome'],
                nome_display=ana['cognome'],
                materia=ana['materia'] or None,
                ore_contratto=ana['ore_contratto'],
                attivo=ana['attivo'],
            )
            db_session.add(doc)
            db_session.flush()
            docenti_db[ana['cognome']] = doc
            stats['docenti_nuovi'] += 1
        else:
            changed = False
            if not doc.materia and ana['materia']:
                doc.materia = ana['materia']; changed = True
            if not doc.nome and ana['nome']:
                doc.nome = ana['nome']; changed = True
            if changed:
                stats['aggiornati'] += 1

    # Ricrea orario
    OrarioDocente.query.delete()
    db_session.flush()

    seen = set()
    for slot in parsed['slots']:
        doc = risolvi(slot['cognome_file'])
        if doc is None:
            stats['non_riconosciuti'].add(slot['cognome_file'])
            continue
        key = (doc.id, slot['giorno'], slot['ora'],
               'comp' if slot['tipo_ora'] == 'compresenza' else 'norm')
        if key in seen:
            continue
        seen.add(key)
        db_session.add(OrarioDocente(
            id_docente=doc.id, giorno=slot['giorno'], ora=slot['ora'],
            classe=slot['classe'], materia=slot['materia'],
            tipo_ora=slot['tipo_ora'],
        ))
        stats['slot_totali'] += 1

    # Salva log
    nr_list = list(stats['non_riconosciuti'])
    log = LogImportazione(
        file_nome=os.path.basename(excel_path),
        slot_totali=stats['slot_totali'],
        docenti_nuovi=stats['docenti_nuovi'],
        non_riconosciuti=json.dumps(nr_list),
        esito='warning' if nr_list else 'ok',
    )
    db_session.add(log)
    db_session.commit()

    stats['non_riconosciuti'] = nr_list
    return stats
