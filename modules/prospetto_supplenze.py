"""
modules/prospetto_supplenze.py
Genera il Prospetto Supplenze giornaliero nel formato MATRICE 25_26 aggiornato.
"""
import io, re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Mappatura ora -> (col_assente, col_sostituto) — celle merged D:E, F:G, H:I...
ORA_COLS = {
    1: (4,  6),   2: (8,  10),  3: (12, 14),
    4: (16, 18),  5: (20, 22),  6: (24, 26),
    7: (28, 30),  8: (32, 34),  9: (36, 38),
}

# Righe classi (col B:C merged)
CLASSE_RIGHE_RAW = {
    '1A AFM': 9,  '2A AFM': 10, '2B AFM': 11,
    '3A RIM': 12, '4A RIM': 13, '5A RIM': 14,
    '1A CAT': 15, '2A CAT': 16, '3A CAT': 17, '4A CAT': 18, '5A CAT': 19,
    '1A LSC': 22, '2A LSC': 23, '2B LSC': 24, '3A LSC': 25, '4A LSC': 26, '5A LSC': 27,
    '1A LSU': 28, '2A LSU': 29, '3A LSU': 30, '4A LSU': 31, '5A LSU': 32, '5B LSU': 33,
    '1A LLI': 34, '1B LLI': 35, '2A LLI': 36, '2B LLI': 37, '3A LLI': 38,
    '4A LLI': 39, '5A LLI': 40, '5B LLI': 41,
    '1A LSP': 42, '2A LSP': 43, '3A LSP': 44, '4A LSP': 45, '5A LSP': 46,
}

# Seconda intestazione righe 20-21 (stessa struttura)
CLASSE_RIGHE_2 = {k: v for k, v in CLASSE_RIGHE_RAW.items() if v >= 22}

# Firme: (col_docente, col_euro, col_c, col_r, col_p) righe 51-59
# Template aggiornato: B:E=docente, J=€, K=C, L=R, M=P per col1
#                      N:Q=docente, V=€, W=C, X=R, Y=P per col2
#                      Z:AC=docente, AH=€, AI=C, AJ=R, AK=P per col3
FIRME_COLS = [
    (2,  10, 11, 12, 13),   # col1: docente=B(2), €=J(10), C=K(11), R=L(12), P=M(13)
    (14, 22, 23, 24, 25),   # col2: docente=N(14), €=V(22), C=W(23), R=X(24), P=Y(25)
    (26, 34, 35, 36, 37),   # col3: docente=Z(26), €=AH(34), C=AI(35), R=AJ(36), P=AK(37)
]
FIRME_START = 51
FIRME_END   = 59

FILL_TIPO = {
    'recupero':      PatternFill('solid', fgColor='C6EFCE'),
    'pagamento':     PatternFill('solid', fgColor='FFEB9C'),
    'completamento': PatternFill('solid', fgColor='D9D9D9'),
    'potenziamento': PatternFill('solid', fgColor='DDEBF7'),
    'disposizione':  PatternFill('solid', fgColor='E2EFDA'),
}

TIPO_SIGLA = {
    'recupero': 'R', 'pagamento': '€',
    'completamento': 'C', 'potenziamento': 'P', 'disposizione': 'D',
}

GIORNI_IT = ['Lunedì','Martedì','Mercoledì','Giovedì','Venerdì','Sabato','Domenica']
MESI_IT   = ['','gennaio','febbraio','marzo','aprile','maggio','giugno',
             'luglio','agosto','settembre','ottobre','novembre','dicembre']

FONT_BOLD = Font(name='Calibri', size=10, bold=True)
FONT_DATA = Font(name='Calibri', size=11, bold=True)


def _norm(c):
    return re.sub(r'\s+', ' ', str(c).strip().upper())


def _data_label(d):
    return f"{GIORNI_IT[d.weekday()]} {d.day} {MESI_IT[d.month]} {d.year}"


def genera_prospetto(data_sel, supplenze, template_path, save_dir=None, attivita_ist=None):
    wb = load_workbook(template_path)

    if 'MATRICE 25_26' in wb.sheetnames:
        ws = wb.copy_worksheet(wb['MATRICE 25_26'])
        nome_foglio = data_sel.strftime('%Y%m%d') + '_Prospetto supplenze'
        ws.title = nome_foglio
    else:
        ws = wb.active
        nome_foglio = ws.title

    # ── 1. Data ───────────────────────────────────────────────
    data_str = _data_label(data_sel)
    ws.cell(4, 5).value = data_str
    ws.cell(4, 5).font  = FONT_DATA

    # ── 2. Docenti assenti/indisponibili in riga 4 e riga 5 ──
    # Raccogli cognomi distinti di assenti + indisponibili
    assenti_set = set()
    for s in supplenze:
        if s.stato != 'annullata' and s.assente:
            assenti_set.add(s.assente.cognome)

    # Recupera anche indisponibili del giorno dal DB
    from models.indisponibilita import Indisponibilita
    from models.assenza import Assenza
    from models import db
    indisp_list = Indisponibilita.query.filter_by(data=data_sel).all()
    for i in indisp_list:
        if i.docente:
            assenti_set.add(i.docente.cognome)

    assenti_str = ', '.join(sorted(assenti_set))

    # Riga 4: M4 (col 13) — spazio tra H4:L4 e il resto
    ws.cell(4, 13).value = assenti_str
    ws.cell(4, 13).font  = FONT_BOLD

    # Riga 5: B5 (col 2) — cella unita B5:AB5
    ws.cell(5, 2).value = assenti_str
    ws.cell(5, 2).font  = FONT_BOLD

    # ── 3. Svuota celle dati ──────────────────────────────────
    for riga in CLASSE_RIGHE_RAW.values():
        for ora in range(1, 10):
            ca, cs = ORA_COLS[ora]
            try:
                ws.cell(riga, ca).value = None
                ws.cell(riga, cs).value = None
                ws.cell(riga, cs).fill  = PatternFill('none')
            except Exception:
                pass

    for col_doc, col_euro, col_c, col_r, col_p in FIRME_COLS:
        for riga in range(FIRME_START, FIRME_END + 1):
            for c in [col_doc, col_euro, col_c, col_r, col_p]:
                try:
                    ws.cell(riga, c).value = None
                except Exception:
                    pass

    # ── 4. Compila supplenze ──────────────────────────────────
    for s in supplenze:
        if s.stato == 'annullata' or s.ora not in ORA_COLS:
            continue

        riga = CLASSE_RIGHE_RAW.get(_norm(s.classe))
        if riga is None:
            continue

        ca, cs = ORA_COLS[s.ora]

        nome_ass  = s.assente.cognome if s.assente else ''
        nome_sost = ''

        if s.sostituto:
            sigla = TIPO_SIGLA.get(s.tipo or '', '')
            nome_sost = s.sostituto.cognome + (f' ({sigla})' if sigla else '')
        elif s.stato == 'scoperta':
            nome_sost = ''         # lascia vuota invece di ???
        elif s.stato == 'non_assegnabile':
            nome_sost = 'N/A'
            nome_ass  = ''

        try:
            if nome_ass:
                ws.cell(riga, ca).value = nome_ass
                ws.cell(riga, ca).font  = FONT_BOLD
            if nome_sost:
                ws.cell(riga, cs).value = nome_sost
                ws.cell(riga, cs).font  = FONT_BOLD
                if s.sostituto and s.tipo in FILL_TIPO:
                    ws.cell(riga, cs).fill = FILL_TIPO[s.tipo]
        except Exception:
            pass

    # ── 5. Tabella firme ──────────────────────────────────────
    sostituti = {}
    for s in supplenze:
        if s.stato == 'annullata' or not s.sostituto:
            continue
        nome = s.sostituto.cognome
        sostituti.setdefault(nome, set())
        if s.tipo:
            sostituti[nome].add(s.tipo)

    idx = 0
    for col_doc, col_euro, col_c, col_r, col_p in FIRME_COLS:
        for riga in range(FIRME_START, FIRME_END + 1):
            if idx >= len(sostituti):
                break
            nome, tipi = sorted(sostituti.items())[idx]
            try:
                ws.cell(riga, col_doc).value = nome
                ws.cell(riga, col_doc).font  = FONT_BOLD
                if 'pagamento'     in tipi: ws.cell(riga, col_euro).value = 'X'
                if 'completamento' in tipi: ws.cell(riga, col_c).value    = 'X'
                if 'recupero'      in tipi: ws.cell(riga, col_r).value    = 'X'
                if 'potenziamento' in tipi: ws.cell(riga, col_p).value    = 'X'
            except Exception:
                pass
            idx += 1

    # ── 6. Attività Istituzionali del giorno ────────────────
    if attivita_ist:
        from openpyxl.styles import Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        # Aggiungi un foglio separato "Att. Istituzionali"
        ws_ist = wb.create_sheet(title='Att. Istituzionali')
        _s = Side(style='thin', color='AAAAAA')
        _border = Border(left=_s, right=_s, top=_s, bottom=_s)
        _fill_hdr = PatternFill('solid', fgColor='1F3864')
        _font_hdr = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        _font_ist = Font(name='Calibri', size=10)

        ws_ist.cell(1, 1).value = f'Attività Istituzionali — {_data_label(data_sel)}'
        ws_ist.cell(1, 1).font  = Font(name='Calibri', size=12, bold=True, color='1F3864')

        hdrs = ['Tipo', 'Titolo', 'Orario', 'Ore', 'Classe/Dip.', 'Partecipanti']
        for c, h in enumerate(hdrs, 1):
            cell = ws_ist.cell(3, c)
            cell.value = h; cell.fill = _fill_hdr
            cell.font  = _font_hdr; cell.border = _border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for r, ev in enumerate(attivita_ist, 4):
            orario = ''
            if ev.ora_inizio:
                orario = ev.ora_inizio + ('–' + ev.ora_fine if ev.ora_fine else '')
            classe_dip = ev.classe or ''
            if ev.dipartimento:
                classe_dip = ev.dipartimento.sigla
            n_part = len(ev.partecipanti) if ev.partecipanti else 0
            n_pres = sum(1 for p in ev.presenze if p.stato == 'presente') if ev.presenze else '—'

            vals = [
                ev.tipo_label,
                ev.titolo,
                orario,
                f'{ev.durata_ore:.1f}h' if ev.durata_ore else '—',
                classe_dip,
                f'{n_pres}/{n_part}' if ev.presenze else f'{n_part} previsti',
            ]
            for c, v in enumerate(vals, 1):
                cell = ws_ist.cell(r, c)
                cell.value = v; cell.font = _font_ist; cell.border = _border

        ws_ist.column_dimensions['A'].width = 22
        ws_ist.column_dimensions['B'].width = 40
        ws_ist.column_dimensions['C'].width = 12
        ws_ist.column_dimensions['D'].width = 8
        ws_ist.column_dimensions['E'].width = 14
        ws_ist.column_dimensions['F'].width = 18

    # ── 7. Rimuovi altri fogli e salva ────────────────────────
    fogli_da_tenere = {nome_foglio}
    if attivita_ist:
        fogli_da_tenere.add('Att. Istituzionali')
    for nome in [s for s in wb.sheetnames if s not in fogli_da_tenere]:
        del wb[nome]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    xlsx_bytes = buf.getvalue()

    if save_dir:
        import os
        os.makedirs(save_dir, exist_ok=True)
        nome_file = nome_foglio.replace(' ', '_') + '.xlsx'
        with open(os.path.join(save_dir, nome_file), 'wb') as f:
            f.write(xlsx_bytes)

    return xlsx_bytes
