def _build_xlsx_singolo(docente, saldi, storico, supplenze, saldo_netto, saldo_effettivo, oggi):
    """
    Genera XLSX singolo docente replicando esattamente il formato originale.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"DOC_{docente.cognome[:24]}".replace('/', '_').replace('\\', '_')

    # ── STILI ────────────────────────────────────────────────
    FONT_TITLE  = Font(name="Calibri", bold=True,  size=15)
    FONT_HDR    = Font(name="Calibri", bold=True,  size=12)
    FONT_NORM   = Font(name="Calibri", bold=False, size=12)
    FONT_TOT    = Font(name="Calibri", bold=True,  size=12)
    FONT_TOT_WH = Font(name="Calibri", bold=True,  size=12, color="FFFFFF")
    FONT_LINK   = Font(name="Calibri", size=12, color="0563C1", underline="single")

    def fill(hex6):
        return PatternFill("solid", fgColor=hex6)

    NO_FILL     = PatternFill(fill_type=None)
    GREY_HDR    = fill("D9D9D9")
    GREEN_LIGHT = fill("C6EFCE")
    RED_LIGHT   = fill("FFC7CE")
    GREEN_DARK  = fill("006100")
    RED_DARK    = fill("C00000")
    YELLOW_PAY  = fill("FFEB9C")

    def thin():
        s = Side(style="thin")
        return Border(left=s, right=s, top=s, bottom=s)

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    def color_saldo(val):
        if val is None or val == 0: return NO_FILL
        return GREEN_LIGHT if val > 0 else RED_LIGHT

    # ── LARGHEZZE COLONNE ────────────────────────────────────
    ws.column_dimensions['A'].width = 60.83
    ws.column_dimensions['B'].width = 25.83
    ws.column_dimensions['C'].width = 5.66
    ws.column_dimensions['D'].width = 14.83
    ws.column_dimensions['E'].width = 17.33
    ws.column_dimensions['F'].width = 17.66
    ws.column_dimensions['G'].width = 9.50
    ws.column_dimensions['H'].width = 45.83

    # ── RIGA 1 — TITOLO ──────────────────────────────────────
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value     = (f"Report ore di supplenza/da recuperare — {docente.cognome}"
                   f" — aggiornamento {oggi.strftime('%d/%m/%Y')}")
    c.font      = FONT_TITLE
    c.alignment = CENTER
    ws.row_dimensions[1].height = 20

    # ── RIGHE 3-5 — TESTATA ──────────────────────────────────
    for r in [3, 4, 5]:
        ws.row_dimensions[r].height = 17

    ws.cell(3, 1).value = "Docente";               ws.cell(3, 1).font = FONT_HDR
    ws.cell(3, 2).value = docente.cognome;          ws.cell(3, 2).font = FONT_NORM
    ws.cell(3, 8).value = "Torna all'Indice";       ws.cell(3, 8).font = FONT_LINK

    ws.cell(4, 1).value = "Saldo lordo da Riepilogo"; ws.cell(4, 1).font = FONT_HDR
    ws.cell(4, 2).value = saldo_netto
    ws.cell(4, 2).fill  = color_saldo(saldo_netto)
    ws.cell(4, 2).font  = FONT_NORM

    ws.cell(5, 1).value = "Saldo effettivo";        ws.cell(5, 1).font = FONT_HDR
    ws.cell(5, 2).value = saldo_effettivo
    ws.cell(5, 2).fill  = color_saldo(saldo_effettivo)
    ws.cell(5, 2).font  = FONT_NORM

    # ── RIGA 9 — INTESTAZIONI TABELLA ────────────────────────
    headers = ["Giorno", "Periodo", "Delta", "Supplenza svolta",
               "Permessi/richieste", "Ore libere Ed. Civica", "Voce", "Note"]
    ws.row_dimensions[9].height = 17
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(9, c_idx)
        cell.value     = h
        cell.font      = FONT_HDR
        cell.fill      = GREY_HDR
        cell.border    = thin()
        cell.alignment = CENTER if c_idx > 2 else LEFT

    # ── RIGHE DATI ────────────────────────────────────────────
    out_row = 10
    for riga in storico:
        ws.row_dimensions[out_row].height = 17
        delta = riga['supplenze'] - riga['permessi'] - riga['civica']

        data_label = riga['data'].strftime('%d/%m/%Y') if riga.get('data') else ''
        periodo    = riga.get('periodo', '')

        vals = [
            (1, data_label,          LEFT,   FONT_NORM, NO_FILL),
            (2, periodo,             LEFT,   FONT_NORM, NO_FILL),
            (3, delta if delta != 0 else 0,
                                     CENTER, FONT_NORM,
                                     GREEN_LIGHT if delta > 0 else (RED_LIGHT if delta < 0 else NO_FILL)),
            (4, riga['supplenze'] or None,
                                     CENTER, FONT_NORM,
                                     GREEN_LIGHT if riga['supplenze'] > 0 else NO_FILL),
            (5, riga['permessi'] or None,
                                     CENTER, FONT_NORM,
                                     RED_LIGHT if riga['permessi'] > 0 else NO_FILL),
            (6, riga['civica'] or None,
                                     CENTER, FONT_NORM,
                                     RED_LIGHT if riga['civica'] > 0 else NO_FILL),
            (7, ("Supplenza" if delta > 0 else ("Recupero" if delta < 0 else "OK")),
                                     LEFT,   FONT_NORM, NO_FILL),
            (8, None,                LEFT,   FONT_NORM, NO_FILL),
        ]

        for c_idx, val, align, fnt, fll in vals:
            cell = ws.cell(out_row, c_idx)
            cell.value     = val
            cell.font      = fnt
            cell.fill      = fll
            cell.border    = thin()
            cell.alignment = align

        out_row += 1

    out_row += 1  # riga vuota

    # ── TOTALI ────────────────────────────────────────────────
    recupero_acc = saldi['permessi'] + saldi['civica']

    totali = [
        # (label, valore, fill_val, bold_label, bianco)
        ("Totale ore di supplenza svolte",         saldi['supplenze'],
         GREEN_DARK if saldi['supplenze'] > 0 else NO_FILL,           True,  True),
        ("Totale ore a recupero accumulate",        recupero_acc,
         RED_DARK   if recupero_acc > 0 else NO_FILL,                 True,  True),
        ("Totale permessi/richieste",               saldi['permessi'] or None,
         RED_DARK   if saldi['permessi'] > 0 else NO_FILL,            False, True),
        ("Totale ore libere Ed. Civica",            saldi['civica'] or None,
         RED_DARK   if saldi['civica'] > 0 else NO_FILL,              False, True),
        ("Ore richieste a pagamento",               saldi['pagamento'] or None,
         YELLOW_PAY if saldi['pagamento'] > 0 else NO_FILL,           True,  False),
        ("Saldo lordo da settimane",                saldo_netto,
         color_saldo(saldo_netto),                                     True,  False),
        ("Saldo effettivo dopo pagamento",          saldo_effettivo,
         color_saldo(saldo_effettivo),                                 True,  False),
        ("Confronto con saldo effettivo Riepilogo", 0,
         NO_FILL,                                                      True,  False),
    ]

    for label, val, val_fill, bold_label, bianco_font in totali:
        ws.row_dimensions[out_row].height = 17

        # Colonna A — etichetta
        cell_a = ws.cell(out_row, 1)
        cell_a.value     = label
        cell_a.font      = FONT_TOT if bold_label else FONT_NORM
        cell_a.border    = thin()
        cell_a.alignment = LEFT

        # Colonna B — valore
        cell_b = ws.cell(out_row, 2)
        cell_b.value     = val
        cell_b.fill      = val_fill

        # Font bianco se sfondo scuro (verde/rosso dark)
        is_dark = val_fill in (GREEN_DARK, RED_DARK)
        if bold_label:
            cell_b.font = FONT_TOT_WH if is_dark else FONT_TOT
        else:
            cell_b.font = Font(name="Calibri", size=12,
                               color="FFFFFF" if is_dark else "000000")

        cell_b.border    = thin()
        cell_b.alignment = CENTER

        out_row += 1

    # Freeze panes
    ws.freeze_panes = "A10"

    return wb
