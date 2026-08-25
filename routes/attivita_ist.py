from flask import Blueprint, render_template, request, redirect, url_for, flash
from models import db
from models.attivita_ist import (AttivitaIst, AttivitaIstPartecipante,
                                  AttivitaIstPresenza, TIPI_ATTIVITA)
from models.materia import Dipartimento, Materia, DocenteMateria
from models.docente import Docente
from models.assenza import Assenza
from datetime import date, datetime
import json
import re

attivita_ist_bp = Blueprint('attivita_ist', __name__)

def _anno_scolastico(d=None):
    d = d or date.today()
    return f'{d.year}-{d.year+1}' if d.month >= 9 else f'{d.year-1}-{d.year}'


def _non_in_servizio_per_data(data_evento):
    """
    Docenti non disponibili per un evento in una data specifica.
    Tre controlli distinti, applicati insieme:

    1. Non in servizio nell'anno scolastico dell'evento (uscita già
       segnalata, AP uscente, aspettativa) — vedi
       routes/docenti.py::_docenti_non_in_servizio.

    2. Non ancora arrivato: anno_scol_inizio successivo all'anno
       dell'evento (es. un docente con anno_scol_inizio 2026-2027 non è
       in servizio per un evento del 2025-2026). _docenti_non_in_servizio
       non lo controlla — quel campo lì non c'entra col suo scopo
       principale (elenco "non attivi" in anagrafica docenti, dove serve
       lo stato più recente, non un anno specifico) — verificato che
       mancasse anche qui dopo la segnalazione di un caso analogo nelle
       prove di recupero di agosto (Sessione 62).

    3. SOLO per eventi di luglio/agosto: il contratto potrebbe essere
       già scaduto pur restando nello stesso anno scolastico. I supplenti
       brevi e i TD "fino a GS" (giorno degli scrutini, CCNL — contratto
       prorogato fino al termine delle operazioni di scrutinio, fine
       giugno) NON sono in servizio a luglio/agosto; solo TI e TD annuale
       lo sono, fino al 31 agosto compreso. Stessa regola già in uso in
       routes/recupero_costanti.py::CONTRATTI_OK per le prove di recupero
       di agosto — riusata qui invece di reinventarla.

       Il contratto usato per questo controllo è quello STORICO
       dell'anno dell'evento (models.docente.DocenteContrattoAnno), se
       registrato — non Docente.tipo_contratto "corrente", che può
       essere già stato aggiornato al contratto del prossimo anno
       mentre si prepara la transizione (es. un TD che entra in ruolo:
       tipo_contratto diventa 'TI' per il nuovo anno, ma per l'agosto
       dell'anno che si sta chiudendo va comunque valutato col
       contratto che aveva allora — segnalato da Roberto, caso Agrò).
       Se non esiste una riga storica per quell'anno, si ricade sul
       campo corrente (comportamento invariato per tutti gli altri).
    """
    from routes.docenti import _docenti_non_in_servizio
    anno_evento = _anno_scolastico(data_evento)
    esclusi = {d.id for d in _docenti_non_in_servizio(anno_evento)}
    esclusi |= {d.id for d in Docente.query.filter(
        Docente.attivo == True,
        Docente.anno_scol_inizio != None,
        Docente.anno_scol_inizio > anno_evento,
    ).all()}

    if data_evento.month in (7, 8):
        from routes.recupero_costanti import CONTRATTI_OK
        from models.docente import DocenteContrattoAnno
        contratti_storici = {
            c.id_docente: c.tipo_contratto for c in
            DocenteContrattoAnno.query.filter_by(anno_scol=anno_evento).all()
        }
        for d in Docente.query.filter_by(attivo=True).all():
            tipo = contratti_storici.get(d.id, d.tipo_contratto)
            if not tipo or tipo not in CONTRATTI_OK:
                esclusi.add(d.id)

    return esclusi



def _preset_partecipanti(attivita):
    """
    Genera lista docenti previsti per l'evento in base al tipo, escludendo
    chi non è in servizio alla data dell'evento — vedi
    _non_in_servizio_per_data() (uscita già segnalata, AP uscente/
    aspettativa, e per luglio/agosto anche il tipo di contratto scaduto:
    supplenti brevi e TD fino a GS). Senza questo controllo, un docente
    non più disponibile continuerebbe a comparire come partecipante
    previsto anche dopo che ha lasciato la scuola o dopo la scadenza del
    contratto.

    Sessione 57: per gli eventi non-scrutinio (bucket A/B, vedi
    models/attivita_ist.py), un docente con un Piano delle Attività
    Personale attivo per l'anno (vedi models/piano_attivita_personale.py
    — cattedra non completa in istituto) NON segue più questo preset
    "per tutti/per classe/per dipartimento": la sua selezione personale
    sostituisce interamente la sua partecipazione prevista a questi
    eventi, in proporzione alle sue ore di contratto. Gli scrutini
    (bucket None) restano invece sempre calcolati come per chiunque
    altro, fuori da questo meccanismo.
    """
    esclusi_ids = _non_in_servizio_per_data(attivita.data)

    docenti_attivi = [d for d in Docente.query.filter_by(attivo=True).all()
                      if d.id not in esclusi_ids]
    tipo = attivita.tipo

    if tipo in ('collegio', 'incontro_famiglie', 'formazione'):
        risultato = [d.id for d in docenti_attivi]

    elif tipo in ('consiglio_classe', 'scrutinio') and attivita.classe:
        from models.orario_docente import OrarioDocente
        ids = {s.id_docente for s in OrarioDocente.query.filter_by(
            classe=attivita.classe).all()}
        risultato = [i for i in ids if i not in esclusi_ids]

    elif tipo in ('dipartimento', 'riunione_materia', 'riunione_referenti') \
            and attivita.id_dipartimento:
        # Anno scolastico dell'EVENTO (dalla sua data), non "oggi": una
        # riunione di dipartimento programmata per marzo 2027 deve
        # guardare le materie del 2026-2027, indipendentemente da
        # quando la si sta creando.
        ids = {dm.id_docente for dm in DocenteMateria.query.join(Materia).filter(
            Materia.id_dipartimento == attivita.id_dipartimento,
            DocenteMateria.anno_scol == _anno_scolastico(attivita.data)
        ).all()}
        risultato = [i for i in ids if i not in esclusi_ids]

    elif tipo == 'glo':
        risultato = []  # solo manuale

    else:
        risultato = [d.id for d in docenti_attivi]

    if attivita.bucket is not None:
        from models.piano_attivita_personale import PianoAttivitaPersonale
        anno_ev = _anno_scolastico(attivita.data)
        piani = {p.id_docente: p for p in
                 PianoAttivitaPersonale.query.filter_by(anno_scol=anno_ev).all()}
        if piani:
            selezionati = {did for did, p in piani.items()
                           if attivita.id in p.ids_attivita_scelte and did not in esclusi_ids}
            risultato = [i for i in risultato if i not in piani] + list(selezionati)

    return risultato


def _iscrivi_docente_a_eventi(id_docente, eventi):
    """
    Nucleo comune a iscrivi_docente_a_obbligatori/_classe/_dipartimento:
    aggiunge id_docente come partecipante (preset=True) a ciascun evento
    della lista, saltando chi non è in servizio a quella data e chi è
    già iscritto. Ritorna il numero di righe aggiunte.
    """
    from models.attivita_ist import AttivitaIstPartecipante

    esclusi_per_data = {}
    aggiunti = 0
    for ev in eventi:
        if ev.data not in esclusi_per_data:
            esclusi_per_data[ev.data] = _non_in_servizio_per_data(ev.data)
        if id_docente in esclusi_per_data[ev.data]:
            continue
        dup = AttivitaIstPartecipante.query.filter_by(
            id_attivita=ev.id, id_docente=id_docente).first()
        if not dup:
            db.session.add(AttivitaIstPartecipante(
                id_attivita=ev.id, id_docente=id_docente, preset=True))
            aggiunti += 1

    if aggiunti:
        db.session.commit()
    return aggiunti


def iscrivi_docente_a_eventi_classe(id_docente, classi_label, anno_scol=None):
    """
    Iscrive un docente agli eventi futuri di Consiglio di classe/
    scrutinio già creati per le classi indicate — chiamata da
    routes/assegnazioni.py (salva/aggiorna_ore/nomina) quando gli si
    assegnano ore su una classe. Usa le Assegnazioni come segnale (non
    l'orario, che _preset_partecipanti usa per un motivo diverso — dare
    la lista più aggiornata possibile vicino alla data dell'evento — ma
    che si stabilizza troppo tardi per essere un buon trigger qui,
    vedi CLAUDE.md): quando Roberto assegna una classe, è già il
    segnale giusto, non serve aspettare l'orario.

    classi_label: iterabile di stringhe nel formato di
    AttivitaIst.classe / OrarioDocente.classe (es. "3A LLI" —
    AssegnazioneClasse.label_classe è già in questo formato).

    anno_scol: l'anno scolastico dell'assegnazione che ha scatenato la
    chiamata (es. "2026-2027" per un'assegnazione in preparazione del
    nuovo anno). Un'etichetta classe come "3A LLI" è la STESSA sia per
    la classe uscente sia per quella entrante, e "eventi futuri"
    (data >= oggi) da sola non basta a distinguerle — segnalato da
    Roberto: assegnando un docente per il 2026-2027 durante la
    preparazione, comparirebbe come partecipante previsto anche negli
    scrutini del 31/08/2026, che appartengono ancora al 2025-2026.
    Se valorizzato, iscrive solo gli eventi il cui anno scolastico
    (dalla data dell'evento) coincide con quello dell'assegnazione.
    """
    if not classi_label:
        return 0
    from datetime import date
    from models.attivita_ist import AttivitaIst

    oggi = date.today()
    eventi = AttivitaIst.query.filter(
        AttivitaIst.data >= oggi,
        AttivitaIst.tipo.in_(('consiglio_classe', 'scrutinio')),
        AttivitaIst.classe.in_(list(classi_label)),
    ).all()
    if anno_scol:
        eventi = [ev for ev in eventi if _anno_scolastico(ev.data) == anno_scol]
    return _iscrivi_docente_a_eventi(id_docente, eventi)


def iscrivi_docente_a_eventi_dipartimento(id_docente, id_dipartimento):
    """
    Iscrive un docente agli eventi futuri di dipartimento/riunione
    materia/riunione referenti già creati per il dipartimento indicato
    — chiamata da routes/assegnazioni.py::_sync_docente_materie() dopo
    aver sincronizzato una nuova DocenteMateria (stesso segnale, non
    serve aspettare altro).

    GLO resta fuori anche da questo: il preset del GLO è sempre vuoto
    (solo manuale, vedi _preset_partecipanti) perché la partecipazione
    dipende dall'alunno seguito, non dalla classe o dal dipartimento —
    non esiste un dato di assegnazione da cui derivarla automaticamente.
    """
    if not id_dipartimento:
        return 0
    from datetime import date
    from models.attivita_ist import AttivitaIst

    oggi = date.today()
    eventi = AttivitaIst.query.filter(
        AttivitaIst.data >= oggi,
        AttivitaIst.tipo.in_(('dipartimento', 'riunione_materia', 'riunione_referenti')),
        AttivitaIst.id_dipartimento == id_dipartimento,
    ).all()
    return _iscrivi_docente_a_eventi(id_docente, eventi)


def iscrivi_docente_a_obbligatori(docente):
    """
    Iscrive un docente appena attivato (nuovo o riattivato, vedi
    routes/docenti.py::nuovo/riattiva) a tutti gli eventi istituzionali
    futuri "per tutti i docenti" già creati — senza questo, un evento
    creato PRIMA che il docente esistesse in anagrafica non lo
    includerebbe mai: il preset (_preset_partecipanti) viene calcolato
    solo alla creazione/modifica dell'evento, non ricalcolato quando
    cambia l'anagrafica.

    Copre solo i tipi il cui preset è "tutti i docenti attivi" senza
    scelta (collegio, incontro_famiglie, altro) e i corsi di Formazione
    con obbligatorio_tutti=True (models/formazione.py — i corsi
    volontari restano esclusi, l'iscrizione lì è sempre una scelta).
    Eventi scoped su classe/dipartimento (Consigli di classe,
    dipartimenti, riunioni materia, GLO) restano fuori di proposito:
    dipendono da orario/assegnazioni che un docente appena creato non
    ha ancora — non c'è nulla di corretto da preimpostare qui.

    Ritorna il numero di iscrizioni aggiunte (0 se il docente è escluso
    per la data, o se non ci sono eventi futuri di questi tipi).
    """
    from datetime import date
    from models.attivita_ist import AttivitaIst
    from models.formazione import CorsoFormazione

    oggi = date.today()
    eventi = AttivitaIst.query.filter(
        AttivitaIst.data >= oggi,
        AttivitaIst.tipo.in_(('collegio', 'incontro_famiglie', 'altro', 'formazione')),
    ).all()
    if not eventi:
        return 0

    corsi_volontari_evento_ids = {
        c.id_attivita for c in CorsoFormazione.query.filter_by(obbligatorio_tutti=False).all()}
    eventi = [ev for ev in eventi
              if not (ev.tipo == 'formazione' and ev.id in corsi_volontari_evento_ids)]

    return _iscrivi_docente_a_eventi(docente.id, eventi)


def _auto_presenze(attivita):
    """Crea record presenze per i partecipanti, pre-compilando assenti noti."""
    assenze_giorno = {a.id_docente: a for a in
                      Assenza.query.filter_by(data=attivita.data).all()}
    for part in attivita.partecipanti:
        dup = AttivitaIstPresenza.query.filter_by(
            id_attivita=attivita.id, id_docente=part.id_docente).first()
        if dup:
            continue
        assenza = assenze_giorno.get(part.id_docente)
        stato = 'assente' if assenza else 'presente'
        db.session.add(AttivitaIstPresenza(
            id_attivita          = attivita.id,
            id_docente           = part.id_docente,
            stato                = stato,
            id_assenza_collegata = assenza.id if assenza else None,
        ))


# ── LISTA ────────────────────────────────────────────────────────────────────

@attivita_ist_bp.route('/attivita-ist')
def lista():
    from datetime import date
    oggi = date.today()
    anno = request.args.get('anno', _anno_scolastico())
    anno_ini = date(int(anno[:4]), 9, 1)
    anno_fin = date(int(anno[:4]) + 1, 8, 31)
    tipo_f   = request.args.get('tipo', '')
    mese_f   = request.args.get('mese', '')

    q = AttivitaIst.query.filter(
        AttivitaIst.data >= anno_ini,
        AttivitaIst.data <= anno_fin,
    )
    if tipo_f:
        q = q.filter_by(tipo=tipo_f)
    if mese_f:
        q = q.filter(db.func.strftime('%m', AttivitaIst.data) == mese_f.zfill(2))
    eventi = q.order_by(AttivitaIst.data, AttivitaIst.ora_inizio).all()

    # Separa le attività già svolte (data passata) da quelle future/odierne:
    # le prime finiscono in una tabella a parte, in fondo alla pagina,
    # più recenti per prime.
    eventi_futuri  = [e for e in eventi if e.data >= oggi]
    eventi_passati = [e for e in eventi if e.data < oggi]
    eventi_passati.reverse()

    dipartimenti = Dipartimento.query.order_by(Dipartimento.ordine).all()
    return render_template('attivita_ist/lista.html',
        eventi_futuri=eventi_futuri, eventi_passati=eventi_passati,
        oggi=oggi, anno=anno,
        tipi=TIPI_ATTIVITA, tipo_f=tipo_f, mese_f=mese_f,
        dipartimenti=dipartimenti,
    )


def _righe_piano_annuale(anno):
    """
    Costruisce la struttura dati condivisa fra la vista a schermo e
    l'export PDF del Piano Annuale — stesso identico modello del
    foglio fornito da Roberto (BOZZA_PIANO_ATTIVITA_2026_27.xlsx):
    colonne Attività | Indirizzo | Classe | Inizio | Fine | Ore |
    Categoria, righe raggruppate per giorno dentro i fogli mensili, con
    le sospensioni/vacanze e il termine lezioni segnati come nel
    calendario colorato del foglio "Introduzione" — non solo gli
    eventi, anche i giorni di non-lezione compaiono nel piano.

    Ritorna (mesi, anni_disponibili, n_eventi) dove mesi è
    [(etichetta_mese, [(data, tipo_giorno, contenuto), ...])] — righe
    già ordinate cronologicamente, tipo_giorno è 'eventi' o
    'sospensione' o 'termine_lezioni'.
    """
    from config_anno import intervallo_anno_scolastico
    from config_calendario import get_data_fine_lezioni
    from models.sospensione import SospensioneDidattica
    from modules.prospetto_supplenze import MESI_IT

    anni_disponibili = sorted(
        {_anno_scolastico(e.data) for e in AttivitaIst.query.all()}, reverse=True)
    if anno not in anni_disponibili:
        anni_disponibili.insert(0, anno)

    ini, fine = intervallo_anno_scolastico(anno)
    eventi = AttivitaIst.query.filter(
        AttivitaIst.data >= ini, AttivitaIst.data <= fine
    ).order_by(AttivitaIst.data, AttivitaIst.ora_inizio).all()

    # Arricchisce ogni evento con le colonne esatte del foglio: Indirizzo
    # e Classe separati (il modello li tiene insieme in un'unica label
    # "3A LLI"), Categoria = etichetta del tipo.
    for ev in eventi:
        m = re.match(r'(\d+[AB]?)\s+(.+)', ev.classe) if ev.classe else None
        ev.col_classe = m.group(1) if m else ''
        ev.col_indirizzo = m.group(2) if m else ''
        ev.col_categoria = ev.tipo_label

    marcatori = []  # (data, tipo, contenuto)
    for s in SospensioneDidattica.query.filter(
            SospensioneDidattica.data_fine >= ini,
            SospensioneDidattica.data_inizio <= fine).order_by(SospensioneDidattica.data_inizio).all():
        marcatori.append((max(s.data_inizio, ini), 'sospensione', s))

    termine = get_data_fine_lezioni(anno)
    if termine and ini <= termine <= fine:
        marcatori.append((termine, 'termine_lezioni', None))

    per_giorno = {}  # data -> {'eventi': [...], 'marcatori': [...]}
    for ev in eventi:
        per_giorno.setdefault(ev.data, {'eventi': [], 'marcatori': []})['eventi'].append(ev)
    for data_m, tipo_m, contenuto in marcatori:
        per_giorno.setdefault(data_m, {'eventi': [], 'marcatori': []})['marcatori'].append(
            (tipo_m, contenuto))

    mesi = []
    chiave_mese_corrente = None
    righe_mese_corrente = None
    for data_g in sorted(per_giorno.keys()):
        chiave = (data_g.year, data_g.month)
        if chiave != chiave_mese_corrente:
            chiave_mese_corrente = chiave
            righe_mese_corrente = []
            mesi.append((f'{MESI_IT[data_g.month]} {data_g.year}', righe_mese_corrente))
        giorno = per_giorno[data_g]
        for tipo_m, contenuto in giorno['marcatori']:
            righe_mese_corrente.append((data_g, tipo_m, contenuto))
        if giorno['eventi']:
            righe_mese_corrente.append((data_g, 'eventi', giorno['eventi']))

    return mesi, anni_disponibili, len(eventi)


# ── VISTA PIANO ANNUALE (Fase 2 del Piano Annuale delle Attività) ────────────

@attivita_ist_bp.route('/attivita-ist/piano-annuale')
def piano_annuale():
    """
    Vista mensile di AttivitaIst raggruppata per giorno, con intestazioni
    di sezione — stesso modello del foglio fornito da Roberto
    (BOZZA_PIANO_ATTIVITA_2026_27.xlsx): colonne Attività | Indirizzo |
    Classe | Inizio | Fine | Ore | Categoria, con sospensioni/vacanze e
    termine lezioni segnati come nel calendario colorato originale —
    non solo gli eventi (Fase 2 del project plan, poi Fase 4
    "pubblicazione" per l'export PDF, vedi piano_annuale_pdf()).
    """
    from routes.impostazione_anno import _anno_default_piano
    anno = request.args.get('anno', _anno_default_piano())
    mesi, anni_disponibili, n_eventi = _righe_piano_annuale(anno)

    return render_template('attivita_ist/piano_annuale.html',
        mesi=mesi, anno=anno, anni_disponibili=anni_disponibili,
        tipi=TIPI_ATTIVITA, oggi=date.today(), n_eventi=n_eventi)


@attivita_ist_bp.route('/attivita-ist/piano-annuale/pdf')
def piano_annuale_pdf():
    """
    Export PDF dello stesso modello mostrato a schermo — Fase 4
    (pubblicazione) del project plan: solo PDF scaricabile, confermato
    con Roberto, niente pagina pubblica in stile "display".
    """
    import io
    from flask import send_file
    from routes.impostazione_anno import _anno_default_piano
    from modules.pdf_fonts import contesto_open_sans

    anno = request.args.get('anno', _anno_default_piano())
    mesi, _anni_disponibili, n_eventi = _righe_piano_annuale(anno)

    html_content = render_template('attivita_ist/piano_annuale_print.html',
        mesi=mesi, anno=anno, tipi=TIPI_ATTIVITA, oggi=date.today(),
        n_eventi=n_eventi, **contesto_open_sans())

    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html_content).write_pdf()
        return send_file(
            io.BytesIO(pdf_bytes), mimetype='application/pdf', as_attachment=True,
            download_name=f'piano_annuale_{anno}_{date.today().isoformat()}.pdf')
    except (ImportError, OSError):
        # WeasyPrint assente (ImportError) o le sue librerie di sistema
        # native mancano (OSError da cffi — caso tipico in sandbox
        # Linux senza pango/cairo, non solo "non pip-installato") —
        # fallback HTML in entrambi i casi, non è un bug fuori dal Mac.
        return html_content


@attivita_ist_bp.route('/attivita-ist/piano-annuale/xlsx')
def piano_annuale_xlsx():
    """
    Export Excel dello stesso modello, questa volta replicando lo stile
    esatto del foglio originale di Roberto
    (BOZZA_PIANO_ATTIVITA_2026_27_CORRETTO.xlsx): titoli e intestazioni
    colorate, colonna mese verticale, banner di giorno/sospensione/
    termine lezioni, più il foglio "Riepilogo ore" per classe — vedi
    modules/export_piano_xlsx.py per i dettagli e per cosa NON viene
    replicato (testo libero/agenda, non un dato strutturato qui).
    """
    import io
    from flask import send_file
    from routes.impostazione_anno import _anno_default_piano
    from config_anno import intervallo_anno_scolastico
    from models.assegnazione import AssegnazioneDocente, AssegnazioneClasse
    from modules.export_piano_xlsx import genera_xlsx_piano_annuale

    anno = request.args.get('anno', _anno_default_piano())
    mesi, _anni_disponibili, n_eventi = _righe_piano_annuale(anno)

    ini, fine = intervallo_anno_scolastico(anno)

    # Elenco ufficiale delle classi dell'anno da Assegnazioni (non solo
    # dagli eventi già calendarizzati): così una classe senza ancora
    # nessun CdC/scrutinio compare comunque con 0 ore, non sparisce.
    righe_classi = (db.session.query(AssegnazioneClasse.indirizzo,
                                      AssegnazioneClasse.anno_corso,
                                      AssegnazioneClasse.sezione)
                     .join(AssegnazioneDocente,
                           AssegnazioneDocente.id == AssegnazioneClasse.id_assegnazione)
                     .filter(AssegnazioneDocente.anno_scol == anno)
                     .distinct().all())
    classi_ore = {}
    for indirizzo, anno_corso, sezione in righe_classi:
        classi_ore[(indirizzo, f'{anno_corso}{sezione}')] = {'cdc': 0.0, 'scrutinio': 0.0}

    eventi_classe = AttivitaIst.query.filter(
        AttivitaIst.data >= ini, AttivitaIst.data <= fine,
        AttivitaIst.tipo.in_(('consiglio_classe', 'scrutinio')),
        AttivitaIst.classe.isnot(None)).all()
    for ev in eventi_classe:
        m = re.match(r'(\d+[AB]?)\s+(.+)', ev.classe)
        if not m:
            continue
        classe, indirizzo = m.group(1), m.group(2)
        acc = classi_ore.setdefault((indirizzo, classe), {'cdc': 0.0, 'scrutinio': 0.0})
        chiave = 'cdc' if ev.tipo == 'consiglio_classe' else 'scrutinio'
        acc[chiave] += ev.durata_ore

    classi_ore_lista = [
        (indirizzo, classe, v['cdc'], v['scrutinio'])
        for (indirizzo, classe), v in sorted(classi_ore.items())
    ]

    wb = genera_xlsx_piano_annuale(mesi, classi_ore_lista, anno)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name=f'piano_annuale_{anno}_{date.today().isoformat()}.xlsx')


# ── RIEPILOGO ORE (Fase 2) ────────────────────────────────────────────────────

@attivita_ist_bp.route('/attivita-ist/riepilogo-ore')
def riepilogo_ore():
    """
    Due report, stesso ruolo del foglio "Riepilogo ore" del piano
    cartaceo (Fase 2 del project plan):
    - ore per classe (Consigli + Scrutini), da AttivitaIst.durata_ore;
    - ore per docente su Collegio/Consigli/Formazione confrontate col
      bucket A/B (40h), per far emergere un'eccedenza prima che si
      presenti — riusa quota_ore_bucket() già scritta per il Piano
      Attività Personale (Sessione 57), non un calcolo parallelo.
    """
    from config_anno import intervallo_anno_scolastico
    from routes.impostazione_anno import _anno_default_piano
    from models.attivita_ist import BUCKET_A, BUCKET_B
    from models.piano_attivita_personale import quota_ore_bucket

    anno = request.args.get('anno', _anno_default_piano())
    anni_disponibili = sorted(
        {_anno_scolastico(e.data) for e in AttivitaIst.query.all()}, reverse=True)
    if anno not in anni_disponibili:
        anni_disponibili.insert(0, anno)

    ini, fine = intervallo_anno_scolastico(anno)

    # Ore per classe (Consigli + Scrutini)
    eventi_classe = AttivitaIst.query.filter(
        AttivitaIst.data >= ini, AttivitaIst.data <= fine,
        AttivitaIst.tipo.in_(('consiglio_classe', 'scrutinio')),
        AttivitaIst.classe.isnot(None),
    ).all()
    ore_per_classe = {}
    for ev in eventi_classe:
        ore_per_classe[ev.classe] = ore_per_classe.get(ev.classe, 0) + ev.durata_ore
    ore_per_classe = sorted(ore_per_classe.items())

    # Ore per docente vs bucket A/B
    righe = (db.session.query(AttivitaIst, AttivitaIstPartecipante.id_docente)
             .join(AttivitaIstPartecipante, AttivitaIstPartecipante.id_attivita == AttivitaIst.id)
             .filter(AttivitaIst.data >= ini, AttivitaIst.data <= fine).all())
    ore_docente = {}  # {id_docente: [ore_a, ore_b]}
    for ev, id_doc in righe:
        if ev.bucket not in (BUCKET_A, BUCKET_B):
            continue
        acc = ore_docente.setdefault(id_doc, [0.0, 0.0])
        acc[0 if ev.bucket == BUCKET_A else 1] += ev.durata_ore

    # Esclude chi non è (più) in servizio nell'anno mostrato: le righe
    # AttivitaIstPartecipante non si aggiornano da sole quando un
    # docente viene segnalato uscente DOPO essere stato convocato (es.
    # un Collegio di inizio anno creato prima che l'uscita fosse nota) —
    # senza questo filtro il riepilogo mostrerebbe ore per docenti già
    # non più in servizio per l'anno selezionato.
    esclusi_rif = _non_in_servizio_per_data(ini)
    docenti = {d.id: d for d in Docente.query.filter(
        Docente.id.in_(ore_docente.keys())).all()
        if d.id not in esclusi_rif}
    riepilogo_docenti = []
    for id_doc, (ore_a, ore_b) in ore_docente.items():
        doc = docenti.get(id_doc)
        if not doc:
            continue
        quota_a, quota_b = quota_ore_bucket(doc, anno)
        riepilogo_docenti.append(dict(
            docente=doc, is_placeholder=False,
            ore_a=round(ore_a, 1), ore_b=round(ore_b, 1),
            quota_a=quota_a, quota_b=quota_b,
            eccede_a=ore_a > quota_a, eccede_b=ore_b > quota_b,
        ))

    # Placeholder di Assegnazioni ancora da nominare (supplente non
    # ancora individuato): mostrano le ore di bucket B (Consigli di
    # classe, non gli scrutini — fuori bucket) già prevedibili dalle
    # classi assegnate, così da vedere il carico atteso prima ancora
    # che arrivi il titolare. Spariscono da sole non appena la
    # AssegnazioneDocente viene nominata (routes/assegnazioni.py::
    # nomina) — a quel punto le ore ricompaiono sotto il nome del
    # docente reale, tramite gli AttivitaIstPartecipante creati da
    # iscrivi_docente_a_eventi_classe().
    from models.assegnazione import AssegnazioneDocente
    from models.formazione import CorsoFormazione
    ore_classe_bucket_b = {}
    for ev in eventi_classe:
        if ev.tipo != 'consiglio_classe':
            continue
        ore_classe_bucket_b[ev.classe] = ore_classe_bucket_b.get(ev.classe, 0) + ev.durata_ore

    # Ore di Formazione obbligatoria (bucket A): valgono per chiunque
    # sia in servizio a prescindere da chi sarà — anche un placeholder
    # ancora da nominare le dovrà, quindi contano già ora (richiesta di
    # Roberto). Stesso valore per tutti i placeholder, calcolato una
    # sola volta fuori dal loro ciclo.
    corsi_obbligatori_ids = {c.id_attivita for c in CorsoFormazione.query.filter_by(
        anno_scol=anno, obbligatorio_tutti=True).all()}
    ore_a_formazione_obbl = round(sum(
        ev.durata_ore for ev in AttivitaIst.query.filter(
            AttivitaIst.id.in_(corsi_obbligatori_ids)).all()
    ), 1) if corsi_obbligatori_ids else 0.0

    placeholder_asgn = AssegnazioneDocente.query.filter(
        AssegnazioneDocente.anno_scol == anno,
        AssegnazioneDocente.id_docente.is_(None),
        AssegnazioneDocente.nome_placeholder.isnot(None),
    ).all()
    for asgn in placeholder_asgn:
        classi_label = {ac.label_classe for ac in asgn.classi}
        ore_b_ph = round(sum(ore_classe_bucket_b.get(lbl, 0) for lbl in classi_label), 1)
        if ore_b_ph <= 0 and ore_a_formazione_obbl <= 0:
            continue
        codice_cc = asgn.classe_concorso.codice if asgn.classe_concorso else '?'
        riepilogo_docenti.append(dict(
            docente=None, etichetta=f'{asgn.nome_placeholder} — {codice_cc}',
            is_placeholder=True,
            ore_a=ore_a_formazione_obbl, ore_b=ore_b_ph, quota_a=None, quota_b=None,
            eccede_a=False, eccede_b=False,
        ))

    riepilogo_docenti.sort(key=lambda r: (r['docente'].cognome, r['docente'].nome)
                            if r['docente'] else (r['etichetta'], ''))

    return render_template('attivita_ist/riepilogo_ore.html',
        anno=anno, anni_disponibili=anni_disponibili,
        ore_per_classe=ore_per_classe, riepilogo_docenti=riepilogo_docenti)


# ── NUOVO / MODIFICA ─────────────────────────────────────────────────────────

@attivita_ist_bp.route('/attivita-ist/nuova', methods=['GET', 'POST'])
@attivita_ist_bp.route('/attivita-ist/<int:id>/modifica', methods=['GET', 'POST'])
def form(id=None):
    evento = AttivitaIst.query.get_or_404(id) if id else None
    docenti = Docente.query.filter_by(attivo=True).order_by(Docente.cognome).all()
    dipartimenti = Dipartimento.query.order_by(Dipartimento.ordine).all()
    classi_db = sorted({s.classe for s in
                        __import__('models.orario_docente', fromlist=['OrarioDocente'])
                        .OrarioDocente.query.all()
                        if s.classe and s.classe not in ('POTENZIAMENTO','---','-x-','')})

    if request.method == 'POST':
        tipo        = request.form['tipo']
        titolo      = request.form['titolo'].strip()
        data_s      = request.form['data']
        ora_ini     = request.form.get('ora_inizio', '').strip() or None
        ora_fin     = request.form.get('ora_fine', '').strip() or None
        note        = request.form.get('note', '').strip() or None
        classe      = request.form.get('classe', '').strip() or None
        id_dip      = request.form.get('id_dipartimento') or None
        doc_ids_raw = request.form.getlist('docenti_ids')
        doc_ids     = [int(x) for x in doc_ids_raw if x.isdigit()]

        if evento:
            # Modifica
            evento.tipo = tipo; evento.titolo = titolo
            evento.data = date.fromisoformat(data_s)
            evento.ora_inizio = ora_ini; evento.ora_fine = ora_fin
            evento.note = note; evento.classe = classe
            evento.id_dipartimento = int(id_dip) if id_dip else None
            # Ricrea partecipanti
            AttivitaIstPartecipante.query.filter_by(id_attivita=evento.id).delete()
        else:
            evento = AttivitaIst(
                tipo=tipo, titolo=titolo,
                data=date.fromisoformat(data_s),
                ora_inizio=ora_ini, ora_fine=ora_fin,
                note=note, classe=classe,
                id_dipartimento=int(id_dip) if id_dip else None,
                origine='manuale',
            )
            db.session.add(evento)
        db.session.flush()

        # Partecipanti: usa preset se nessuna selezione manuale
        if not doc_ids:
            doc_ids = _preset_partecipanti(evento)
        for did in doc_ids:
            db.session.add(AttivitaIstPartecipante(
                id_attivita=evento.id, id_docente=did, preset=True))

        db.session.commit()
        flash(f'Evento {"aggiornato" if id else "registrato"}: {titolo}', 'success')
        return redirect(url_for('attivita_ist.lista'))

    # Pre-selezione docenti per preset
    preset_ids = _preset_partecipanti(evento) if evento else []
    docenti_selezionati = {p.id_docente for p in evento.partecipanti} if evento else set()

    # Esclude dall'elenco selezionabile chi non è in servizio alla data
    # dell'evento (uscita già segnalata per quell'anno, non ancora
    # arrivato...), a meno che non sia già preset/selezionato — per non
    # farlo mai sparire silenziosamente da una selezione esistente, solo
    # evitare di proporne di nuovi non più in servizio (vedi anche
    # routes/formazione.py::form(), stesso principio).
    data_rif = evento.data if evento else date.fromisoformat(
        request.args.get('data', date.today().isoformat()))
    esclusi_rif = _non_in_servizio_per_data(data_rif)
    gia_coinvolti = set(preset_ids) | docenti_selezionati
    docenti = [d for d in docenti if d.id not in esclusi_rif or d.id in gia_coinvolti]

    return render_template('attivita_ist/form.html',
        evento=evento, docenti=docenti, dipartimenti=dipartimenti,
        classi=classi_db, tipi=TIPI_ATTIVITA,
        preset_ids=preset_ids,
        docenti_selezionati=docenti_selezionati,
        data_sel=(evento.data.isoformat() if evento
                  else request.args.get('data', date.today().isoformat())),
        next_url=request.args.get('next', '').strip(),
    )


# ── ELIMINA ──────────────────────────────────────────────────────────────────

@attivita_ist_bp.route('/attivita-ist/<int:id>/elimina', methods=['POST'])
def elimina(id):
    e = AttivitaIst.query.get_or_404(id)

    # AttivitaIstPartecipante/AttivitaIstPresenza sono in cascade
    # 'delete-orphan' sul modello, spariscono da sole con l'evento. Le
    # sostituzioni scrutinio no (SostituzioneScrutinio non ha una
    # relazione con cascade dal lato AttivitaIst) — senza ripulirle qui
    # a mano restano righe orfane che puntano a un id_attivita ormai
    # inesistente. Registra la lapide per ognuna (come le altre
    # cancellazioni di questa tabella nel sync automatico, vedi
    # modules/auto_sync.py) prima di cancellarle, altrimenti una
    # postazione non ancora allineata potrebbe farle risuscitare.
    from models.sostituzione_scrutinio import SostituzioneScrutinio
    from modules.auto_sync import registra_eliminazione
    sostituzioni = SostituzioneScrutinio.query.filter_by(id_attivita=id).all()
    for s in sostituzioni:
        registra_eliminazione('sostituzioni_scrutinio',
                               {'id_attivita': s.id_attivita, 'id_assente': s.id_assente})
        db.session.delete(s)

    db.session.delete(e)
    db.session.commit()
    flash('Evento eliminato.', 'warning')

    # Torna alla pagina da cui si è arrivati (es. Piano delle attività),
    # non sempre a "Attività istituzionali" — Roberto: eliminando da lì
    # perdeva il contesto (mese/filtri) su cui stava lavorando. Il
    # chiamante passa la propria URL come campo nascosto "next"; per
    # sicurezza si accetta solo un percorso relativo di questa stessa
    # app (mai un URL assoluto/esterno).
    next_url = request.form.get('next', '').strip()
    if next_url.startswith('/') and not next_url.startswith('//'):
        return redirect(next_url)
    return redirect(url_for('attivita_ist.lista'))


# ── PRESENZE ─────────────────────────────────────────────────────────────────

@attivita_ist_bp.route('/attivita-ist/<int:id>/presenze', methods=['GET', 'POST'])
def presenze(id):
    evento = AttivitaIst.query.get_or_404(id)

    # Auto-genera presenze mancanti (anche se alcune già esistono)
    presenti_ids = {p.id_docente for p in evento.presenze}
    mancanti = [p for p in evento.partecipanti if p.id_docente not in presenti_ids]
    if mancanti:
        for part in mancanti:
            db.session.add(AttivitaIstPresenza(
                id_attivita=evento.id,
                id_docente=part.id_docente,
                stato='presente',
            ))
        db.session.commit()

    if request.method == 'POST':
        for p in evento.presenze:
            stato    = request.form.get(f'stato_{p.id_docente}', 'presente')
            nota     = request.form.get(f'nota_{p.id_docente}', '').strip() or None
            ora_ini  = request.form.get(f'ora_ini_{p.id_docente}', '').strip() or None
            ora_fin  = request.form.get(f'ora_fin_{p.id_docente}', '').strip() or None
            p.stato          = stato
            p.note           = nota
            # Ore parziali: salva solo se diverse dall'intero evento
            p.ora_inizio_eff = ora_ini if (ora_ini and ora_ini != evento.ora_inizio) else None
            p.ora_fine_eff   = ora_fin if (ora_fin and ora_fin != evento.ora_fine)   else None
        db.session.commit()
        flash('Presenze salvate.', 'success')
        return redirect(url_for('attivita_ist.presenze', id=id))

    assenze_giorno = {a.id_docente: a for a in
                      Assenza.query.filter_by(data=evento.data).all()}

    # Indisponibilità dichiarate per la stessa data (impegni già noti:
    # colloqui, uscite, gare, formazione, ecc.) — non escludono di per sé
    # dalla convocazione (a differenza delle assenze), ma vanno segnalate
    # perché indicano un possibile conflitto da verificare.
    from models.indisponibilita import Indisponibilita
    indisponibilita_giorno = {}
    for i in Indisponibilita.query.filter_by(data=evento.data).all():
        indisponibilita_giorno.setdefault(i.id_docente, []).append(i)

    presenze_map  = {p.id_docente: p for p in evento.presenze}

    # Docenti convocati (già presenti in elenco) ma non più in servizio alla
    # data dell'evento: segnalati con un avviso per ricordare di nominare un
    # sostituto (per gli scrutini si può usare la funzione "Sostituzioni").
    # Include anche, per gli eventi di luglio/agosto, chi ha un contratto già
    # scaduto (supplenti brevi, TD fino a GS) — vedi _non_in_servizio_per_data.
    non_in_servizio_ids = _non_in_servizio_per_data(evento.data)

    # Stesso filtro applicato anche alla tendina "+ Aggiungi docente": non
    # ha senso proporre di aggiungere qualcuno già non in servizio a quella
    # data (chi è già stato convocato prima di uscire resta comunque
    # visibile in elenco, solo segnalato — vedi non_in_servizio_ids sopra).
    docenti_extra = [d for d in Docente.query.filter_by(attivo=True).all()
                     if d.id not in non_in_servizio_ids]
    docenti_extra.sort(key=lambda d: d.cognome)

    # Navigazione al precedente/successivo: scorre TUTTI gli eventi in
    # ordine cronologico (stesso ordinamento della Lista), anche su
    # giorni diversi — richiesto da Roberto per gestire le presenze di
    # una classe dopo l'altra (es. i 27 scrutini di un giorno) senza
    # tornare ogni volta alla lista completa.
    ids_ordinati = [r[0] for r in db.session.query(AttivitaIst.id)
                     .order_by(AttivitaIst.data, AttivitaIst.ora_inizio, AttivitaIst.id).all()]
    idx = ids_ordinati.index(evento.id)
    evento_prec = db.session.get(AttivitaIst, ids_ordinati[idx - 1]) if idx > 0 else None
    evento_succ = db.session.get(AttivitaIst, ids_ordinati[idx + 1]) if idx < len(ids_ordinati) - 1 else None

    # Card "Sostituti individuati" (solo scrutini): quanti dei docenti
    # da sostituire (assenti/giustificati o non in servizio — stessa
    # regola di sostituzione_scrutinio()::presenze_assenti) hanno già
    # una SostituzioneScrutinio con sostituto nominato — richiesto da
    # Roberto per vedere a colpo d'occhio quanto resta da fare senza
    # aprire la pagina Sostituzioni.
    n_da_sostituire = n_sostituti_individuati = 0
    if evento.tipo == 'scrutinio':
        from models.sostituzione_scrutinio import SostituzioneScrutinio
        da_sostituire_ids = {p.id_docente for p in evento.presenze
                              if p.stato in ('assente', 'giustificato')
                              or p.id_docente in non_in_servizio_ids}
        n_da_sostituire = len(da_sostituire_ids)
        if da_sostituire_ids:
            n_sostituti_individuati = SostituzioneScrutinio.query.filter(
                SostituzioneScrutinio.id_attivita == evento.id,
                SostituzioneScrutinio.id_assente.in_(da_sostituire_ids),
                SostituzioneScrutinio.id_sostituto.isnot(None),
            ).count()

    return render_template('attivita_ist/presenze.html',
        evento=evento, presenze_map=presenze_map,
        assenze_giorno=assenze_giorno,
        indisponibilita_giorno=indisponibilita_giorno,
        docenti_extra=docenti_extra,
        non_in_servizio_ids=non_in_servizio_ids,
        evento_prec=evento_prec, evento_succ=evento_succ,
        n_da_sostituire=n_da_sostituire,
        n_sostituti_individuati=n_sostituti_individuati,
    )


@attivita_ist_bp.route('/attivita-ist/<int:id>/presenze/aggiungi', methods=['POST'])
def aggiungi_partecipante(id):
    evento = AttivitaIst.query.get_or_404(id)
    did = int(request.form['id_docente'])
    dup = AttivitaIstPartecipante.query.filter_by(
        id_attivita=id, id_docente=did).first()
    if not dup:
        db.session.add(AttivitaIstPartecipante(
            id_attivita=id, id_docente=did, preset=False))
        db.session.flush()
    # Aggiunge presenza se non c'è
    dup_p = AttivitaIstPresenza.query.filter_by(
        id_attivita=id, id_docente=did).first()
    if not dup_p:
        assenza = Assenza.query.filter_by(
            id_docente=did, data=evento.data).first()
        db.session.add(AttivitaIstPresenza(
            id_attivita=id, id_docente=did,
            stato='assente' if assenza else 'presente',
            id_assenza_collegata=assenza.id if assenza else None,
        ))
    db.session.commit()
    flash('Docente aggiunto.', 'success')
    return redirect(url_for('attivita_ist.presenze', id=id))


# ── IMPORT PIANO ANNUALE ─────────────────────────────────────────────────────

@attivita_ist_bp.route('/attivita-ist/import-piano', methods=['GET', 'POST'])
def import_piano():
    if request.method == 'POST':
        n = _import_piano_2025_26()
        flash(f'Importati {n} eventi dal Piano delle Attività 2025/26.', 'success')
        return redirect(url_for('attivita_ist.lista'))
    # Controlla se già importato
    gia = AttivitaIst.query.filter_by(origine='import_piano').count()
    return render_template('attivita_ist/import_piano.html', gia_importati=gia)


def _import_piano_2025_26():
    """
    Importa gli eventi rimanenti del Piano 2025/26 (da giugno 2026 in poi,
    dato che i mesi precedenti sono già passati).
    Solo eventi non ancora presenti (controllo per tipo+data+titolo).
    """
    EVENTI = [
        # ── GIUGNO 2026 ──────────────────────────────────────────────────
        # Scrutini finali 6/6 — ordine da piano attività
        # LL: col6 → LLI
        {'tipo':'scrutinio','titolo':'Scrutinio finale LLI V A','data':'2026-06-06','ora_ini':'10:30','ora_fin':'11:15','classe':'5A LLI'},
        {'tipo':'scrutinio','titolo':'Scrutinio finale LLI V B','data':'2026-06-06','ora_ini':'11:15','ora_fin':'12:00','classe':'5B LLI'},
        # LSC: col4 → 12:45
        {'tipo':'scrutinio','titolo':'Scrutinio finale LSC V A','data':'2026-06-06','ora_ini':'12:45','ora_fin':'13:15','classe':'5A LSC'},
        # LSP: col7 → 14:00
        {'tipo':'scrutinio','titolo':'Scrutinio finale LSP V A','data':'2026-06-06','ora_ini':'14:00','ora_fin':'14:45','classe':'5A LSP'},
        # CAT: col3 → 14:45
        {'tipo':'scrutinio','titolo':'Scrutinio finale CAT V A','data':'2026-06-06','ora_ini':'14:45','ora_fin':'15:30','classe':'5A CAT'},
        # AFM/RIM: col2 → 15:30
        {'tipo':'scrutinio','titolo':'Scrutinio finale RIM V A','data':'2026-06-06','ora_ini':'15:30','ora_fin':'16:15','classe':'5A RIM'},
        # LSU: col5 → 16:15 e 17:00
        {'tipo':'scrutinio','titolo':'Scrutinio finale LSU V A','data':'2026-06-06','ora_ini':'16:15','ora_fin':'17:00','classe':'5A LSU'},
        {'tipo':'scrutinio','titolo':'Scrutinio finale LSU V B','data':'2026-06-06','ora_ini':'17:00','ora_fin':'17:45','classe':'5B LSU'},
        # Scrutini classi intermedie (8-10/6)
        {'tipo':'scrutinio','titolo':'Scrutinio RIM IV A','data':'2026-06-08','ora_ini':'08:00','ora_fin':'08:45','classe':'4A RIM'},
        {'tipo':'scrutinio','titolo':'Scrutinio AFM II A','data':'2026-06-08','ora_ini':'08:45','ora_fin':'09:30','classe':'2A AFM'},
        {'tipo':'scrutinio','titolo':'Scrutinio AFM II B','data':'2026-06-08','ora_ini':'09:30','ora_fin':'10:15','classe':'2B AFM'},
        {'tipo':'scrutinio','titolo':'Scrutinio RIM III A','data':'2026-06-08','ora_ini':'10:15','ora_fin':'11:00','classe':'3A RIM'},
        {'tipo':'scrutinio','titolo':'Scrutinio AFM I A','data':'2026-06-08','ora_ini':'11:00','ora_fin':'11:45','classe':'1A AFM'},
        {'tipo':'scrutinio','titolo':'Scrutinio CAT IV A','data':'2026-06-08','ora_ini':'11:45','ora_fin':'12:30','classe':'4A CAT'},
        {'tipo':'scrutinio','titolo':'Scrutinio CAT III A','data':'2026-06-08','ora_ini':'13:30','ora_fin':'14:15','classe':'3A CAT'},
        {'tipo':'scrutinio','titolo':'Scrutinio CAT II A','data':'2026-06-08','ora_ini':'14:15','ora_fin':'15:00','classe':'2A CAT'},
        {'tipo':'scrutinio','titolo':'Scrutinio CAT I A','data':'2026-06-08','ora_ini':'15:00','ora_fin':'15:45','classe':'1A CAT'},
        {'tipo':'scrutinio','titolo':'Scrutinio LSC IV A','data':'2026-06-09','ora_ini':'08:00','ora_fin':'08:45','classe':'4A LSC'},
        {'tipo':'scrutinio','titolo':'Scrutinio LSC III A','data':'2026-06-09','ora_ini':'08:45','ora_fin':'09:30','classe':'3A LSC'},
        {'tipo':'scrutinio','titolo':'Scrutinio LSC II B','data':'2026-06-09','ora_ini':'09:30','ora_fin':'10:15','classe':'2B LSC'},
        {'tipo':'scrutinio','titolo':'Scrutinio LSC II A','data':'2026-06-09','ora_ini':'10:15','ora_fin':'11:00','classe':'2A LSC'},
        {'tipo':'scrutinio','titolo':'Scrutinio LSC I A','data':'2026-06-09','ora_ini':'11:00','ora_fin':'11:45','classe':'1A LSC'},
        {'tipo':'scrutinio','titolo':'Scrutinio LSU IV A','data':'2026-06-09','ora_ini':'11:45','ora_fin':'12:30','classe':'4A LSU'},
        {'tipo':'scrutinio','titolo':'Scrutinio LSU III A','data':'2026-06-09','ora_ini':'13:30','ora_fin':'14:15','classe':'3A LSU'},
        {'tipo':'scrutinio','titolo':'Scrutinio LL II B','data':'2026-06-09','ora_ini':'14:15','ora_fin':'15:00','classe':'2B LLI'},
        {'tipo':'scrutinio','titolo':'Scrutinio LL II A','data':'2026-06-09','ora_ini':'15:00','ora_fin':'15:45','classe':'2A LLI'},
        {'tipo':'scrutinio','titolo':'Scrutinio LL I A','data':'2026-06-09','ora_ini':'15:45','ora_fin':'16:30','classe':'1A LLI'},
        # 10/6: col H=LSP (IV,III,II,I), col G=LL/LLI (I A, I B, II B, II A, III A, IV A)
        {'tipo':'scrutinio','titolo':'Scrutinio LSP IV A','data':'2026-06-10','ora_ini':'08:00','ora_fin':'08:45','classe':'4A LSP'},
        {'tipo':'scrutinio','titolo':'Scrutinio LSP III A','data':'2026-06-10','ora_ini':'08:45','ora_fin':'09:30','classe':'3A LSP'},
        {'tipo':'scrutinio','titolo':'Scrutinio LSP II A','data':'2026-06-10','ora_ini':'09:30','ora_fin':'10:15','classe':'2A LSP'},
        {'tipo':'scrutinio','titolo':'Scrutinio LSP I A','data':'2026-06-10','ora_ini':'10:15','ora_fin':'11:00','classe':'1A LSP'},
        {'tipo':'scrutinio','titolo':'Scrutinio LLI I A','data':'2026-06-10','ora_ini':'11:00','ora_fin':'11:45','classe':'1A LLI'},
        {'tipo':'scrutinio','titolo':'Scrutinio LLI I B','data':'2026-06-10','ora_ini':'11:45','ora_fin':'12:30','classe':'1B LLI'},
        {'tipo':'scrutinio','titolo':'Scrutinio LLI II B','data':'2026-06-10','ora_ini':'13:30','ora_fin':'14:15','classe':'2B LLI'},
        {'tipo':'scrutinio','titolo':'Scrutinio LLI II A','data':'2026-06-10','ora_ini':'14:15','ora_fin':'15:00','classe':'2A LLI'},
        {'tipo':'scrutinio','titolo':'Scrutinio LLI III A','data':'2026-06-10','ora_ini':'15:00','ora_fin':'15:45','classe':'3A LLI'},
        {'tipo':'scrutinio','titolo':'Scrutinio LLI IV A','data':'2026-06-10','ora_ini':'15:45','ora_fin':'16:30','classe':'4A LLI'},
        # Collegio docenti 12/6
        {'tipo':'collegio','titolo':'Collegio dei docenti — giugno 2026','data':'2026-06-12','ora_ini':'11:00','ora_fin':'12:30'},
        # Incontro per materie 12/6
        {'tipo':'riunione_materia','titolo':'Incontro per materie — recupero giugno','data':'2026-06-12','ora_ini':'09:00','ora_fin':'10:30'},
        # Incontro famiglie 12/6
        {'tipo':'incontro_famiglie','titolo':'Incontro con le famiglie — giugno','data':'2026-06-12','ora_ini':'14:00','ora_fin':'15:00'},
        # ── SETTEMBRE 2026 ────────────────────────────────────────────────
        {'tipo':'collegio','titolo':'Collegio dei docenti — settembre 2026','data':'2026-09-01','ora_ini':'08:30','ora_fin':'13:00'},
    ]

    count = 0
    for ev in EVENTI:
        dup = AttivitaIst.query.filter_by(
            tipo=ev['tipo'], data=date.fromisoformat(ev['data']),
            titolo=ev['titolo']).first()
        if dup:
            continue
        obj = AttivitaIst(
            tipo       = ev['tipo'],
            titolo     = ev['titolo'],
            data       = date.fromisoformat(ev['data']),
            ora_inizio = ev.get('ora_ini'),
            ora_fine   = ev.get('ora_fin'),
            classe     = ev.get('classe'),
            origine    = 'import_piano',
        )
        db.session.add(obj)
        db.session.flush()

        # Preset partecipanti
        for did in _preset_partecipanti(obj):
            db.session.add(AttivitaIstPartecipante(
                id_attivita=obj.id, id_docente=did, preset=True))
        count += 1

    db.session.commit()
    return count


# ── IMPORT PIANO ANNUALE DA FILE .XLSX (standard corrente) ─────────────────────

@attivita_ist_bp.route('/attivita-ist/import-xlsx', methods=['GET', 'POST'])
def import_piano_xlsx():
    """
    Importa il Piano Annuale delle Attività da un file .xlsx costruito
    secondo lo standard adottato dalla scuola (banner-giorno a piena
    larghezza, righe-slot per Consigli/GLO, righe-evento per Collegio/
    Formazione/Incontri). Flusso in due passi: anteprima poi conferma,
    per evitare importazioni accidentali.
    """
    from modules.import_piano_xlsx import parse_piano_xlsx

    if request.method == 'POST' and 'eventi_json' in request.form:
        # ── passo 2: conferma import ──
        try:
            eventi = json.loads(request.form['eventi_json'])
        except Exception:
            flash('Dati di importazione non validi. Ripeti il caricamento del file.', 'danger')
            return redirect(url_for('attivita_ist.import_piano_xlsx'))

        n_importati = 0
        n_duplicati = 0
        dipartimenti_map = {d.nome.upper(): d.id for d in Dipartimento.query.all()}

        for ev in eventi:
            data_ev = date.fromisoformat(ev['data'])
            dup = AttivitaIst.query.filter_by(
                tipo=ev['tipo'], data=data_ev, titolo=ev['titolo'],
                classe=ev.get('classe'), ora_inizio=ev.get('ora_ini')).first()
            if dup:
                n_duplicati += 1
                continue

            obj = AttivitaIst(
                tipo       = ev['tipo'],
                titolo     = ev['titolo'],
                data       = data_ev,
                ora_inizio = ev.get('ora_ini'),
                ora_fine   = ev.get('ora_fin'),
                classe     = ev.get('classe'),
                note       = ev.get('note'),
                origine    = 'import_piano',
            )
            db.session.add(obj)
            db.session.flush()

            for did in _preset_partecipanti(obj):
                db.session.add(AttivitaIstPartecipante(
                    id_attivita=obj.id, id_docente=did, preset=True))
            n_importati += 1

        db.session.commit()
        flash(f'Importati {n_importati} eventi. {n_duplicati} già presenti (saltati).', 'success')
        return redirect(url_for('attivita_ist.lista'))

    if request.method == 'POST':
        # ── passo 1: caricamento file → anteprima ──
        f = request.files.get('file_xlsx')
        if not f or not f.filename:
            flash('Nessun file selezionato.', 'warning')
            return redirect(url_for('attivita_ist.import_piano_xlsx'))

        try:
            risultato = parse_piano_xlsx(f.read())
        except Exception as e:
            flash(f'Errore nella lettura del file: {e}', 'danger')
            return redirect(url_for('attivita_ist.import_piano_xlsx'))

        eventi = risultato['eventi']
        if not eventi:
            flash('Nessun evento riconosciuto nel file. Verifica che sia nel formato standard del Piano Annuale.', 'warning')
            return redirect(url_for('attivita_ist.import_piano_xlsx'))

        conteggio_tipi = {}
        for ev in eventi:
            conteggio_tipi[ev['tipo']] = conteggio_tipi.get(ev['tipo'], 0) + 1

        return render_template('attivita_ist/import_piano_xlsx_preview.html',
            fogli=risultato['fogli'], eventi=eventi, avvisi=risultato['avvisi'],
            conteggio_tipi=conteggio_tipi, tipi=TIPI_ATTIVITA,
            eventi_json=json.dumps(eventi),
        )

    return render_template('attivita_ist/import_piano_xlsx.html')


# ── DIPARTIMENTI E MATERIE ────────────────────────────────────────────────────

@attivita_ist_bp.route('/attivita-ist/dipartimenti', methods=['GET', 'POST'])
def dipartimenti():
    dips = Dipartimento.query.order_by(Dipartimento.ordine).all()
    # Materie non ancora assegnate = quelle nel dipartimento "Non assegnato" (sigla '—')
    dip_non_ass = Dipartimento.query.filter_by(sigla='—').first()
    if dip_non_ass:
        materie_da_assegnare = (Materia.query
                                .filter_by(id_dipartimento=dip_non_ass.id)
                                .order_by(Materia.sigla).all())
    else:
        materie_da_assegnare = []
    # Filtra i dipartimenti reali (escludi "Non assegnato" dall'elenco principale)
    dips_reali = [d for d in dips if d.sigla != '—']

    # Referenti di dipartimento — a differenza di Materie/Dipartimenti
    # (catalogo stabile, non legato all'anno), IncaricaDocente è per
    # anno scolastico. Prima usava sempre get_anno_corrente() (l'anno
    # "di sistema" configurato a mano, spesso disallineato dall'anno su
    # cui si sta effettivamente lavorando — vedi Task 23) senza
    # possibilità di scegliere: ora c'è un selettore, come nelle altre
    # pagine anno-scoped, di default sull'anno con dati reali.
    from routes.impostazione_anno import _anno_default_piano
    from models.incarico import IncaricaDocente, TipoIncarico
    anno_c = request.args.get('anno', _anno_default_piano())
    anni_disponibili = sorted(
        {r.anno_scol for r in IncaricaDocente.query.all()}, reverse=True)
    if anno_c not in anni_disponibili:
        anni_disponibili.insert(0, anno_c)
    tipo_ref = TipoIncarico.query.filter_by(nome='Referente di dipartimento').first()
    referenti = {}  # {id_dipartimento: Docente}
    if tipo_ref:
        nomine = IncaricaDocente.query.filter_by(
            anno_scol=anno_c, id_tipo_incarico=tipo_ref.id).all()
        for n in nomine:
            if n.id_dipartimento:
                referenti[n.id_dipartimento] = n.docente

    return render_template('attivita_ist/dipartimenti.html',
                           dipartimenti=dips_reali,
                           tutte_materie=materie_da_assegnare,
                           referenti=referenti,
                           anno_c=anno_c,
                           anni_disponibili=anni_disponibili,
                           tipi=TIPI_ATTIVITA)


@attivita_ist_bp.route('/attivita-ist/dipartimenti/assegna-materia', methods=['POST'])
def assegna_materia_dipartimento():
    """Assegna una materia esistente a un dipartimento (cambia id_dipartimento)."""
    id_materia = request.form.get('id_materia', type=int)
    id_dip     = request.form.get('id_dipartimento', type=int)
    if id_materia and id_dip:
        m = Materia.query.get(id_materia)
        if m:
            m.id_dipartimento = id_dip
            db.session.commit()
            flash(f'Materia "{m.nome_breve or m.nome}" assegnata al dipartimento.', 'success')
    return redirect(url_for('attivita_ist.dipartimenti'))





@attivita_ist_bp.route('/attivita-ist/dipartimenti/salva', methods=['POST'])
def salva_dipartimento():
    id_d  = request.form.get('id')
    nome  = request.form['nome'].strip()
    sigla = request.form['sigla'].strip().upper()
    ordine = int(request.form.get('ordine', 0))
    if id_d:
        d = Dipartimento.query.get_or_404(int(id_d))
        d.nome = nome; d.sigla = sigla; d.ordine = ordine
    else:
        db.session.add(Dipartimento(nome=nome, sigla=sigla, ordine=ordine))
    db.session.commit()
    flash('Dipartimento salvato.', 'success')
    return redirect(url_for('attivita_ist.dipartimenti'))


@attivita_ist_bp.route('/attivita-ist/materie/salva', methods=['POST'])
def salva_materia():
    import json
    id_m      = request.form.get('id')
    nome      = request.form['nome'].strip()
    sigla     = request.form['sigla'].strip().upper()
    id_dip    = int(request.form['id_dipartimento'])
    cod_or    = request.form.get('codice_orario', '').strip() or None
    indirizzi = request.form.getlist('indirizzi')
    nome_breve = request.form.get('nome_breve', '').strip() or None
    alias      = request.form.get('alias', '').strip().upper() or None

    if id_m:
        m = Materia.query.get_or_404(int(id_m))
        m.nome=nome; m.sigla=sigla; m.id_dipartimento=id_dip
        m.codice_orario=cod_or; m.indirizzi_json=json.dumps(indirizzi)
        m.nome_breve=nome_breve; m.alias=alias
    else:
        db.session.add(Materia(nome=nome, sigla=sigla, id_dipartimento=id_dip,
                               codice_orario=cod_or,
                               indirizzi_json=json.dumps(indirizzi),
                               nome_breve=nome_breve, alias=alias))
    db.session.commit()
    flash('Materia salvata.', 'success')
    return redirect(url_for('attivita_ist.dipartimenti'))


# ── ROSTER DOCENTI-MATERIE ────────────────────────────────────────────────────

@attivita_ist_bp.route('/attivita-ist/roster', methods=['GET', 'POST'])
def assegnazioni():
    from routes.impostazione_anno import _anno_default_piano
    anno   = request.args.get('anno', _anno_default_piano())
    docenti = Docente.query.filter_by(attivo=True).order_by(Docente.cognome).all()
    materie = Materia.query.join(Dipartimento).order_by(
        Dipartimento.ordine, Materia.nome).all()
    # Assegnazioni correnti
    assegn = {(dm.id_docente, dm.id_materia)
              for dm in DocenteMateria.query.filter_by(anno_scol=anno).all()}

    if request.method == 'POST':
        # Tocca solo le righe 'manuale' (quelle dichiarate da questa
        # pagina): non cancella mai le righe 'auto', sincronizzate
        # automaticamente da Assegnazioni classi -> docenti quando si
        # inseriscono ore su una materia. Prima questo salvataggio
        # cancellava TUTTO il roster dell'anno (tutti i docenti, incluse
        # le righe 'auto') per poi ricrearlo solo dalle caselle di questa
        # pagina — un salvataggio qui avrebbe potuto silenziosamente
        # perdere dati derivati da un'altra pagina. Per togliere una
        # materia 'auto' bisogna farlo da Assegnazioni (togliendo le ore)
        # o dal passo 10 "Docenti <-> Materie", che gestisce entrambe le
        # origini esplicitamente.
        DocenteMateria.query.filter_by(anno_scol=anno, origine='manuale').delete()
        coppie = set()
        for key in request.form:
            if key.startswith('dm_'):
                _, did, mid = key.split('_')
                coppie.add((int(did), int(mid)))
        n_nuove = 0
        for did, mid in coppie:
            esiste = DocenteMateria.query.filter_by(
                id_docente=did, id_materia=mid, anno_scol=anno).first()
            if not esiste:
                db.session.add(DocenteMateria(
                    id_docente=did, id_materia=mid, anno_scol=anno,
                    origine='manuale'))
                n_nuove += 1
        db.session.commit()
        flash(f'Roster aggiornato ({len(coppie)} assegnazioni, {n_nuove} nuove).', 'success')
        return redirect(url_for('attivita_ist.assegnazioni', anno=anno))

    docenti_lista = Docente.query.filter_by(attivo=True).order_by(Docente.cognome).all()
    return render_template('attivita_ist/roster.html',
        docenti=docenti_lista, materie=materie, assegn=assegn, anno=anno,
        dipartimenti=Dipartimento.query.order_by(Dipartimento.ordine).all())


# ── SOSTITUZIONI SCRUTINIO ────────────────────────────────────────────────────

@attivita_ist_bp.route('/attivita-ist/<int:id>/sostituzioni', methods=['GET', 'POST'])
def sostituzione_scrutinio(id):
    from models.sostituzione_scrutinio import SostituzioneScrutinio
    from models.materia import DocenteMateria, Materia
    from models.orario_docente import OrarioDocente
    from models.assenza import Assenza as AssenzaM
    evento = AttivitaIst.query.get_or_404(id)

    if evento.tipo != 'scrutinio':
        flash('Questa funzione è disponibile solo per gli scrutini.', 'warning')
        return redirect(url_for('attivita_ist.presenze', id=id))

    if request.method == 'POST':
        id_assente   = int(request.form['id_assente'])
        id_sostituto = request.form.get('id_sostituto') or None
        n_prot       = request.form.get('n_protocollo', '').strip() or None
        data_nomina  = request.form.get('data_nomina') or None
        note         = request.form.get('note', '').strip() or None

        # Un sostituto non può coprire due assenti nella stessa riunione:
        # controllo anche qui, non solo escludendolo dal menu, perché il
        # menu è filtrato al caricamento della pagina — due form inviati
        # da schede diverse aperte insieme scavalcherebbero il filtro lato
        # client.
        if id_sostituto:
            conflitto = SostituzioneScrutinio.query.filter(
                SostituzioneScrutinio.id_attivita == id,
                SostituzioneScrutinio.id_assente != id_assente,
                SostituzioneScrutinio.id_sostituto == int(id_sostituto),
            ).first()
            if conflitto:
                flash(f'{conflitto.sostituto.cognome} è già stato nominato sostituto di '
                      f'{conflitto.assente.cognome} in questa riunione: non può sostituire '
                      f'due docenti contemporaneamente.', 'error')
                return redirect(url_for('attivita_ist.sostituzione_scrutinio', id=id))

        sost = SostituzioneScrutinio.query.filter_by(
            id_attivita=id, id_assente=id_assente).first()
        if sost:
            sost.id_sostituto = int(id_sostituto) if id_sostituto else None
            sost.n_protocollo = n_prot
            sost.data_nomina  = date.fromisoformat(data_nomina) if data_nomina else None
            sost.note         = note
        else:
            db.session.add(SostituzioneScrutinio(
                id_attivita  = id,
                id_assente   = id_assente,
                id_sostituto = int(id_sostituto) if id_sostituto else None,
                n_protocollo = n_prot,
                data_nomina  = date.fromisoformat(data_nomina) if data_nomina else None,
                note         = note,
            ))

        # Se questa stessa coppia (evento, assente) era stata cancellata
        # in passato e aveva una lapide (SyncTombstone) — es. il blocco
        # di sostituzioni del 31/08 ricancellato su richiesta di Roberto
        # per ricominciare — va rimossa qui: altrimenti il prossimo giro
        # del sync automatico, trovando in locale una riga la cui chiave
        # risulta ancora lapidata, la cancella di nuovo (bug reale
        # riscontrato: nominare di nuovo su una macchina faceva sparire
        # la nomina appena inserita al giro successivo del sync,
        # indipendentemente da quale postazione l'avesse fatta). Una
        # nuova nomina inserita apposta dall'utente prevale sempre sulla
        # lapide di una cancellazione precedente per la stessa chiave.
        from models.sync_tombstone import SyncTombstone
        chiave_json = json.dumps({'id_attivita': id, 'id_assente': id_assente}, sort_keys=True)
        SyncTombstone.query.filter_by(
            tabella='sostituzioni_scrutinio', chiave_logica=chiave_json).delete()

        db.session.commit()
        flash('Sostituzione registrata.', 'success')
        return redirect(url_for('attivita_ist.sostituzione_scrutinio', id=id))

    # Docenti non in servizio a questa data (stesso controllo di
    # _preset_partecipanti). Usato sia per i candidati sostituti sotto,
    # sia qui per gli assenti: un partecipante non più in servizio va
    # sempre considerato "da sostituire", anche se la sua riga presenze
    # è rimasta sullo stato di default 'presente' perché nessuno l'ha
    # ancora aggiornata a mano — altrimenti la pagina "Sostituzioni"
    # risulta vuota nonostante il badge "non più in servizio" mostrato
    # nella pagina presenze (bug segnalato da Roberto: Agrò, non in
    # servizio dal 31/08, non compariva come assente da sostituire).
    esclusi_servizio = _non_in_servizio_per_data(evento.data)

    # Docenti assenti a questo scrutinio
    presenze_assenti = [p for p in evento.presenze
                        if p.stato in ('assente', 'giustificato')
                        or p.id_docente in esclusi_servizio]

    # Classi dello scrutinio (per escludere i docenti di quella classe)
    classe_scrutinio = evento.classe  # es. '3A LSC'

    # Docenti della classe (da escludere)
    docenti_classe_ids = set()
    if classe_scrutinio:
        docenti_classe_ids = {
            s.id_docente for s in OrarioDocente.query.filter_by(
                classe=classe_scrutinio).all()
        }

    # Tutti i docenti attivi NON della classe, esclusi quelli non in
    # servizio a questa data (stesso controllo di _preset_partecipanti —
    # senza, un docente con anno_scol_inizio futuro comparirebbe come
    # possibile sostituto per uno scrutinio prima di essere arrivato).
    tutti = Docente.query.filter_by(attivo=True).order_by(Docente.cognome).all()
    candidati_base = [d for d in tutti
                      if d.id not in docenti_classe_ids and d.id not in esclusi_servizio]

    # Lookup id -> Docente, usato nel fallback su Docente.materia (testo
    # libero) quando DocenteMateria è vuota — include anche gli assenti,
    # che potrebbero non essere in "tutti" se non più attivi.
    _docenti_by_id = {dd.id: dd for dd in tutti}
    _docenti_by_id.update({p.docente.id: p.docente for p in presenze_assenti})

    # Assenti quel giorno (non disponibili)
    assenti_giorno = {a.id_docente for a in AssenzaM.query.filter_by(data=evento.data).all()}

    # Già impegnati in un altro scrutinio contemporaneo
    altri_scrutini = AttivitaIst.query.filter(
        AttivitaIst.data == evento.data,
        AttivitaIst.tipo == 'scrutinio',
        AttivitaIst.id != id
    ).all()
    def _to_min(t):
        try:
            h, m = map(int, t.split(':'))
            return h * 60 + m
        except Exception:
            return 0

    impegnati_altri = set()
    ev_ini = _to_min(evento.ora_inizio) if evento.ora_inizio else 0
    ev_fin = _to_min(evento.ora_fine)   if evento.ora_fine   else ev_ini + 45
    for alt in altri_scrutini:
        if not alt.ora_inizio:
            continue
        alt_ini = _to_min(alt.ora_inizio)
        alt_fin = _to_min(alt.ora_fine) if alt.ora_fine else alt_ini + 45
        # Sovrapposizione reale: i due intervalli si intersecano
        if alt_ini < ev_fin and alt_fin > ev_ini:
            impegnati_altri.update(p.id_docente for p in alt.partecipanti)

    # Altre attività istituzionali dello stesso giorno, indicizzate per
    # docente partecipante: servono sia al punteggio "altra riunione" sia,
    # per ogni candidato, a trovare la riunione immediatamente precedente
    # o successiva a questo scrutinio — utile per scegliere chi è comunque
    # già a scuola in quella fascia oraria.
    from collections import defaultdict
    altri_ev_giorno = AttivitaIst.query.filter(
        AttivitaIst.data == evento.data,
        AttivitaIst.id != id
    ).all()
    riunioni_per_docente = defaultdict(list)
    for ev2 in altri_ev_giorno:
        if not ev2.ora_inizio:
            continue
        for p2 in ev2.partecipanti:
            riunioni_per_docente[p2.id_docente].append(ev2)

    def _riunione_prec_succ(docente_id):
        """Tra le altre riunioni del giorno a cui partecipa questo
        docente, la più vicina prima e dopo questo scrutinio (per
        orario), o None se non ce ne sono."""
        prec = succ = None
        prec_fin = -1
        succ_ini = 10**9
        for ev2 in riunioni_per_docente.get(docente_id, []):
            ev2_ini = _to_min(ev2.ora_inizio)
            ev2_fin = _to_min(ev2.ora_fine) if ev2.ora_fine else ev2_ini + 45
            if ev2_fin <= ev_ini and ev2_fin > prec_fin:
                prec, prec_fin = ev2, ev2_fin
            elif ev2_ini >= ev_fin and ev2_ini < succ_ini:
                succ, succ_ini = ev2, ev2_ini
        return prec, succ

    def _segnali_candidato(d, assente_id, riun_prec, riun_succ):
        """
        Calcola TUTTI i segnali di priorità applicabili a questo
        candidato, non solo il migliore — prima, un return anticipato
        al primo segnale trovato (es. stessa materia) impediva di
        vedere se ne fossero presenti anche altri (es. anche una
        riunione lo stesso giorno): un candidato finiva sempre
        etichettato con un solo segnale, quello con priorità più alta,
        anche quando ne aveva altri (segnalato da Roberto: vedeva quasi
        sempre solo "③ riunione" e voleva sapere se c'erano anche gli
        altri segnali).

        Ritorna (score_ordinamento, segnali) dove segnali è un set di
        interi tra {1,2,3,4}:
        1. Stessa materia dell'assente
        2. Stesso dipartimento
        3. Libero e ha un'altra riunione lo stesso giorno (prima o
           dopo questo scrutinio)
        4. Generico disponibile — SOLO se nessun altro segnale è
           presente, altrimenti non aggiunge informazione.

        Lo score di ordinamento resta identico a prima (stesso segnale
        migliore decide, con lo stesso dettaglio fine di vicinanza
        oraria per "riunione" — Task 47): serve solo a ordinare la
        lista, i segnali INFORMATIVI da mostrare sono tutti quelli
        presenti in "segnali", non solo quello scelto per l'ordine.
        """
        if d.id in assenti_giorno: return 999, set()  # non disponibile
        if d.id in impegnati_altri: return 998, set()

        segnali = set()

        # Materia dell'assente — anno scolastico dell'evento (evento.data),
        # non "oggi": uno scrutinio programmato per un anno diverso da
        # quello corrente deve confrontare le materie di quell'anno.
        anno_evento = _anno_scolastico(evento.data)
        assente_mat_ids = {dm.id_materia for dm in DocenteMateria.query.filter_by(
            id_docente=assente_id, anno_scol=anno_evento).all()}
        cand_mat_ids = {dm.id_materia for dm in DocenteMateria.query.filter_by(
            id_docente=d.id, anno_scol=anno_evento).all()}

        # DocenteMateria (struttura per-anno, popolata da Assegnazioni/
        # checkbox scheda docente) copre solo una minoranza dei docenti
        # (~54 su 97): per gli altri i segnali 1/2 non scattavano mai,
        # anche quando la materia era nota — segnalato da Roberto (vedeva
        # quasi solo "③ riunione" in sostituzioni scrutinio). Fallback sul
        # campo libero Docente.materia (compilato per la quasi totalità
        # dei docenti, es. anagrafica di importazione) quando la parte
        # strutturata è vuota per uno dei due lati.
        assente_doc = _docenti_by_id.get(assente_id)
        assente_mat_txt = {m.strip().upper() for m in (assente_doc.materia or '').split(',') if m.strip()} \
            if assente_doc else set()
        cand_mat_txt = {m.strip().upper() for m in (d.materia or '').split(',') if m.strip()}

        stessa_materia = bool(assente_mat_ids & cand_mat_ids)
        if not stessa_materia and not assente_mat_ids and not cand_mat_ids:
            stessa_materia = bool(assente_mat_txt & cand_mat_txt)
        # Ulteriore fallback: stessa classe di concorso — segnalato da
        # Roberto (2B LSC, Del Curto assente: Boffi non risultava "①
        # stessa materia" nonostante stessa CC e stessa materia
        # insegnata, solo perché il campo libero "materia" era scritto
        # in modo diverso — "Scienze motorie" contro "Discipline
        # sportive"). La classe di concorso non è legata all'anno
        # scolastico nel modello (campo unico su Docente), quindi è un
        # segnale più affidabile del testo libero quando entrambi i
        # docenti ce l'hanno e coincide — non lo sostituisce, si
        # aggiunge come ulteriore modo di rilevare la stessa materia.
        if not stessa_materia and assente_doc and assente_doc.id_classe_concorso \
                and assente_doc.id_classe_concorso == d.id_classe_concorso:
            stessa_materia = True
        if stessa_materia:
            segnali.add(1)  # stessa materia

        assente_dips = cand_dips = set()
        if assente_mat_ids and cand_mat_ids:
            assente_dips = {m.id_dipartimento for m in Materia.query.filter(
                Materia.id.in_(assente_mat_ids)).all()}
            cand_dips = {m.id_dipartimento for m in Materia.query.filter(
                Materia.id.in_(cand_mat_ids)).all()}
        elif assente_mat_txt and cand_mat_txt:
            assente_dips = {m.id_dipartimento for m in Materia.query.filter(
                db.func.upper(Materia.nome).in_(assente_mat_txt)).all()}
            cand_dips = {m.id_dipartimento for m in Materia.query.filter(
                db.func.upper(Materia.nome).in_(cand_mat_txt)).all()}
        if assente_dips & cand_dips:
            segnali.add(2)  # stesso dipartimento

        if riun_prec or riun_succ:
            segnali.add(3)  # altra riunione lo stesso giorno

        if not segnali:
            segnali.add(4)  # generico disponibile

        # Ordinamento: la vicinanza oraria (③) è il criterio dominante —
        # un docente già a scuola per un altro impegno vicino è la scelta
        # più pratica, va richiamato apposta solo se nessuno è nei
        # paraggi. Materia (①) e dipartimento (②) restano visibili come
        # segnali ma pesano solo come bonus fine, a parità (o quasi) di
        # comodità oraria — prima di questo fix erano il criterio
        # dominante, e un match di materia scavalcava un candidato molto
        # più comodo in orario, cosa che Roberto non voleva più vedere.
        # Fasce ben separate (nessuna si sovrappone anche col bonus
        # massimo) cosi' il bonus non fa mai scavalcare una fascia oraria
        # peggiore:
        #   riunione prima:  0-4   (comodissimo)
        #   riunione dopo:  10-14  (comodo, ma dopo è meno pratico di prima)
        #   nessuna riunione quel giorno: 40 (va richiamato apposta)
        if riun_prec:
            prec_fin = _to_min(riun_prec.ora_fine) if riun_prec.ora_fine else _to_min(riun_prec.ora_inizio) + 45
            gap = max(0, ev_ini - prec_fin)
            orario_score = min(gap / 480, 1) * 4
        elif riun_succ:
            succ_ini = _to_min(riun_succ.ora_inizio)
            gap = max(0, succ_ini - ev_fin)
            orario_score = 10 + min(gap / 480, 1) * 4
        else:
            orario_score = 40

        if 1 in segnali:
            bonus = -2  # stessa materia
        elif 2 in segnali:
            bonus = -1  # stesso dipartimento
        else:
            bonus = 0

        score = orario_score + bonus

        return score, segnali

    # Calcola score per ogni candidato rispetto a ogni assente
    sostituzioni_attuali = {s.id_assente: s for s in
                            SostituzioneScrutinio.query.filter_by(id_attivita=id).all()}

    righe = []
    for p in presenze_assenti:
        assente = p.docente
        # Sostituti già nominati per UN ALTRO assente in questa stessa
        # riunione: non possono comparire come candidati anche qui, non
        # possono coprire due assenti contemporaneamente.
        gia_impegnati_riunione = {
            s.id_sostituto for a_id, s in sostituzioni_attuali.items()
            if a_id != assente.id and s.id_sostituto
        }
        candidati_riga = [d for d in candidati_base
                          if d.id != assente.id
                          and d.id not in assenti_giorno
                          and d.id not in gia_impegnati_riunione]
        cands_prec_succ = {d.id: _riunione_prec_succ(d.id) for d in candidati_riga}
        cands_scored = sorted(
            [(d, *_segnali_candidato(d, assente.id, *cands_prec_succ[d.id]),
              *cands_prec_succ[d.id])
             for d in candidati_riga],
            key=lambda x: x[1]
        )
        sost_att = sostituzioni_attuali.get(assente.id)
        docenti_disp_riga = [d for d in candidati_base
                             if d.id not in assenti_giorno
                             and d.id not in impegnati_altri
                             and d.id not in gia_impegnati_riunione]
        # Stesso assente, già sostituito lo stesso giorno in un'altra
        # riunione — richiesto da Roberto: aprendo la pagina di una
        # classe dove questo docente è ancora da sostituire, vedere se
        # è già stato coperto altrove aiuta a orientarsi rapidamente
        # (es. per scegliere lo stesso sostituto se disponibile, o solo
        # per contesto) senza dover controllare evento per evento.
        altre_sostituzioni = (SostituzioneScrutinio.query
            .join(AttivitaIst, SostituzioneScrutinio.id_attivita == AttivitaIst.id)
            .filter(SostituzioneScrutinio.id_assente == assente.id,
                    SostituzioneScrutinio.id_attivita != evento.id,
                    SostituzioneScrutinio.id_sostituto.isnot(None),
                    AttivitaIst.data == evento.data)
            .all())
        righe.append({
            'presenza': p,
            'assente': assente,
            'candidati': cands_scored[:8],
            'docenti_disponibili': docenti_disp_riga,
            'sostituzione': sost_att,
            'altre_sostituzioni': altre_sostituzioni,
        })

    return render_template('attivita_ist/sostituzioni_scrutinio.html',
        evento=evento,
        righe=righe,
        oggi=date.today(),
    )


def _righe_protocollazione(data_da, data_a):
    """Righe (SostituzioneScrutinio con sostituto assegnato) per la
    pagina di protocollazione, filtrate per intervallo di date sugli
    scrutini. Senza filtro data, mostra solo quelle NON ancora
    protocollate (il caso d'uso principale: compilare i protocolli
    mancanti) — con un filtro data esplicito mostra tutte, anche quelle
    già fatte, per poter rivedere un blocco specifico."""
    from models.sostituzione_scrutinio import SostituzioneScrutinio
    q = (SostituzioneScrutinio.query
         .join(AttivitaIst, SostituzioneScrutinio.id_attivita == AttivitaIst.id)
         .filter(AttivitaIst.tipo == 'scrutinio',
                 SostituzioneScrutinio.id_sostituto.isnot(None)))
    if data_da:
        q = q.filter(AttivitaIst.data >= data_da)
    if data_a:
        q = q.filter(AttivitaIst.data <= data_a)
    if not data_da and not data_a:
        q = q.filter(db.or_(SostituzioneScrutinio.n_protocollo.is_(None),
                             SostituzioneScrutinio.n_protocollo == ''))
    righe = q.all()
    righe.sort(key=lambda s: (s.attivita.data, s.attivita.classe or '', s.assente.cognome))
    return righe


def _gruppi_protocollazione(data_da, data_a):
    """Raggruppa le righe per SOSTITUTO, non per assente — Roberto: in
    segreteria il decreto si fa per ogni docente che sostituisce
    qualcuno, elencando TUTTE le sue coperture (classi/orari diversi,
    anche di assenti diversi) sotto lo stesso numero di protocollo — un
    solo documento per persona, non uno per ogni singola sostituzione.
    La vista "una riga per assente" precedente costringeva a scrivere
    lo stesso protocollo più volte per lo stesso sostituto.

    Ogni gruppo espone 'protocollo' (il valore comune se tutte le righe
    del gruppo ce l'hanno uguale, altrimenti stringa vuota) e
    'protocolli_diversi' (True se il gruppo ha valori diversi/parziali
    — caso residuo di dati inseriti prima di questo cambio, segnalato
    invece di deciso in automatico)."""
    righe = _righe_protocollazione(data_da, data_a)
    gruppi = {}
    ordine = []
    for s in righe:
        if s.id_sostituto not in gruppi:
            gruppi[s.id_sostituto] = {'sostituto': s.sostituto, 'righe': []}
            ordine.append(s.id_sostituto)
        gruppi[s.id_sostituto]['righe'].append(s)

    risultato = []
    for sid in ordine:
        g = gruppi[sid]
        g['righe'].sort(key=lambda r: (r.attivita.data, r.attivita.ora_inizio or '', r.attivita.classe or ''))
        protocolli = {r.n_protocollo for r in g['righe']}
        risultato.append({
            'sostituto': g['sostituto'],
            'righe': g['righe'],
            'protocollo': next(iter(protocolli)) if len(protocolli) == 1 else '',
            'protocolli_diversi': len(protocolli) > 1,
            'ids': [r.id for r in g['righe']],
        })
    risultato.sort(key=lambda g: (g['sostituto'].cognome if g['sostituto'] else '',
                                   g['sostituto'].nome if g['sostituto'] else ''))
    return risultato


@attivita_ist_bp.route('/attivita-ist/protocollazione', methods=['GET', 'POST'])
def protocollazione_scrutini():
    """Riepilogo di tutte le sostituzioni assegnate su un blocco di
    scrutini (o su tutte quelle non ancora protocollate se nessun
    intervallo è specificato), raggruppate per sostituto (vedi
    _gruppi_protocollazione) — un numero di protocollo per persona,
    applicato a tutte le sue coperture nel gruppo in un solo salvataggio,
    invece di aprire ogni scrutinio singolarmente. Il campo N. protocollo
    è lo stesso SostituzioneScrutinio.n_protocollo usato in
    sostituzione_scrutinio(): non serve nessuna sincronizzazione, sono
    le stesse righe viste da due pagine."""
    if request.method == 'POST':
        ids = [int(x) for x in request.form.get('ids_sostituzione', '').split(',') if x.strip()]
        protocollo = request.form.get('n_protocollo', '').strip() or None
        from models.sostituzione_scrutinio import SostituzioneScrutinio
        righe_gruppo = SostituzioneScrutinio.query.filter(SostituzioneScrutinio.id.in_(ids)).all()
        for r in righe_gruppo:
            r.n_protocollo = protocollo
        db.session.commit()
        n = len(righe_gruppo)
        flash(f'Protocollo aggiornato per {n} sostituzion{"e" if n == 1 else "i"}.', 'success')
        return redirect(url_for('attivita_ist.protocollazione_scrutini',
                                 data_da=request.form.get('data_da') or None,
                                 data_a=request.form.get('data_a') or None))

    data_da = request.args.get('data_da') or None
    data_a  = request.args.get('data_a') or None
    gruppi = _gruppi_protocollazione(data_da, data_a)

    return render_template('attivita_ist/protocollazione_scrutini.html',
        gruppi=gruppi, data_da=data_da, data_a=data_a,
    )


@attivita_ist_bp.route('/attivita-ist/protocollazione/export')
def protocollazione_scrutini_export():
    from openpyxl.utils import get_column_letter
    from routes.export_xlsx import _wb, _hdr, _row, _border_all, _send

    data_da = request.args.get('data_da') or None
    data_a  = request.args.get('data_a') or None
    gruppi = _gruppi_protocollazione(data_da, data_a)

    wb = _wb()
    ws = wb.create_sheet('Protocollazione scrutini')
    larghezze = [28, 12, 8, 28, 12, 18]
    for i, larg in enumerate(larghezze, 1):
        ws.column_dimensions[get_column_letter(i)].width = larg

    r = 1
    ws.cell(r, 1, 'Protocollazione sostituzioni scrutini')
    r += 1
    if data_da or data_a:
        ws.cell(r, 1, f"Periodo: {data_da or '…'} — {data_a or '…'}")
    else:
        ws.cell(r, 1, 'Solo sostituzioni non ancora protocollate')
    r += 2

    # Una riga per copertura, ma raggruppate per sostituto (come il
    # decreto che la segreteria emette per ciascuno di loro, con lo
    # stesso protocollo per tutte le sue coperture — Roberto) invece
    # che una lista piatta ordinata per assente.
    r = _hdr(ws, r, ['Sostituto', 'Data', 'Ora', 'Docente assente', 'Classe', 'N. protocollo'])
    for g in gruppi:
        for s in g['righe']:
            r = _row(ws, r, [
                g['sostituto'].nome_completo if g['sostituto'] else '',
                s.attivita.data.strftime('%d/%m/%Y'),
                s.attivita.ora_inizio or '',
                s.assente.nome_completo,
                s.attivita.classe or '',
                s.n_protocollo or '',
            ])
    _border_all(ws, 4, max(r - 1, 4), 1, 6)

    return _send(wb, 'protocollazione_scrutini.xlsx')
