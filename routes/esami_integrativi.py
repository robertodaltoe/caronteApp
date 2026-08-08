"""
Modulo Esami Integrativi (passaggi e trasferimenti di settembre).
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash
from models import db
from models.esami_integrativi import EsameIntegrativoCandidato, EsameIntegrativoMateria
from models.docente import Docente
from models.orario_docente import OrarioDocente, classi_attive as _classi_attive
from datetime import datetime

esami_integrativi_bp = Blueprint('esami_integrativi', __name__)

from config_anno import get_anno_corrente as _get_anno
ANNO = _get_anno()

# Stesse famiglie di sinonimi materia già usate nel modulo rientro — per
# proporre automaticamente i docenti idonei a esaminare una materia.
_FAMIGLIE_MATERIE = [
    {'LATINO', 'LINGUA LATINA', 'LINGUA E CULTURA LATINA'},
    {'ITALIANO', 'LINGUA E LETTERATURA ITALIANA', 'LINGUA E CULTURA ITALIANA'},
    {'MATEMATICA', 'MATEMATICA CON INFORMATICA', 'MATEMATICA E COMPLEMENTI DI MATEMATICA'},
    {'INFORMATICA', 'TECNOLOGIE INFORMATICHE'},
    {'STORIA', 'STORIA E GEOGRAFIA'},
    {'FISICA', 'SCIENZE INTEGRATE (FISICA)'},
    {'INGLESE', 'LINGUA INGLESE', 'LINGUA E CULTURA STRANIERA (INGLESE)'},
    {'TEDESCO', 'LINGUA TEDESCA', 'LINGUA E CULTURA STRANIERA TEDESCO'},
]

def _materia_canonica(materia):
    mu = (materia or '').strip().upper()
    for fam in _FAMIGLIE_MATERIE:
        if mu in fam:
            return sorted(fam)[0]
    return mu


def _docenti_idonei_materia(materia):
    """Tutti i docenti che insegnano quella materia (famiglia sinonimi), in qualsiasi classe."""
    mat_can = _materia_canonica(materia)
    ids = set()
    for o in OrarioDocente.query.filter(OrarioDocente.materia.isnot(None)).all():
        if _materia_canonica(o.materia) == mat_can:
            ids.add(o.id_docente)
    return Docente.query.filter(Docente.id.in_(ids)).order_by(Docente.cognome).all() if ids else []


def _docente_disponibile(id_docente, data):
    """Contratto attivo + nessuna assenza registrata in quella data."""
    if not id_docente or not data:
        return True
    d = Docente.query.get(id_docente)
    if not d or not d.attivo:
        return False
    from models.assenza import Assenza
    ass = Assenza.query.filter_by(id_docente=id_docente, data=data).first()
    return ass is None


# ── INDICE ──────────────────────────────────────────────────────────
@esami_integrativi_bp.route('/esami-integrativi')
def index():
    candidati = EsameIntegrativoCandidato.query.filter_by(anno_scol=ANNO).order_by(
        EsameIntegrativoCandidato.cognome).all()

    n_materie = sum(len(c.materie) for c in candidati)
    n_calendarizzate = sum(1 for c in candidati for m in c.materie if m.data)

    return render_template('esami_integrativi/index.html',
        candidati=candidati,
        n_candidati=len(candidati),
        n_materie=n_materie,
        n_calendarizzate=n_calendarizzate,
        anno=ANNO)


# ── CANDIDATI + MATERIE ────────────────────────────────────────────
@esami_integrativi_bp.route('/esami-integrativi/candidati', methods=['GET', 'POST'])
def candidati():
    if request.method == 'POST':
        azione = request.form.get('azione')

        if azione == 'aggiungi':
            cognome = request.form.get('cognome', '').strip().upper()
            nome = request.form.get('nome', '').strip().title()
            classe_dest = request.form.get('classe_destinazione', '').strip().upper()
            provenienza = request.form.get('provenienza', '').strip() or None
            note = request.form.get('note', '').strip() or None
            if not (cognome and nome and classe_dest):
                flash('Cognome, nome e classe di destinazione sono obbligatori.', 'warning')
                return redirect(url_for('esami_integrativi.candidati'))
            db.session.add(EsameIntegrativoCandidato(
                anno_scol=ANNO, cognome=cognome, nome=nome,
                classe_destinazione=classe_dest, provenienza=provenienza, note=note))
            db.session.commit()
            flash(f'Candidato {cognome} {nome} aggiunto.', 'success')

        elif azione == 'elimina':
            cand_id = request.form.get('id')
            if cand_id:
                cand = EsameIntegrativoCandidato.query.get(int(cand_id))
                if cand:
                    nome_completo = cand.nome_completo
                    db.session.delete(cand)
                    db.session.commit()
                    flash(f'Candidato {nome_completo} rimosso.', 'warning')

        elif azione == 'aggiungi_materia':
            cand_id = request.form.get('id_candidato')
            materia = request.form.get('materia', '').strip().upper()
            tipologia = request.form.get('tipologia') or None
            if tipologia not in (None, 'scritta', 'orale'):
                tipologia = None
            if cand_id and materia:
                db.session.add(EsameIntegrativoMateria(
                    id_candidato=int(cand_id), materia=materia, tipologia=tipologia))
                db.session.commit()
                flash(f'Materia {materia} aggiunta.', 'success')

        elif azione == 'elimina_materia':
            mat_id = request.form.get('id_materia')
            if mat_id:
                mat = EsameIntegrativoMateria.query.get(int(mat_id))
                if mat:
                    db.session.delete(mat)
                    db.session.commit()
                    flash('Materia rimossa.', 'warning')

        elif azione == 'imposta_tipologia':
            mat_id = request.form.get('id_materia')
            tipologia = request.form.get('tipologia') or None
            if tipologia not in (None, 'scritta', 'orale'):
                tipologia = None
            if mat_id:
                mat = EsameIntegrativoMateria.query.get(int(mat_id))
                if mat:
                    mat.tipologia = tipologia
                    db.session.commit()

        return redirect(url_for('esami_integrativi.candidati'))

    righe = EsameIntegrativoCandidato.query.filter_by(anno_scol=ANNO).order_by(
        EsameIntegrativoCandidato.cognome).all()

    # Elenco materie note dall'orario, per il datalist di autocompletamento.
    materie_disponibili = sorted({
        o.materia.strip().upper() for o in
        OrarioDocente.query.filter(OrarioDocente.materia.isnot(None)).all() if o.materia
    })

    return render_template('esami_integrativi/candidati.html',
        candidati=righe, materie_disponibili=materie_disponibili,
        classi_attive=_classi_attive(), anno=ANNO)


# ── COMMISSIONE + CALENDARIO ────────────────────────────────────────
@esami_integrativi_bp.route('/esami-integrativi/calendario', methods=['GET', 'POST'])
def calendario():
    if request.method == 'POST':
        azione = request.form.get('azione')

        if azione == 'imposta_membro':
            id_mat = int(request.form['id_materia'])
            campo = request.form.get('campo')  # docente_1 | docente_2
            id_doc = request.form.get('id_docente') or None
            mat = EsameIntegrativoMateria.query.get_or_404(id_mat)
            if campo in ('docente_1', 'docente_2'):
                setattr(mat, f'id_{campo}', int(id_doc) if id_doc else None)
                db.session.commit()
            return redirect(url_for('esami_integrativi.calendario'))

        elif azione == 'modifica_orario':
            id_mat = int(request.form['id_materia'])
            mat = EsameIntegrativoMateria.query.get_or_404(id_mat)
            data_str = request.form.get('data', '')
            mat.data = datetime.strptime(data_str, '%Y-%m-%d').date() if data_str else None
            mat.ora_inizio = request.form.get('ora_inizio') or None
            mat.ora_fine = request.form.get('ora_fine') or None
            db.session.commit()
            flash('Orario aggiornato.', 'success')
            return redirect(url_for('esami_integrativi.calendario'))

        elif azione == 'azzera_tutto':
            for mat in EsameIntegrativoMateria.query.all():
                mat.data = None
                mat.ora_inizio = None
                mat.ora_fine = None
            db.session.commit()
            flash('Calendario azzerato.', 'warning')
            return redirect(url_for('esami_integrativi.calendario'))

    candidati_list = EsameIntegrativoCandidato.query.filter_by(anno_scol=ANNO).order_by(
        EsameIntegrativoCandidato.cognome).all()

    righe = []
    for cand in candidati_list:
        materie_riga = []
        for mat in cand.materie:
            disponibile_1 = _docente_disponibile(mat.id_docente_1, mat.data) if mat.docente_1 else None
            disponibile_2 = _docente_disponibile(mat.id_docente_2, mat.data) if mat.docente_2 else None
            docenti_idonei = _docenti_idonei_materia(mat.materia)
            materie_riga.append({
                'materia_obj': mat,
                'disponibile_1': disponibile_1,
                'disponibile_2': disponibile_2,
                'docenti_idonei': docenti_idonei,
            })
        righe.append({'candidato': cand, 'materie': materie_riga})

    conflitti = _calcola_conflitti(righe)

    return render_template('esami_integrativi/calendario.html',
        righe=righe, conflitti=conflitti, anno=ANNO)


def _calcola_conflitti(righe):
    """Sovrapposizioni tra esami diversi che condividono un membro commissione."""
    def _t(s):
        try:
            h, m = map(int, s.split(':'))
            return h * 60 + m
        except Exception:
            return None

    esami_validi = []
    for r in righe:
        for mr in r['materie']:
            mat = mr['materia_obj']
            if mat.data and mat.ora_inizio and mat.ora_fine:
                esami_validi.append((r['candidato'], mat))

    conflitti = []
    for i, (cand1, m1) in enumerate(esami_validi):
        ini1, fin1 = _t(m1.ora_inizio), _t(m1.ora_fine)
        membri1 = {d.id for d in m1.membri_commissione}
        if not membri1 or ini1 is None:
            continue
        for cand2, m2 in esami_validi[i+1:]:
            if m2.data != m1.data:
                continue
            ini2, fin2 = _t(m2.ora_inizio), _t(m2.ora_fine)
            if ini2 is None:
                continue
            if ini1 < fin2 and ini2 < fin1:
                membri2 = {d.id for d in m2.membri_commissione}
                comuni = membri1 & membri2
                if comuni:
                    nomi = ', '.join(Docente.query.get(did).cognome for did in comuni)
                    conflitti.append({
                        'msg': f'{nomi}: esame {cand1.cognome} ({m1.materia}) e '
                               f'{cand2.cognome} ({m2.materia}) sovrapposti il '
                               f'{m1.data.strftime("%d/%m")}',
                    })
    return conflitti


@esami_integrativi_bp.route('/esami-integrativi/export-xlsx')
def export_xlsx():
    """
    Export del calendario esami integrativi: un foglio unico ordinato
    per giornata, poi per candidato e orario. Stesso impianto grafico
    degli export dei moduli recupero/rientro.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file

    candidati_list = EsameIntegrativoCandidato.query.filter_by(anno_scol=ANNO).order_by(
        EsameIntegrativoCandidato.cognome).all()

    GIORNI = ['Lunedì', 'Martedì', 'Mercoledì', 'Giovedì', 'Venerdì', 'Sabato', 'Domenica']

    BLU    = PatternFill('solid', start_color='7a1c17')
    AZZUR  = PatternFill('solid', start_color='dbeafe')
    BOLD   = Font(bold=True)
    BOLD_W = Font(bold=True, color='FFFFFF')
    THIN   = Border(left=Side(style='thin', color='d1d5db'),
                     right=Side(style='thin', color='d1d5db'),
                     top=Side(style='thin', color='d1d5db'),
                     bottom=Side(style='thin', color='d1d5db'))
    CENTER = Alignment(horizontal='center', vertical='center')
    WRAP   = Alignment(wrap_text=True, vertical='center')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Calendario'

    from config_istituto import get_dati_istituto
    ws['A1'] = get_dati_istituto()['nome_istituto']
    ws['A1'].font = Font(bold=True, size=13)
    ws['A2'] = f'ESAMI INTEGRATIVI — A.S. {ANNO}'
    ws['A2'].font = Font(bold=True, size=11)
    ws.append([])
    ws.append([])

    esami_per_giorno = {}
    esami_senza_data = []
    for cand in candidati_list:
        for mat in cand.materie:
            if mat.data and mat.ora_inizio and mat.ora_fine:
                esami_per_giorno.setdefault(mat.data, []).append((mat, cand))
            else:
                esami_senza_data.append((mat, cand))

    HEADER_COLS = ['Orario', 'Candidato', 'Classe dest.', 'Provenienza', 'Materia',
                   'Tipologia', 'Docente 1', 'Docente 2']
    N_COLS = len(HEADER_COLS)

    def _nome_doc(d):
        return f'{d.cognome} {d.nome or ""}'.strip() if d else '—'

    def _tipologia_label(t):
        return {'scritta': 'Scritta', 'orale': 'Orale'}.get(t, '—')

    def _scrivi_header_giorno(titolo):
        row = ws.max_row + 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
        cell = ws.cell(row=row, column=1, value=titolo)
        cell.font = BOLD_W
        cell.fill = BLU
        cell.alignment = CENTER
        row += 1
        for col, h in enumerate(HEADER_COLS, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = BOLD
            c.fill = AZZUR
            c.alignment = CENTER
            c.border = THIN
        return row + 1

    def _scrivi_riga(row, mat, cand, mostra_orario=True):
        orario_str = f'{mat.ora_inizio}–{mat.ora_fine}' if mostra_orario and mat.ora_inizio else '—'
        vals = [orario_str, cand.nome_completo, cand.classe_destinazione,
                cand.provenienza or '—', mat.materia, _tipologia_label(mat.tipologia),
                _nome_doc(mat.docente_1), _nome_doc(mat.docente_2)]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = THIN
            c.alignment = CENTER if col == 1 else WRAP
        return row + 1

    for data in sorted(esami_per_giorno.keys()):
        coppie = sorted(esami_per_giorno[data], key=lambda mc: (mc[1].cognome, mc[0].ora_inizio))
        titolo = f'{GIORNI[data.weekday()]} {data.strftime("%d/%m/%Y")}'
        row = _scrivi_header_giorno(titolo)
        for mat, cand in coppie:
            row = _scrivi_riga(row, mat, cand)
        ws.append([])

    if esami_senza_data:
        row = _scrivi_header_giorno('⚠︎ Da calendarizzare')
        for mat, cand in sorted(esami_senza_data, key=lambda mc: (mc[1].cognome, mc[0].materia)):
            row = _scrivi_riga(row, mat, cand, mostra_orario=False)
        ws.append([])

    if not esami_per_giorno and not esami_senza_data:
        ws.append(['Nessun candidato inserito.'])

    larghezze = [13, 22, 12, 22, 18, 12, 20, 20]
    for i, w in enumerate(larghezze, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'calendario_esami_integrativi_{ANNO}.xlsx',
    )
