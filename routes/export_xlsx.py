"""
Export XLSX per ogni passo dell'hub impostazione anno + export per classe.
Route: /export/<passo>?anno=2026-2027
       /export/classe/<label>?anno=2026-2027
"""
import io
from flask import Blueprint, request, send_file
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

export_bp = Blueprint('export_xlsx', __name__)

# ── Stili comuni ────────────────────────────────────────────────────
BLU   = '1e3a5f'
BLU_L = 'dbe9f6'
GRIG  = 'f3f4f6'
VERD  = '166534'
VERD_L= 'dcfce7'
ROSS  = 'dc2626'
GIAL  = 'fef9c3'

def _hdr(ws, row, cols, color=BLU, font_color='FFFFFF', bold=True, size=10):
    fill = PatternFill('solid', fgColor=color)
    font = Font(bold=bold, color=font_color, size=size, name='Arial')
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row, c, val)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                    wrap_text=True)
    return row + 1

def _row(ws, row, vals, bold=False, bg=None, num_fmt=None):
    fill = PatternFill('solid', fgColor=bg) if bg else None
    for c, val in enumerate(vals, 1):
        cell = ws.cell(row, c, val)
        cell.font = Font(bold=bold, size=10, name='Arial')
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(vertical='center')
        if num_fmt:
            cell.number_format = num_fmt
    return row + 1

def _border_all(ws, min_row, max_row, min_col, max_col):
    thin = Side(style='thin', color='D1D5DB')
    for r in range(min_row, max_row+1):
        for c in range(min_col, max_col+1):
            cell = ws.cell(r, c)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def _title(ws, testo, anno):
    ws.cell(1, 1, testo).font = Font(bold=True, size=13, color=BLU, name='Arial')
    ws.cell(2, 1, f'Anno scolastico: {anno}').font = Font(size=10, color='6B7280', name='Arial')
    ws.row_dimensions[1].height = 20
    return 4  # prima riga dati

def _wb():
    wb = Workbook()
    wb.remove(wb.active)
    return wb

def _send(wb, nome):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=nome,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ══ PASSO 1 — Classi di concorso ════════════════════════════════════
def _export_p1(anno):
    from models.classe_concorso import ClasseConcorso
    wb = _wb()
    ws = wb.create_sheet('Classi di concorso')
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10

    r = _title(ws, 'Classi di concorso', anno)
    r = _hdr(ws, r, ['Codice', 'Nome', 'Tipo', 'Attiva'])
    for cc in ClasseConcorso.query.order_by(ClasseConcorso.codice).all():
        tipo = 'ITP (B)' if cc.codice.startswith('B-') else 'Titolare (A)'
        bg = GRIG if not cc.attiva else None
        r = _row(ws, r, [cc.codice, cc.nome, tipo, 'Sì' if cc.attiva else 'No'],
                 bg=bg)
    _border_all(ws, 4, r-1, 1, 4)
    return wb


# ══ PASSO 2 — Piano di studi ════════════════════════════════════════
def _export_p2(anno, indirizzo=None):
    from models.piano_studi import PianoStudi, ClasseSezione
    from models.classe_concorso import ClasseConcorso
    from config_anno import get_anno_corrente

    wb = _wb()
    indirizzi = [indirizzo] if indirizzo else sorted({
        p.indirizzo for p in PianoStudi.query.filter_by(anno_scol=anno).all()})

    for ind in indirizzi:
        ws = wb.create_sheet(ind)
        anni_attivi = sorted({
            s.anno_corso for s in ClasseSezione.query.filter_by(
                anno_scol=anno, indirizzo=ind, attiva=True).all()})
        if not anni_attivi:
            anni_attivi = list(range(1, 6))

        # Intestazione
        r = _title(ws, f'Piano di studi — {ind}', anno)
        hdrs = ['CC', 'Materia'] + [f'{ac}° anno\n(h/sett)' for ac in anni_attivi] + \
               [f'{ac}° anno\n(h/anno×33)' for ac in anni_attivi]
        r = _hdr(ws, r, hdrs, size=9)

        # Raggruppamento per CC
        cc_ids = sorted({p.id_classe_concorso for p in
                         PianoStudi.query.filter_by(anno_scol=anno, indirizzo=ind).all()})
        for cc_id in cc_ids:
            cc = ClasseConcorso.query.get(cc_id)
            righe = PianoStudi.query.filter_by(
                anno_scol=anno, indirizzo=ind,
                id_classe_concorso=cc_id).all()
            for p in righe:
                ore_sett = [p.ore_settimanali if p.anno_corso == ac else '' for ac in anni_attivi]
                ore_ann  = [p.ore_settimanali * 33 if p.anno_corso == ac else '' for ac in anni_attivi]
                bg = GIAL if p.compresenza else None
                r = _row(ws, r,
                         [cc.codice, p.nome_materia_locale] + ore_sett + ore_ann,
                         bg=bg)

        # Totali per anno
        ws.cell(r, 2, 'TOTALE ORE SETTIMANALI').font = Font(bold=True, name='Arial', size=10)
        ws.cell(r, 2).fill = PatternFill('solid', fgColor=BLU_L)
        for i, ac in enumerate(anni_attivi):
            col = 3 + i
            righe_ac = PianoStudi.query.filter_by(
                anno_scol=anno, indirizzo=ind, anno_corso=ac, compresenza=False).all()
            tot = sum(p.ore_settimanali for p in righe_ac)
            ws.cell(r, col, tot).font = Font(bold=True, name='Arial')
            ws.cell(r, col).fill = PatternFill('solid', fgColor=BLU_L)
            ws.cell(r, 3 + len(anni_attivi) + i, tot * 33).font = Font(bold=True, name='Arial')
            ws.cell(r, 3 + len(anni_attivi) + i).fill = PatternFill('solid', fgColor=BLU_L)

        # Larghezze colonne
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        for i in range(len(anni_attivi) * 2):
            ws.column_dimensions[get_column_letter(3 + i)].width = 11
        _border_all(ws, 4, r, 1, 2 + len(anni_attivi) * 2)
    return wb


# ══ PASSO 3 — Materie ↔ CC ══════════════════════════════════════════
def _export_p3(anno):
    from models.materia import Materia, Dipartimento
    from models.classe_concorso import MateriaClasseConcorso
    wb = _wb()
    ws = wb.create_sheet('Materie e CC')
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 25

    r = _title(ws, 'Materie ↔ Classi di concorso', anno)
    r = _hdr(ws, r, ['Sigla', 'Nome ministeriale', 'Nome breve', 'Alias', 'Classi di concorso'])
    for dip in Dipartimento.query.filter(Dipartimento.sigla != '—').order_by(Dipartimento.ordine).all():
        # Riga dipartimento
        ws.cell(r, 1, dip.nome).font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.cell(r, 1).fill = PatternFill('solid', fgColor='374151')
        r += 1
        for m in Materia.query.filter_by(id_dipartimento=dip.id).order_by(Materia.sigla).all():
            ccs = ', '.join(mcc.classe_concorso.codice for mcc in
                           MateriaClasseConcorso.query.filter_by(id_materia=m.id).all())
            r = _row(ws, r, [m.sigla, m.nome, m.nome_breve or '', m.alias or '', ccs])
    _border_all(ws, 4, r-1, 1, 5)
    return wb


# ══ PASSO 4 — Classi attive ═════════════════════════════════════════
def _export_p4(anno):
    from models.piano_studi import ClasseSezione
    from models.aula import Aula
    wb = _wb()
    ws = wb.create_sheet('Classi attive')
    for w, col in zip([12, 10, 8, 10, 25], 'ABCDE'):
        ws.column_dimensions[col].width = w

    r = _title(ws, 'Classi attive', anno)
    r = _hdr(ws, r, ['Indirizzo', 'Anno', 'Sezione', 'Classe', 'Aula'])
    sezioni = ClasseSezione.query.filter_by(anno_scol=anno, attiva=True).order_by(
        ClasseSezione.indirizzo, ClasseSezione.anno_corso, ClasseSezione.sezione).all()
    for s in sezioni:
        lbl = f'{s.anno_corso}{s.sezione} {s.indirizzo}'
        aula = Aula.query.filter_by(anno_scol=anno, classe=lbl).first()
        r = _row(ws, r, [s.indirizzo, s.anno_corso, s.sezione, lbl,
                          f'{aula.aula} — {aula.sede}' if aula else ''])
    _border_all(ws, 4, r-1, 1, 5)
    return wb


# ══ PASSO 5 — Calcolo organico ══════════════════════════════════════
def _export_p5(anno):
    from models.piano_studi import CalcoloOrganico
    from models.classe_concorso import ClasseConcorso
    wb = _wb()
    ws = wb.create_sheet('Calcolo organico')
    for w, col in zip([14, 40, 14, 12, 12, 12], 'ABCDEF'):
        ws.column_dimensions[col].width = w

    r = _title(ws, 'Calcolo organico richiesto', anno)
    r = _hdr(ws, r, ['CC', 'Nome', 'Tipo', 'Ore tot.', 'N. docenti', 'Ore residue'])
    for calc in CalcoloOrganico.query.filter_by(anno_scol=anno).join(
            ClasseConcorso, CalcoloOrganico.id_classe_concorso == ClasseConcorso.id).order_by(
            ClasseConcorso.codice).all():
        r = _row(ws, r, [calc.classe_concorso.codice, calc.classe_concorso.nome,
                          calc.tipo_calcolato or '', calc.ore_totali_calcolate or 0,
                          calc.n_coi_calcolato or 0, calc.ore_resto_calcolato or 0])
    _border_all(ws, 4, r-1, 1, 6)
    return wb


# ══ PASSO 6 — Organico USR ══════════════════════════════════════════
def _export_p6(anno):
    from models.classe_concorso import ClasseConcorso, CattedraOrganico
    wb = _wb()
    for tipo, label in [('diritto', 'Organico di diritto'), ('fatto', 'Organico di fatto')]:
        ws = wb.create_sheet(label)
        for w, col in zip([14, 35, 8, 8, 12, 8, 12], 'ABCDEFG'):
            ws.column_dimensions[col].width = w
        r = _title(ws, label, anno)
        r = _hdr(ws, r, ['CC', 'Nome', 'COI', 'COE', 'Ore residue', 'DOC', 'Note'])
        for cat in CattedraOrganico.query.filter_by(anno_scol=anno, tipo=tipo).join(
                ClasseConcorso, CattedraOrganico.id_classe_concorso==ClasseConcorso.id).order_by(
                ClasseConcorso.codice).all():
            r = _row(ws, r, [cat.classe_concorso.codice, cat.classe_concorso.nome,
                              cat.n_coi or 0, cat.n_coe or 0, cat.ore_residue or 0,
                              cat.n_docenti or 0, cat.note or ''])
        _border_all(ws, 4, r-1, 1, 7)
    return wb


# ══ PASSO 7 — Docenti per anno ══════════════════════════════════════
def _export_p7(anno):
    from routes.impostazione_anno import _docenti_per_anno
    wb = _wb()
    ws = wb.create_sheet('Docenti')
    for w, col in zip([18, 14, 14, 12, 8, 12, 14], 'ABCDEFG'):
        ws.column_dimensions[col].width = w

    r = _title(ws, 'Docenti per anno scolastico', anno)
    r = _hdr(ws, r, ['Cognome', 'Nome', 'Tipo contratto', 'Status',
                      'Ore max', 'CC principale', 'Scuola AP'])
    for d in _docenti_per_anno(anno):
        from models.classe_concorso import ClasseConcorso
        cc = ClasseConcorso.query.get(d.id_classe_concorso) if d.id_classe_concorso else None
        status_map = {'presente': 'Presente', 'ap_entrante': 'AP Entrante',
                      'ap_uscente': 'AP Uscente', 'aspettativa': 'Aspettativa'}
        bg = GIAL if d.status_presenza in ('ap_uscente', 'aspettativa') else None
        r = _row(ws, r, [d.cognome, d.nome, d.tipo_contratto or '',
                          status_map.get(d.status_presenza or 'presente', ''),
                          d.ore_max_effettive, cc.codice if cc else '', d.scuola_ap or ''],
                 bg=bg)
    _border_all(ws, 4, r-1, 1, 7)
    return wb


# ══ PASSO 8 — Docenti ↔ CC (con confronto USR) ══════════════════════
def _export_p8(anno):
    from models.docente import Docente
    from models.classe_concorso import (ClasseConcorso, DocenteClasseConcorso,
                                         CattedraOrganico)
    wb = _wb()
    ws = wb.create_sheet('Docenti CC')
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 8

    r = _title(ws, 'Docenti ↔ Classi di concorso', anno)
    r = _hdr(ws, r, ['Docente', 'CC', 'TI in app', 'DOC USR', 'Scarto'])
    for cc in ClasseConcorso.query.filter_by(attiva=True).order_by(ClasseConcorso.codice).all():
        docenti_cc = (Docente.query.join(DocenteClasseConcorso,
                       DocenteClasseConcorso.id_docente == Docente.id)
                      .filter(DocenteClasseConcorso.id_classe_concorso == cc.id,
                              Docente.attivo == True)
                      .order_by(Docente.cognome).all())
        if not docenti_cc:
            continue
        cat = CattedraOrganico.query.filter_by(
            anno_scol=anno, id_classe_concorso=cc.id, tipo='fatto').first() or \
              CattedraOrganico.query.filter_by(
            anno_scol=anno, id_classe_concorso=cc.id, tipo='diritto').first()
        n_usr = cat.n_docenti if cat else 0
        n_ti  = sum(1 for d in docenti_cc if d.tipo_contratto == 'TI')
        scarto = n_ti - n_usr
        bg = VERD_L if scarto == 0 else GIAL if abs(scarto) == 1 else 'fde8e8'
        for i, d in enumerate(docenti_cc):
            r = _row(ws, r,
                     [f'{d.cognome} {d.nome}', cc.codice,
                      n_ti if i == 0 else '', n_usr if i == 0 else '',
                      scarto if i == 0 else ''],
                     bg=bg)
    _border_all(ws, 4, r-1, 1, 5)
    return wb


# ══ PASSO 9 — Assegnazioni classi (stile "ASSEGNAZIONI CLASSI" ufficiale) ══
#
# Replica lo stile del file di riferimento fornito dall'utente:
# - un foglio "organico compless." con TUTTE le classi di concorso,
#   colonna RICHIESTA (somma ore) e colonna titolari (docenti assegnati)
# - un foglio per ogni area disciplinare (stessa griglia classi, senza
#   RICHIESTA/titolari), stessi nomi/ordine delle aree già usati altrove
# - colonne raggruppate per indirizzo con gli stessi colori del file
#   originale; solo le classi ATTIVE vengono incluse
# - aggiunta la colonna POT (ore di potenziamento) subito dopo la colonna
#   nome-materia, non presente nel file originale ma richiesta esplicitamente

GRUPPI_INDIRIZZO_P9 = [
    {'label': 'AFM-RIM',     'indirizzi': ['AFM', 'RIM'], 'color': 'CCECFF'},
    {'label': 'CAT',         'indirizzi': ['CAT'],        'color': 'FFFFCC'},
    {'label': 'LSU',         'indirizzi': ['LSU'],        'color': 'FFCC99'},
    {'label': 'LSC',         'indirizzi': ['LSC'],        'color': 'CCCCCC'},
    {'label': 'LINGUISTICO', 'indirizzi': ['LLI'],        'color': 'CCFFCC'},
    {'label': 'SPORTIVO',    'indirizzi': ['LSP'],        'color': 'FFCCFF'},
]
POT_COLOR_P9 = 'E5D4F5'


def _p9_anno_corto(anno):
    try:
        a1, a2 = anno.split('-')
        return f'{a1}/{a2[-2:]}'
    except Exception:
        return anno


def _p9_classi_gruppi(anno):
    """[(gruppo, [label_classe,...]), ...] con solo classi ATTIVE per l'anno."""
    from models.piano_studi import ClasseSezione
    out = []
    for gruppo in GRUPPI_INDIRIZZO_P9:
        secs = (ClasseSezione.query
                .filter(ClasseSezione.anno_scol == anno,
                        ClasseSezione.indirizzo.in_(gruppo['indirizzi']),
                        ClasseSezione.attiva == True)
                .all())
        ordine_ind = {ind: i for i, ind in enumerate(gruppo['indirizzi'])}
        secs.sort(key=lambda s: (ordine_ind.get(s.indirizzo, 99), s.anno_corso, s.sezione))
        labels = [f'{s.anno_corso}{s.sezione} {s.indirizzo}' for s in secs]
        out.append((gruppo, labels))
    return out


def _p9_scrivi_intestazione(ws, anno, gruppi_classi, con_richiesta, start_row=1):
    """
    Scrive due righe di intestazione a partire da start_row (titolo anno,
    gruppi indirizzo colorati, nomi classi, colonna POT). Ritorna:
    label_col, label_color, ultima_classe_col, col_pot, col_richiesta,
    col_titolari.
    """
    thin = Side(style='thin', color='FF000000')
    box  = Border(left=thin, right=thin, top=thin, bottom=thin)
    r1, r2 = start_row, start_row + 1

    ws.cell(r1, 1, f'AS {_p9_anno_corto(anno)}').font = Font(bold=True, size=18, name='Calibri')

    col = 2
    col_pot = col
    for rr in (r1, r2):
        c = ws.cell(rr, col_pot, 'POT' if rr == r1 else 'ore pot.')
        c.font = Font(bold=True, size=8, name='Calibri')
        c.fill = PatternFill('solid', fgColor=POT_COLOR_P9)
        c.alignment = Alignment(horizontal='center')
        c.border = box
    col += 1

    label_col = {}
    label_color = {}
    for gruppo, labels in gruppi_classi:
        if not labels:
            continue
        start = col
        for lbl in labels:
            c = ws.cell(r2, col, lbl)
            c.font = Font(bold=True, size=8, name='Calibri')
            c.fill = PatternFill('solid', fgColor=gruppo['color'])
            c.alignment = Alignment(horizontal='center')
            c.border = box
            label_col[lbl] = col
            label_color[lbl] = gruppo['color']
            col += 1
        end = col - 1
        h = ws.cell(r1, start, gruppo['label'])
        h.font = Font(bold=True, size=8, name='Calibri')
        h.fill = PatternFill('solid', fgColor=gruppo['color'])
        h.alignment = Alignment(horizontal='center')
        h.border = box
        if end > start:
            ws.merge_cells(start_row=r1, start_column=start, end_row=r1, end_column=end)
    ultima_classe_col = col - 1

    col_richiesta = col_titolari = None
    if con_richiesta:
        col_richiesta = col
        ws.cell(r1, col, 'RICHIESTA').font = Font(bold=True, size=8, name='Calibri')
        col += 1
        col_titolari = col
        ws.cell(r1, col, 'titolari').font = Font(bold=True, size=8, name='Calibri')

    if start_row == 1:
        ws.freeze_panes = ws.cell(3, col_pot).coordinate
    return label_col, label_color, ultima_classe_col, col_pot, col_richiesta, col_titolari


def _p9_scrivi_intestazione2(ws, start_row, anno, gruppi_classi):
    """Seconda intestazione (griglia assegnazioni), stessa struttura, più in basso."""
    return _p9_scrivi_intestazione(ws, anno, gruppi_classi, con_richiesta=False,
                                    start_row=start_row)


def _p9_scrivi_blocco_cc(ws, r, anno, cc, label_col, label_color, ultima_classe_col,
                          col_pot, col_richiesta, col_titolari):
    """Scrive il blocco di una classe di concorso: titolo + una riga per materia."""
    from models.piano_studi import PianoStudi, ClasseSezione
    from models.assegnazione import CattedraPotenziamento, AssegnazioneDocente
    from collections import OrderedDict
    from openpyxl.utils import get_column_letter as gcl

    thin = Side(style='thin', color='FF000000')
    box  = Border(left=thin, right=thin, top=thin, bottom=thin)

    materie = OrderedDict()  # nome_materia -> {label_classe: ore}
    for p in PianoStudi.query.filter_by(anno_scol=anno, id_classe_concorso=cc.id,
                                          compresenza=False).all():
        secs = ClasseSezione.query.filter_by(anno_scol=anno, indirizzo=p.indirizzo,
                                              anno_corso=p.anno_corso, attiva=True).all()
        for s in secs:
            lbl = f'{p.anno_corso}{s.sezione} {p.indirizzo}'
            if lbl not in label_col:
                continue
            materie.setdefault(p.nome_materia_locale, {})
            materie[p.nome_materia_locale][lbl] = (
                materie[p.nome_materia_locale].get(lbl, 0) + p.ore_settimanali)

    pot = CattedraPotenziamento.query.filter_by(anno_scol=anno, id_classe_concorso=cc.id).first()
    if not materie and not pot:
        return r  # nessun dato per questa CC nell'anno: salta il blocco

    nome_cc = f'{cc.codice} {cc.nome}'.upper()
    c1 = ws.cell(r, 1, nome_cc)
    c1.font = Font(bold=True, size=10, name='Calibri')
    c1.border = box
    if ultima_classe_col >= 2:
        c2 = ws.cell(r, 2, nome_cc)
        c2.font = Font(bold=True, size=10, name='Calibri')
        c2.border = box
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=ultima_classe_col)
    r += 1

    if pot:
        ws.cell(r, 1, 'Potenziamento').font = Font(size=8, name='Calibri')
        ws.cell(r, 1).border = box
        pcell = ws.cell(r, col_pot, pot.ore)
        pcell.font = Font(bold=True, size=8, name='Calibri', color='FF000000')
        pcell.fill = PatternFill('solid', fgColor=POT_COLOR_P9)
        pcell.alignment = Alignment(horizontal='center')
        pcell.border = box
        for lbl, col in label_col.items():
            e = ws.cell(r, col, '-')
            e.font = Font(size=8, name='Calibri')
            e.fill = PatternFill('solid', fgColor=label_color[lbl])
            e.alignment = Alignment(horizontal='center')
            e.border = box
        r += 1

    riga_start_materie = r
    for nome_mat, ore_per_classe in materie.items():
        ws.cell(r, 1, nome_mat).font = Font(size=8, name='Calibri')
        ws.cell(r, 1).border = box
        pc = ws.cell(r, col_pot, '-')
        pc.font = Font(size=8, name='Calibri')
        pc.fill = PatternFill('solid', fgColor=POT_COLOR_P9)
        pc.alignment = Alignment(horizontal='center')
        pc.border = box
        for lbl, col in label_col.items():
            ore = ore_per_classe.get(lbl)
            cell = ws.cell(r, col, ore if ore else '-')
            cell.font = Font(bold=True, size=8, name='Calibri',
                              color='FF000000')
            cell.fill = PatternFill('solid', fgColor=label_color[lbl])
            cell.alignment = Alignment(horizontal='center')
            cell.border = box
        r += 1
    riga_end_materie = r - 1
    if riga_end_materie < riga_start_materie:
        riga_end_materie = riga_start_materie

    if col_richiesta and materie:
        col_last = gcl(ultima_classe_col)
        formula = f'=SUM(B{riga_start_materie}:{col_last}{riga_end_materie})'
        rc = ws.cell(riga_end_materie, col_richiesta, formula)
        rc.font = Font(size=8, name='Calibri')

    if col_titolari:
        assegnazioni = AssegnazioneDocente.query.filter_by(
            anno_scol=anno, id_classe_concorso=cc.id).all()
        riga_max = max(riga_end_materie, riga_start_materie)
        for i, a in enumerate(assegnazioni):
            rr = riga_start_materie + i
            if rr > riga_max:
                break  # blocco pieno: evita di scrivere sopra il blocco successivo
            nome = a.display_name
            ore_doc = a.docente.ore_max_effettive if a.docente else ''
            tc = ws.cell(rr, col_titolari, f'{nome} {ore_doc}'.strip())
            tc.font = Font(size=8, name='Calibri')

    return r + 1


def _p9_scrivi_blocco_cc_assegnazioni(ws, r, anno, cc, label_col, label_color,
                                        ultima_classe_col, col_pot):
    """
    Seconda griglia (sotto quella di RICHIESTA): una riga per DOCENTE con le
    ore effettivamente assegnate per classe, invece che una riga per materia.
    Aggiunge una riga 'ORE RESIDUE' quando un docente non copre l'intero
    monte ore contrattuale, per segnalare la necessità di un supplente.
    """
    from models.assegnazione import AssegnazioneDocente

    thin = Side(style='thin', color='FF000000')
    box  = Border(left=thin, right=thin, top=thin, bottom=thin)

    assegnazioni = AssegnazioneDocente.query.filter_by(
        anno_scol=anno, id_classe_concorso=cc.id).all()
    if not assegnazioni:
        return r

    nome_cc = f'{cc.codice} {cc.nome}'.upper()
    c1 = ws.cell(r, 1, nome_cc)
    c1.font = Font(bold=True, size=10, name='Calibri')
    c1.border = box
    if ultima_classe_col >= 2:
        c2 = ws.cell(r, 2, nome_cc)
        c2.font = Font(bold=True, size=10, name='Calibri')
        c2.border = box
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=ultima_classe_col)
    r += 1

    supplente_n = 0
    for a in assegnazioni:
        ore_cl = {}
        pot_ore = 0
        for ac in a.classi:
            if ac.indirizzo == 'POT':
                pot_ore += ac.ore
                continue
            lbl = ac.label_classe
            ore_cl[lbl] = ore_cl.get(lbl, 0) + ac.ore

        max_ore = a.docente.ore_max_effettive if a.docente else None
        nome = a.display_name
        etichetta = f'{nome} {int(max_ore)} ore' if max_ore else nome
        ws.cell(r, 1, etichetta).font = Font(size=8, name='Calibri')
        ws.cell(r, 1).border = box

        pc = ws.cell(r, col_pot, pot_ore if pot_ore else '-')
        pc.font = Font(bold=True, size=8, name='Calibri', color='FF000000')
        pc.fill = PatternFill('solid', fgColor=POT_COLOR_P9)
        pc.alignment = Alignment(horizontal='center')
        pc.border = box

        for lbl, col in label_col.items():
            ore = ore_cl.get(lbl)
            cell = ws.cell(r, col, ore if ore else '-')
            cell.font = Font(bold=True, size=8, name='Calibri', color='FF000000')
            cell.fill = PatternFill('solid', fgColor=label_color[lbl])
            cell.alignment = Alignment(horizontal='center')
            cell.border = box
        r += 1

        if max_ore:
            residuo = max_ore - sum(ore_cl.values()) - pot_ore
            if residuo > 0:
                supplente_n += 1
                rc = ws.cell(r, 1, f'ORE RESIDUE {residuo:g} - SUPPLENTE {supplente_n}')
                rc.font = Font(size=8, name='Calibri', bold=True, color='FFDC2626')
                rc.border = box
                for lbl, col in label_col.items():
                    e = ws.cell(r, col, '-')
                    e.font = Font(size=8, name='Calibri')
                    e.fill = PatternFill('solid', fgColor=label_color[lbl])
                    e.alignment = Alignment(horizontal='center')
                    e.border = box
                epc = ws.cell(r, col_pot, '-')
                epc.fill = PatternFill('solid', fgColor=POT_COLOR_P9)
                epc.border = box
                r += 1

    return r + 1


def _p9_imposta_colonne(ws, col_pot, ultima_classe_col):
    ws.column_dimensions['A'].width = 35.14
    ws.column_dimensions[get_column_letter(col_pot)].width = 7
    for c in range(col_pot + 1, ultima_classe_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 5.3
    ws.sheet_view.showGridLines = False


def _export_p9(anno):
    from routes.assegnazioni import AREE
    from models.classe_concorso import ClasseConcorso
    wb = _wb()
    gruppi_classi = _p9_classi_gruppi(anno)

    # ── Foglio principale: tutte le CC, con RICHIESTA + titolari ──────
    ws_main = wb.create_sheet('organico compless.')
    label_col, label_color, ultima_col, col_pot, col_ric, col_tit = \
        _p9_scrivi_intestazione(ws_main, anno, gruppi_classi, con_richiesta=True)
    r = 3
    for area in AREE:
        for codice in area['cc']:
            cc = ClasseConcorso.query.filter_by(codice=codice).first()
            if not cc:
                continue
            r = _p9_scrivi_blocco_cc(ws_main, r, anno, cc, label_col, label_color,
                                       ultima_col, col_pot, col_ric, col_tit)

    # ── Seconda griglia: assegnazione per docente (sotto, stessa struttura) ──
    r += 1
    label_col2, label_color2, _, col_pot2, _, _ = \
        _p9_scrivi_intestazione2(ws_main, r, anno, gruppi_classi)
    r += 2
    for area in AREE:
        for codice in area['cc']:
            cc = ClasseConcorso.query.filter_by(codice=codice).first()
            if not cc:
                continue
            r = _p9_scrivi_blocco_cc_assegnazioni(ws_main, r, anno, cc, label_col2,
                                                    label_color2, ultima_col, col_pot2)
    _p9_imposta_colonne(ws_main, col_pot, ultima_col)

    # ── Un foglio per area disciplinare (senza RICHIESTA/titolari) ────
    for area in AREE:
        nome_foglio = area['nome'][:31].replace('/', '-').replace('*', '').replace('?', '')
        ws = wb.create_sheet(nome_foglio)
        label_col, label_color, ultima_col, col_pot, _, _ = \
            _p9_scrivi_intestazione(ws, anno, gruppi_classi, con_richiesta=False)
        r = 3
        for codice in area['cc']:
            cc = ClasseConcorso.query.filter_by(codice=codice).first()
            if not cc:
                continue
            r = _p9_scrivi_blocco_cc(ws, r, anno, cc, label_col, label_color,
                                       ultima_col, col_pot, None, None)

        r += 1
        label_col2, label_color2, _, col_pot2, _, _ = \
            _p9_scrivi_intestazione2(ws, r, anno, gruppi_classi)
        r += 2
        for codice in area['cc']:
            cc = ClasseConcorso.query.filter_by(codice=codice).first()
            if not cc:
                continue
            r = _p9_scrivi_blocco_cc_assegnazioni(ws, r, anno, cc, label_col2,
                                                    label_color2, ultima_col, col_pot2)
        _p9_imposta_colonne(ws, col_pot, ultima_col)


    return wb


# ══ PASSO 10 — Docenti ↔ Materie ════════════════════════════════════
def _export_p10(anno):
    from models.materia import DocenteMateria
    from models.docente import Docente
    from models.materia import Materia
    wb = _wb()
    ws = wb.create_sheet('Docenti Materie')
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 35

    r = _title(ws, 'Docenti ↔ Materie', anno)
    r = _hdr(ws, r, ['Docente', 'Tipo contratto', 'Materie assegnate'])
    docenti = Docente.query.filter_by(attivo=True).order_by(Docente.cognome).all()
    for d in docenti:
        materie = DocenteMateria.query.filter_by(
            id_docente=d.id, anno_scol=anno).all()
        if not materie:
            continue
        nomi = ', '.join(m.materia.nome_breve or m.materia.nome for m in materie)
        r = _row(ws, r, [f'{d.cognome} {d.nome}', d.tipo_contratto or '', nomi])
    _border_all(ws, 4, r-1, 1, 3)
    return wb


# ══ EXPORT CLASSE — docenti + materie + incarichi ═══════════════════
#
# _export_classe() (un file per una classe) e _aggiungi_foglio_classe()
# (un foglio per classe dentro il file "tutte le classi") disegnavano
# esattamente la stessa tabella con una piccola duplicazione di codice.
# Consolidato in _riempi_foglio_classe(), che riempie un foglio già
# creato: le due funzioni pubbliche ora si limitano a creare il workbook
# (o il foglio) e a delegarle il contenuto.
def _riempi_foglio_classe(ws, anno, label_classe):
    """
    Riempie il foglio ws con lo schema di una classe (es. '1A AFM'):
    docenti assegnati con materie/ore e incarichi di classe.
    Ritorna False (foglio lasciato vuoto) se label_classe non è nel
    formato atteso "<anno><sezione> <indirizzo>".
    """
    import re
    m = re.match(r'(\d+)([AB]?)\s+(.+)', label_classe)
    if not m:
        return False
    anno_corso = int(m.group(1))
    sezione    = m.group(2) or 'A'
    indirizzo  = m.group(3).strip()

    from models.assegnazione import AssegnazioneDocente, AssegnazioneClasse
    from models.incarico import IncaricaDocente
    from models.materia import Materia
    from collections import defaultdict

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 24

    r = _title(ws, f'Classe {anno_corso}{sezione} {indirizzo}', anno)
    r = _hdr(ws, r, ['Docente', 'Materia', 'Ore', 'Incarico nella classe'])

    # Assegnazioni docenti per questa classe
    assegnazioni_classe = AssegnazioneClasse.query.filter_by(
        indirizzo=indirizzo, anno_corso=anno_corso, sezione=sezione).join(
        AssegnazioneDocente, AssegnazioneClasse.id_assegnazione == AssegnazioneDocente.id).filter(
        AssegnazioneDocente.anno_scol == anno).all()

    # Raggruppa per docente
    doc_map = defaultdict(list)
    for ac in assegnazioni_classe:
        a = ac.assegnazione
        mat = Materia.query.get(ac.id_materia) if ac.id_materia else None
        nome_mat = (mat.nome_breve or mat.nome) if mat else (a.tipo if a.tipo else '—')
        doc_map[a].append((nome_mat, ac.ore))

    # Incarichi per questa classe (interrogati una sola volta, riusati
    # sia per la mappa id_docente->incarico sia per la sezione dedicata)
    incarichi_classe = IncaricaDocente.query.filter_by(
        anno_scol=anno,
        indirizzo=indirizzo,
        anno_corso=anno_corso,
        sezione=sezione).all()
    incarichi_map = {inc.id_docente: inc.tipo.nome for inc in incarichi_classe}

    if doc_map:
        for a, materie_ore in sorted(doc_map.items(),
                                      key=lambda x: x[0].docente.cognome if x[0].docente else 'zzz'):
            nome_doc = a.display_name
            incarico = incarichi_map.get(a.id_docente, '')
            for i, (mat, ore) in enumerate(materie_ore):
                r = _row(ws, r, [nome_doc if i == 0 else '', mat, ore,
                                  incarico if i == 0 else ''])
    else:
        ws.cell(r, 1, 'Nessuna assegnazione presente per questa classe.').font = \
            Font(italic=True, color='9CA3AF', name='Arial')
        r += 1

    # Incarichi senza docente assegnato (es. coordinatore non ancora inserito)
    if incarichi_classe:
        r += 1
        ws.cell(r, 1, 'Incarichi di classe').font = Font(bold=True, color=BLU, name='Arial', size=10)
        r += 1
        r = _hdr(ws, r, ['Tipo incarico', 'Docente', '', ''], color='374151')
        for inc in incarichi_classe:
            r = _row(ws, r, [inc.tipo.nome,
                              f'{inc.docente.cognome} {inc.docente.nome}', '', ''])
    _border_all(ws, 4, r-1, 1, 4)
    return True


def _export_classe(anno, label_classe):
    """Per una classe (es. '1A AFM'): file a sé stante con docenti/materie/incarichi."""
    wb = _wb()
    sheet_name = label_classe[:31].replace('/', '-')
    ws = wb.create_sheet(sheet_name)
    if not _riempi_foglio_classe(ws, anno, label_classe):
        from flask import abort
        abort(400)
    return wb


# ══ ROUTE DISPATCHER ════════════════════════════════════════════════
@export_bp.route('/export/<passo>')
def export_passo(passo):
    from config_anno import get_anno_corrente
    anno = request.args.get('anno', get_anno_corrente())
    ind  = request.args.get('indirizzo')

    mapping = {
        'p1': (_export_p1,  f'classi_concorso_{anno}.xlsx',      lambda: _export_p1(anno)),
        'p2': (_export_p2,  f'piano_studi_{anno}.xlsx',           lambda: _export_p2(anno, ind)),
        'p3': (_export_p3,  f'materie_cc_{anno}.xlsx',            lambda: _export_p3(anno)),
        'p4': (_export_p4,  f'classi_attive_{anno}.xlsx',         lambda: _export_p4(anno)),
        'p5': (_export_p5,  f'calcolo_organico_{anno}.xlsx',      lambda: _export_p5(anno)),
        'p6': (_export_p6,  f'organico_usr_{anno}.xlsx',          lambda: _export_p6(anno)),
        'p7': (_export_p7,  f'docenti_{anno}.xlsx',               lambda: _export_p7(anno)),
        'p8': (_export_p8,  f'docenti_cc_{anno}.xlsx',            lambda: _export_p8(anno)),
        'p9': (_export_p9,  f'assegnazioni_{anno}.xlsx',          lambda: _export_p9(anno)),
        'p10':(_export_p10, f'docenti_materie_{anno}.xlsx',       lambda: _export_p10(anno)),
    }
    if passo not in mapping:
        from flask import abort
        abort(404)
    _, nome, fn = mapping[passo]
    wb = fn()
    return _send(wb, nome)


@export_bp.route('/export/classe/<path:label>')
def export_classe(label):
    from config_anno import get_anno_corrente
    anno = request.args.get('anno', get_anno_corrente())
    wb = _export_classe(anno, label)
    nome = f'classe_{label.replace(" ", "_")}_{anno}.xlsx'
    return _send(wb, nome)


# ══ EXPORT TUTTE LE CLASSI in un unico file ══════════════════════════

def _aggiungi_foglio_classe(wb, anno, label_classe):
    """Aggiunge un foglio classe al workbook esistente (usato da 'tutte le classi')."""
    sheet_name = label_classe[:31].replace('/', '-')
    ws = wb.create_sheet(sheet_name)
    _riempi_foglio_classe(ws, anno, label_classe)


@export_bp.route('/export/tutte-classi')
def tutte_classi():
    from config_anno import get_anno_corrente
    from models.piano_studi import ClasseSezione
    anno = request.args.get('anno', get_anno_corrente())

    sezioni = ClasseSezione.query.filter_by(
        anno_scol=anno, attiva=True).order_by(
        ClasseSezione.indirizzo,
        ClasseSezione.anno_corso,
        ClasseSezione.sezione).all()

    wb = _wb()
    for s in sezioni:
        label = f'{s.anno_corso}{s.sezione} {s.indirizzo}'
        _aggiungi_foglio_classe(wb, anno, label)

    if not wb.sheetnames:
        ws = wb.create_sheet('Nessuna classe')
        ws.cell(1, 1, 'Nessuna classe attiva trovata.')

    return _send(wb, f'schede_classi_{anno}.xlsx')
