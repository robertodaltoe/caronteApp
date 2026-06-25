from flask import Blueprint, render_template, request, redirect, url_for, flash, make_response
from models import db
from models.recupero import RecuperoDocente, RecuperoGruppo, RecuperoLezione, RecuperoAlunno, RecuperoVincolo
from models.docente import Docente
from models.materia import Materia, DocenteMateria
from datetime import date, timedelta

# Costanti e helper condivisi: vivono in recupero_costanti.py (foglia
# senza dipendenze da questo file), per evitare import circolari con i
# sotto-moduli (recupero_export, recupero_agosto) che le importano a
# loro volta dalla stessa fonte.
from routes.recupero_costanti import (
    ANNO, DATA_INIZIO, DATA_FINE,
    ANNO_AGO, PERIODO_AGO, CONTRATTI_OK, TIPO_PROVA_LABEL,
    _FAMIGLIE_MATERIE, _materia_canonica, _norm_materia,
    _split_cognome_nome, _parse_tipo_prova,
)

recupero_bp = Blueprint('recupero', __name__)


# ── INDICE GENERALE (staging condiviso giugno+agosto) ─────────────────
@recupero_bp.route('/recupero')
def index():
    """
    Home del modulo recupero: import unico del file Excel (condiviso tra
    corsi di giugno e prove di agosto), riepilogo alunni con giudizio
    sospeso, e i due percorsi operativi separati.
    """
    from models.recupero import RecuperoImport

    imports = RecuperoImport.query.filter_by(anno_scol=ANNO).all()
    tot_alunni_import = len(imports)

    conteggi = {'aderisce':0,'studio_ind':0,'non_risposto':0,'non_aderisce':0,'sconosciuto':0}
    for imp in imports:
        conteggi[imp.stato_adesione] = conteggi.get(imp.stato_adesione, 0) + 1

    n_materie = len({imp.materia_norm for imp in imports})

    return render_template('recupero/index.html',
        tot_alunni_import=tot_alunni_import,
        conteggi=conteggi,
        n_materie=n_materie,
        anno=ANNO,
    )


# ── INDICE CORSI DI GIUGNO ──────────────────────────────────────────────
@recupero_bp.route('/recupero/giugno')
def giugno_index():
    docenti_disp = (RecuperoDocente.query
                    .filter_by(anno_scol=ANNO)
                    .join(Docente)
                    .order_by(Docente.cognome)
                    .all())
    # Solo i gruppi dei corsi di recupero di giugno — esclude le prove di
    # agosto, che hanno periodo_codice='prove_agosto' e vanno conteggiate
    # separatamente nella loro pagina dedicata.
    gruppi = (RecuperoGruppo.query
              .join(RecuperoDocente)
              .filter(RecuperoDocente.anno_scol == ANNO,
                      RecuperoGruppo.periodo_codice == 'corsi_giugno')
              .order_by(RecuperoGruppo.materia)
              .all())
    # Statistiche
    tot_ore = sum(g.ore_pianificate for g in gruppi)
    tot_alunni = sum(g.n_alunni or 0 for g in gruppi)

    return render_template('recupero/giugno_index.html',
        docenti_disp=docenti_disp,
        gruppi=gruppi,
        tot_ore=tot_ore,
        tot_alunni=tot_alunni,
        anno=ANNO,
    )


# ── VERIFICA COPERTURA ────────────────────────────────────────────────
def _export_copertura_xlsx(righe, titolo):
    """
    Foglio firme: una riga per ogni LEZIONE pianificata per ogni studente,
    raggruppato per materia. Uno studente con 3 lezioni di Matematica
    pianificate avrà 3 righe (3 caselle firma), una per ciascuna data.
    Gli studenti senza gruppo (no_gruppo, non_iscritto, no_corso, non_aderisce)
    restano con una sola riga, perché non hanno lezioni a cui riferirsi.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, PatternFill)
    from openpyxl.utils import get_column_letter
    from collections import defaultdict
    import io

    GIORNI = ['Lun','Mar','Mer','Gio','Ven','Sab','Dom']

    wb = Workbook()
    ws = wb.active
    ws.title = 'Foglio firme'

    # Stili
    BLU     = '1e3a5f'
    BLU_SEZ = '2F4F8C'
    VERDE   = 'dcfce7'; VERDE_T  = '166534'
    ROSSO   = 'fee2e2'; ROSSO_T  = 'dc2626'
    GIALLO  = 'fef9c3'; GIALLO_T = '92400e'
    GRIGIO  = 'f3f4f6'
    GRIGIO2 = 'f3f4f6'; GRIGIO2_T = '6b7280'

    def fill(hex_bg):
        return PatternFill('solid', start_color=hex_bg, fgColor=hex_bg)
    def border():
        s = Side(style='thin', color='d1d5db')
        return Border(left=s, right=s, top=s, bottom=s)
    def hdr_font():
        return Font(bold=True, color='FFFFFF', name='Arial', size=9)
    def cell_font(bold=False, color='000000'):
        return Font(bold=bold, color=color, name='Arial', size=9)

    STATO_CFG = {
        'ok':           (VERDE,   VERDE_T,   '✓ ok'),
        'no_gruppo':    (ROSSO,   ROSSO_T,   '✗ no gruppo'),
        'non_iscritto': (ROSSO,   ROSSO_T,   '❓ non iscritto'),
        'no_corso':     (GIALLO,  GIALLO_T,  '📚 no corso'),
        'non_aderisce': (GRIGIO2, GRIGIO2_T, '✗ non aderisce'),
    }

    # Riga titolo generale
    ws.merge_cells('A1:H1')
    ws['A1'] = titolo or 'Verifica copertura recuperi — foglio firme'
    ws['A1'].font = Font(bold=True, color='FFFFFF', name='Arial', size=12)
    ws['A1'].fill = fill(BLU)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    hdrs = ['Data lezione', 'Classe', 'Cognome', 'Nome', 'Docente assegnante',
            'Stato', 'Firma']
    widths = [16, 9, 18, 16, 28, 14, 22]

    # Raggruppa le righe per materia (usa il nome del gruppo se presente,
    # altrimenti la materia grezza dello studente)
    per_materia = defaultdict(list)
    for r in righe:
        if r.get('gruppo'):
            chiave = r['gruppo'].materia
        else:
            chiave = r['imp'].materia_raw or r['imp'].materia_norm or 'Senza materia'
        per_materia[chiave].append(r)

    row = 3
    tot_firme_generale = 0

    for materia in sorted(per_materia.keys()):
        righe_mat = per_materia[materia]

        # Header sezione materia
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = materia.upper()
        ws[f'A{row}'].font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        ws[f'A{row}'].fill = fill(BLU_SEZ)
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 18
        row += 1

        # Intestazioni colonne per questa sezione
        for col, h in enumerate(hdrs, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = hdr_font()
            cell.fill = fill(BLU)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border()
        ws.row_dimensions[row].height = 16
        row += 1

        n_firme_materia = 0
        i_riga_colore = 0

        for r in righe_mat:
            stato = r.get('stato', '')
            bg, fg, label = STATO_CFG.get(stato, (GRIGIO, '374151', stato))
            gruppo = r.get('gruppo')

            # Lezioni pianificate per il gruppo di questo studente (se presente)
            lezioni = sorted(gruppo.lezioni, key=lambda l: (l.data, l.ora_inizio)) if gruppo else []

            if lezioni:
                righe_da_scrivere = [
                    f'{GIORNI[l.data.weekday()]} {l.data.strftime("%d/%m/%Y")}  {l.ora_inizio}-{l.ora_fine}'
                    for l in lezioni
                ]
            else:
                # Nessuna lezione pianificata (o nessun gruppo): una sola riga senza data
                righe_da_scrivere = ['—']

            for data_str in righe_da_scrivere:
                vals = [
                    data_str,
                    r['imp'].classe,
                    r['imp'].cognome,
                    r['imp'].nome,
                    r['imp'].docente_raw[:40] if r['imp'].docente_raw else '',
                    label,
                    '',  # Firma
                ]
                row_fill = fill(bg) if stato != 'ok' else None
                for col, val in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = border()
                    cell.alignment = Alignment(vertical='center', wrap_text=(col == 5))
                    if col == 6:  # Stato
                        cell.font = Font(bold=True, color=fg, name='Arial', size=9)
                        cell.fill = fill(bg)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif col == 7:  # Firma
                        cell.fill = fill('FFFFFF')
                    else:
                        cell.font = cell_font(bold=(col == 3))
                        if row_fill:
                            cell.fill = fill(GRIGIO)
                        elif i_riga_colore % 2 == 1:
                            cell.fill = fill('F4F7FC')
                ws.row_dimensions[row].height = 15
                row += 1
                n_firme_materia += 1
                i_riga_colore += 1

        # Riga totale firme per materia
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = f'TOTALE FIRME — {materia.upper()}'
        ws[f'A{row}'].font = Font(bold=True, color='FFFFFF', name='Arial', size=9)
        ws[f'A{row}'].fill = fill('1F3864')
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        cell_tot = ws.cell(row=row, column=7, value=n_firme_materia)
        cell_tot.font = Font(bold=True, color='FFFFFF', name='Arial', size=9)
        cell_tot.fill = fill('1F3864')
        cell_tot.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 16
        row += 2  # riga vuota tra sezioni

        tot_firme_generale += n_firme_materia

    # Larghezze colonne
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@recupero_bp.route('/recupero/copertura')
def copertura():
    """
    Verifica copertura UNIFICATA: per ogni alunno+materia mostra sia lo
    stato del corso di recupero di giugno (dipende dallo stato_adesione:
    chi non aderisce non viene conteggiato come "da seguire" a giugno)
    sia lo stato della prova di agosto (sempre rilevante: anche chi non
    ha aderito al corso, o ha scelto studio individuale, deve comunque
    sostenere la prova se il debito non risulta sanato).
    """
    from models.recupero import RecuperoImport
    from collections import defaultdict
    from flask import send_file

    imports = RecuperoImport.query.filter_by(anno_scol=ANNO).order_by(
        RecuperoImport.cognome, RecuperoImport.nome,
        RecuperoImport.materia_norm).all()

    # Famiglie sinonimi per copertura (condivise tra giugno e agosto)
    _FAM_COV = [
        {'LATINO', 'LINGUA LATINA', 'LINGUA E CULTURA LATINA'},
        {'ITALIANO', 'LINGUA E LETTERATURA ITALIANA', 'LINGUA E CULTURA ITALIANA'},
        {'MATEMATICA', 'MATEMATICA CON INFORMATICA', 'MATEMATICA E COMPLEMENTI DI MATEMATICA'},
        {'INFORMATICA', 'TECNOLOGIE INFORMATICHE'},
        {'STORIA', 'STORIA E GEOGRAFIA'},
        {'FISICA', 'SCIENZE INTEGRATE (FISICA)'},
        {'INGLESE', 'LINGUA INGLESE', 'LINGUA E CULTURA STRANIERA (INGLESE)'},
        {'TEDESCO', 'LINGUA TEDESCA', 'LINGUA E CULTURA STRANIERA TEDESCO'},
    ]
    def _match_cov(m1, m2):
        m1u,m2u = m1.strip().upper(), m2.strip().upper()
        if m1u == m2u: return True
        for f in _FAM_COV:
            fs = {x.upper() for x in f}
            if m1u in fs and m2u in fs: return True
        return False

    def _trova_gruppo(imp, gruppi_pool):
        for g in gruppi_pool:
            if not _match_cov(g.materia, imp.materia_norm or ''): continue
            classi_g = {cl.strip().upper() for cl in g.classi.split(',')}
            if imp.classe.upper() in classi_g:
                return g
        return None

    if not imports:
        return render_template('recupero/copertura.html',
            righe=[], n_ok=0, n_no_corso=0, n_no_iscritto=0, n_no_gruppo=0,
            n_non_aderisce=0)

    gruppi_giugno = (RecuperoGruppo.query.join(RecuperoDocente)
                     .filter(RecuperoDocente.anno_scol == ANNO,
                             RecuperoGruppo.periodo_codice == 'corsi_giugno').all())
    gruppi_agosto = (RecuperoGruppo.query.join(RecuperoDocente)
                     .filter(RecuperoDocente.anno_scol == ANNO_AGO,
                             RecuperoGruppo.periodo_codice == PERIODO_AGO).all())

    righe = []
    n_ok = n_no_corso = n_no_iscritto = n_no_gruppo = n_non_aderisce = 0

    for imp in imports:
        adesione = imp.stato_adesione  # aderisce|studio_ind|non_risposto|non_aderisce|sconosciuto

        # ── Stato GIUGNO: dipende dall'adesione (chi non aderisce non
        # viene seguito a giugno, è una scelta legittima per il corso) ──
        if adesione == 'non_risposto':
            stato_giu = 'non_iscritto'; n_no_iscritto += 1; gruppo_giu = None
        elif adesione == 'studio_ind':
            stato_giu = 'no_corso'; n_no_corso += 1; gruppo_giu = None
        elif adesione == 'non_aderisce':
            stato_giu = 'non_aderisce'; n_non_aderisce += 1; gruppo_giu = None
        else:  # aderisce | sconosciuto
            gruppo_giu = _trova_gruppo(imp, gruppi_giugno)
            if gruppo_giu:
                stato_giu = 'ok'; n_ok += 1
            else:
                stato_giu = 'no_gruppo'; n_no_gruppo += 1

        # ── Stato AGOSTO: sempre rilevante. Anche chi non ha aderito al
        # corso o ha scelto studio individuale deve sostenere la prova,
        # quindi si verifica comunque se esiste un gruppo/calendario. ──
        gruppo_ago = _trova_gruppo(imp, gruppi_agosto)
        stato_ago = 'ok' if gruppo_ago else 'no_gruppo'
        n_lezioni_ago = len(gruppo_ago.lezioni) if gruppo_ago else 0

        righe.append({
            'imp': imp,
            # Compatibilità con l'export XLSX esistente (usa 'gruppo'/'stato'
            # riferiti a giugno, comportamento storico)
            'gruppo': gruppo_giu, 'stato': stato_giu,
            'n_lezioni': len(gruppo_giu.lezioni) if gruppo_giu else 0,
            # Nuovi campi per la vista unificata
            'gruppo_giugno': gruppo_giu, 'stato_giugno': stato_giu,
            'gruppo_agosto': gruppo_ago, 'stato_agosto': stato_ago,
            'n_lezioni_agosto': n_lezioni_ago,
        })

    titolo = request.args.get('titolo', 'Verifica copertura recuperi')
    if request.args.get('export') == 'xlsx':
        buf = _export_copertura_xlsx(righe, titolo)
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='copertura_recuperi.xlsx'
        )

    return render_template('recupero/copertura.html',
        righe=righe, n_ok=n_ok,
        n_no_corso=n_no_corso,
        n_no_iscritto=n_no_iscritto,
        n_no_gruppo=n_no_gruppo,
        n_non_aderisce=n_non_aderisce)


# ══════════════════════════════════════════════════════════════════════
# Le route specifiche di giugno e agosto sono in sotto-moduli separati
# (vedi import qui sotto). Qui restano solo le route condivise:
# index, giugno_index, copertura.
# ══════════════════════════════════════════════════════════════════════

# Importa i sotto-moduli con le route spostate fuori da questo file
# (export e agosto), per tenerlo più piccolo. Gli import vanno in fondo
# (dopo che recupero_bp esiste già in questo modulo) cosi' i sotto-moduli
# possono importare recupero_bp da qui senza creare un ciclo: quando
# Python esegue queste righe, il modulo routes.recupero è già
# parzialmente inizializzato con recupero_bp definito, quindi i
# sotto-moduli lo trovano.
from routes import recupero_export  # noqa: E402,F401
from routes import recupero_agosto  # noqa: E402,F401
from routes import recupero_giugno  # noqa: E402,F401
