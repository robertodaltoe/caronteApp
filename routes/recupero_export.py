"""
Export XLSX del modulo recupero: calendario completo (Famiglie+Docenti),
schede individuali per docente/coppia somministratore-assistente.

Le route qui definite sono registrate sullo stesso blueprint recupero_bp
importato da routes.recupero — questo file viene importato in coda a
routes/recupero.py apposta per evitare un ciclo: questo modulo importa
SOLO da recupero_costanti (foglia), mai da routes.recupero stesso.
"""
from flask import render_template, request, redirect, url_for, flash, send_file
from models.recupero import RecuperoDocente, RecuperoGruppo
from routes.recupero_costanti import ANNO, ANNO_AGO, PERIODO_AGO

from routes.recupero import recupero_bp


def _nome_istituto():
    """Nome istituto configurabile (Impostazioni > Dati istituto), invece
    di ripetere la stringa a mano in ogni foglio Excel generato qui."""
    from config_istituto import get_dati_istituto
    return get_dati_istituto()['nome_istituto']


# ── EXPORT XLSX ───────────────────────────────────────────────────────
@recupero_bp.route('/recupero/export-xlsx')
def export_xlsx():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file

    gruppi_list = (RecuperoGruppo.query
                   .join(RecuperoDocente)
                   .filter(RecuperoDocente.anno_scol == ANNO,
                           RecuperoGruppo.periodo_codice == 'corsi_giugno')
                   .order_by(RecuperoGruppo.materia)
                   .all())

    GIORNI = ['Lunedì','Martedì','Mercoledì','Giovedì','Venerdì','Sabato','Domenica']
    MESI   = ['','Gennaio','Febbraio','Marzo','Aprile','Maggio','Giugno',
              'Luglio','Agosto','Settembre','Ottobre','Novembre','Dicembre']

    def fmt_data(d):
        return f"{GIORNI[d.weekday()]} {d.day} {MESI[d.month]}"

    # Stili
    BLU   = PatternFill('solid', start_color='1e3a5f')
    AZZUR = PatternFill('solid', start_color='dbeafe')
    VERDE = PatternFill('solid', start_color='dcfce7')
    GRAY  = PatternFill('solid', start_color='f3f4f6')
    BOLD  = Font(bold=True)
    BOLD_W= Font(bold=True, color='FFFFFF')
    THIN  = Border(
        left=Side(style='thin', color='d1d5db'),
        right=Side(style='thin', color='d1d5db'),
        top=Side(style='thin', color='d1d5db'),
        bottom=Side(style='thin', color='d1d5db'),
    )
    CENTER = Alignment(horizontal='center', vertical='center')
    WRAP   = Alignment(wrap_text=True, vertical='center')

    wb = Workbook()

    # ── FOGLIO FAMIGLIE ───────────────────────────────────────────────
    wsF = wb.active
    wsF.title = 'Famiglie'

    wsF.append([])
    wsF['A1'] = _nome_istituto()
    wsF['A1'].font = Font(bold=True, size=13)
    wsF['A2'] = f'CALENDARIO CORSI DI RECUPERO — A.S. {ANNO}'
    wsF['A2'].font = Font(bold=True, size=11)
    wsF['A3'] = 'Periodo: 18 giugno – 1 luglio 2026'
    wsF['A3'].font = Font(italic=True, color='6b7280')
    wsF.append([])

    # Raggruppa lezioni per materia
    from collections import defaultdict
    per_materia = defaultdict(list)
    for g in gruppi_list:
        for l in g.lezioni:
            per_materia[g.materia].append((l, g))

    row = 5
    for materia in sorted(per_materia.keys()):
        lezioni = sorted(per_materia[materia], key=lambda x: (x[0].data, x[0].ora_inizio))

        # Header materia
        wsF.merge_cells(f'A{row}:E{row}')
        wsF[f'A{row}'] = materia.upper()
        wsF[f'A{row}'].font = BOLD_W
        wsF[f'A{row}'].fill = BLU
        wsF[f'A{row}'].alignment = CENTER
        row += 1

        # Header colonne
        for col, h in enumerate(['Giorno', 'Data', 'Orario', 'Durata', 'Classi'], 1):
            cell = wsF.cell(row=row, column=col, value=h)
            cell.font = BOLD
            cell.fill = AZZUR
            cell.alignment = CENTER
            cell.border = THIN
        row += 1

        for l, g in lezioni:
            vals = [
                GIORNI[l.data.weekday()],
                l.data.strftime('%d/%m/%Y'),
                f'{l.ora_inizio}–{l.ora_fine}',
                f'{l.durata_ore}h',
                g.classi,
            ]
            for col, v in enumerate(vals, 1):
                cell = wsF.cell(row=row, column=col, value=v)
                cell.border = THIN
                cell.alignment = WRAP
            row += 1

        row += 1  # spazio tra materie

    # Larghezze famiglie
    for i, w in enumerate([14, 14, 14, 8, 30], 1):
        wsF.column_dimensions[get_column_letter(i)].width = w

    # ── FOGLIO DOCENTI ────────────────────────────────────────────────
    wsD = wb.create_sheet('Docenti')

    wsD['A1'] = _nome_istituto()
    wsD['A1'].font = Font(bold=True, size=13)
    wsD['A2'] = f'CALENDARIO CORSI DI RECUPERO — A.S. {ANNO} — USO INTERNO'
    wsD['A2'].font = Font(bold=True, size=11, color='dc2626')
    wsD.append([])
    wsD.append([])

    # Indice staging: (cognome, nome, classe, materia_norm) →︎ stato_adesione
    from models.recupero import RecuperoImport
    _FAM_EXP = [
        {'LATINO', 'LINGUA LATINA', 'LINGUA E CULTURA LATINA'},
        {'ITALIANO', 'LINGUA E LETTERATURA ITALIANA', 'LINGUA E CULTURA ITALIANA'},
        {'MATEMATICA', 'MATEMATICA CON INFORMATICA', 'MATEMATICA E COMPLEMENTI DI MATEMATICA'},
        {'INFORMATICA', 'TECNOLOGIE INFORMATICHE'},
        {'STORIA', 'STORIA E GEOGRAFIA'},
        {'FISICA', 'SCIENZE INTEGRATE (FISICA)'},
        {'INGLESE', 'LINGUA INGLESE', 'LINGUA E CULTURA STRANIERA (INGLESE)'},
        {'TEDESCO', 'LINGUA TEDESCA', 'LINGUA E CULTURA STRANIERA TEDESCO'},
    ]
    def _fam_match_exp(m1, m2):
        m1u,m2u = m1.strip().upper(), m2.strip().upper()
        if m1u == m2u: return True
        for f in _FAM_EXP:
            fs = {x.upper() for x in f}
            if m1u in fs and m2u in fs: return True
        return False

    # Carica tutti gli alunni con debito in questa materia+classe dallo staging
    imports_all_exp = RecuperoImport.query.filter_by(anno_scol=ANNO).all()

    STATO_LABEL = {
        'aderisce':     '✓︎ aderisce',
        'sconosciuto':  '✓︎ aderisce',
        'non_risposto': '? non risposto',
        'non_aderisce': '✕︎ non aderisce',
        'studio_ind':   '▥︎ studio ind.',
    }
    STATO_COLOR = {
        'aderisce':     '166534',  # verde scuro
        'sconosciuto':  '166534',
        'non_risposto': 'dc2626',  # rosso
        'non_aderisce': '6b7280',  # grigio
        'studio_ind':   '92400e',  # arancio
    }
    STATO_BG = {
        'aderisce':     'dcfce7',
        'sconosciuto':  'dcfce7',
        'non_risposto': 'fee2e2',
        'non_aderisce': 'f3f4f6',
        'studio_ind':   'fef9c3',
    }

    row = 5
    for materia in sorted(per_materia.keys()):
        lezioni = sorted(per_materia[materia], key=lambda x: (x[0].data, x[0].ora_inizio))
        # Gruppi di questa materia
        gruppi_mat = list({g.id: g for (l,g) in lezioni}.values())

        wsD.merge_cells(f'A{row}:I{row}')
        wsD[f'A{row}'] = materia.upper()
        wsD[f'A{row}'].font = BOLD_W
        wsD[f'A{row}'].fill = BLU
        wsD[f'A{row}'].alignment = CENTER
        row += 1

        for col, h in enumerate(['Giorno','Data','Orario','Docente','Classe','Cognome','Nome','Aula','Adesione'], 1):
            cell = wsD.cell(row=row, column=col, value=h)
            cell.font = BOLD
            cell.fill = VERDE
            cell.alignment = CENTER
            cell.border = THIN
        row += 1

        for l, g in lezioni:
            doc = g.docente
            nome_doc = f'{doc.cognome} {doc.nome or ""}'.strip() if doc else '—'
            classi_g = {cl.strip().upper() for cl in g.classi.split(',')}

            # Tutti gli alunni con debito in questa materia+classi dallo staging
            alunni_staging = [
                imp for imp in imports_all_exp
                if imp.classe.upper() in classi_g
                and _fam_match_exp(materia, imp.materia_norm or '')
            ]
            # Ordina: prima aderiscono, poi non risposto, poi non aderisce, poi studio_ind
            ordine = {'aderisce':0,'sconosciuto':0,'non_risposto':1,'non_aderisce':2,'studio_ind':3}
            alunni_staging = sorted(alunni_staging,
                key=lambda a: (ordine.get(a.stato_adesione,9), a.classe, a.cognome))

            row_inizio_blocco = row

            if alunni_staging:
                for i_al, al in enumerate(alunni_staging):
                    stato = al.stato_adesione or 'sconosciuto'
                    label = STATO_LABEL.get(stato, stato)
                    col_t = STATO_COLOR.get(stato, '374151')
                    col_b = STATO_BG.get(stato, 'ffffff')

                    # Colonne 1-4 (Giorno/Data/Orario/Docente) solo sulla prima riga,
                    # verranno unite verticalmente dopo il ciclo
                    vals = [
                        GIORNI[l.data.weekday()] if i_al == 0 else None,
                        l.data.strftime('%d/%m/%Y') if i_al == 0 else None,
                        f'{l.ora_inizio}–{l.ora_fine}' if i_al == 0 else None,
                        nome_doc if i_al == 0 else None,
                        al.classe,
                        al.cognome,
                        al.nome,
                        l.aula or '—',
                        label,
                    ]
                    for col, v in enumerate(vals, 1):
                        cell = wsD.cell(row=row, column=col, value=v)
                        cell.border = THIN
                        cell.alignment = WRAP
                        if col == 9:  # Adesione: colorata per stato
                            cell.font = Font(bold=True, color=col_t, name='Arial', size=9)
                            cell.fill = PatternFill('solid', start_color=col_b)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        elif col <= 4:
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            cell.fill = PatternFill('solid', start_color='fdebd3')
                        elif i_al % 2 == 1:
                            cell.fill = PatternFill('solid', start_color='f8fafc')
                    row += 1

                # Merge verticale colonne Giorno/Data/Orario/Docente per tutto il blocco
                if row - 1 > row_inizio_blocco:
                    for col_letter in ('A', 'B', 'C', 'D'):
                        wsD.merge_cells(f'{col_letter}{row_inizio_blocco}:{col_letter}{row-1}')
            else:
                # Nessun alunno dallo staging — mostra la lezione vuota
                vals = [
                    GIORNI[l.data.weekday()],
                    l.data.strftime('%d/%m/%Y'),
                    f'{l.ora_inizio}–{l.ora_fine}',
                    nome_doc, g.classi, '—', '—', l.aula or '—', '—',
                ]
                for col, v in enumerate(vals, 1):
                    cell = wsD.cell(row=row, column=col, value=v)
                    cell.border = THIN
                    if col <= 4:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.fill = PatternFill('solid', start_color='fdebd3')
                row += 1

        row += 1

    for i, w in enumerate([12, 12, 12, 22, 10, 18, 16, 8, 14], 1):
        wsD.column_dimensions[get_column_letter(i)].width = w

    # ── FOGLIO GIORNATE ───────────────────────────────────────────────
    # Una sezione per ogni giorno — solo materie e orari, senza nomi
    from collections import defaultdict

    # Raggruppa lezioni per data
    lezioni_per_data = defaultdict(list)
    for g in gruppi_list:
        for l in g.lezioni:
            lezioni_per_data[l.data].append((l, g))

    if lezioni_per_data:
        wsG = wb.create_sheet('Giornate')
        wsG['A1'] = _nome_istituto()
        wsG['A1'].font = Font(bold=True, size=13)
        wsG['A2'] = f'CALENDARIO GIORNALIERO — CORSI DI RECUPERO — A.S. {ANNO}'
        wsG['A2'].font = Font(bold=True, size=11)
        wsG.append([])

        row_g = 4
        for data in sorted(lezioni_per_data.keys()):
            coppie = sorted(lezioni_per_data[data], key=lambda x: x[0].ora_inizio)

            # Header giorno
            giorno_str = f'{GIORNI[data.weekday()]} {data.strftime("%d/%m/%Y")}'
            wsG.merge_cells(f'A{row_g}:G{row_g}')
            wsG[f'A{row_g}'] = giorno_str.upper()
            wsG[f'A{row_g}'].font = BOLD_W
            wsG[f'A{row_g}'].fill = BLU
            wsG[f'A{row_g}'].alignment = CENTER
            wsG[f'A{row_g}'].border = THIN
            row_g += 1

            # Intestazioni colonne
            for col, h in enumerate(['Orario','Materia','Docente','Classi','N. alunni','Ore','Aula'], 1):
                cell = wsG.cell(row=row_g, column=col, value=h)
                cell.font = BOLD
                cell.fill = VERDE
                cell.alignment = CENTER
                cell.border = THIN
            row_g += 1

            # Righe lezioni
            for i_r, (l, g) in enumerate(coppie):
                doc = g.docente
                nome_doc = f'{doc.cognome} {(doc.nome or "")[0]}.' if doc else '—'
                n_alunni = len(g.alunni) or '—'
                try:
                    h1,m1 = map(int, l.ora_inizio.split(':'))
                    h2,m2 = map(int, l.ora_fine.split(':'))
                    ore_h = (h2*60+m2 - h1*60-m1) / 60
                    ore_str = f'{ore_h:.1f}h'.replace('.0h','h')
                except Exception:
                    ore_str = '—'

                vals = [
                    f'{l.ora_inizio}–{l.ora_fine}',
                    g.materia,
                    nome_doc,
                    g.classi,
                    n_alunni,
                    ore_str,
                    l.aula or '—',
                ]
                for col, v in enumerate(vals, 1):
                    cell = wsG.cell(row=row_g, column=col, value=v)
                    cell.border = THIN
                    cell.alignment = Alignment(vertical='center',
                                               wrap_text=(col == 2))
                    if i_r % 2 == 1:
                        cell.fill = PatternFill('solid', start_color='f0f4ff')
                row_g += 1

            row_g += 1  # spazio tra giorni

        for i, w in enumerate([14, 35, 18, 22, 10, 8, 10], 1):
            wsG.column_dimensions[get_column_letter(i)].width = w
        wsG.freeze_panes = 'A4'

    # ── FOGLIO RIEPILOGO ORE ──────────────────────────────────────────
    # Una riga per docente+materia, con ore totali per materia e per docente.
    # Ogni sessione (data+orario+docente) è contata una sola volta.
    wsR = wb.create_sheet('Riepilogo Ore')
    wsR['A1'] = _nome_istituto()
    wsR['A1'].font = Font(bold=True, size=13)
    wsR['A2'] = f'RIEPILOGO ORE CORSI DI RECUPERO — A.S. {ANNO}'
    wsR['A2'].font = Font(bold=True, size=11)
    wsR.append([])

    for col, h in enumerate(['DOCENTE', 'MATERIA', 'ORE (per materia)', 'ORE TOTALI'], 1):
        cell = wsR.cell(row=4, column=col, value=h)
        cell.font = BOLD_W
        cell.fill = BLU
        cell.alignment = CENTER
        cell.border = THIN

    # Calcola ore per docente+materia: conta sessioni unique (data, ora_inizio, ora_fine, id_gruppo)
    ore_per_doc_mat = defaultdict(float)  # (docente_id, materia) -> ore
    nome_docente_map = {}
    for g in gruppi_list:
        if not g.docente: continue
        sessioni_viste = set()
        for l in g.lezioni:
            key = (l.data, l.ora_inizio, l.ora_fine, g.id)
            if key in sessioni_viste: continue
            sessioni_viste.add(key)
            try:
                h1,m1 = map(int, l.ora_inizio.split(':'))
                h2,m2 = map(int, l.ora_fine.split(':'))
                ore = (h2*60+m2 - h1*60-m1) / 60
            except Exception:
                ore = 0
            ore_per_doc_mat[(g.docente.id, g.materia)] += ore
        nome_docente_map[g.docente.id] = f'{g.docente.cognome} {g.docente.nome or ""}'.strip()

    # Raggruppa per docente, ordina materie
    per_doc = defaultdict(list)
    for (doc_id, materia), ore in ore_per_doc_mat.items():
        per_doc[doc_id].append((materia, ore))

    row_r = 5
    totale_generale = 0
    for doc_id in sorted(per_doc.keys(), key=lambda d: nome_docente_map.get(d, '')):
        materie = sorted(per_doc[doc_id])
        tot_doc = sum(o for _, o in materie)
        totale_generale += tot_doc
        for i_m, (materia, ore) in enumerate(materie):
            vals = [
                nome_docente_map.get(doc_id, '?') if i_m == 0 else None,
                materia,
                int(ore) if ore == int(ore) else ore,
                tot_doc if i_m == 0 else None,
            ]
            for col, v in enumerate(vals, 1):
                cell = wsR.cell(row=row_r, column=col, value=v)
                cell.border = THIN
                if col in (1, 4):
                    cell.font = BOLD
                cell.alignment = Alignment(horizontal='left' if col<=2 else 'center', vertical='center')
            row_r += 1

    # Riga totale generale
    wsR.merge_cells(f'A{row_r}:C{row_r}')
    wsR[f'A{row_r}'] = 'TOTALE GENERALE'
    wsR[f'A{row_r}'].font = BOLD_W
    wsR[f'A{row_r}'].fill = PatternFill('solid', start_color='1f3864')
    cell_tot = wsR.cell(row=row_r, column=4, value=int(totale_generale) if totale_generale==int(totale_generale) else totale_generale)
    cell_tot.font = BOLD_W
    cell_tot.fill = PatternFill('solid', start_color='1f3864')
    cell_tot.alignment = CENTER
    row_r += 2

    wsR.cell(row=row_r, column=1, value=(
        'Nota: ogni sessione (data + orario + docente) è contata una sola volta, '
        'indipendentemente dal numero di studenti o classi presenti.'
    )).font = Font(italic=True, size=9, color='6b7280')

    for i, w in enumerate([22, 42, 16, 14], 1):
        wsR.column_dimensions[get_column_letter(i)].width = w

    # Salva in memoria e invia
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'corsi_recupero_{ANNO}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _genera_scheda_docente_xlsx(docente, gruppi_docente, anno_scol):
    """
    Genera un workbook XLSX con la scheda calendario di un singolo docente:
    sessioni (data+orario+materia) con elenco alunni e stato adesione.
    Replica lo stile della scheda di esempio fornita da Roberto.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from models.recupero import RecuperoImport

    GIORNI = ['Lunedì','Martedì','Mercoledì','Giovedì','Venerdì','Sabato','Domenica']

    _FAM_SCHEDA = [
        {'LATINO', 'LINGUA LATINA', 'LINGUA E CULTURA LATINA'},
        {'ITALIANO', 'LINGUA E LETTERATURA ITALIANA', 'LINGUA E CULTURA ITALIANA'},
        {'MATEMATICA', 'MATEMATICA CON INFORMATICA', 'MATEMATICA E COMPLEMENTI DI MATEMATICA'},
        {'INFORMATICA', 'TECNOLOGIE INFORMATICHE'},
        {'STORIA', 'STORIA E GEOGRAFIA'},
        {'FISICA', 'SCIENZE INTEGRATE (FISICA)'},
        {'INGLESE', 'LINGUA INGLESE', 'LINGUA E CULTURA STRANIERA (INGLESE)'},
        {'TEDESCO', 'LINGUA TEDESCA', 'LINGUA E CULTURA STRANIERA TEDESCO'},
    ]
    def _match_scheda(m1, m2):
        m1u,m2u = m1.strip().upper(), m2.strip().upper()
        if m1u == m2u: return True
        for f in _FAM_SCHEDA:
            fs = {x.upper() for x in f}
            if m1u in fs and m2u in fs: return True
        return False

    STATO_LABEL = {
        'aderisce':'✓︎ aderisce', 'sconosciuto':'✓︎ aderisce',
        'non_risposto':'? non risposto', 'non_aderisce':'✕︎ non aderisce',
        'studio_ind':'▥︎ studio ind.',
    }
    STATO_FILL = {
        'aderisce':'C6E0B4', 'sconosciuto':'C6E0B4',
        'non_risposto':'FFE699', 'non_aderisce':'F4B6B6',
        'studio_ind':'FFF2CC',
    }

    imports_all = RecuperoImport.query.filter_by(anno_scol=anno_scol).all()

    # Stili
    BLU_FILL   = PatternFill('solid', start_color='2F4F8C')
    HDR_FILL   = PatternFill('solid', start_color='D9E1F2')
    COLHDR_FILL= PatternFill('solid', start_color='EDEDED')
    TOT_FILL   = PatternFill('solid', start_color='1F3864')
    ROW_ALT    = PatternFill('solid', start_color='F4F7FC')
    WHITE_FILL = PatternFill('solid', start_color='FFFFFF')
    BOLD_W     = Font(bold=True, color='FFFFFF', size=11)
    BOLD_W10   = Font(bold=True, color='FFFFFF', size=10)
    BOLD       = Font(bold=True, size=9)
    NORMAL     = Font(size=9)
    THIN_SIDE  = Side(style='thin', color='B0B8CC')
    THIN       = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
    CENTER     = Alignment(horizontal='center', vertical='center')
    LEFT       = Alignment(horizontal='left', vertical='center')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Calendario'

    nome_doc = f'{docente.cognome} {docente.nome or ""}'.strip()

    ws.merge_cells('A1:H1')
    ws['A1'] = _nome_istituto()
    ws['A1'].font = BOLD_W; ws['A1'].fill = BLU_FILL; ws['A1'].alignment = CENTER

    ws.merge_cells('A2:H2')
    ws['A2'] = f'CALENDARIO CORSI DI RECUPERO — A.S. {anno_scol}'
    ws['A2'].font = BOLD_W10; ws['A2'].fill = BLU_FILL; ws['A2'].alignment = CENTER

    ws.merge_cells('A3:H3')
    ws['A3'] = f'Docente: {nome_doc}'
    ws['A3'].font = BOLD_W10; ws['A3'].fill = BLU_FILL; ws['A3'].alignment = CENTER

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 6

    # Raccogli tutte le sessioni (data, ora_inizio, ora_fine, materia, gruppo)
    sessioni = []
    for g in gruppi_docente:
        for l in g.lezioni:
            sessioni.append((l.data, l.ora_inizio, l.ora_fine, g.materia, g, l))
    sessioni.sort(key=lambda x: (x[0], x[1]))

    row = 5
    ore_totali_docente = 0.0

    for data, ora_ini, ora_fine, materia, g, l in sessioni:
        try:
            h1,m1 = map(int, ora_ini.split(':'))
            h2,m2 = map(int, ora_fine.split(':'))
            durata_h = (h2*60+m2 - h1*60-m1) / 60
        except Exception:
            durata_h = 0
        ore_totali_docente += durata_h

        ws.row_dimensions[row].height = 16
        ws.merge_cells(f'A{row}:H{row}')
        header_str = (f'{GIORNI[data.weekday()]} {data.strftime("%d/%m/%Y")}   '
                      f'{ora_ini}–{ora_fine}   ({durata_h:.1f}h)   —   {materia.upper()}')
        ws[f'A{row}'] = header_str
        ws[f'A{row}'].font = BOLD; ws[f'A{row}'].fill = HDR_FILL; ws[f'A{row}'].alignment = LEFT
        row += 1

        # Header colonne (E:H)
        ws.row_dimensions[row].height = 14
        for col, h in zip(['E','F','G','H'], ['Classe','Cognome','Nome','Adesione']):
            cell = ws[f'{col}{row}']
            cell.value = h
            cell.font = BOLD; cell.fill = COLHDR_FILL; cell.alignment = CENTER
            cell.border = THIN
        row += 1

        # Alunni: tutti quelli con debito in questa materia+classi, dallo staging
        classi_g = {cl.strip().upper() for cl in g.classi.split(',')}
        alunni_sess = [
            imp for imp in imports_all
            if imp.classe.upper() in classi_g
            and _match_scheda(materia, imp.materia_norm or '')
        ]
        ordine = {'aderisce':0,'sconosciuto':0,'non_risposto':1,'non_aderisce':2,'studio_ind':3}
        alunni_sess.sort(key=lambda a: (ordine.get(a.stato_adesione,9), a.classe, a.cognome))

        for i_al, al in enumerate(alunni_sess):
            ws.row_dimensions[row].height = 14
            stato = al.stato_adesione or 'sconosciuto'
            bg = STATO_FILL.get(stato, 'FFFFFF')
            label = STATO_LABEL.get(stato, stato)
            row_fill = ROW_ALT if i_al % 2 == 1 else WHITE_FILL

            vals = [al.classe, al.cognome, al.nome, label]
            for col, v in zip(['E','F','G','H'], vals):
                cell = ws[f'{col}{row}']
                cell.value = v
                cell.border = THIN
                cell.alignment = CENTER if col in ('E','H') else LEFT
                cell.font = NORMAL
                cell.fill = PatternFill('solid', start_color=bg) if col == 'H' else row_fill
            row += 1

        row += 1  # riga vuota tra sessioni

    # Riga totale ore
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = 'TOTALE ORE DOCENTE'
    ws[f'A{row}'].font = BOLD_W; ws[f'A{row}'].fill = TOT_FILL; ws[f'A{row}'].alignment = LEFT
    cell_tot = ws[f'H{row}']
    cell_tot.value = int(ore_totali_docente) if ore_totali_docente == int(ore_totali_docente) else ore_totali_docente
    cell_tot.font = BOLD_W; cell_tot.fill = TOT_FILL; cell_tot.alignment = CENTER
    row += 2

    # Legenda
    ws[f'A{row}'] = 'Legenda:'
    ws[f'A{row}'].font = Font(bold=True, size=9)
    row += 1
    legenda = [('A','✓︎ aderisce','C6E0B4'), ('B','? non risposto','FFE699'), ('C','✕︎ non aderisce','F4B6B6')]
    for col, label, color in legenda:
        cell = ws[f'{col}{row}']
        cell.value = label
        cell.fill = PatternFill('solid', start_color=color)
        cell.font = Font(size=9)

    # Larghezze colonne
    for col, w in zip(['A','B','C','D','E','F','G','H'], [12,12,13,32,9,18,18,16]):
        ws.column_dimensions[col].width = w

    ws.freeze_panes = 'A5'

    return wb


@recupero_bp.route('/recupero/export-schede-docenti')
def export_schede_docenti():
    """
    Genera un file ZIP con una scheda XLSX per ciascun docente che ha
    almeno una lezione pianificata — calendario individuale + alunni + stato adesione.
    """
    import io, zipfile
    from flask import send_file
    from collections import defaultdict

    gruppi_list = (RecuperoGruppo.query
                   .join(RecuperoDocente)
                   .filter(RecuperoDocente.anno_scol == ANNO,
                           RecuperoGruppo.periodo_codice == 'corsi_giugno')
                   .order_by(RecuperoGruppo.materia)
                   .all())

    # Raggruppa gruppi per docente
    gruppi_per_docente = defaultdict(list)
    for g in gruppi_list:
        if g.docente and g.lezioni:
            gruppi_per_docente[g.docente.id].append(g)

    if not gruppi_per_docente:
        flash('Nessuna lezione pianificata: genera prima il calendario.', 'warning')
        return redirect(url_for('recupero.calendario'))

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for doc_id, gruppi_doc in gruppi_per_docente.items():
            docente = gruppi_doc[0].docente
            wb_doc = _genera_scheda_docente_xlsx(docente, gruppi_doc, ANNO)
            file_buf = io.BytesIO()
            wb_doc.save(file_buf)
            file_buf.seek(0)
            cognome_safe = docente.cognome.replace(' ', '_')
            nome_safe = (docente.nome or '').split()[0] if docente.nome else ''
            filename = f'{cognome_safe}_{nome_safe}.xlsx'.strip('_')
            zf.writestr(filename, file_buf.read())

    zip_buf.seek(0)
    return send_file(zip_buf, as_attachment=True,
                     download_name=f'schede_docenti_recupero_{ANNO}.zip',
                     mimetype='application/zip')


def _genera_scheda_coppia_agosto_xlsx(somministratore, assistente, gruppi_coppia, anno_scol):
    """
    Scheda XLSX per agosto, una per ogni coppia somministratore+assistente:
    data, orario, durata, materia/classi e nominativi dei candidati per
    ogni prova che quella coppia segue insieme. Stesso stile grafico della
    scheda docente di giugno, senza colonna stato adesione (la prova si
    sostiene comunque, indipendentemente dall'adesione al corso).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    GIORNI = ['Lunedì','Martedì','Mercoledì','Giovedì','Venerdì','Sabato','Domenica']

    BLU_FILL    = PatternFill('solid', start_color='2F4F8C')
    HDR_FILL    = PatternFill('solid', start_color='D9E1F2')
    COLHDR_FILL = PatternFill('solid', start_color='EDEDED')
    TOT_FILL    = PatternFill('solid', start_color='1F3864')
    ROW_ALT     = PatternFill('solid', start_color='F4F7FC')
    WHITE_FILL  = PatternFill('solid', start_color='FFFFFF')
    BOLD_W      = Font(bold=True, color='FFFFFF', size=11)
    BOLD_W10    = Font(bold=True, color='FFFFFF', size=10)
    BOLD        = Font(bold=True, size=9)
    NORMAL      = Font(size=9)
    THIN_SIDE   = Side(style='thin', color='B0B8CC')
    THIN        = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
    CENTER      = Alignment(horizontal='center', vertical='center')
    LEFT        = Alignment(horizontal='left', vertical='center')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Calendario'

    nome_somm = f'{somministratore.cognome} {somministratore.nome or ""}'.strip() if somministratore else '—'
    nome_assist = f'{assistente.cognome} {assistente.nome or ""}'.strip() if assistente else '—'

    ws.merge_cells('A1:G1')
    ws['A1'] = _nome_istituto()
    ws['A1'].font = BOLD_W; ws['A1'].fill = BLU_FILL; ws['A1'].alignment = CENTER

    ws.merge_cells('A2:G2')
    ws['A2'] = f'CALENDARIO PROVE DI RECUPERO — A.S. {anno_scol}'
    ws['A2'].font = BOLD_W10; ws['A2'].fill = BLU_FILL; ws['A2'].alignment = CENTER

    ws.merge_cells('A3:G3')
    ws['A3'] = f'Somministratore: {nome_somm}   —   Assistente: {nome_assist}'
    ws['A3'].font = BOLD_W10; ws['A3'].fill = BLU_FILL; ws['A3'].alignment = CENTER

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 6

    # Raccogli tutte le sessioni (data, ora_inizio, ora_fine, materia, gruppo, lezione)
    sessioni = []
    for g in gruppi_coppia:
        for l in g.lezioni:
            sessioni.append((l.data, l.ora_inizio, l.ora_fine, g.materia, g, l))
    sessioni.sort(key=lambda x: (x[0], x[1]))

    row = 5
    ore_totali = 0.0

    for data, ora_ini, ora_fine, materia, g, l in sessioni:
        try:
            h1,m1 = map(int, ora_ini.split(':'))
            h2,m2 = map(int, ora_fine.split(':'))
            durata_h = (h2*60+m2 - h1*60-m1) / 60
        except Exception:
            durata_h = 0
        ore_totali += durata_h

        ws.row_dimensions[row].height = 16
        ws.merge_cells(f'A{row}:G{row}')
        header_str = (f'{GIORNI[data.weekday()]} {data.strftime("%d/%m/%Y")}   '
                      f'{ora_ini}–{ora_fine}   ({durata_h:.1f}h)   —   {materia.upper()}   '
                      f'({g.classi})')
        ws[f'A{row}'] = header_str
        ws[f'A{row}'].font = BOLD; ws[f'A{row}'].fill = HDR_FILL; ws[f'A{row}'].alignment = LEFT
        row += 1

        # Header colonne candidati (D:G) — G = colonna da compilare a mano
        ws.row_dimensions[row].height = 14
        for col, h in zip(['D','E','F','G'], ['Classe','Cognome','Nome','Presenza si/no']):
            cell = ws[f'{col}{row}']
            cell.value = h
            cell.font = BOLD; cell.fill = COLHDR_FILL; cell.alignment = CENTER
            cell.border = THIN
        row += 1

        candidati = sorted(g.alunni, key=lambda a: (a.classe, a.cognome))
        for i_al, al in enumerate(candidati):
            ws.row_dimensions[row].height = 14
            row_fill = ROW_ALT if i_al % 2 == 1 else WHITE_FILL
            vals = [al.classe, al.cognome, al.nome, '']
            for col, v in zip(['D','E','F','G'], vals):
                cell = ws[f'{col}{row}']
                cell.value = v
                cell.border = THIN
                cell.alignment = CENTER if col in ('D','G') else LEFT
                cell.font = NORMAL
                cell.fill = row_fill
            row += 1

        if not candidati:
            ws.row_dimensions[row].height = 14
            cell = ws[f'D{row}']
            cell.value = '— nessun candidato collegato —'
            cell.font = Font(italic=True, size=9, color='9CA3AF')
            cell.border = THIN
            row += 1

        row += 1  # riga vuota tra sessioni

    # Riga totale ore
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'TOTALE ORE COPPIA'
    ws[f'A{row}'].font = BOLD_W; ws[f'A{row}'].fill = TOT_FILL; ws[f'A{row}'].alignment = LEFT
    cell_tot = ws[f'F{row}']
    cell_tot.value = int(ore_totali) if ore_totali == int(ore_totali) else ore_totali
    cell_tot.font = BOLD_W; cell_tot.fill = TOT_FILL; cell_tot.alignment = CENTER

    # Larghezze colonne
    for col, w in zip(['A','B','C','D','E','F','G'], [14,14,14,9,18,18,16]):
        ws.column_dimensions[col].width = w

    ws.freeze_panes = 'A5'

    return wb


@recupero_bp.route('/recupero/agosto/export-schede-coppie')
def agosto_export_schede_coppie():
    """
    File ZIP con una scheda XLSX per ogni coppia somministratore+assistente
    delle prove di agosto — data, durata e nominativi dei candidati per
    ogni prova che quella coppia segue insieme.
    """
    import io, zipfile
    from flask import send_file
    from collections import defaultdict

    gruppi_list = (RecuperoGruppo.query
                   .join(RecuperoDocente)
                   .filter(RecuperoDocente.anno_scol == ANNO_AGO,
                           RecuperoGruppo.periodo_codice == PERIODO_AGO)
                   .all())

    gruppi_per_coppia = defaultdict(list)
    for g in gruppi_list:
        if g.docente and g.lezioni:
            id_assist = g.id_sorvegliante  # può essere None
            gruppi_per_coppia[(g.docente.id, id_assist)].append(g)

    if not gruppi_per_coppia:
        flash('Nessuna prova pianificata: genera prima il calendario agosto.', 'warning')
        return redirect(url_for('recupero.agosto_calendario'))

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for (id_somm, id_assist), gruppi_coppia in gruppi_per_coppia.items():
            somministratore = gruppi_coppia[0].docente
            assistente = gruppi_coppia[0].sorvegliante if id_assist else None

            wb_coppia = _genera_scheda_coppia_agosto_xlsx(
                somministratore, assistente, gruppi_coppia, ANNO_AGO)
            file_buf = io.BytesIO()
            wb_coppia.save(file_buf)
            file_buf.seek(0)

            cogn_somm = somministratore.cognome.replace(' ', '_')
            cogn_assist = assistente.cognome.replace(' ', '_') if assistente else 'SENZA_ASSISTENTE'
            filename = f'{cogn_somm}_e_{cogn_assist}.xlsx'
            zf.writestr(filename, file_buf.read())

    zip_buf.seek(0)
    return send_file(zip_buf, as_attachment=True,
                     download_name=f'schede_coppie_prove_agosto_{ANNO_AGO}.zip',
                     mimetype='application/zip')



# ── EXPORT XLSX AGOSTO ────────────────────────────────────────────────
@recupero_bp.route('/recupero/agosto/export-xlsx')
def agosto_export_xlsx():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file

    gruppi = (RecuperoGruppo.query.join(RecuperoDocente)
              .filter(RecuperoDocente.anno_scol == ANNO_AGO,
                      RecuperoGruppo.periodo_codice == PERIODO_AGO)
              .order_by(RecuperoGruppo.materia).all())

    GIORNI = ['Lunedì','Martedì','Mercoledì','Giovedì','Venerdì','Sabato','Domenica']
    MESI   = ['','Gennaio','Febbraio','Marzo','Aprile','Maggio','Giugno',
              'Luglio','Agosto','Settembre','Ottobre','Novembre','Dicembre']
    TIPO_LABEL = {'scritto':'Scritto','orale':'Orale',
                  'pratico':'Pratico','scritto_orale':'Scritto+Orale'}

    BLU   = PatternFill('solid', start_color='1e3a5f')
    AZZUR = PatternFill('solid', start_color='dbeafe')
    VERDE = PatternFill('solid', start_color='dcfce7')
    BOLD  = Font(bold=True)
    BOLD_W= Font(bold=True, color='FFFFFF')
    THIN  = Border(left=Side(style='thin',color='d1d5db'),
                   right=Side(style='thin',color='d1d5db'),
                   top=Side(style='thin',color='d1d5db'),
                   bottom=Side(style='thin',color='d1d5db'))
    CENTER = Alignment(horizontal='center', vertical='center')
    WRAP   = Alignment(wrap_text=True, vertical='center')

    wb = Workbook()
    ws_fam = wb.active
    ws_fam.title = 'Famiglie'
    ws_doc = wb.create_sheet('Docenti')

    def sheet_header(ws, interno=False):
        ws['A1'] = _nome_istituto()
        ws['A1'].font = Font(bold=True, size=13)
        ws['A2'] = f'CALENDARIO PROVE DI RECUPERO — A.S. {ANNO_AGO}'
        ws['A2'].font = Font(bold=True, size=11,
                             color='dc2626' if interno else '000000')
        ws.append([])
        ws.append([])

    sheet_header(ws_fam)
    sheet_header(ws_doc, interno=True)

    # ── Foglio Famiglie: raggruppato per GIORNATA, non per materia.
    # Tutte le prove dello stesso giorno stanno nella stessa sottotabella,
    # ordinate per orario di inizio (le prove con lo stesso orario si
    # susseguono una sotto l'altra, come per il calendario interno).
    lezioni_per_giorno = {}
    for g in gruppi:
        for l in g.lezioni:
            lezioni_per_giorno.setdefault(l.data, []).append((l, g))

    for data in sorted(lezioni_per_giorno.keys()):
        coppie = sorted(lezioni_per_giorno[data], key=lambda lg: lg[0].ora_inizio)

        row_f = ws_fam.max_row + 1
        ws_fam.merge_cells(f'A{row_f}:F{row_f}')
        ws_fam[f'A{row_f}'] = f'{GIORNI[data.weekday()]} {data.strftime("%d/%m/%Y")}'
        ws_fam[f'A{row_f}'].font = BOLD_W
        ws_fam[f'A{row_f}'].fill = BLU
        ws_fam[f'A{row_f}'].alignment = CENTER
        row_f += 1
        for col, h in enumerate(['Orario','Materia','Tipo prova','Durata','Classi'], 1):
            c = ws_fam.cell(row=row_f, column=col, value=h)
            c.font = BOLD; c.fill = AZZUR; c.alignment = CENTER; c.border = THIN
        row_f += 1
        for l, g in coppie:
            tipo_str = TIPO_LABEL.get(g.tipo_prova or 'scritto', '—')
            vals = [f'{l.ora_inizio}–{l.ora_fine}', g.materia.upper(),
                    tipo_str, f'{l.durata_ore}h', g.classi]
            for col, v in enumerate(vals, 1):
                c = ws_fam.cell(row=row_f, column=col, value=v)
                c.border = THIN; c.alignment = WRAP
            row_f += 1
        ws_fam.append([])  # spazio tra giornate

    for g in gruppi:
        lezioni = sorted(g.lezioni, key=lambda l: (l.data, l.ora_inizio))
        if not lezioni:
            continue

        doc = g.docente
        nome_doc = f'{doc.cognome} {doc.nome or ""}'.strip() if doc else '—'
        assist = g.sorvegliante
        nome_assist = f'{assist.cognome} {assist.nome or ""}'.strip() if assist else '—'
        tipo_str = TIPO_LABEL.get(g.tipo_prova or 'scritto', '—')

        # ── Foglio Docenti: una riga per alunno per prova ────────────
        row_d = ws_doc.max_row + 1
        ws_doc.merge_cells(f'A{row_d}:H{row_d}')
        ws_doc[f'A{row_d}'] = f'{g.materia.upper()} — {tipo_str} — Somministratore: {nome_doc} — Assistente: {nome_assist}'
        ws_doc[f'A{row_d}'].font = BOLD_W
        ws_doc[f'A{row_d}'].fill = BLU
        ws_doc[f'A{row_d}'].alignment = CENTER
        row_d += 1
        for col, h in enumerate(['Giorno','Data','Orario','Somministratore','Assistente','Classe','Cognome','Nome'], 1):
            c = ws_doc.cell(row=row_d, column=col, value=h)
            c.font = BOLD; c.fill = VERDE; c.alignment = CENTER; c.border = THIN
        row_d += 1
        for l in lezioni:
            alunni_g = sorted(g.alunni, key=lambda a: (a.classe, a.cognome))
            row_inizio_blocco = row_d
            if not alunni_g:
                # Nessun alunno collegato: una riga sola, niente da unire
                vals = [GIORNI[l.data.weekday()], l.data.strftime('%d/%m/%Y'),
                        f'{l.ora_inizio}–{l.ora_fine}', nome_doc, nome_assist,
                        g.classi, '—', '—']
                for col, v in enumerate(vals, 1):
                    c = ws_doc.cell(row=row_d, column=col, value=v)
                    c.border = THIN; c.alignment = WRAP
                row_d += 1
                continue
            for i_al, al in enumerate(alunni_g):
                # Giorno/Data/Orario/Somministratore/Assistente sono identici
                # per tutte le righe di questa prova: si scrivono solo sulla
                # prima riga e si uniscono verticalmente dopo il blocco.
                vals = [
                    GIORNI[l.data.weekday()] if i_al == 0 else None,
                    l.data.strftime('%d/%m/%Y') if i_al == 0 else None,
                    f'{l.ora_inizio}–{l.ora_fine}' if i_al == 0 else None,
                    nome_doc if i_al == 0 else None,
                    nome_assist if i_al == 0 else None,
                    al.classe, al.cognome, al.nome,
                ]
                for col, v in enumerate(vals, 1):
                    c = ws_doc.cell(row=row_d, column=col, value=v)
                    c.border = THIN
                    c.alignment = CENTER if col <= 5 else WRAP
                    if col <= 5:
                        c.fill = PatternFill('solid', start_color='fdebd3')
                    elif i_al % 2 == 1:
                        c.fill = PatternFill('solid', start_color='f8fafc')
                row_d += 1
            if row_d - 1 > row_inizio_blocco:
                for col_letter in ('A', 'B', 'C', 'D', 'E'):
                    ws_doc.merge_cells(f'{col_letter}{row_inizio_blocco}:{col_letter}{row_d-1}')
        ws_doc.append([])

    # Larghezze (Famiglie: Orario, Materia, Tipo prova, Durata, Classi)
    for i, w in enumerate([14,28,14,10,18], 1):
        ws_fam.column_dimensions[get_column_letter(i)].width = w
    for i, w in enumerate([12,12,12,22,22,10,18,16], 1):
        ws_doc.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'prove_recupero_agosto_{ANNO_AGO}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
