"""
Route specifiche dei corsi di recupero di giugno: docenti disponibili,
creazione gruppi, calendario lezioni con generazione bozza, vincoli
orari per docente, import staging alunni, proposte automatiche.

Registrate sullo stesso blueprint recupero_bp importato da
routes.recupero — questo file importa SOLO da recupero_costanti
(foglia), mai da routes.recupero stesso, per evitare un ciclo (lo
stesso schema già usato per recupero_export.py e recupero_agosto.py).
"""
from flask import render_template, request, redirect, url_for, flash
from models import db
from models.recupero import RecuperoDocente, RecuperoGruppo, RecuperoLezione, RecuperoAlunno, RecuperoVincolo
from models.docente import Docente
from models.orario_docente import classi_attive as _classi_attive
from datetime import date, timedelta
from routes.recupero_costanti import (
    ANNO, DATA_INIZIO, DATA_FINE, TIPO_PROVA_LABEL,
    _materia_canonica, _norm_materia,
    _split_cognome_nome, _parse_tipo_prova, docenti_in_servizio_query,
)

from routes.recupero import recupero_bp


# ── DOCENTI DISPONIBILI ───────────────────────────────────────────────
@recupero_bp.route('/recupero/docenti', methods=['GET', 'POST'])
def docenti():
    if request.method == 'POST':
        azione = request.form.get('azione')

        if azione == 'aggiungi':
            id_doc = int(request.form['id_docente'])
            note   = request.form.get('note', '').strip() or None
            exists = RecuperoDocente.query.filter_by(
                id_docente=id_doc, anno_scol=ANNO).first()
            materie_extra = request.form.get('materie_extra','').strip() or None
            if not exists:
                db.session.add(RecuperoDocente(
                    id_docente=id_doc, anno_scol=ANNO, note=note,
                    materie_extra=materie_extra))
                db.session.commit()
                d = Docente.query.get(id_doc)
                flash(f'{d.cognome} aggiunto ai disponibili.', 'success')
            else:
                flash('Docente già presente.', 'warning')

        elif azione == 'modifica_docente':
            rid = int(request.form['id'])
            rd  = RecuperoDocente.query.get_or_404(rid)
            rd.note          = request.form.get('note','').strip() or None
            rd.materie_extra = request.form.get('materie_extra','').strip() or None
            db.session.commit()
            flash('Aggiornato.', 'success')

        elif azione == 'rimuovi':
            rid = int(request.form['id'])
            rd  = RecuperoDocente.query.get_or_404(rid)
            db.session.delete(rd)
            db.session.commit()
            flash('Docente rimosso.', 'warning')

        return redirect(url_for('recupero.docenti'))

    disponibili = (RecuperoDocente.query.filter_by(anno_scol=ANNO)
                   .join(Docente).order_by(Docente.cognome).all())
    disp_ids = {rd.id_docente for rd in disponibili}
    # In servizio nell'anno dei corsi di recupero — vedi
    # routes/recupero_costanti.py::docenti_in_servizio_query() (Sessione
    # 62: senza questo filtro comparivano anche docenti non ancora
    # arrivati, arruolabili per errore in un anno prima del loro inizio
    # effettivo di servizio).
    tutti = docenti_in_servizio_query(ANNO).order_by(Docente.cognome).all()
    non_ancora = [d for d in tutti if d.id not in disp_ids]

    return render_template('recupero/docenti.html',
        disponibili=disponibili,
        non_ancora=non_ancora,
        anno=ANNO,
    )


# ── GRUPPI ────────────────────────────────────────────────────────────
@recupero_bp.route('/recupero/gruppi', methods=['GET', 'POST'])
def gruppi():
    if request.method == 'POST':
        azione = request.form.get('azione')

        if azione == 'aggiungi':
            id_recdoc      = int(request.form['id_rec_docente'])
            materia        = request.form.get('materia', '').strip()
            note           = request.form.get('note', '').strip() or None
            max_ore        = int(request.form.get('max_ore', '10') or 10)
            max_ore_giorno = int(request.form.get('max_ore_giorno', '2') or 2)
            # Classi da checkbox o testo libero
            classi_list = request.form.getlist('classi_cb')
            classi = ','.join(cl.strip() for cl in classi_list if cl.strip()) if classi_list else request.form.get('classi','').strip()
            db.session.add(RecuperoGruppo(
                id_rec_docente=id_recdoc, materia=materia,
                classi=classi, max_ore=max_ore,
                max_ore_giorno=max_ore_giorno, note=note,
            ))
            db.session.commit()
            flash('Gruppo aggiunto.', 'success')

        elif azione == 'modifica':
            gid = int(request.form['id'])
            g   = RecuperoGruppo.query.get_or_404(gid)
            g.materia        = request.form.get('materia','').strip() or g.materia
            g.id_rec_docente = int(request.form['id_rec_docente'])
            g.max_ore        = int(request.form.get('max_ore', g.max_ore or 10) or 10)
            g.max_ore_giorno = int(request.form.get('max_ore_giorno', g.max_ore_giorno or 2) or 2)
            g.note           = request.form.get('note','').strip() or None
            # Classi: da checkbox (classi_cb) o testo libero (classi)
            classi_cb = request.form.getlist('classi_cb')
            if classi_cb:
                # Checkbox presenti: usa quelli (anche se la lista è vuota = nessuna classe)
                g.classi = ','.join(cl.strip() for cl in classi_cb if cl.strip())
            else:
                # Fallback testo libero
                classi_txt = request.form.get('classi','').strip()
                if classi_txt:
                    g.classi = classi_txt
                # Se entrambi vuoti: mantieni le classi esistenti
            db.session.commit()
            flash('Gruppo aggiornato.', 'success')

        elif azione == 'elimina':
            gid = int(request.form['id'])
            g   = RecuperoGruppo.query.get_or_404(gid)
            db.session.delete(g)
            db.session.commit()
            flash('Gruppo eliminato.', 'warning')

        elif azione == 'crea_da_proposta':
            from models.recupero import RecuperoImport
            materia     = request.form.get('materia','').strip()
            cognome_doc = request.form.get('cognome_doc','').strip()
            classi_list = request.form.getlist('classi_cb')
            classi      = ','.join(classi_list) if classi_list else request.form.get('classi','').strip()
            id_recdoc   = int(request.form['id_rec_docente'])
            max_ore     = int(request.form.get('max_ore', 10) or 10)
            max_ore_g   = int(request.form.get('max_ore_giorno', 2) or 2)
            g = RecuperoGruppo(
                id_rec_docente=id_recdoc, materia=materia,
                classi=classi, max_ore=max_ore, max_ore_giorno=max_ore_g,
            )
            db.session.add(g); db.session.flush()
            # Collega alunni dal staging
            imports = RecuperoImport.query.filter_by(
                anno_scol=ANNO, materia_norm=materia,
                cognome_docente=cognome_doc).all()
            for imp in imports:
                if imp.classe.upper() in [cl.strip().upper() for cl in classi_list]:
                    exists = RecuperoAlunno.query.filter_by(
                        id_gruppo=g.id, cognome=imp.cognome,
                        nome=imp.nome, classe=imp.classe).first()
                    if not exists:
                        db.session.add(RecuperoAlunno(
                            id_gruppo=g.id, classe=imp.classe,
                            cognome=imp.cognome, nome=imp.nome,
                            codice_fisc=imp.codice_fisc, email=imp.email,
                        ))
            db.session.commit()
            flash(f'Gruppo creato: {materia} ({classi}).', 'success')

        return redirect(url_for('recupero.gruppi'))

    # GET ─────────────────────────────────────────────────────────────
    from models.recupero import RecuperoImport
    from collections import defaultdict

    disponibili = (RecuperoDocente.query.filter_by(anno_scol=ANNO)
                   .join(Docente).order_by(Docente.cognome).all())
    gruppi_list = (RecuperoGruppo.query
                   .join(RecuperoDocente)
                   .join(Docente, Docente.id == RecuperoDocente.id_docente)
                   .filter(RecuperoDocente.anno_scol == ANNO,
                           RecuperoGruppo.periodo_codice == 'corsi_giugno')
                   .order_by(RecuperoGruppo.materia)
                   .all())

    # Famiglie di sinonimi — solo quelli confermati, nessuna inferenza per somiglianza
    # Ogni famiglia è un insieme: se due materie appartengono alla stessa famiglia, matchano.
    _FAMIGLIE = [
        # Latino
        {'LATINO', 'LINGUA LATINA', 'LINGUA E CULTURA LATINA'},
        # Italiano
        {'ITALIANO', 'LINGUA E LETTERATURA ITALIANA', 'LINGUA E CULTURA ITALIANA'},
        # Matematica
        {'MATEMATICA', 'MATEMATICA CON INFORMATICA', 'MATEMATICA E COMPLEMENTI DI MATEMATICA'},
        {'INFORMATICA', 'TECNOLOGIE INFORMATICHE'},
        # Storia
        {'STORIA', 'STORIA E GEOGRAFIA'},
        # Fisica
        {'FISICA', 'SCIENZE INTEGRATE (FISICA)'},
        # Inglese
        {'INGLESE', 'LINGUA INGLESE', 'LINGUA E CULTURA STRANIERA (INGLESE)'},
        # Tedesco (solo esatta, nessun alias noto)
        {'TEDESCO', 'LINGUA TEDESCA', 'LINGUA E CULTURA STRANIERA TEDESCO'},
        # Le altre materie (topografia, filosofia, economia, scienze naturali, informatica, ecc.)
        # NON hanno alias — vengono confrontate solo per corrispondenza esatta o sottostringa.
    ]

    def _match_sinonimi(m1, m2):
        m1u = m1.strip().upper()
        m2u = m2.strip().upper()
        if m1u == m2u:
            return True
        # Match solo se ENTRAMBE le materie sono membri esatti della stessa famiglia
        # Nessuna inferenza per sottostringa — evita MATEMATICA vs INFORMATICA
        for famiglia in _FAMIGLIE:
            f = {x.upper() for x in famiglia}
            if m1u in f and m2u in f:
                return True
        return False

    # Proposte dai dati importati
    imports = RecuperoImport.query.filter_by(anno_scol=ANNO).all()
    per_proposta = defaultdict(lambda: {'classi':set(),'alunni':[],'docente_raw':'',
                                        'cognome_doc':'','n_non_risposto':0,'n_non_aderisce':0})
    for imp in imports:
        key = (imp.materia_norm, imp.cognome_docente, imp.nome_ini_docente)
        per_proposta[key]['classi'].add(imp.classe)
        per_proposta[key]['alunni'].append(imp)
        per_proposta[key]['docente_raw'] = imp.docente_raw
        per_proposta[key]['cognome_doc'] = imp.cognome_docente
        if imp.stato_adesione == 'non_risposto': per_proposta[key]['n_non_risposto'] += 1
        elif imp.stato_adesione == 'non_aderisce': per_proposta[key]['n_non_aderisce'] += 1

    def trova_rd(cognome_doc, nome_ini=''):
        cogn = cognome_doc.upper()
        ini  = nome_ini.upper()
        if ini:
            for rd in disponibili:
                rc = rd.docente.cognome.upper()
                rn_ini = (rd.docente.nome or '').strip().upper()[:1]
                if rc == cogn and rn_ini == ini:
                    return rd
        trovati = [rd for rd in disponibili if rd.docente.cognome.upper() == cogn]
        if len(trovati) == 1: return trovati[0]
        for rd in disponibili:
            if rd.docente.cognome.upper() in cogn or cogn in rd.docente.cognome.upper():
                return rd
        return None

    # Mappa (materia_upper, cognome_upper) →︎ lista gruppi creati
    gruppi_per_proposta = {}
    for g in gruppi_list:
        cogn_g = (g.docente_rec.docente.cognome.upper()
                  if g.docente_rec and g.docente_rec.docente else '')
        key = (g.materia.upper(), cogn_g)
        gruppi_per_proposta.setdefault(key, []).append(g)

    proposte = []
    for (materia, cogn_doc, ini_doc), dati in sorted(per_proposta.items()):
        rd = trova_rd(cogn_doc, ini_doc)
        classi_ord = sorted(dati['classi'])
        n_ader = len([a for a in dati['alunni']
                      if a.stato_adesione in ('aderisce','sconosciuto')])
        # Per ogni classe della proposta, trova il gruppo che la copre (se esiste)
        mat_up = materia.upper()

        def _gruppi_per_materia(mat):
            result = []
            for g in gruppi_list:
                if _match_sinonimi(mat, g.materia) and g not in result:
                    result.append(g)
            return result

        gruppi_materia = _gruppi_per_materia(mat_up)

        # copertura_classi: {classe: gruppo_che_la_copre | None}
        # Una classe è coperta solo se:
        # 1. Esiste un gruppo con quella materia+classe, E
        # 2. Il docente del gruppo coincide con il docente assegnante (stesso cognome)
        #    OPPURE tutti gli alunni aderenti di quella classe sono già nel gruppo
        copertura_classi = {}
        for cls in classi_ord:
            cls_up = cls.upper()
            copertura_classi[cls] = None
            for g in gruppi_materia:
                classi_g = {cl.strip().upper() for cl in g.classi.split(',')}
                if cls_up not in classi_g:
                    continue
                # Verifica che il docente del gruppo corrisponda all'assegnante
                # oppure che gli alunni aderenti di questa classe siano già coperti
                doc_g = g.docente.cognome.upper() if g.docente else ''
                alunni_cls = [a for a in dati['alunni']
                              if a.classe.upper() == cls_up
                              and a.stato_adesione in ('aderisce','sconosciuto')]
                alunni_nel_gruppo = {(a.cognome, a.nome, a.classe) for a in g.alunni}
                alunni_aderenti_cls = {(a.cognome, a.nome, a.classe) for a in alunni_cls}
                # Coperto se: stesso cognome assegnante O tutti gli alunni già nel gruppo
                stesso_doc = cogn_doc.upper() == doc_g or cogn_doc.upper() in doc_g
                tutti_coperti = bool(alunni_aderenti_cls) and alunni_aderenti_cls.issubset(alunni_nel_gruppo)
                if stesso_doc or tutti_coperti:
                    copertura_classi[cls] = g
                    break

        gruppi_creati = list({g for g in copertura_classi.values() if g})
        classi_libere = [cls for cls, g in copertura_classi.items() if g is None]
        if n_ader > 0 or gruppi_creati:
            proposte.append({
                'materia': materia, 'cognome_doc': cogn_doc,
                'docente_raw': dati['docente_raw'],
                'classi': classi_ord,
                'n_alunni': n_ader, 'n_tot': len(dati['alunni']),
                'n_non_risposto': dati['n_non_risposto'],
                'n_non_aderisce': dati['n_non_aderisce'],
                'rd': rd,
                'gruppi_creati': gruppi_creati,
                'copertura_classi': copertura_classi,  # {cls: gruppo|None}
                'classi_libere': classi_libere,         # classi senza gruppo
            })

    # Classi disponibili per selettore manuale
    all_classi = sorted({imp.classe for imp in imports if imp.classe})
    materie_imp = sorted({imp.materia_norm for imp in imports if imp.materia_norm})

    # Classi per gruppo: quelle della materia del gruppo
    def _classi_per_materia(materia):
        mat = materia.strip().upper()
        classi = set()
        for i in imports:
            if i.classe and i.materia_norm and _match_sinonimi(mat, i.materia_norm):
                classi.add(i.classe)
        return sorted(classi) if classi else all_classi

    classi_per_gruppo = {g.id: _classi_per_materia(g.materia) for g in gruppi_list}

    # Conteggio aderenti per gruppo: conta gli import che matchano materia+classe del gruppo
    def _conta_aderenti(g):
        classi_g = {cl.strip().upper() for cl in g.classi.split(',')}
        tot = ader = 0
        for imp in imports:
            if imp.classe.upper() not in classi_g: continue
            if not _match_sinonimi(g.materia, imp.materia_norm or ''): continue
            tot += 1
            if imp.stato_adesione == 'aderisce':
                ader += 1
        return tot, ader

    aderenti_per_gruppo = {g.id: _conta_aderenti(g) for g in gruppi_list}

    tot_alunni_gruppi = sum(v[0] for v in aderenti_per_gruppo.values())
    tot_aderenti_gruppi = sum(v[1] for v in aderenti_per_gruppo.values())

    return render_template('recupero/gruppi.html',
        disponibili=disponibili, gruppi=gruppi_list,
        proposte=proposte, all_classi=all_classi,
        materie_imp=materie_imp,
        classi_per_gruppo=classi_per_gruppo,
        aderenti_per_gruppo=aderenti_per_gruppo,
        tot_alunni_gruppi=tot_alunni_gruppi,
        tot_aderenti_gruppi=tot_aderenti_gruppi,
    )


# ── CALENDARIO ────────────────────────────────────────────────────────
@recupero_bp.route('/recupero/calendario', methods=['GET', 'POST'])
def calendario():
    from modules.recupero_giugno_calendario import (
        azione_aggiungi, azione_modifica, azione_elimina,
        azione_elimina_giorno, azione_elimina_tutto, azione_completa_bozza,
        costruisci_dati_calendario,
    )

    if request.method == 'POST':
        azione = request.form.get('azione')
        handlers = {
            'aggiungi':       lambda: azione_aggiungi(request.form),
            'modifica':       lambda: azione_modifica(request.form),
            'elimina':        lambda: azione_elimina(request.form),
            'elimina_giorno': lambda: azione_elimina_giorno(request.form),
            'elimina_tutto':  lambda: azione_elimina_tutto(),
            'completa_bozza': lambda: azione_completa_bozza(),
        }
        handler = handlers.get(azione)
        if handler:
            risultato = handler()
            flash(risultato['msg'], risultato['cat'])
        return redirect(url_for('recupero.calendario'))

    # Tutto il calcolo (date disponibili, sync alunni da staging, conflitti)
    # è in modules/recupero_giugno_calendario.py: la route si limita a
    # chiamarlo e a passare il risultato al template.
    dati = costruisci_dati_calendario()
    return render_template('recupero/calendario.html', **dati)





# ── VINCOLI DOCENTE ───────────────────────────────────────────────────
@recupero_bp.route('/recupero/vincoli', methods=['GET', 'POST'])
def vincoli():
    if request.method == 'POST':
        azione = request.form.get('azione')

        if azione in ('aggiungi', 'aggiungi_multi'):
            from datetime import date as _date
            id_recdoc  = int(request.form['id_rec_docente'])
            giorni     = [int(g) for g in request.form.getlist('giorni')]
            if not giorni and 'giorno' in request.form:
                giorni = [int(request.form['giorno'])]
            ora_ini    = request.form.get('ora_inizio','08:00').strip()
            ora_fin    = request.form.get('ora_fine','13:00').strip()
            note       = request.form.get('note','').strip() or None
            # Vincolo date
            d_ini_s    = request.form.get('data_inizio','').strip()
            d_fin_s    = request.form.get('data_fine','').strip()
            d_ini = _date.fromisoformat(d_ini_s) if d_ini_s else None
            d_fin = _date.fromisoformat(d_fin_s) if d_fin_s else None

            if not giorni:
                # Nessun giorno selezionato →︎ vincolo solo su date (tutti i gg)
                giorni = [None]

            # classi_vincolo può arrivare come checkbox multipli o testo libero
            classi_list = request.form.getlist('classi_vincolo_cb')
            if classi_list:
                classi_v = ','.join(c.strip().upper() for c in classi_list if c.strip()) or None
            else:
                raw = request.form.get('classi_vincolo','').strip()
                classi_v = ','.join(c.strip().upper() for c in raw.split(',') if c.strip()) or None
            materia_v  = request.form.get('materia_vincolo','').strip() or None
            for giorno in giorni:
                db.session.add(RecuperoVincolo(
                    id_rec_docente=id_recdoc, anno_scol=ANNO,
                    giorno=giorno, ora_inizio=ora_ini,
                    ora_fine=ora_fin, note=note,
                    data_inizio=d_ini, data_fine=d_fin,
                    classi_vincolo=classi_v,
                    materia_vincolo=materia_v,
                ))
            db.session.commit()
            n = len([g for g in giorni if g is not None]) or 1
            flash(f'{"Fascia aggiunta" if n==1 else str(n)+" fasce aggiunte"}.', 'success')

        elif azione == 'elimina':
            vid = int(request.form['id'])
            v = RecuperoVincolo.query.get_or_404(vid)
            db.session.delete(v)
            db.session.commit()
            flash('Fascia eliminata.', 'warning')

        elif azione == 'modifica':
            from datetime import date as _date
            vid      = int(request.form['id'])
            v        = RecuperoVincolo.query.get_or_404(vid)
            v.ora_inizio  = request.form.get('ora_inizio','08:00').strip()
            v.ora_fine    = request.form.get('ora_fine','13:00').strip()
            d_ini_s  = request.form.get('data_inizio','').strip()
            d_fin_s  = request.form.get('data_fine','').strip()
            v.data_inizio = _date.fromisoformat(d_ini_s) if d_ini_s else None
            v.data_fine   = _date.fromisoformat(d_fin_s) if d_fin_s else None
            v.note           = request.form.get('note','').strip() or None
            classi_list2 = request.form.getlist('classi_vincolo_cb')
            if classi_list2:
                v.classi_vincolo = ','.join(c.strip().upper() for c in classi_list2 if c.strip()) or None
            else:
                raw2 = request.form.get('classi_vincolo','').strip()
                v.classi_vincolo = ','.join(c.strip().upper() for c in raw2.split(',') if c.strip()) or None
            v.materia_vincolo = request.form.get('materia_vincolo','').strip() or None
            db.session.commit()
            flash('Fascia aggiornata.', 'success')

        elif azione == 'elimina_tutti':
            id_recdoc = int(request.form['id_rec_docente'])
            RecuperoVincolo.query.filter_by(id_rec_docente=id_recdoc).delete()
            db.session.commit()
            flash('Disponibilità azzerata — il sistema userà il default (tutti i giorni 08:00–13:00).', 'info')

        # Torna al box del docente modificato
        anchor = ''
        if 'id_rec_docente' in request.form:
            anchor = f'#docente-{request.form["id_rec_docente"]}'
        elif 'id' in request.form:
            vid = request.form.get('id','')
            if vid.isdigit():
                vobj = RecuperoVincolo.query.get(int(vid))
                if vobj: anchor = f'#docente-{vobj.id_rec_docente}'
        return redirect(url_for('recupero.vincoli') + anchor)

    disponibili = (RecuperoDocente.query.filter_by(anno_scol=ANNO)
                   .join(Docente).order_by(Docente.cognome).all())
    GIORNI_NOMI = ['Lunedì','Martedì','Mercoledì','Giovedì','Venerdì']

    # Mappa id_recdoc →︎ classi con debiti nelle materie che il docente può coprire
    from models.recupero import RecuperoImport
    from collections import defaultdict

    classi_per_docente = {}
    for rd in disponibili:
        # Materie coperte: principale (può essere "ITALIANO, STORIA") + extra
        mat_principale_raw = (rd.docente.materia or '').strip().upper()
        # Splitta materie multiple nella colonna principale (es. "ITALIANO, STORIA")
        mat_principali = [m.strip() for m in mat_principale_raw.split(',') if m.strip()]
        materie_extra  = [m.strip().upper() for m in rd.materie_extra.split(',')
                          if m.strip()] if rd.materie_extra else []
        materie_coperte = mat_principali + materie_extra

        # Carica una volta sola tutte le combinazioni classe+materia
        tutti = (RecuperoImport.query
                 .filter_by(anno_scol=ANNO)
                 .with_entities(RecuperoImport.classe, RecuperoImport.materia_norm)
                 .distinct().all())

        # Sinonimi: materia breve/DB →︎ varianti presenti nel registro elettronico
        # Aggiornati sulle materie effettive del file importato
        _SINONIMI = {
            # Italiano
            'ITALIANO':        ('LINGUA E LETTERATURA ITALIANA','LETTERATURA ITALIANA','ITALIANA'),
            'LETTERATURA':     ('LINGUA E LETTERATURA ITALIANA','LETTERATURA ITALIANA','ITALIANA'),
            # Latino
            'LATINO':          ('LINGUA E CULTURA LATINA','LINGUA LATINA'),
            'LINGUA LATINA':   ('LINGUA E CULTURA LATINA','LINGUA LATINA'),
            # Storia
            'STORIA':          ('STORIA','STORIA E GEOGRAFIA'),
            'STORIA E GEOGRAFIA': ('STORIA','STORIA E GEOGRAFIA'),
            # Filosofia
            'FILOSOFIA':       ('FILOSOFIA',),
            # Matematica
            'MATEMATICA':      ('MATEMATICA','MATEMATICA E COMPLEMENTI DI MATEMATICA',
                                'MATEMATICA CON INFORMATICA'),
            # Fisica
            'FISICA':          ('FISICA','SCIENZE INTEGRATE (FISICA)'),
            # Scienze
            'SCIENZE':         ('SCIENZE NATURALI','BIOLOGIA','CHIMICA',
                                'SCIENZE INTEGRATE'),
            'SCIENZE NATURALI':('SCIENZE NATURALI','BIOLOGIA','CHIMICA'),
            'BIOLOGIA':        ('SCIENZE NATURALI','BIOLOGIA','CHIMICA'),
            'CHIMICA':         ('SCIENZE NATURALI','BIOLOGIA','CHIMICA'),
            # Inglese
            'INGLESE':         ('LINGUA E CULTURA STRANIERA (INGLESE)','LINGUA INGLESE','INGLESE'),
            'LINGUA INGLESE':  ('LINGUA E CULTURA STRANIERA (INGLESE)','LINGUA INGLESE','INGLESE'),
            # Tedesco
            'TEDESCO':         ('LINGUA E CULTURA STRANIERA TEDESCO','LINGUA TEDESCA','TEDESCO'),
            'LINGUA TEDESCA':  ('LINGUA E CULTURA STRANIERA TEDESCO','LINGUA TEDESCA','TEDESCO'),
            # Francese / Spagnolo
            'FRANCESE':        ('LINGUA FRANCESE','FRANCESE'),
            'SPAGNOLO':        ('LINGUA SPAGNOLA','SPAGNOLO'),
            # Informatica (NON includere 'MATEMATICA CON INFORMATICA': quella
            # materia resta di titolarità Matematica, non Informatica)
            'INFORMATICA':     ('INFORMATICA','TECNOLOGIE INFORMATICHE'),
            'TECNOLOGIE INFORMATICHE': ('INFORMATICA','TECNOLOGIE INFORMATICHE'),
            # Economia / Diritto / Estimo
            'ECONOMIA':        ('ECONOMIA','GEOPEDOLOGIA, ECONOMIA ED ESTIMO'),
            'ESTIMO':          ('GEOPEDOLOGIA, ECONOMIA ED ESTIMO',),
            'GEOPEDOLOGIA':    ('GEOPEDOLOGIA, ECONOMIA ED ESTIMO',),
            # Topografia
            'TOPOGRAFIA':      ('TOPOGRAFIA',),
            # Arte
            'ARTE':            ('ARTE','STORIA ARTE','DISEGNO'),
        }

        def _match_materie(mat_doc, mat_imp):
            md = mat_doc.strip().upper()
            mi = mat_imp.strip().upper()
            # Match diretto sottostringa
            if md in mi or mi in md:
                return True
            # Match tramite sinonimi
            for chiave, varianti in _SINONIMI.items():
                # Se la materia del docente è questa chiave o una variante
                doc_match = chiave in md or any(v in md for v in varianti)
                # E la materia dell'import è questa chiave o una variante
                imp_match = chiave in mi or any(v in mi for v in varianti)
                if doc_match and imp_match:
                    return True
            return False

        classi = set()
        for mat in materie_coperte:
            for cls, mat_imp in tutti:
                if cls and mat_imp and _match_materie(mat, mat_imp):
                    classi.add(cls)

        classi_per_docente[rd.id] = sorted(classi)

    return render_template('recupero/vincoli.html',
        disponibili=disponibili, GIORNI=GIORNI_NOMI,
        classi_per_docente=classi_per_docente, enumerate=enumerate)


# ── ALUNNI: import XLSX ───────────────────────────────────────────────
@recupero_bp.route('/recupero/alunni', methods=['GET', 'POST'])
def alunni():
    if request.method == 'POST':
        azione = request.form.get('azione')

        if azione == 'modifica_stato':
            imp_id = request.form.get('id')
            nuovo_stato = request.form.get('stato_adesione')
            stati_validi = ('aderisce','studio_ind','non_risposto','non_aderisce','sconosciuto')
            if imp_id and nuovo_stato in stati_validi:
                from models.recupero import RecuperoImport
                imp = RecuperoImport.query.get(int(imp_id))
                if imp:
                    imp.stato_adesione = nuovo_stato
                    db.session.commit()
            return redirect(url_for('recupero.alunni'))

        if azione == 'modifica_tipo_prova':
            # Modifica manuale del tipo prova per un singolo studente — utile
            # quando un docente comunica un cambio (es. da scritto a orale)
            # dopo l'import del file. Il valore salvato è già normalizzato
            # (scritto/orale/pratico/scritto_orale), cosi' _parse_tipo_prova
            # lo rilegge stabilmente senza ambiguita'.
            imp_id = request.form.get('id')
            nuovo_tipo = request.form.get('tipo_prova')
            tipi_validi = ('scritto', 'orale', 'pratico', 'scritto_orale')
            if imp_id and nuovo_tipo in tipi_validi:
                from models.recupero import RecuperoImport
                imp = RecuperoImport.query.get(int(imp_id))
                if imp:
                    imp.tipo_prova_raw = nuovo_tipo
                    db.session.commit()
            return redirect(url_for('recupero.alunni'))

        if azione == 'aggiungi_alunno':
            # Inserimento manuale di un singolo studente, con la stessa
            # normalizzazione usata per l'import del file Excel — cosi'
            # il nuovo record si comporta in modo identico a quelli
            # caricati da file (riconoscimento materia/docente, famiglie
            # sinonimi, ecc.) in entrambi i percorsi giugno e agosto.
            from models.recupero import RecuperoImport

            classe   = request.form.get('classe', '').strip().upper()
            cognome  = request.form.get('cognome', '').strip().upper()
            nome     = request.form.get('nome', '').strip().title()
            materia_in = request.form.get('materia', '').strip()
            docente_in = request.form.get('docente', '').strip()
            stato    = request.form.get('stato_adesione', 'sconosciuto')
            tipo_prova_in = request.form.get('tipo_prova', '').strip() or None

            stati_validi = ('aderisce','studio_ind','non_risposto','non_aderisce','sconosciuto')
            if stato not in stati_validi:
                stato = 'sconosciuto'

            if not (classe and cognome and nome and materia_in):
                flash('Classe, cognome, nome e materia sono obbligatori.', 'warning')
                return redirect(url_for('recupero.alunni'))

            materia_norm = _norm_materia(materia_in)
            cogn_doc, ini_doc = _split_cognome_nome(docente_in) if docente_in else ('', '')

            db.session.add(RecuperoImport(
                anno_scol=ANNO, classe=classe,
                cognome=cognome, nome=nome,
                materia_raw=materia_in[:200], materia_norm=materia_norm,
                docente_raw=docente_in[:200] or None,
                cognome_docente=cogn_doc or None,
                nome_ini_docente=ini_doc or None,
                stato_adesione=stato,
                tipo_prova_raw=tipo_prova_in,
            ))
            db.session.commit()
            flash(f'Alunno {cognome} {nome} aggiunto.', 'success')
            return redirect(url_for('recupero.alunni'))

        if azione == 'elimina_alunno':
            # Rimuove un singolo studente dallo staging E i collegamenti
            # RecuperoAlunno già fatti per QUELLA STESSA MATERIA (giugno e
            # agosto) — usa la famiglia di sinonimi materia, non solo
            # cognome+nome, perché lo stesso studente può comparire in più
            # righe per materie diverse (es. Matematica + Geopedologia):
            # eliminare una riga non deve toccare i gruppi delle altre.
            from models.recupero import RecuperoImport
            imp_id = request.form.get('id')
            if imp_id:
                imp = RecuperoImport.query.get(int(imp_id))
                if imp:
                    nome_completo = f'{imp.cognome} {imp.nome}'
                    mat_can_elim = _materia_canonica(imp.materia_norm or '')

                    candidati = RecuperoAlunno.query.filter_by(
                        cognome=imp.cognome, nome=imp.nome, classe=imp.classe).all()
                    n_rimossi = 0
                    for al in candidati:
                        g = RecuperoGruppo.query.get(al.id_gruppo)
                        if g and _materia_canonica(g.materia) == mat_can_elim:
                            db.session.delete(al)
                            n_rimossi += 1

                    db.session.delete(imp)
                    db.session.commit()
                    msg = f'Alunno {nome_completo} rimosso dall\'elenco.'
                    if n_rimossi:
                        msg += f' Rimosso anche da {n_rimossi} gruppo/i già calendarizzato/i per la stessa materia.'
                    flash(msg, 'warning')
            return redirect(url_for('recupero.alunni'))

        if azione == 'import':
            f = request.files.get('file_xlsx')
            if not f:
                flash('Nessun file selezionato.', 'warning')
                return redirect(url_for('recupero.alunni'))

            import pandas as pd, io
            from openpyxl import load_workbook
            from models.recupero import RecuperoImport

            file_bytes = f.read()

            # Leggi la formattazione (colori/barrato) con openpyxl
            wb = load_workbook(io.BytesIO(file_bytes))
            ws = wb.active
            # Mappa riga_excel →︎ stato_adesione
            stati_per_riga = {}
            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                cell = row[1]  # colonna B = cognome
                strike = cell.font.strike if cell.font else False
                fill   = cell.fill
                rgb    = fill.fgColor.rgb if fill and fill.fgColor and fill.fgColor.type == 'rgb' else None
                tema   = fill.fgColor.theme if fill and fill.fgColor and fill.fgColor.type == 'theme' else None
                if strike:
                    stato = 'non_aderisce'
                elif rgb == 'FFFF0000':
                    stato = 'non_risposto'
                elif tema == 9:
                    stato = 'aderisce'
                elif tema == 7:
                    stato = 'studio_ind'
                else:
                    stato = 'sconosciuto'
                stati_per_riga[i] = stato

            # Leggi dati con pandas (tutte le righe, non solo corso di recupero)
            df = pd.read_excel(io.BytesIO(file_bytes))

            # Svuota staging precedente
            RecuperoImport.query.filter_by(anno_scol=ANNO).delete()
            db.session.commit()

            # Cognomi noti dal DB (inclusi quelli composti da più parole, es.
            # "DEL PAPA", "DELLA MARIANNA"), ordinati per lunghezza decrescente
            # in numero di parole — cosi' il match prova prima le combinazioni
            # piu' lunghe e non spezza erroneamente un cognome composto.
            # (usa le funzioni globali _split_cognome_nome / _norm_materia,
            # definite più sotto nel modulo e condivise con l'inserimento
            # manuale dell'alunno)

            inseriti = {'aderisce':0,'studio_ind':0,'non_risposto':0,'non_aderisce':0,'sconosciuto':0}
            for idx, row in df.iterrows():
                # idx è 0-based, riga excel = idx+2
                riga_excel = idx + 2
                stato = stati_per_riga.get(riga_excel, 'sconosciuto')

                classe   = str(row['classe']).strip().upper()
                cognome  = str(row['cognome']).strip().upper()
                nome     = str(row['nome']).strip()
                recupero      = str(row.get('recupero','')).strip().lower()
                tipo_prova_raw = str(row.get('tipo prova','')).strip() or None
                materia  = _norm_materia(row['materia'])
                doc_raw  = str(row.get('docente','')).strip()
                cogn_doc, nome_ini_doc = _split_cognome_nome(doc_raw)
                cf       = str(row.get('codice_fisc','')).strip() or None
                email    = str(row.get('email','')).strip() or None
                if email and '@' not in email: email = None

                db.session.add(RecuperoImport(
                    anno_scol=ANNO, classe=classe,
                    cognome=cognome, nome=nome,
                    codice_fisc=cf, email=email,
                    materia_raw=str(row['materia']).strip(),
                    materia_norm=materia,
                    docente_raw=doc_raw,
                    cognome_docente=cogn_doc,
                    nome_ini_docente=nome_ini_doc,
                    stato_adesione=stato,
                    tipo_prova_raw=tipo_prova_raw,
                ))
                inseriti[stato] = inseriti.get(stato, 0) + 1

            db.session.commit()
            flash(
                f'Importati {sum(inseriti.values())} alunni — '
                f'✓︎ {inseriti["aderisce"]} aderiscono, '
                f'▥︎ {inseriti["studio_ind"]} studio individuale, '
                f'? {inseriti["non_risposto"]} non hanno risposto, '
                f'✕︎ {inseriti["non_aderisce"]} non aderiscono.',
                'success'
            )

        elif azione == 'elimina_tutti':
            from models.recupero import RecuperoImport
            # Lo staging (RecuperoImport) e' condiviso tra giugno e agosto:
            # eliminandolo, gli alunni collegati ai gruppi di ENTRAMBI i
            # periodi restano agganciati a dati che non esistono piu' nello
            # staging — quindi vanno puliti insieme, non solo quelli di giugno.
            RecuperoImport.query.filter_by(anno_scol=ANNO).delete()
            anno_ids = [rd.id for rd in RecuperoDocente.query.filter_by(anno_scol=ANNO).all()]
            gruppi_ids = [g.id for g in RecuperoGruppo.query.filter(
                RecuperoGruppo.id_rec_docente.in_(anno_ids)).all()]
            n = RecuperoAlunno.query.filter(RecuperoAlunno.id_gruppo.in_(gruppi_ids)).delete(synchronize_session=False)
            db.session.commit()
            flash(f'Eliminati {n} alunni (corsi giugno + prove agosto) e staging pulito.', 'warning')

        elif azione == 'elimina':
            aid = int(request.form['id'])
            a   = RecuperoAlunno.query.get_or_404(aid)
            db.session.delete(a)
            db.session.commit()

        return redirect(url_for('recupero.alunni'))

    # GET: mostra alunni dallo staging (RecuperoImport)
    from models.recupero import RecuperoImport
    from collections import defaultdict

    imports = (RecuperoImport.query
               .filter_by(anno_scol=ANNO)
               .order_by(RecuperoImport.materia_norm, RecuperoImport.cognome)
               .all())

    # Raggruppa per materia+docente per visualizzazione
    per_materia = defaultdict(list)
    for imp in imports:
        key = (imp.materia_norm, imp.cognome_docente, imp.nome_ini_docente)
        per_materia[key].append(imp)

    tot = len(imports)
    # Conteggi per stato
    conteggi = {'aderisce':0,'studio_ind':0,'non_risposto':0,'non_aderisce':0,'sconosciuto':0}
    for imp in imports:
        conteggi[imp.stato_adesione] = conteggi.get(imp.stato_adesione, 0) + 1

    return render_template('recupero/alunni.html',
        per_materia=per_materia, tot=tot, conteggi=conteggi,
        parse_tipo_prova=_parse_tipo_prova, TIPO_PROVA_LABEL=TIPO_PROVA_LABEL,
        classi_attive=_classi_attive())


# ── GENERA BOZZA CALENDARIO ───────────────────────────────────────────
@recupero_bp.route('/recupero/genera-bozza', methods=['POST'])
def genera_bozza():
    """
    Genera automaticamente una bozza di calendario che:
    - Rispetta i vincoli di disponibilità di ogni docente
    - Garantisce che nessun alunno abbia due lezioni sovrapposte
    - Rispetta max 2h/giorno per gruppo e max ore totali per gruppo
    """
    import json
    from datetime import timedelta

    # Elimina lezioni esistenti (solo bozza)
    conferma = request.form.get('conferma_elimina') == '1'
    if not conferma:
        flash('Seleziona la casella di conferma prima di generare la bozza.', 'warning')
        return redirect(url_for('recupero.calendario'))

    # Elimina SOLO le lezioni dei corsi di giugno (mai quelle delle prove
    # di agosto, che sono un periodo completamente separato).
    anno_ids = [rd.id for rd in RecuperoDocente.query.filter_by(anno_scol=ANNO).all()]
    gruppi_ids = [g.id for g in RecuperoGruppo.query.filter(
        RecuperoGruppo.id_rec_docente.in_(anno_ids),
        RecuperoGruppo.periodo_codice == 'corsi_giugno').all()]
    RecuperoLezione.query.filter(
        RecuperoLezione.id_gruppo.in_(gruppi_ids)).delete(synchronize_session=False)
    db.session.commit()

    gruppi = (RecuperoGruppo.query
              .join(RecuperoDocente)
              .filter(RecuperoDocente.anno_scol == ANNO,
                      RecuperoGruppo.periodo_codice == 'corsi_giugno')
              .all())

    # Date disponibili (lun-ven, 18/6-1/7)
    date_disp = []
    cur = DATA_INIZIO
    while cur <= DATA_FINE:
        if cur.weekday() < 5:
            date_disp.append(cur)
        cur += timedelta(days=1)

    # Slot già occupati per alunno: {(cf o nome_classe): set di (data, ora_ini, ora_fine)}
    slot_alunni = {}  # chiave: (cognome, nome, classe) →︎ set di (data, ini_min, fin_min)

    def _t(s):
        try:
            h, m = map(int, s.split(':'))
            return h * 60 + m
        except Exception:
            return 0

    def _sovrappone(d, ini, fin, occupied):
        ini_m, fin_m = _t(ini), _t(fin)
        for od, oi, of in occupied:
            if od == d and oi < fin_m and of > ini_m:
                return True
        return False

    def _priorita_gruppo(g):
        """
        Calcola punteggio di priorità (più alto = più difficile = va schedulato prima).
        Criteri:
          1. Vincoli più stretti: meno ore disponibili totali nel periodo
          2. Numero alunni aderenti: più alunni = più urgente
          3. Numero di vincoli: più vincoli = meno flessibilità
        """
        vincoli = g.docente_rec.vincoli if g.docente_rec else []

        # Ore disponibili totali nel periodo = somma delle finestre orarie per data
        ore_disp_totali = 0.0
        giorni_disp = set()
        for data in date_disp:
            wd = data.weekday()
            for v in vincoli:
                if v.data_inizio and data < v.data_inizio: continue
                if v.data_fine   and data > v.data_fine:   continue
                if v.giorno is not None and v.giorno != wd: continue
                try:
                    h1,m1 = map(int,v.ora_inizio.split(':'))
                    h2,m2 = map(int,v.ora_fine.split(':'))
                    ore_disp_totali += (h2*60+m2 - h1*60-m1) / 60
                    giorni_disp.add(data)
                except Exception:
                    pass
        if not vincoli:
            # Nessun vincolo = massima disponibilità →︎ bassa priorità
            ore_disp_totali = 999
            giorni_disp = set(date_disp)

        n_alunni   = len(g.alunni)
        n_vincoli  = len(vincoli)

        # Score: meno ore disponibili →︎ score alto; più alunni →︎ score alto
        # Usiamo inverso delle ore disponibili normalizzato
        score_ore      = 1.0 / (ore_disp_totali + 1)    # alto se poche ore
        score_alunni   = n_alunni / 50.0                  # normalizzato su 50
        score_vincoli  = n_vincoli / 20.0                 # normalizzato su 20

        # Peso: ore_disponibili conta di più
        return score_ore * 3 + score_alunni * 2 + score_vincoli * 1

    # Ordina gruppi: priorità decrescente (più difficile prima)
    gruppi_ordinati = sorted(gruppi, key=_priorita_gruppo, reverse=True)

    inserite = 0
    saltati  = []
    # Tracking slot occupati per docente: {id_rec_docente: {data: fine_ultima_lezione_min}}
    slot_docente = {}
    for g in gruppi_ordinati:
        vincoli_doc = g.docente_rec.vincoli
        ha_vincoli = bool(vincoli_doc)

        def _slot_per_data(data):
            # Restituisce [(ora_ini, ora_fine)] validi per questa data e questo gruppo
            wd = data.weekday()
            if not ha_vincoli:
                return [('08:00','13:00')]
            classi_gruppo = {c.strip().upper() for c in g.classi.split(',')}
            slots = []
            for v in vincoli_doc:
                # Controlla vincolo date
                if v.data_inizio and data < v.data_inizio: continue
                if v.data_fine   and data > v.data_fine:   continue
                # Controlla giorno
                if v.giorno is not None and v.giorno != wd: continue
                # Controlla vincolo classi
                if v.classi_vincolo:
                    classi_v = {cv.strip().upper() for cv in v.classi_vincolo.split(',')}
                    if not classi_gruppo.intersection(classi_v):
                        continue
                # Controlla vincolo materia
                if v.materia_vincolo:
                    mat_v = v.materia_vincolo.strip().upper()
                    mat_g = g.materia.strip().upper()
                    if mat_v not in mat_g and mat_g not in mat_v:
                        continue
                slots.append((v.ora_inizio, v.ora_fine))
            return slots if slots else []

        max_ore_tot  = g.max_ore or 10
        max_ore_g    = g.max_ore_giorno or 2
        ore_pian     = 0
        ore_per_data = {}

        alunni_g = g.alunni

        for data in date_disp:
            if ore_pian >= max_ore_tot:
                break
            wd = data.weekday()
            ore_ok = _slot_per_data(data)
            if not ore_ok:
                continue

            ore_oggi = ore_per_data.get(data, 0)
            if ore_oggi >= max_ore_g:
                continue

            # Prova ogni fascia disponibile per questo giorno
            for fascia_ini, fascia_fin in ore_ok:
                fascia_durata = (_t(fascia_fin) - _t(fascia_ini)) / 60
                if fascia_durata <= 0:
                    continue

                durata_h = min(fascia_durata, max_ore_g - ore_oggi,
                               max_ore_tot - ore_pian, 2)
                if durata_h <= 0:
                    continue

                # Ora di inizio: dopo l'ultima lezione del docente in questa data
                doc_id = g.id_rec_docente
                fine_doc = slot_docente.get(doc_id, {}).get(data, _t(fascia_ini))
                # Se il docente ha già una lezione, inizia subito dopo (con pausa 30min)
                if fine_doc > _t(fascia_ini):
                    ini_m = fine_doc + 30
                else:
                    ini_m = _t(fascia_ini)

                # Verifica che rimanga dentro la fascia
                fin_max = _t(fascia_fin)
                durata_m = int(durata_h * 60)
                if ini_m + durata_m > fin_max:
                    continue

                ini = f'{ini_m // 60:02d}:{ini_m % 60:02d}'
                fin_m_val = ini_m + durata_m
                fin = f'{fin_m_val // 60:02d}:{fin_m_val % 60:02d}'

                # Verifica no sovrapposizione alunni
                conflitto = False
                for al in alunni_g:
                    key = (al.cognome, al.nome, al.classe)
                    if _sovrappone(data, ini, fin, slot_alunni.get(key, set())):
                        conflitto = True
                        break

                if conflitto:
                    continue

                db.session.add(RecuperoLezione(
                    id_gruppo=g.id, data=data,
                    ora_inizio=ini, ora_fine=fin,
                ))
                ore_pian += durata_h
                ore_per_data[data] = ore_per_data.get(data, 0) + durata_h
                inserite += 1

                # Aggiorna slot alunni e slot docente
                for al in alunni_g:
                    key = (al.cognome, al.nome, al.classe)
                    slot_alunni.setdefault(key, set()).add((data, ini_m, fin_m_val))
                slot_docente.setdefault(doc_id, {})[data] = fin_m_val
                break  # una fascia per giorno per questo gruppo

                break  # già gestito sotto
            else:
                # Nessuna fascia andata a buon fine per questa data →︎ continua
                pass

        if ore_pian == 0:
            saltati.append(g)

    db.session.commit()

    msg = f'Bozza generata: {inserite} lezioni inserite.'
    if saltati:
        nomi = ', '.join(f'{g.materia[:20]} ({g.docente.cognome if g.docente else "?"})' for g in saltati)
        msg += f' ⚠︎ {len(saltati)} gruppi senza slot: {nomi}.'
    flash(msg, 'success' if not saltati else 'warning')
    return redirect(url_for('recupero.calendario'))


# ── TABELLA STAGING IMPORT ────────────────────────────────────────────
# (usata da proposte e import alunni)

# ── PROPOSTE GRUPPI DA RECUPERI IMPORTATI ────────────────────────────
@recupero_bp.route('/recupero/proposte')
def proposte():
    from collections import defaultdict
    from models.recupero import RecuperoImport

    imports = RecuperoImport.query.filter_by(anno_scol=ANNO).all()
    disponibili = (RecuperoDocente.query.filter_by(anno_scol=ANNO)
                   .join(Docente).order_by(Docente.cognome).all())

    if not imports:
        return render_template('recupero/proposte.html',
            proposte=[], disponibili=disponibili, anno=ANNO)

    per_gruppo = defaultdict(lambda: {
        'classi': set(), 'alunni': [], 'docente_raw': '',
        'n_non_risposto': 0, 'n_non_aderisce': 0
    })
    for imp in imports:
        key = (imp.materia_norm, imp.cognome_docente, imp.nome_ini_docente)
        per_gruppo[key]['classi'].add(imp.classe)
        per_gruppo[key]['alunni'].append(imp)
        per_gruppo[key]['docente_raw'] = imp.docente_raw
        if imp.stato_adesione == 'non_risposto':
            per_gruppo[key]['n_non_risposto'] += 1
        elif imp.stato_adesione == 'non_aderisce':
            per_gruppo[key]['n_non_aderisce'] += 1

    def trova_disponibile(cognome_doc, nome_ini=''):
        cogn = cognome_doc.upper()
        ini  = nome_ini.upper()
        # Match preciso cognome + iniziale
        if ini:
            for rd in disponibili:
                rc = rd.docente.cognome.upper()
                rn_ini = (rd.docente.nome or '').strip().upper()[:1]
                if rc == cogn and rn_ini == ini:
                    return rd
        # Fallback solo cognome se non ambiguo
        trovati = [rd for rd in disponibili if rd.docente.cognome.upper() == cogn]
        if len(trovati) == 1:
            return trovati[0]
        # Ultimo tentativo: cognome contenuto
        for rd in disponibili:
            if rd.docente.cognome.upper() in cogn or cogn in rd.docente.cognome.upper():
                return rd
        return None

    gruppi_esistenti = {}
    for g in (RecuperoGruppo.query.join(RecuperoDocente)
              .filter(RecuperoDocente.anno_scol == ANNO,
                      RecuperoGruppo.periodo_codice == 'corsi_giugno').all()):
        gruppi_esistenti[(g.materia.upper(), g.id_rec_docente)] = g

    proposte_list = []
    for (materia, cogn_doc, ini_doc), dati in sorted(per_gruppo.items()):
        rd_sug = trova_disponibile(cogn_doc, ini_doc)
        alunni_aderiscono = [a for a in dati['alunni']
                              if a.stato_adesione in ('aderisce','sconosciuto','studio_ind')]
        proposte_list.append({
            'materia':          materia,
            'docente_raw':      dati['docente_raw'],
            'cognome_doc':      cogn_doc,
            'classi':           ', '.join(sorted(dati['classi'])),
            'n_alunni':         len(alunni_aderiscono),
            'n_tot':            len(dati['alunni']),
            'n_non_risposto':   dati['n_non_risposto'],
            'n_non_aderisce':   dati['n_non_aderisce'],
            'rd_suggerito':     rd_sug,
            'gruppo_esistente': gruppi_esistenti.get(
                (materia.upper(), rd_sug.id)) if rd_sug else None,
        })

    return render_template('recupero/proposte.html',
        proposte=proposte_list, disponibili=disponibili, anno=ANNO)


@recupero_bp.route('/recupero/proposte/crea', methods=['POST'])
def crea_da_proposta():
    from models.recupero import RecuperoImport

    materia     = request.form.get('materia','').strip()
    cognome_doc = request.form.get('cognome_doc','').strip()
    classi      = request.form.get('classi','').strip()
    id_recdoc   = int(request.form['id_rec_docente'])
    max_ore     = int(request.form.get('max_ore', 10) or 10)
    max_ore_g   = int(request.form.get('max_ore_giorno', 2) or 2)

    # Permette gruppi multipli per stessa materia+docente (turni)
    g = RecuperoGruppo(
        id_rec_docente=id_recdoc, materia=materia,
        classi=classi, max_ore=max_ore, max_ore_giorno=max_ore_g,
    )
    db.session.add(g)
    db.session.flush()
    flash(f'Gruppo creato: {materia} ({classi}).', 'success')

    imports = RecuperoImport.query.filter_by(
        anno_scol=ANNO, materia_norm=materia,
        cognome_docente=cognome_doc).all()

    for imp in imports:
        exists = RecuperoAlunno.query.filter_by(
            id_gruppo=g.id, cognome=imp.cognome,
            nome=imp.nome, classe=imp.classe).first()
        if not exists:
            db.session.add(RecuperoAlunno(
                id_gruppo=g.id, classe=imp.classe,
                cognome=imp.cognome, nome=imp.nome,
                codice_fisc=imp.codice_fisc, email=imp.email,
            ))

    db.session.commit()
    return redirect(url_for('recupero.proposte'))
