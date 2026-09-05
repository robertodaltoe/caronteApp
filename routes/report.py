import os, io
from flask import Blueprint, render_template, request, send_file, abort, flash, redirect, url_for
from models import db
from models.docente import Docente
from models.movimento_banca_ore import MovimentoBancaOre
from models.supplenza import Supplenza
from models.assenza import Assenza
from sqlalchemy import func
from datetime import date


def get_ore_ist_docente(id_docente, anno=None):
    """
    Calcola ore bucket A e B per un docente nell'a.s. (solo presenze=presente).
    Usato per contatore CCNL art.44.
    """
    from models.attivita_ist import AttivitaIst, AttivitaIstPresenza, TIPI_ATTIVITA
    if anno is None:
        oggi = date.today()
        anno_ini = date(oggi.year if oggi.month >= 9 else oggi.year - 1, 9, 1)
        anno_fin = date(anno_ini.year + 1, 8, 31)
    else:
        anno_ini = date(int(anno[:4]), 9, 1)
        anno_fin = date(int(anno[:4]) + 1, 8, 31)

    from config_istituto import get_dati_istituto
    limite = get_dati_istituto()['ore_ist_limite']

    try:
        presenze = (AttivitaIstPresenza.query
                    .join(AttivitaIst,
                          AttivitaIst.id == AttivitaIstPresenza.id_attivita)
                    .filter(AttivitaIstPresenza.id_docente == id_docente,
                            AttivitaIstPresenza.stato == 'presente',
                            AttivitaIst.data >= anno_ini,
                            AttivitaIst.data <= anno_fin)
                    .all())
    except Exception:
        return {'A': 0.0, 'B': 0.0, 'limite': limite}

    ore_a = round(sum(p.ore_effettive for p in presenze
                      if TIPI_ATTIVITA.get(p.attivita.tipo, {}).get('bucket') == 'A'), 1)
    ore_b = round(sum(p.ore_effettive for p in presenze
                      if TIPI_ATTIVITA.get(p.attivita.tipo, {}).get('bucket') == 'B'), 1)

    # Dettaglio per il prospetto: lista presenze ordinate per data
    dettaglio = sorted(presenze, key=lambda p: p.attivita.data)

    return {'A': ore_a, 'B': ore_b, 'limite': limite, 'dettaglio': dettaglio}

report_bp = Blueprint('report', __name__)

# ── Costanti ──────────────────────────────────────────────────
TIPI_SUPPLENZA = ('supplenza_recupero', 'supplenza_completamento',
                   'supplenza_potenziamento', 'supplenza_disposizione')
TIPI_PERMESSO  = ('permesso', 'assenza', 'permesso_orario')
TIPI_PERM_IST  = ('permesso_ist',)   # permesso orario su att. istituzionali
TIPI_CIVICA    = ('civica', 'ed_civica')
TIPI_PAGAMENTO = ('supplenza_pagamento',)


def get_saldi_docente(id_docente, anno_scol=None):
    """
    Calcola saldi per un docente nell'anno scolastico indicato (default:
    l'anno corrente), separando effettivo (<=oggi) e previsto (>oggi).
    Ritorna dict con chiavi:
      supplenze, permessi, civica, pagamento  — tutto (effettivo + previsto)
      sup_svolte, perm_svolte, civ_svolte     — solo fino a oggi
      sup_prev, perm_prev, civ_prev           — solo future

    NB: il saldo è sempre limitato a un singolo anno scolastico (colonna
    anno_scol) — non è più un totale cumulativo su tutta la storia del
    docente. Per consultare un anno diverso da quello corrente, passare
    esplicitamente anno_scol.
    """
    from config_anno import get_anno_corrente
    if anno_scol is None:
        anno_scol = get_anno_corrente()

    oggi = date.today()
    movimenti = MovimentoBancaOre.query.filter_by(
        id_docente=id_docente, anno_scol=anno_scol).all()

    def somma(movs, tipi, abs_val=False):
        vals = [abs(m.minuti) if abs_val else m.minuti
                for m in movs if m.tipo in tipi]
        return sum(vals) // 60

    svolti  = [m for m in movimenti if m.data and m.data <= oggi]
    previsti = [m for m in movimenti if m.data and m.data > oggi]

    return {
        # Totale (per retrocompatibilità)
        'supplenze':   somma(movimenti, TIPI_SUPPLENZA),
        'permessi':    somma(movimenti, TIPI_PERMESSO, abs_val=True),
        'perm_ist':    somma(movimenti, TIPI_PERM_IST, abs_val=True),
        'civica':      somma(movimenti, TIPI_CIVICA,   abs_val=True),
        'pagamento':   somma(movimenti, TIPI_PAGAMENTO, abs_val=True),
        # Effettivo (svolto — <= oggi)
        'sup_svolte':  somma(svolti,  TIPI_SUPPLENZA),
        'perm_svolte': somma(svolti,  TIPI_PERMESSO, abs_val=True),
        'civ_svolte':  somma(svolti,  TIPI_CIVICA,   abs_val=True),
        # Previsto (futuro — > oggi)
        'sup_prev':  somma(previsti, TIPI_SUPPLENZA),
        'perm_prev': somma(previsti, TIPI_PERMESSO, abs_val=True),
        'civ_prev':  somma(previsti, TIPI_CIVICA,   abs_val=True),
    }


def get_storico_settimanale(id_docente, anno_scol=None):
    """
    Raggruppa i movimenti per settimana (data), limitati all'anno
    scolastico indicato (default: quello corrente).
    Ritorna lista di dict ordinata per data.
    """
    from config_anno import get_anno_corrente
    if anno_scol is None:
        anno_scol = get_anno_corrente()

    movimenti = (MovimentoBancaOre.query
                 .filter_by(id_docente=id_docente, anno_scol=anno_scol)
                 .order_by(MovimentoBancaOre.data)
                 .all())

    # Raggruppa per data
    from collections import defaultdict
    per_data = defaultdict(lambda: {
        'supplenze': 0, 'permessi': 0, 'civica': 0, 'pagamento': 0, 'altro': 0
    })
    for m in movimenti:
        d = m.data
        if m.tipo in TIPI_SUPPLENZA:
            per_data[d]['supplenze'] += m.minuti // 60
        elif m.tipo in TIPI_PERMESSO:
            per_data[d]['permessi'] += abs(m.minuti) // 60
        elif m.tipo in TIPI_CIVICA:
            per_data[d]['civica'] += abs(m.minuti) // 60
        elif m.tipo in TIPI_PAGAMENTO:
            per_data[d]['pagamento'] += abs(m.minuti) // 60
        else:
            per_data[d]['altro'] += m.minuti // 60

    return [{'data': d, **v} for d, v in sorted(per_data.items())]


# ── SCADENZA 3 MESI (accordo sindacale: ore da recuperare/richiedere) ──
def _lotti_aperti_docente(id_docente, anno_scol, oggi=None):
    """
    Da accordo sindacale, ogni singola ora a debito (da recuperare) o a
    credito (da richiedere — pagamento o recupero) ha una propria scadenza
    di 3 mesi dalla data in cui è maturata — non il saldo complessivo del
    docente. Esempio: un'ora di permesso presa il 31/01 scade il 30/04;
    un'altra presa il 4/02 scade il 4/05, indipendentemente l'una dall'altra.

    Per calcolarlo servono lotti distinti per data, abbinati in ordine FIFO
    quando arrivano movimenti di segno opposto che li compensano (es. una
    supplenza svolta compensa, a partire dal permesso più vecchio ancora
    aperto, il debito accumulato con i permessi). Non un semplice saldo
    cumulato: due movimenti dello stesso segno (es. due permessi) NON si
    compensano fra loro e restano due lotti distinti, ciascuno con la
    propria scadenza.

    Ritorna la lista dei lotti ancora aperti (non ancora compensati) per il
    docente, con: data_apertura, minuti (con segno: negativo=debito,
    positivo=credito), scadenza (data_apertura + 3 mesi di calendario),
    scaduto (bool).
    """
    from dateutil.relativedelta import relativedelta
    from collections import deque

    if oggi is None:
        oggi = date.today()

    from config_istituto import get_dati_istituto
    mesi_scadenza = get_dati_istituto()['scadenza_saldo_mesi']

    movs = (MovimentoBancaOre.query
            .filter_by(id_docente=id_docente, anno_scol=anno_scol)
            .filter(MovimentoBancaOre.data <= oggi)
            .order_by(MovimentoBancaOre.data, MovimentoBancaOre.id)
            .all())

    lotti = deque()  # ciascun elemento: [data_apertura, minuti_residui_con_segno]
    for m in movs:
        resto = m.minuti
        if resto == 0:
            continue
        # Compensa FIFO con i lotti più vecchi di segno opposto
        while resto != 0 and lotti and (lotti[0][1] > 0) != (resto > 0):
            piu_vecchio = lotti[0]
            if abs(piu_vecchio[1]) <= abs(resto):
                resto += piu_vecchio[1]
                lotti.popleft()
            else:
                piu_vecchio[1] += resto
                resto = 0
        if resto != 0:
            lotti.append([m.data, resto])

    risultati = []
    for data_apertura, minuti in lotti:
        scadenza = data_apertura + relativedelta(months=mesi_scadenza)
        risultati.append({
            'data_apertura': data_apertura,
            'minuti': minuti,
            'ore': abs(minuti) // 60,
            'tipo': 'credito' if minuti > 0 else 'debito',
            'scadenza': scadenza,
            'scaduto': oggi > scadenza,
        })
    return risultati


def _scadenza_saldi(docenti, anno_scol):
    """
    Applica `_lotti_aperti_docente` a tutti i docenti e restituisce, per
    ciascuno con almeno un lotto scaduto, l'elenco dei lotti scaduti.
    Indicatore pensato per informare (non bloccare): segnale di
    qualità/puntualità nella gestione della banca ore per il DS, non un
    blocco operativo.
    """
    oggi = date.today()
    risultati = []
    for d in docenti:
        lotti = _lotti_aperti_docente(d.id, anno_scol, oggi=oggi)
        scaduti = [l for l in lotti if l['scaduto']]
        if not scaduti:
            continue
        scaduti.sort(key=lambda l: l['data_apertura'])
        risultati.append({
            'docente': d,
            'lotti_scaduti': scaduti,
            'n_scaduti': len(scaduti),
            'piu_vecchio': scaduti[0]['data_apertura'],
            'eta_giorni_piu_vecchio': (oggi - scaduti[0]['data_apertura']).days,
        })
    return risultati


# ── CRUSCOTTO DI MONITORAGGIO (hub /report) ──────────────────
def _dati_cruscotto(docenti, saldi, costo_ora):
    """
    Calcola i dati del cruscotto semplificato: numeri di sintesi, docenti
    in credito/debito estremo, andamento assenze/supplenze del mese,
    alert su soglie (ore istituzionali vicine al limite, supplenze
    scoperte nei prossimi giorni). Riusa i saldi già calcolati da index()
    invece di ricalcolarli.
    """
    from datetime import timedelta

    oggi = date.today()

    tot_credito = tot_debito = 0
    critici = []   # netto <= -5h → debito rilevante
    alti    = []   # netto >= 8h  → credito rilevante
    for d in docenti:
        s = saldi.get(d.id, {})
        netto = s.get('netto_eff', 0)
        if netto > 0:
            tot_credito += netto
        elif netto < 0:
            tot_debito += netto
        if netto <= -5:
            critici.append({'docente': d, 'saldo': netto})
        if netto >= 8:
            alti.append({'docente': d, 'saldo': netto})
    critici.sort(key=lambda x: x['saldo'])
    alti.sort(key=lambda x: -x['saldo'])

    # Andamento assenze/supplenze: mese corrente (fino a oggi) vs stesso
    # periodo del mese precedente, per un confronto omogeneo.
    primo_mese = oggi.replace(day=1)
    if primo_mese.month == 1:
        primo_mese_prec = primo_mese.replace(year=primo_mese.year - 1, month=12)
    else:
        primo_mese_prec = primo_mese.replace(month=primo_mese.month - 1)
    giorni_trascorsi = (oggi - primo_mese).days
    fine_confronto_prec = primo_mese_prec + timedelta(days=giorni_trascorsi)

    n_assenze_mese = Assenza.query.filter(
        Assenza.data >= primo_mese, Assenza.data <= oggi).count()
    n_assenze_mese_prec = Assenza.query.filter(
        Assenza.data >= primo_mese_prec, Assenza.data <= fine_confronto_prec).count()

    n_supplenze_mese = Supplenza.query.filter(
        Supplenza.data >= primo_mese, Supplenza.data <= oggi,
        Supplenza.stato == 'assegnata').count()
    n_supplenze_mese_prec = Supplenza.query.filter(
        Supplenza.data >= primo_mese_prec, Supplenza.data <= fine_confronto_prec,
        Supplenza.stato == 'assegnata').count()

    # Alert soglie: supplenze scoperte nei prossimi 7 giorni + docenti
    # vicini al limite ore istituzionali CCNL art.44 (soglia 32h su 40h).
    supplenze_scoperte_7gg = Supplenza.query.filter(
        Supplenza.stato == 'scoperta',
        Supplenza.data >= oggi,
        Supplenza.data <= oggi + timedelta(days=7)).count()

    from config_istituto import get_dati_istituto as _get_dati_istituto_soglia
    soglia_alert_ist = _get_dati_istituto_soglia()['ore_ist_soglia_alert']
    alert_ore_ist = []
    for d in docenti:
        oi = get_ore_ist_docente(d.id)
        if oi.get('A', 0) >= soglia_alert_ist or oi.get('B', 0) >= soglia_alert_ist:
            alert_ore_ist.append({'docente': d, 'A': oi.get('A', 0), 'B': oi.get('B', 0)})
    alert_ore_ist.sort(key=lambda x: -max(x['A'], x['B']))

    # Scadenza 3 mesi (accordo sindacale) — informativo, non bloccante.
    # Ogni voce riguarda un docente con almeno un'ora (lotto) scaduta;
    # 'n_scaduti' è il numero di lotti/ore scadute per quel docente,
    # 'eta_giorni_piu_vecchio' l'età del lotto scaduto più vecchio.
    from config_anno import get_anno_corrente
    scadenze_oltre_termine = sorted(
        _scadenza_saldi(docenti, get_anno_corrente()),
        key=lambda x: -x['eta_giorni_piu_vecchio'])

    return {
        'tot_credito': tot_credito, 'tot_debito': tot_debito,
        'costo_stimato_credito': tot_credito * costo_ora,
        'n_docenti': len(docenti),
        'critici': critici, 'alti': alti,
        'n_assenze_mese': n_assenze_mese, 'n_assenze_mese_prec': n_assenze_mese_prec,
        'n_supplenze_mese': n_supplenze_mese, 'n_supplenze_mese_prec': n_supplenze_mese_prec,
        'supplenze_scoperte_7gg': supplenze_scoperte_7gg,
        'alert_ore_ist': alert_ore_ist,
        'scadenze_oltre_termine': scadenze_oltre_termine,
        'oggi': oggi,
    }


# ── INDICE REPORT — HUB con tab (Cruscotto / Dirigente / Docenti / Segreteria) ──
@report_bp.route('/report')
def index():
    tab = request.args.get('tab', 'cruscotto')
    if tab not in ('cruscotto', 'dirigente', 'docenti', 'segreteria'):
        tab = 'cruscotto'

    docenti = Docente.query.filter_by(attivo=True).order_by(Docente.cognome).all()

    # Calcola saldi per tutti (serve sia al tab Docenti sia al Cruscotto)
    saldi = {}
    for d in docenti:
        s = get_saldi_docente(d.id)
        lordo_eff   = s['sup_svolte'] - s['perm_svolte'] - s['civ_svolte']
        netto_eff   = lordo_eff - s['pagamento']
        lordo_prev  = s['sup_prev'] - s['perm_prev'] - s['civ_prev']
        saldo_lordo = s['supplenze'] - s['permessi'] - s['civica']
        saldo_netto = saldo_lordo - s['pagamento']
        saldi[d.id] = {**s,
            'lordo': saldo_lordo, 'netto': saldo_netto,
            'netto_eff': netto_eff, 'lordo_eff': lordo_eff,
            'netto_prev': lordo_prev,
        }

    # Costo ora supplenza — configurabile in Impostazioni > Dati istituto
    from config_istituto import get_costo_ora
    COSTO_ORA = get_costo_ora()

    # Ore istituzionali per tutti (solo ruoli interni)
    from flask import session as _sess2
    ruolo_idx = _sess2.get('ruolo', 'collaboratore')
    ore_ist_idx = {}
    if ruolo_idx in ('ds', 'dsga', 'segreteria'):
        for d in docenti:
            ore_ist_idx[d.id] = get_ore_ist_docente(d.id)

    cruscotto = None
    if tab == 'cruscotto':
        cruscotto = _dati_cruscotto(docenti, saldi, COSTO_ORA)

    # Selettore "archivio" (Roberto: "mettilo per ogni report della
    # pagina report") — solo per il tab Docenti, l'unico che qui mostra
    # dati realmente riferiti a un anno scolastico. Il Cruscotto resta
    # sempre sull'anno corrente di proposito: è un pannello in tempo
    # reale (andamento mese in corso, supplenze scoperte nei prossimi
    # 7 giorni) — "consultare il cruscotto dell'anno scorso" non ha un
    # significato coerente, un selettore lì sarebbe fuorviante. Il tab
    # Segreteria è solo un elenco di link ad altri strumenti, nessun
    # dato proprio da filtrare per anno.
    from config_anno import get_anno_corrente
    anno_corrente = get_anno_corrente()
    anno_docenti = anno_corrente
    docenti_report = docenti
    saldi_report = saldi
    ore_ist_report = ore_ist_idx
    anni_disponibili_docenti = [anno_corrente]
    if tab == 'docenti':
        from routes.impostazione_anno import _docenti_per_anno
        from routes.banca_ore import _anni_disponibili
        anno_docenti = request.args.get('anno', anno_corrente)
        anni_disponibili_docenti = _anni_disponibili()
        docenti_report = _docenti_per_anno(anno_docenti)
        saldi_report = {}
        for d in docenti_report:
            s = get_saldi_docente(d.id, anno_scol=anno_docenti)
            lordo_eff   = s['sup_svolte'] - s['perm_svolte'] - s['civ_svolte']
            netto_eff   = lordo_eff - s['pagamento']
            lordo_prev  = s['sup_prev'] - s['perm_prev'] - s['civ_prev']
            saldo_lordo = s['supplenze'] - s['permessi'] - s['civica']
            saldo_netto = saldo_lordo - s['pagamento']
            saldi_report[d.id] = {**s,
                'lordo': saldo_lordo, 'netto': saldo_netto,
                'netto_eff': netto_eff, 'lordo_eff': lordo_eff,
                'netto_prev': lordo_prev,
            }
        ore_ist_report = {}
        if ruolo_idx in ('ds', 'dsga', 'segreteria'):
            for d in docenti_report:
                ore_ist_report[d.id] = get_ore_ist_docente(d.id, anno=anno_docenti)

    return render_template('report/index.html',
        tab=tab,
        docenti=docenti_report, saldi=saldi_report, oggi=date.today(),
        costo_ora=COSTO_ORA,
        ore_ist_idx=ore_ist_report,
        ruolo_utente=ruolo_idx,
        cruscotto=cruscotto,
        anno=anno_docenti, anno_corrente=anno_corrente,
        anni_disponibili=anni_disponibili_docenti)


# ── REPORT SINGOLO DOCENTE ───────────────────────────────────
@report_bp.route('/report/docente/<int:id>')
def singolo(id):
    d = Docente.query.get_or_404(id)
    saldi   = get_saldi_docente(id)
    storico = get_storico_settimanale(id)

    # Saldo effettivo (ore già svolte, <= oggi)
    saldo_lordo_eff  = saldi['sup_svolte'] - saldi['perm_svolte'] - saldi['civ_svolte']
    saldo_netto_eff  = saldo_lordo_eff - saldi['pagamento']

    # Saldo previsto (include anche ore future)
    saldo_lordo_prev = saldi['sup_prev'] - saldi['perm_prev'] - saldi['civ_prev']
    saldo_netto_prev = saldo_lordo_prev  # pagamento non si applica al futuro

    # Totale complessivo (retrocompatibilità)
    saldo_lordo   = saldi['supplenze'] - saldi['permessi'] - saldi.get('perm_ist',0) - saldi['civica']
    saldo_netto   = saldo_lordo - saldi['pagamento']

    # Supplenze dettaglio
    supplenze = (Supplenza.query
                 .filter_by(id_sostituto=id)
                 .filter(Supplenza.stato == 'assegnata')
                 .order_by(Supplenza.data)
                 .all())

    # Orario settimanale del docente
    from models.orario_docente import OrarioDocente
    GIORNI = ['Lunedì','Martedì','Mercoledì','Giovedì','Venerdì','Sabato']
    slots = OrarioDocente.query.filter_by(id_docente=id).order_by(
        OrarioDocente.giorno, OrarioDocente.ora).all()
    # Struttura: {giorno: {ora: slot}}
    orario = {}
    ore_usate = set()
    for s in slots:
        if s.classe and s.classe not in ('---', '-x-', ''):
            orario.setdefault(s.giorno, {})[s.ora] = s
            ore_usate.add(s.ora)
    ore_list = sorted(ore_usate) if ore_usate else list(range(1, 6))
    giorni_usati = sorted(set(s.giorno for s in slots))

    # Ore CCNL istituzionali (solo per ruoli interni)
    from flask import session as _sess
    ruolo_utente = _sess.get('ruolo', 'collaboratore')
    ore_ist = None
    if ruolo_utente in ('ds', 'dsga', 'segreteria'):
        ore_ist = get_ore_ist_docente(id)

    return render_template('report/singolo.html',
        docente=d,
        saldi=saldi,
        saldo_lordo=saldo_lordo,
        saldo_netto=saldo_netto,
        saldo_lordo_eff=saldo_lordo_eff,
        saldo_netto_eff=saldo_netto_eff,
        saldo_lordo_prev=saldo_lordo_prev,
        saldo_netto_prev=saldo_netto_prev,
        storico=storico,
        supplenze=supplenze,
        orario=orario,
        ore_list=ore_list,
        giorni_usati=giorni_usati,
        giorni_nomi=GIORNI,
        oggi=date.today(),
        ore_ist=ore_ist,
        ruolo_utente=ruolo_utente,
    )


# ── REPORT SINGOLO — EXPORT PDF ──────────────────────────────
@report_bp.route('/report/docente/<int:id>/pdf')
def singolo_pdf(id):
    """Genera PDF del report singolo via WeasyPrint o fallback HTML.
    Rispetta l'anno scolastico passato in query string (?anno=AAAA-AAAA),
    così l'export scaricato da una vista archivio riflette l'anno che si
    sta consultando invece di tornare sempre all'anno corrente."""
    from config_anno import get_anno_corrente
    anno_corrente = get_anno_corrente()
    anno = request.args.get('anno', anno_corrente)

    d = Docente.query.get_or_404(id)
    saldi   = get_saldi_docente(id, anno_scol=anno)
    storico = get_storico_settimanale(id, anno_scol=anno)

    saldo_lordo     = saldi['supplenze'] - saldi['permessi'] - saldi['civica']
    saldo_netto     = saldo_lordo - saldi['pagamento']

    from config_anno import intervallo_anno_scolastico
    _inizio_anno, _fine_anno = intervallo_anno_scolastico(anno)
    supplenze = (Supplenza.query
                 .filter_by(id_sostituto=id)
                 .filter(Supplenza.stato == 'assegnata')
                 .filter(Supplenza.data >= _inizio_anno, Supplenza.data <= _fine_anno)
                 .order_by(Supplenza.data)
                 .all())

    from modules.pdf_fonts import contesto_open_sans
    html_content = render_template('report/singolo_print.html',
        docente=d, saldi=saldi,
        saldo_lordo=saldo_lordo, saldo_netto=saldo_netto,
        storico=storico, supplenze=supplenze,
        oggi=date.today(), anno=anno, anno_corrente=anno_corrente,
        **contesto_open_sans(),
    )

    suffisso_anno = '' if anno == anno_corrente else f'_{anno}'
    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html_content).write_pdf()
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'report_{d.cognome}{suffisso_anno}_{date.today().isoformat()}.pdf'
        )
    except ImportError:
        # WeasyPrint non installato — ritorna HTML con print CSS
        return html_content


# ── REPORT SINGOLO — EXPORT XLSX ────────────────────────────
@report_bp.route('/report/docente/<int:id>/xlsx')
def singolo_xlsx(id):
    """Come singolo_pdf: rispetta ?anno= per esportare l'anno archiviato
    che si sta consultando, invece di tornare sempre all'anno corrente."""
    from modules.xlsx_report import _build_xlsx_singolo
    from config_anno import get_anno_corrente
    anno_corrente = get_anno_corrente()
    anno = request.args.get('anno', anno_corrente)

    d        = Docente.query.get_or_404(id)
    saldi    = get_saldi_docente(id, anno_scol=anno)
    storico  = get_storico_settimanale(id, anno_scol=anno)
    saldo_lordo = saldi['supplenze'] - saldi['permessi'] - saldi['civica']
    netto    = saldo_lordo - saldi['pagamento']
    effettivo= netto
    from config_anno import intervallo_anno_scolastico
    _inizio_anno, _fine_anno = intervallo_anno_scolastico(anno)
    supplenze= (Supplenza.query.filter_by(id_sostituto=id)
                .filter(Supplenza.stato=='assegnata')
                .filter(Supplenza.data >= _inizio_anno, Supplenza.data <= _fine_anno)
                .order_by(Supplenza.data).all())

    wb = _build_xlsx_singolo(d, saldi, storico, supplenze, netto, effettivo, date.today())
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    suffisso_anno = '' if anno == anno_corrente else f'_{anno}'
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'DOC_{d.cognome}{suffisso_anno}_{date.today().isoformat()}.xlsx'
    )


# ── REPORT GLOBALE XLSX ──────────────────────────────────────
@report_bp.route('/report/globale/xlsx')
def globale_xlsx():
    """Export XLSX con foglio indice + un foglio per docente."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        return "openpyxl non disponibile", 500

    wb = openpyxl.Workbook()

    BLU   = PatternFill("solid", fgColor="1F3864")
    VERDE = PatternFill("solid", fgColor="D4F0E0")
    ROSSO = PatternFill("solid", fgColor="FDE8E8")
    GIALL = PatternFill("solid", fgColor="FFF3CD")
    GREY  = PatternFill("solid", fgColor="F0F4F8")

    def hdr(cell, text, fill=BLU):
        cell.value = text
        cell.fill  = fill
        cell.font  = Font(bold=True,
                          color="FFFFFF" if fill == BLU else "000000",
                          size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)

    def thin():
        s = Side(style="thin", color="AAAAAA")
        return Border(left=s, right=s, top=s, bottom=s)

    # ── Foglio Indice ─────────────────────────────────────────
    ws_idx = wb.active
    ws_idx.title = "Indice"

    from config_istituto import get_dati_istituto as _get_dati_istituto_xlsx
    _nome_ist_xlsx = _get_dati_istituto_xlsx()['nome_istituto']
    ws_idx.merge_cells("A1:H1")
    ws_idx["A1"].value = f"BANCA ORE DOCENTI — {_nome_ist_xlsx} — {date.today().strftime('%d/%m/%Y')}"
    ws_idx["A1"].font  = Font(bold=True, size=14, color="1F3864")
    ws_idx.row_dimensions[1].height = 28

    headers = ["Docente", "H/sett", "Supplenze\nsvolte",
               "Permessi\norari", "Ed. Civica\nlibero",
               "Ore a\npagamento", "Saldo\nnetto", "Situazione"]
    for c, h in enumerate(headers, 1):
        hdr(ws_idx.cell(3, c), h)
    ws_idx.row_dimensions[3].height = 32

    docenti = Docente.query.filter_by(attivo=True).order_by(Docente.cognome).all()

    for row_n, d in enumerate(docenti, 4):
        s = get_saldi_docente(d.id)
        lordo = s['supplenze'] - s['permessi'] - s['civica']
        netto = lordo - s['pagamento']

        ws_idx.cell(row_n, 1).value = d.cognome + (f" {d.nome}" if d.nome else "")
        ws_idx.cell(row_n, 2).value = d.ore_contratto
        ws_idx.cell(row_n, 3).value = s['supplenze']
        ws_idx.cell(row_n, 4).value = s['permessi']
        ws_idx.cell(row_n, 5).value = s['civica']
        ws_idx.cell(row_n, 6).value = s['pagamento'] if s['pagamento'] else None
        ws_idx.cell(row_n, 7).value = netto

        if netto > 0:
            sit, fill = "CREDITO", VERDE
        elif netto < 0:
            sit, fill = "RECUPERO", ROSSO
        else:
            sit, fill = "OK", GREY
        ws_idx.cell(row_n, 8).value = sit
        ws_idx.cell(row_n, 8).fill  = fill

        # Colora saldo
        ws_idx.cell(row_n, 7).fill = VERDE if netto > 0 else (ROSSO if netto < 0 else GREY)
        ws_idx.cell(row_n, 3).fill = VERDE if s['supplenze'] > 0 else GREY
        ws_idx.cell(row_n, 4).fill = ROSSO if s['permessi'] > 0 else GREY
        ws_idx.cell(row_n, 5).fill = ROSSO if s['civica']   > 0 else GREY

        for c in range(1, 9):
            ws_idx.cell(row_n, c).border = thin()
            ws_idx.cell(row_n, c).alignment = Alignment(horizontal="center",
                                                          vertical="center")
        ws_idx.cell(row_n, 1).alignment = Alignment(horizontal="left",
                                                      vertical="center")

    # Larghezze colonne indice
    for col, w in zip("ABCDEFGH", [28, 8, 12, 12, 12, 12, 10, 12]):
        ws_idx.column_dimensions[col].width = w

    # ── Foglio per ogni docente ───────────────────────────────
    for d in docenti:
        s       = get_saldi_docente(d.id)
        storico = get_storico_settimanale(d.id)
        lordo     = s['supplenze'] - s['permessi'] - s['civica']
        netto     = lordo - s['pagamento']
        effettivo = netto

        safe_name = d.cognome[:28].replace('/', '_').replace('\\', '_')
        try:
            ws = wb.create_sheet(title=f"DOC_{safe_name}")
        except Exception:
            ws = wb.create_sheet(title=f"DOC_{d.id}")

        # Testata
        ws.merge_cells("A1:G1")
        ws["A1"].value = f"Report banca ore — {d.cognome}{' ' + d.nome if d.nome else ''}"
        ws["A1"].font  = Font(bold=True, size=13, color="1F3864")
        ws.row_dimensions[1].height = 24

        ws["A3"].value = "Saldo netto"
        ws["B3"].value = netto
        ws["A4"].value = "Saldo effettivo (dopo pagamento)"
        ws["B4"].value = effettivo
        ws["A5"].value = "Aggiornamento"
        ws["B5"].value = date.today().strftime('%d/%m/%Y')
        for r in [3,4,5]:
            ws.cell(r,1).font = Font(bold=True, size=10)

        for row_v in [(3, netto), (4, effettivo)]:
            fill = VERDE if row_v[1] > 0 else (ROSSO if row_v[1] < 0 else GREY)
            ws.cell(row_v[0], 2).fill = fill

        # Intestazioni storico
        hdrs_s = ["Data", "Supplenze\n+h", "Permessi\n-h",
                  "Ed. Civica\n-h", "Pagamento\nh", "Delta\ngiornata"]
        for c, h in enumerate(hdrs_s, 1):
            hdr(ws.cell(7, c), h)
        ws.row_dimensions[7].height = 30

        for r_n, riga in enumerate(storico, 8):
            delta = riga['supplenze'] - riga['permessi'] - riga['civica']
            ws.cell(r_n, 1).value = riga['data'].strftime('%d/%m/%Y') if riga['data'] else ''
            ws.cell(r_n, 2).value = riga['supplenze'] or None
            ws.cell(r_n, 3).value = riga['permessi']  or None
            ws.cell(r_n, 4).value = riga['civica']    or None
            ws.cell(r_n, 5).value = riga['pagamento'] or None
            ws.cell(r_n, 6).value = delta             if delta != 0 else None

            ws.cell(r_n, 2).fill = VERDE if riga['supplenze'] > 0 else GREY
            ws.cell(r_n, 3).fill = ROSSO if riga['permessi']  > 0 else GREY
            ws.cell(r_n, 4).fill = ROSSO if riga['civica']    > 0 else GREY
            ws.cell(r_n, 6).fill = VERDE if delta > 0 else (ROSSO if delta < 0 else GREY)

            for c in range(1, 7):
                ws.cell(r_n, c).border = thin()
                ws.cell(r_n, c).alignment = Alignment(horizontal="center")

        # Totali
        last = 8 + len(storico)
        ws.cell(last, 1).value = "TOTALI"
        ws.cell(last, 1).font  = Font(bold=True)
        ws.cell(last, 2).value = s['supplenze']
        ws.cell(last, 3).value = s['permessi']
        ws.cell(last, 4).value = s['civica']
        ws.cell(last, 5).value = s['pagamento']
        ws.cell(last, 6).value = netto
        for c in range(1, 7):
            ws.cell(last, c).font   = Font(bold=True)
            ws.cell(last, c).border = thin()

        for col, w in zip("ABCDEF", [14, 10, 10, 10, 10, 10]):
            ws.column_dimensions[col].width = w

    # ── Salva in buffer ───────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'BancaOre_completa_{date.today().isoformat()}.xlsx'
    )


# ── PROSPETTO SUPPLENZE GIORNALIERO ─────────────────────────
@report_bp.route('/prospetto')
@report_bp.route('/prospetto/<string:data_str>')
def prospetto(data_str=None):
    import os
    from modules.prospetto_supplenze import genera_prospetto

    if data_str is None:
        data_str = date.today().isoformat()
    try:
        data_sel = date.fromisoformat(data_str)
    except ValueError:
        data_sel = date.today()

    template_path = os.path.join(
        os.path.dirname(os.path.dirname(__file__)),
        'data', 'prospetto_template.xlsx'
    )
    if not os.path.exists(template_path):
        return 'Template prospetto non trovato in data/prospetto_template.xlsx', 404

    supplenze = (Supplenza.query
                 .filter_by(data=data_sel)
                 .filter(Supplenza.stato != 'annullata')
                 .order_by(Supplenza.ora)
                 .all())

    # Attività istituzionali del giorno per il prospetto
    try:
        from models.attivita_ist import AttivitaIst
        attivita_ist_giorno = (AttivitaIst.query
                               .filter_by(data=data_sel)
                               .order_by(AttivitaIst.ora_inizio)
                               .all())
    except Exception:
        attivita_ist_giorno = []

    save_dir   = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                              'data', 'prospetti')
    xlsx_bytes = genera_prospetto(data_sel, supplenze, template_path,
                                  save_dir=save_dir,
                                  attivita_ist=attivita_ist_giorno)

    nome_file = f'Prospetto_supplenze_{data_sel.strftime("%d%m%Y")}.xlsx'
    return send_file(
        io.BytesIO(xlsx_bytes),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nome_file
    )



# ── ESPORTA TUTTI PDF ────────────────────────────────────────
@report_bp.route('/report/esporta-tutti-pdf')
def esporta_tutti_pdf():
    """Genera un PDF per ogni docente e li restituisce in uno ZIP."""
    import zipfile
    try:
        from weasyprint import HTML
    except ImportError:
        return 'WeasyPrint non installato — impossibile generare PDF', 500

    docenti = Docente.query.filter_by(attivo=True).order_by(Docente.cognome).all()
    oggi_str = date.today().isoformat()

    # Docenti con lo stesso cognome (es. due omonimi) devono avere nomi di
    # file distinti nello ZIP, altrimenti uno dei due PDF viene perso
    # silenziosamente all'estrazione (bug reale osservato: Ghezzi,
    # Tramontana, Valena sono ciascuno in coppia in questo istituto).
    from collections import Counter
    conteggio_cognomi = Counter(d.cognome for d in docenti)

    zip_buf = io.BytesIO()
    nomi_usati = set()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for d in docenti:
            try:
                saldi   = get_saldi_docente(d.id)
                storico = get_storico_settimanale(d.id)
                saldo_lordo = saldi['supplenze'] - saldi['permessi'] - saldi['civica']
                saldo_netto = saldo_lordo - saldi['pagamento']
                supplenze = (Supplenza.query
                             .filter_by(id_sostituto=d.id)
                             .filter(Supplenza.stato == 'assegnata')
                             .order_by(Supplenza.data)
                             .all())
                from modules.pdf_fonts import contesto_open_sans
                html_content = render_template('report/singolo_print.html',
                    docente=d, saldi=saldi,
                    saldo_lordo=saldo_lordo, saldo_netto=saldo_netto,
                    storico=storico, supplenze=supplenze,
                    oggi=date.today(),
                    **contesto_open_sans(),
                )
                pdf_bytes = HTML(string=html_content).write_pdf()
                if conteggio_cognomi[d.cognome] > 1:
                    # Cognome duplicato: aggiungi l'iniziale del nome per
                    # distinguere i file nello ZIP.
                    iniziale = f'_{d.nome[0]}' if d.nome else f'_id{d.id}'
                    fname = f'DOC_{d.cognome}{iniziale}.pdf'
                else:
                    fname = f'DOC_{d.cognome}.pdf'
                if fname in nomi_usati:
                    # Ulteriore fallback di sicurezza, in caso anche
                    # cognome+iniziale coincidano.
                    fname = f'DOC_{d.cognome}_id{d.id}.pdf'
                nomi_usati.add(fname)
                zf.writestr(fname, pdf_bytes)
            except Exception as e:
                # Aggiungi file di errore invece di fallire tutto
                zf.writestr(f'ERRORE_{d.cognome}_id{d.id}.txt', str(e))

    zip_buf.seek(0)
    return send_file(
        zip_buf,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'BancaOre_PDF_{oggi_str}.zip'
    )

# ── EXPORT EXCEL SETTIMANA ──────────────────────────────────
@report_bp.route('/export/excel', methods=['GET', 'POST'])
def export_excel():
    from modules.export_excel_sett import aggiorna_sett_excel, aggiorna_riepilogo_excel
    from models.movimento_banca_ore import MovimentoBancaOre
    from modules.import_banca_ore import leggi_movimenti_file
    import glob

    file_path = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                             'data', 'Banca_Ore_Docenti_v3.xlsm')

    # Settimane disponibili dal file
    from openpyxl import load_workbook as _lw
    wb = _lw(file_path, data_only=True)
    from modules.import_banca_ore import _parse_data_settimana
    # Rileva TUTTE le settimane presenti nel file (anche se aggiunte dopo)
    settimane = []
    import re as _re
    for nome in wb.sheetnames:
        m = _re.match(r'^sett\.(\d+)$', nome)
        if not m:
            continue
        sn = int(m.group(1))
        titolo = wb[nome].cell(1,1).value
        data   = _parse_data_settimana(str(titolo) if titolo else None)
        settimane.append({'n': sn, 'titolo': str(titolo) if titolo else nome, 'data': data})
    settimane.sort(key=lambda x: x['n'])

    msg = None
    if request.method == 'POST':
        azione = request.form.get('azione')

        if azione == 'sett':
            sett_n = int(request.form.get('sett_n', 30))

            # Determina le date della settimana dal titolo del foglio Excel
            from openpyxl import load_workbook as _lw2
            from modules.import_banca_ore import _parse_data_settimana
            from datetime import timedelta
            _wb2 = _lw2(file_path, data_only=True)
            if f'sett.{sett_n}' not in _wb2.sheetnames:
                flash(f'Foglio sett.{sett_n} non trovato.', 'error')
                return redirect(url_for('report.export_excel'))
            _titolo = _wb2[f'sett.{sett_n}'].cell(1,1).value
            data_inizio = _parse_data_settimana(str(_titolo) if _titolo else None)
            if not data_inizio:
                flash(f'Impossibile determinare la data della sett.{sett_n}', 'error')
                return redirect(url_for('report.export_excel'))
            data_fine = data_inizio + timedelta(days=6)

            # Raccoglie movimenti DB per quella settimana per DATE
            movs = MovimentoBancaOre.query.filter(
                MovimentoBancaOre.data.between(data_inizio, data_fine)
            ).all()

            from models.docente import Docente as _D
            dati = {}
            for m in movs:
                d = db.session.get(_D, m.id_docente)
                if not d:
                    continue
                cog = d.cognome.upper().replace('’', "'").replace('‘', "'")
                if cog not in dati:
                    dati[cog] = {'sup': 0, 'perm': 0, 'civ': 0}
                if m.tipo == 'supplenza_recupero':
                    dati[cog]['sup'] += abs(m.minuti) // 60
                elif m.tipo in ('permesso_orario', 'permesso'):
                    dati[cog]['perm'] += abs(m.minuti) // 60
                elif m.tipo in ('civica', 'ed_civica'):
                    dati[cog]['civ'] += abs(m.minuti) // 60

            ok, msg = aggiorna_sett_excel(sett_n, dati, file_path)
            flash(
                msg + f' | {data_inizio.strftime("%d/%m")}–{data_fine.strftime("%d/%m/%Y")} | {len(movs)} movimenti',
                'success' if ok else 'error'
            )

        elif azione == 'riepilogo':
            # Aggiorna riepilogo completo
            from models.docente import Docente as _D
            from models.movimento_banca_ore import MovimentoBancaOre as _M
            all_movs = _M.query.all()
            saldi = {}
            pagamenti = {}
            for m in all_movs:
                d = _D.query.get(m.id_docente)
                if not d:
                    continue
                cog = d.cognome.upper().replace('’', "'")
                if cog not in saldi:
                    saldi[cog] = {'sup': 0, 'perm': 0, 'civ': 0}
                if m.tipo == 'supplenza_recupero':
                    saldi[cog]['sup'] += abs(m.minuti) // 60
                elif m.tipo in ('permesso_orario', 'permesso'):
                    saldi[cog]['perm'] += abs(m.minuti) // 60
                elif m.tipo in ('civica', 'ed_civica'):
                    saldi[cog]['civ'] += abs(m.minuti) // 60
                elif m.tipo == 'supplenza_pagamento':
                    pagamenti[cog] = pagamenti.get(cog, 0) + abs(m.minuti) // 60

            ok, msg = aggiorna_riepilogo_excel(saldi, pagamenti, file_path)
            flash(msg, 'success' if ok else 'error')

        return redirect(url_for('report.export_excel'))

    return render_template('report/export_excel.html',
        settimane=settimane, oggi=date.today())


# ── OTTIMIZZAZIONE SIMULAZIONI ────────────────────────────────
@report_bp.route('/ottimizzazione-simulazioni')
def ottimizzazione_simulazioni():
    """
    Prospetto ottimizzato per i giorni di simulazione:
    - Per ogni supplenza scoperta, suggerisce il sostituto ideale in base al saldo banca ore
    - Privilegia debitori, libera creditori con permessi
    """
    from models.supplenza import Supplenza
    from models.indisponibilita import Indisponibilita
    from models.orario_docente import OrarioDocente
    from models.assenza import Assenza
    from models.attivita_fuori_aula import AttivitaFuoriAula
    from collections import defaultdict

    oggi = date.today()

    # Date simulazioni (escludi festivi lun/mar 1-2 giugno)
    att_sim_all = AttivitaFuoriAula.query.filter(
        AttivitaFuoriAula.tipo == 'simulazione',
        AttivitaFuoriAula.data_fine >= oggi,
        AttivitaFuoriAula.stato == 'attiva'
    ).all()
    date_sim_set = set()
    for a in att_sim_all:
        cur = a.data_inizio
        from datetime import timedelta
        while cur <= a.data_fine:
            if cur >= oggi and cur.weekday() < 6:
                date_sim_set.add(cur)
            cur += timedelta(days=1)
    date_sim_ordered = sorted(date_sim_set)

    # Saldi effettivi attuali per tutti i docenti (solo anno scolastico corrente)
    from config_anno import get_anno_corrente
    anno_corrente = get_anno_corrente()
    docenti_attivi = Docente.query.filter_by(attivo=True).all()
    saldi_att = {}
    for d in docenti_attivi:
        movs = MovimentoBancaOre.query.filter(
            MovimentoBancaOre.id_docente == d.id,
            MovimentoBancaOre.anno_scol == anno_corrente,
            MovimentoBancaOre.data <= oggi
        ).all()
        sup  = sum(m.minuti for m in movs if m.tipo == 'supplenza_recupero') // 60
        perm = sum(abs(m.minuti) for m in movs if m.tipo in ('permesso_orario','permesso')) // 60
        civ  = sum(abs(m.minuti) for m in movs if m.tipo in ('civica','ed_civica')) // 60
        pag  = sum(abs(m.minuti) for m in movs if m.tipo == 'supplenza_pagamento') // 60
        saldi_att[d.id] = sup - perm - civ - pag

    # Sorveglianza già accreditata nel periodo simulazioni
    if date_sim_ordered:
        movs_sorv = MovimentoBancaOre.query.filter(
            MovimentoBancaOre.data.between(date_sim_ordered[0], date_sim_ordered[-1]),
            MovimentoBancaOre.descrizione.like('Sorveglianza%')
        ).all()
    else:
        movs_sorv = []
    sorv_per_doc = defaultdict(int)
    for m in movs_sorv:
        sorv_per_doc[m.id_docente] += m.minuti // 60

    # Proiezione = saldo attuale + sorveglianza
    saldi_proj = {d.id: saldi_att.get(d.id, 0) + sorv_per_doc.get(d.id, 0)
                  for d in docenti_attivi}

    # Date simulazioni future
    att_sim = AttivitaFuoriAula.query.filter(
        AttivitaFuoriAula.tipo == 'simulazione',
        AttivitaFuoriAula.data_fine >= oggi,
        AttivitaFuoriAula.stato == 'attiva'
    ).all()
    date_sim = sorted(set(
        d for a in att_sim
        for d in (a.data_inizio,) if a.data_inizio >= oggi
    ))

    # (Nota: qui prima c'era un secondo calcolo di "saldi" per tutti i
    # docenti, duplicato rispetto a saldi_att/saldi_proj sopra e privo del
    # filtro anno_scol — verificato che non veniva mai effettivamente
    # utilizzato più sotto (il sort usa saldi_proj), quindi rimosso come
    # codice morto invece di essere anche lui corretto per anno.)

    # Per ogni giorno di simulazione, calcola il prospetto
    prospetto = {}
    for data_sim in date_sim:
        giorno = data_sim.weekday()

        # Supplenze scoperte
        sups_scoperte = Supplenza.query.filter_by(
            data=data_sim, stato='scoperta'
        ).order_by(Supplenza.ora).all()

        # Indisponibili e assenti
        indisp_ids_per_ora = defaultdict(set)
        for i in Indisponibilita.query.filter_by(data=data_sim).all():
            indisp_ids_per_ora[i.ora].add(i.id_docente)
        assenti_ids = {a.id_docente for a in Assenza.query.filter_by(data=data_sim).all()}

        # Per ogni supplenza scoperta, trova i migliori sostituti
        righe = []
        for s in sups_scoperte:
            d_ass = db.session.get(Docente, s.id_assente) if s.id_assente else None

            # Docenti disponibili in quell'ora
            occupati = set()
            # Chi è a scuola in quell'ora (ha lezione)
            for slot in OrarioDocente.query.filter_by(giorno=giorno, ora=s.ora).all():
                if slot.tipo_ora == 'lezione' and slot.classe not in ('POTENZIAMENTO','---','-x-',''):
                    if slot.id_docente != (s.id_assente or -1):
                        occupati.add(slot.id_docente)

            disponibili = []
            for d in docenti_attivi:
                if d.id in assenti_ids: continue
                if d.id in indisp_ids_per_ora.get(s.ora, set()): continue
                if d.id in occupati: continue
                if d.id == (s.id_assente or -1): continue
                disponibili.append(d)

            # Calcola giorni liberi per ogni candidato (nessuna lezione in tutto il giorno)
            def giorni_liberi_docente(doc_id, escludi_giorno=None):
                GIORNI_NOMI = ['Lunedì','Martedì','Mercoledì','Giovedì','Venerdì','Sabato']
                liberi = []
                for g in range(6):
                    if g == escludi_giorno:
                        continue
                    slots = OrarioDocente.query.filter_by(id_docente=doc_id, giorno=g).filter(
                        OrarioDocente.tipo_ora.in_(['lezione','potenziamento'])
                    ).count()
                    if slots == 0:
                        liberi.append(GIORNI_NOMI[g])
                return liberi

            # Controlla se il giorno della supplenza è libero per il candidato
            giorno_supplenza = data_sim.weekday()

            # Ordina per proiezione (saldo attuale + sorveglianza già accreditata)
            disponibili.sort(key=lambda d: saldi_proj.get(d.id, 0))

            # Top 3
            top3 = []
            for d in disponibili[:3]:
                sal_att = saldi_att.get(d.id, 0)
                sal_pro = saldi_proj.get(d.id, 0)
                # Verifica se quel giorno è libero per questo docente
                ha_lezioni_quel_giorno = OrarioDocente.query.filter_by(
                    id_docente=d.id, giorno=giorno_supplenza
                ).filter(
                    OrarioDocente.tipo_ora.in_(['lezione','potenziamento'])
                ).count() > 0
                gg_recupero = giorni_liberi_docente(d.id, escludi_giorno=giorno_supplenza) if not ha_lezioni_quel_giorno else []

                top3.append({
                    'id': d.id,
                    'cognome': d.cognome,
                    'nome': d.nome or '',
                    'saldo': sal_att,
                    'saldo_proj': sal_pro,
                    'tipo': 'debitore' if sal_pro < 0 else ('creditore' if sal_pro > 0 else 'pari'),
                    'giorno_libero': not ha_lezioni_quel_giorno,
                    'giorni_recupero': gg_recupero,
                })

            righe.append({
                'supplenza': s,
                'assente': d_ass,
                'candidati': top3,
            })

        # Creditori da liberare con permesso in quel giorno
        creditori_liberabili = []
        for d in docenti_attivi:
            sal = saldi_proj.get(d.id, 0)  # usa proiezione con sorveglianza
            if sal <= 0: continue
            if d.id in assenti_ids: continue
            # Ha lezione quel giorno?
            slots = OrarioDocente.query.filter_by(id_docente=d.id, giorno=giorno).filter(
                OrarioDocente.tipo_ora == 'lezione'
            ).all()
            if not slots: continue
            # Non è accompagnatore simulazione
            is_acc = any(d in a.accompagnatori for a in att_sim if a.data_inizio == data_sim)
            if is_acc: continue
            creditori_liberabili.append({
                'docente': d,
                'saldo_att': saldi_att.get(d.id, 0),
                'saldo': sal,  # proiezione
                'ore': sorted(set(s.ora for s in slots)),
            })
        creditori_liberabili.sort(key=lambda x: -x['saldo'])

        prospetto[data_sim] = {
            'righe': righe,
            'creditori': creditori_liberabili[:8],
        }

    return render_template('report/ottimizzazione_simulazioni.html',
        prospetto=prospetto,
        oggi=oggi,
    )


# ── PIANIFICAZIONE PERMESSI ──────────────────────────────────
@report_bp.route('/report/pianifica-permessi', methods=['GET', 'POST'])
def pianifica_permessi():
    from config_anno import get_anno_corrente
    from config_calendario import set_data_fine_lezioni, set_ore_ultimo_giorno
    from modules.pianificazione_permessi import calcola_pianificazione
    anno_corrente = get_anno_corrente()

    # Configurazione (via form nella pagina stessa, invece di date hardcoded
    # nel codice) — per anno scolastico, così va aggiornata una volta sola
    # ogni settembre invece di essere dimenticata nel codice. I giorni non
    # didattici (ponti, sospensioni) non hanno più un elenco proprio qui:
    # vengono presi direttamente da Impostazioni > Sospensioni didattiche,
    # così c'è un solo posto dove inserirli invece di due elenchi separati
    # da tenere allineati a mano.
    if request.method == 'POST':
        try:
            data_fine_form = date.fromisoformat(request.form.get('data_fine_lezioni', ''))
            set_data_fine_lezioni(anno_corrente, data_fine_form)
            ore_ultimo_form = request.form.get('ore_ultimo_giorno', '').strip()
            set_ore_ultimo_giorno(anno_corrente, int(ore_ultimo_form) if ore_ultimo_form else None)
            flash(f'Calendario per {anno_corrente} aggiornato.', 'success')
        except ValueError:
            flash('Data non valida (formato atteso: AAAA-MM-GG).', 'danger')
        return redirect(url_for('report.pianifica_permessi'))

    # Il calcolo (chi può chiedere un permesso, quando, quante ore) è
    # interamente in modules/pianificazione_permessi.py: la route si
    # limita a chiamarlo e a passare il risultato al template.
    calc = calcola_pianificazione(anno_corrente)

    return render_template('report/pianifica_permessi.html',
        risultati=calc['risultati'], oggi=calc['oggi'],
        anno_corrente=anno_corrente,
        data_fine_lezioni=calc['fine_anno'],
        n_festivi_extra=calc['n_festivi_extra'],
        ore_ultimo_giorno=calc['ore_ultimo_giorno'])


# ── INCARICHI PER DOCENTE (vista di sola lettura) ────────────
@report_bp.route('/report/incarichi-docenti')
def incarichi_docenti():
    """Vista trasversale di sola lettura: per ogni docente, tutti gli
    incarichi dell'anno. La gestione (assegna/elimina) resta in incarichi.index."""
    from models.incarico import IncaricaDocente
    from config_anno import get_anno_corrente
    anno = request.args.get('anno', get_anno_corrente())

    nomine = (IncaricaDocente.query
              .filter_by(anno_scol=anno)
              .join(Docente, IncaricaDocente.id_docente == Docente.id)
              .order_by(Docente.cognome, Docente.nome)
              .all())

    # Raggruppa per docente
    from collections import defaultdict
    per_doc = defaultdict(list)
    for n in nomine:
        per_doc[n.docente].append(n)

    anni = sorted({r.anno_scol for r in IncaricaDocente.query.all()}, reverse=True)
    if not anni:
        anni = [anno]

    return render_template('report/incarichi_docenti.html',
        anno=anno, anni_disponibili=anni, per_doc=per_doc)


# ── STORICO PROSPETTI ────────────────────────────────────────
@report_bp.route('/prospetti')
def lista_prospetti():
    import glob
    prospetti_dir = os.path.join(
        os.path.dirname(os.path.dirname(__file__)), 'data', 'prospetti'
    )
    os.makedirs(prospetti_dir, exist_ok=True)
    files = sorted(glob.glob(os.path.join(prospetti_dir, '*.xlsx')), reverse=True)
    prospetti = []
    for f in files:
        nome = os.path.basename(f)
        size = os.path.getsize(f)
        mtime = date.fromtimestamp(os.path.getmtime(f))
        prospetti.append({'nome': nome, 'size': size, 'data': mtime})
    return render_template('report/prospetti.html',
        prospetti=prospetti, oggi=date.today())


@report_bp.route('/prospetti/scarica/<string:nome>')
def scarica_prospetto(nome):
    prospetti_dir = os.path.join(
        os.path.dirname(os.path.dirname(__file__)), 'data', 'prospetti'
    )
    path = os.path.join(prospetti_dir, nome)
    if not os.path.exists(path):
        return 'File non trovato', 404
    with open(path, 'rb') as f:
        data_bytes = f.read()
    return send_file(
        io.BytesIO(data_bytes),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nome
    )


@report_bp.route('/prospetti/elimina/<string:nome>', methods=['POST'])
def elimina_prospetto(nome):
    from flask import flash, redirect, url_for
    prospetti_dir = os.path.join(
        os.path.dirname(os.path.dirname(__file__)), 'data', 'prospetti'
    )
    path = os.path.join(prospetti_dir, nome)
    if os.path.exists(path):
        os.remove(path)
        flash(f'Prospetto {nome} eliminato.', 'warning')
    return redirect(url_for('report.lista_prospetti'))


# ── REPORT DIRIGENTE ─────────────────────────────────────────
@report_bp.route('/report/dirigente')
def dirigente():
    # Selettore "archivio" (Roberto: "come vedo il report del dirigente
    # dell'anno precedente?") — stesso pattern già in uso in Banca Ore
    # (routes/banca_ore.py::_anni_disponibili()), riusato qui invece di
    # duplicarlo. _docenti_per_anno() al posto del filtro naive
    # attivo=True: consultando un anno passato vanno inclusi i docenti
    # usciti nel frattempo ed esclusi quelli non ancora in servizio
    # allora — stesso principio già applicato altrove in questa sessione.
    from config_anno import get_anno_corrente
    from routes.impostazione_anno import _docenti_per_anno
    from routes.banca_ore import _anni_disponibili
    anno_corrente = get_anno_corrente()
    anno = request.args.get('anno', anno_corrente)

    docenti = _docenti_per_anno(anno)

    equilibrio = credito = debito = critico = 0
    tot_supplenze = tot_permessi = tot_civica = tot_pagamento = 0

    casi_critici   = []  # debito > 5h
    crediti_alti   = []  # credito > 8h

    for d in docenti:
        s = get_saldi_docente(d.id, anno_scol=anno)
        lordo = s['supplenze'] - s['permessi'] - s['civica']
        netto = lordo - s['pagamento']

        tot_supplenze += s['supplenze']
        tot_permessi  += s['permessi']
        tot_civica    += s['civica']
        tot_pagamento += s['pagamento']

        if netto > 0:
            credito += 1
            if netto >= 8:
                crediti_alti.append({'docente': d, 'saldo': netto})
        elif netto < 0:
            debito += 1
            if netto <= -5:
                critico += 1
                casi_critici.append({'docente': d, 'saldo': netto})
        else:
            equilibrio += 1

    # Situazione complessiva
    perc_ok = round(equilibrio / len(docenti) * 100) if docenti else 0
    if perc_ok >= 75:
        situazione = ('EQUILIBRATA', 'verde')
    elif perc_ok >= 50:
        situazione = ('SOTTO PRESSIONE', 'giallo')
    else:
        situazione = ('CRITICA', 'rosso')

    casi_critici.sort(key=lambda x: x['saldo'])
    crediti_alti.sort(key=lambda x: -x['saldo'])

    # Indicatore di qualità: tasso di puntualità nel saldare la banca ore
    # entro 3 mesi, come da accordo sindacale. Calcolato sulle singole ore
    # (lotti) a debito/credito ancora aperte, ciascuna con la propria
    # scadenza individuale di 3 mesi dalla data di maturazione — non sul
    # saldo complessivo del docente. Indicatore informativo per il
    # Dirigente, non blocca nessuna operazione. Anno selezionato (non
    # sempre quello corrente): consultando un anno passato, un lotto mai
    # regolarizzato entro la scadenza resta un'informazione storica
    # legittima, non solo un indicatore "in tempo reale".
    n_lotti_aperti = 0
    n_lotti_scaduti = 0
    for d in docenti:
        lotti = _lotti_aperti_docente(d.id, anno)
        n_lotti_aperti += len(lotti)
        n_lotti_scaduti += sum(1 for l in lotti if l['scaduto'])
    if n_lotti_aperti > 0:
        tasso_puntualita = round((n_lotti_aperti - n_lotti_scaduti) / n_lotti_aperti * 100)
    else:
        tasso_puntualita = 100
    n_saldi_aperti  = n_lotti_aperti
    n_oltre_termine = n_lotti_scaduti

    from config_istituto import get_costo_ora
    costo_ora_dirigente = get_costo_ora()

    return render_template('report/dirigente.html',
        anno = anno, anno_corrente = anno_corrente,
        anni_disponibili = _anni_disponibili(),
        n_docenti    = len(docenti),
        equilibrio   = equilibrio,
        credito      = credito,
        debito       = debito,
        critico      = critico,
        perc_ok      = perc_ok,
        situazione   = situazione,
        tot_supplenze= tot_supplenze,
        tot_permessi = tot_permessi,
        tot_civica   = tot_civica,
        tot_pagamento= tot_pagamento,
        casi_critici = casi_critici,
        crediti_alti = crediti_alti,
        n_saldi_aperti   = n_saldi_aperti,
        n_oltre_termine  = n_oltre_termine,
        tasso_puntualita = tasso_puntualita,
        costo_ora        = costo_ora_dirigente,
        oggi         = date.today(),
    )


# ── PROSPETTO WEB GIORNALIERO ─────────────────────────────────
@report_bp.route('/prospetto-web')
@report_bp.route('/prospetto-web/<string:data_str>')
def prospetto_web(data_str=None):
    """Prospetto HTML giornaliero con supplenze + attività istituzionali."""
    if data_str is None:
        data_str = date.today().isoformat()
    try:
        data_sel = date.fromisoformat(data_str)
    except ValueError:
        data_sel = date.today()

    supplenze = (Supplenza.query
                 .filter_by(data=data_sel)
                 .filter(Supplenza.stato != 'annullata')
                 .order_by(Supplenza.ora)
                 .all())

    try:
        from models.attivita_ist import AttivitaIst
        eventi_ist = (AttivitaIst.query
                      .filter_by(data=data_sel)
                      .order_by(AttivitaIst.ora_inizio)
                      .all())
    except Exception:
        eventi_ist = []

    GIORNI = ['Lunedì','Martedì','Mercoledì','Giovedì','Venerdì','Sabato','Domenica']
    MESI   = ['','Gennaio','Febbraio','Marzo','Aprile','Maggio','Giugno',
              'Luglio','Agosto','Settembre','Ottobre','Novembre','Dicembre']
    from datetime import timedelta
    return render_template('report/prospetto_web.html',
        data_sel=data_sel,
        supplenze=supplenze,
        eventi_ist=eventi_ist,
        oggi=date.today(),
        timedelta=timedelta,
        giorno_it=GIORNI[data_sel.weekday()],
        mese_it=MESI[data_sel.month],
    )

