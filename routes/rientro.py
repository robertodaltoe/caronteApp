"""
Modulo Colloqui di Rientro dall'estero.
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash
from models import db
from models.rientro import RientroMateriaClasse, RientroCandidato, RientroColloquio, RuoloIstituzionale
from models.docente import Docente
from models.orario_docente import OrarioDocente
from models.recupero import RecuperoPeriodo, RecuperoGruppo, RecuperoDocente as RecGruppoDocente
from datetime import datetime, timedelta

rientro_bp = Blueprint('rientro', __name__)

from config_anno import get_anno_corrente as _get_anno
ANNO = _get_anno()
PERIODO_CODICE = 'colloqui_rientro'
DURATA_MIN = 45

# Stesse famiglie di sinonimi materia già usate nel modulo recupero —
# servono per il matching automatico classe+materia -> docente e per il
# filtro "stessa materia/famiglia, altra classe" nella sostituzione.
_FAMIGLIE_MATERIE = [
    {'LATINO', 'LINGUA LATINA', 'LINGUA E CULTURA LATINA'},
    {'ITALIANO', 'LINGUA E LETTERATURA ITALIANA', 'LINGUA E CULTURA ITALIANA'},
    {'MATEMATICA', 'MATEMATICA CON INFORMATICA', 'MATEMATICA E COMPLEMENTI DI MATEMATICA'},
    {'INFORMATICA', 'TECNOLOGIE INFORMATICHE'},
    {'STORIA', 'STORIA E GEOGRAFIA'},
    {'FISICA', 'SCIENZE INTEGRATE (FISICA)'},
    {'INGLESE', 'LINGUA INGLESE', 'LINGUA E CULTURA STRANIERA (INGLESE)'},
    {'TEDESCO', 'LINGUA TEDESCA', 'LINGUA E CULTURA STRANIERA TEDESCO'},
]

def _materia_canonica(materia):
    mu = (materia or '').strip().upper()
    for fam in _FAMIGLIE_MATERIE:
        if mu in fam:
            return sorted(fam)[0]
    return mu

def _stesso_dipartimento(m1, m2):
    return _materia_canonica(m1) == _materia_canonica(m2)


# ── INDICE ──────────────────────────────────────────────────────────
@rientro_bp.route('/rientro')
def index():
    periodo = RecuperoPeriodo.query.filter_by(
        anno_scol=ANNO, codice=PERIODO_CODICE).first()

    candidati = RientroCandidato.query.filter_by(anno_scol=ANNO).order_by(
        RientroCandidato.classe, RientroCandidato.cognome).all()

    n_calendarizzati = sum(1 for c in candidati if c.colloquio and c.colloquio.data)

    return render_template('rientro/index.html',
        periodo=periodo,
        candidati=candidati,
        n_totale=len(candidati),
        n_calendarizzati=n_calendarizzati,
        anno=ANNO)


# ── MATERIE PER CLASSE ─────────────────────────────────────────────────
@rientro_bp.route('/rientro/materie', methods=['GET', 'POST'])
def materie():
    if request.method == 'POST':
        azione = request.form.get('azione')

        if azione == 'aggiungi':
            classe = request.form.get('classe', '').strip().upper()
            materie_sel = request.form.getlist('materie[]')
            if not classe or not materie_sel:
                flash('Seleziona una classe e almeno una materia.', 'warning')
                return redirect(url_for('rientro.materie'))
            if len(materie_sel) > 4:
                flash('Massimo 4 materie per classe.', 'warning')
                return redirect(url_for('rientro.materie'))

            # Sostituisce sempre l'elenco completo per quella classe
            RientroMateriaClasse.query.filter_by(anno_scol=ANNO, classe=classe).delete()
            for m in materie_sel:
                db.session.add(RientroMateriaClasse(
                    anno_scol=ANNO, classe=classe, materia=m.strip().upper()))
            db.session.commit()
            flash(f'Materie impostate per {classe}.', 'success')

        elif azione == 'elimina_classe':
            classe = request.form.get('classe', '').strip().upper()
            RientroMateriaClasse.query.filter_by(anno_scol=ANNO, classe=classe).delete()
            db.session.commit()
            flash(f'Materie rimosse per {classe}.', 'warning')

        return redirect(url_for('rientro.materie'))

    righe = RientroMateriaClasse.query.filter_by(anno_scol=ANNO).order_by(
        RientroMateriaClasse.classe, RientroMateriaClasse.materia).all()

    per_classe = {}
    for r in righe:
        per_classe.setdefault(r.classe, []).append(r.materia)

    # Tutte le classi note dall'orario (per il select), con le materie
    # effettivamente insegnate in ciascuna.
    materie_per_classe_orario = {}
    for o in OrarioDocente.query.filter(OrarioDocente.classe.isnot(None),
                                         OrarioDocente.materia.isnot(None)).all():
        materie_per_classe_orario.setdefault(o.classe, set()).add(o.materia.strip().upper())
    materie_per_classe_orario = {k: sorted(v) for k, v in materie_per_classe_orario.items()}

    return render_template('rientro/materie.html',
        per_classe=per_classe,
        classi_disponibili=sorted(materie_per_classe_orario.keys()),
        materie_per_classe=materie_per_classe_orario,
        anno=ANNO)


# ── CANDIDATI ───────────────────────────────────────────────────────
@rientro_bp.route('/rientro/candidati', methods=['GET', 'POST'])
def candidati():
    if request.method == 'POST':
        azione = request.form.get('azione')

        if azione == 'aggiungi':
            cognome = request.form.get('cognome', '').strip().upper()
            nome = request.form.get('nome', '').strip().title()
            classe = request.form.get('classe', '').strip().upper()
            note = request.form.get('note', '').strip() or None
            if not (cognome and nome and classe):
                flash('Cognome, nome e classe sono obbligatori.', 'warning')
                return redirect(url_for('rientro.candidati'))
            cand = RientroCandidato(anno_scol=ANNO, cognome=cognome, nome=nome,
                                     classe=classe, note=note)
            db.session.add(cand)
            db.session.commit()
            db.session.add(RientroColloquio(id_candidato=cand.id))
            db.session.commit()
            flash(f'Candidato {cognome} {nome} aggiunto.', 'success')

        elif azione == 'elimina':
            cand_id = request.form.get('id')
            if cand_id:
                cand = RientroCandidato.query.get(int(cand_id))
                if cand:
                    nome_completo = f'{cand.cognome} {cand.nome}'
                    db.session.delete(cand)
                    db.session.commit()
                    flash(f'Candidato {nome_completo} rimosso.', 'warning')

        return redirect(url_for('rientro.candidati'))

    righe = RientroCandidato.query.filter_by(anno_scol=ANNO).order_by(
        RientroCandidato.classe, RientroCandidato.cognome).all()
    classi_con_materie = sorted({r.classe for r in
        RientroMateriaClasse.query.filter_by(anno_scol=ANNO).all()})

    return render_template('rientro/candidati.html',
        candidati=righe, classi_con_materie=classi_con_materie, anno=ANNO)


def _docenti_idonei_classe_materia(classe, materia):
    """Tutti i docenti che insegnano quella materia (famiglia sinonimi) in quella classe."""
    mat_can = _materia_canonica(materia)
    righe = OrarioDocente.query.filter_by(classe=classe).all()
    ids = set()
    for o in righe:
        if o.materia and _materia_canonica(o.materia) == mat_can:
            ids.add(o.id_docente)
    return Docente.query.filter(Docente.id.in_(ids)).all() if ids else []


def _docenti_sostituti(materia, escludi_ids):
    """Tutti i docenti della stessa famiglia materia, in QUALSIASI classe."""
    mat_can = _materia_canonica(materia)
    ids = set()
    for o in OrarioDocente.query.filter(OrarioDocente.materia.isnot(None)).all():
        if _materia_canonica(o.materia) == mat_can:
            ids.add(o.id_docente)
    ids -= set(escludi_ids)
    return Docente.query.filter(Docente.id.in_(ids)).order_by(Docente.cognome).all() if ids else []


def _docente_disponibile(id_docente, data):
    """Contratto attivo + nessuna assenza registrata in quella data."""
    if not id_docente or not data:
        return True
    d = Docente.query.get(id_docente)
    if not d or not d.attivo:
        return False
    from models.assenza import Assenza
    ass = Assenza.query.filter_by(id_docente=id_docente, data=data).first()
    return ass is None


# ── COMMISSIONE + CALENDARIO ────────────────────────────────────────
@rientro_bp.route('/rientro/calendario', methods=['GET', 'POST'])
def calendario():
    if request.method == 'POST':
        azione = request.form.get('azione')

        if azione == 'imposta_membro':
            id_coll = int(request.form['id_colloquio'])
            campo = request.form.get('campo')  # docente_1..4 | membro_ds
            id_doc = request.form.get('id_docente') or None
            coll = RientroColloquio.query.get_or_404(id_coll)
            if campo == 'membro_ds':
                coll.id_ruolo_istituzionale = int(id_doc) if id_doc else None
                db.session.commit()
            elif campo in ('docente_1', 'docente_2', 'docente_3', 'docente_4'):
                setattr(coll, f'id_{campo}', int(id_doc) if id_doc else None)
                db.session.commit()
            return redirect(url_for('rientro.calendario'))

        elif azione == 'modifica_orario':
            id_coll = int(request.form['id_colloquio'])
            coll = RientroColloquio.query.get_or_404(id_coll)
            data_str = request.form.get('data', '')
            coll.data = datetime.strptime(data_str, '%Y-%m-%d').date() if data_str else None
            coll.ora_inizio = request.form.get('ora_inizio') or None
            coll.ora_fine = request.form.get('ora_fine') or None
            db.session.commit()
            flash('Orario aggiornato.', 'success')
            return redirect(url_for('rientro.calendario'))

        elif azione == 'genera_bozza':
            _genera_bozza_rientro()
            flash('Bozza colloqui generata.', 'success')
            return redirect(url_for('rientro.calendario'))

        elif azione == 'completa_bozza':
            _genera_bozza_rientro(solo_vuoti=True)
            flash('Bozza completata: i colloqui già calendarizzati non sono stati modificati.', 'success')
            return redirect(url_for('rientro.calendario'))

        elif azione == 'azzera_tutto':
            for coll in RientroColloquio.query.all():
                coll.data = None
                coll.ora_inizio = None
                coll.ora_fine = None
            db.session.commit()
            flash('Calendario azzerato.', 'warning')
            return redirect(url_for('rientro.calendario'))

    periodo = RecuperoPeriodo.query.filter_by(anno_scol=ANNO, codice=PERIODO_CODICE).first()

    candidati_list = RientroCandidato.query.filter_by(anno_scol=ANNO).order_by(
        RientroCandidato.classe, RientroCandidato.cognome).all()

    # Per ogni candidato: materie della sua classe, docenti idonei proposti
    # per ciascuna materia (con flag disponibilità), e sostituti possibili.
    righe = []
    for cand in candidati_list:
        coll = cand.colloquio
        if coll is None:
            coll = RientroColloquio(id_candidato=cand.id)
            db.session.add(coll)
            db.session.commit()

        materie_classe = [m.materia for m in RientroMateriaClasse.query.filter_by(
            anno_scol=ANNO, classe=cand.classe).order_by(RientroMateriaClasse.materia).all()]

        membri_materia = []
        for i, materia in enumerate(materie_classe[:4], start=1):
            id_attuale = getattr(coll, f'id_docente_{i}')
            doc_attuale = Docente.query.get(id_attuale) if id_attuale else None
            disponibile = _docente_disponibile(id_attuale, coll.data) if doc_attuale else None
            sostituti = _docenti_sostituti(materia, escludi_ids=[id_attuale] if id_attuale else [])
            membri_materia.append({
                'campo': f'docente_{i}',
                'materia': materia,
                'docente_attuale': doc_attuale,
                'disponibile': disponibile,
                'sostituti': sostituti,
            })

        righe.append({
            'candidato': cand,
            'colloquio': coll,
            'membri_materia': membri_materia,
        })

    # Persone idonee a fare da membro DS/vicario: lette dall'anagrafica
    # ruoli istituzionali (DS, Vicario 1, Vicario 2, ...) — Roberto scelgo
    # lui per ciascun candidato, non c'è un default automatico.
    membri_ds_possibili = RuoloIstituzionale.query.filter_by(attivo=True).order_by(
        RuoloIstituzionale.ruolo).all()

    # Conflitti: stesso membro impegnato in due colloqui rientro nello
    # stesso slot, o impegnato in un gruppo prova di agosto nello stesso
    # slot (sola lettura su agosto, nessuna scrittura incrociata).
    conflitti = _calcola_conflitti_rientro(righe)

    return render_template('rientro/calendario.html',
        periodo=periodo,
        righe=righe,
        membri_ds_possibili=membri_ds_possibili,
        conflitti=conflitti,
        durata_min=DURATA_MIN,
        anno=ANNO)


def _calcola_conflitti_rientro(righe):
    """
    Verifica sovrapposizioni tra:
    - colloqui di rientro diversi che condividono un membro commissione
    - colloqui di rientro e gruppi prova di agosto (stesso docente come
      somministratore o assistente), in sola lettura — non si tocca mai
      il modulo recupero da qui.
    """
    from routes.recupero import ANNO_AGO, PERIODO_AGO

    def _t(s):
        try:
            h, m = map(int, s.split(':'))
            return h * 60 + m
        except Exception:
            return None

    conflitti = []

    # Gruppi prova agosto con lezioni, per data
    gruppi_agosto = (RecuperoGruppo.query.join(RecGruppoDocente)
                     .filter(RecGruppoDocente.anno_scol == ANNO_AGO,
                             RecuperoGruppo.periodo_codice == PERIODO_AGO).all())
    lezioni_agosto_per_data = {}
    for g in gruppi_agosto:
        membri_g = set(filter(None, [
            g.docente.id if g.docente else None, g.id_sorvegliante]))
        for l in g.lezioni:
            lezioni_agosto_per_data.setdefault(l.data, []).append(
                (_t(l.ora_inizio), _t(l.ora_fine), membri_g, g))

    # Colloqui rientro con orario fissato
    colloqui_validi = [r for r in righe if r['colloquio'].data
                        and r['colloquio'].ora_inizio and r['colloquio'].ora_fine]

    for i, r1 in enumerate(colloqui_validi):
        c1 = r1['colloquio']
        ini1, fin1 = _t(c1.ora_inizio), _t(c1.ora_fine)
        membri1 = {d.id for d in c1.membri_commissione}
        if not membri1 or ini1 is None:
            continue

        # Conflitto con altri colloqui rientro lo stesso giorno
        for r2 in colloqui_validi[i+1:]:
            c2 = r2['colloquio']
            if c2.data != c1.data:
                continue
            ini2, fin2 = _t(c2.ora_inizio), _t(c2.ora_fine)
            if ini2 is None:
                continue
            if ini1 < fin2 and ini2 < fin1:
                membri2 = {d.id for d in c2.membri_commissione}
                comuni = membri1 & membri2
                if comuni:
                    nomi = ', '.join(Docente.query.get(did).cognome for did in comuni)
                    conflitti.append({
                        'tipo': 'rientro',
                        'msg': f'{nomi}: colloquio {r1["candidato"].cognome} e '
                               f'{r2["candidato"].cognome} sovrapposti il {c1.data.strftime("%d/%m")}',
                    })

        # Conflitto con gruppi prova agosto lo stesso giorno
        for ini_ago, fin_ago, membri_ago, g in lezioni_agosto_per_data.get(c1.data, []):
            if ini_ago is None:
                continue
            if ini1 < fin_ago and ini_ago < fin1:
                comuni = membri1 & membri_ago
                if comuni:
                    nomi = ', '.join(Docente.query.get(did).cognome for did in comuni)
                    conflitti.append({
                        'tipo': 'agosto',
                        'msg': f'{nomi}: colloquio {r1["candidato"].cognome} sovrapposto '
                               f'con prova {g.materia[:20]} il {c1.data.strftime("%d/%m")}',
                    })

    return conflitti


def _genera_bozza_rientro(solo_vuoti=False):
    """
    Sequenza i candidati per classe di provenienza, blocchi da 45 minuti,
    nel periodo configurato. Verifica che nessuno dei membri già assegnati
    sia impegnato in un altro colloquio rientro o in un gruppo prova
    agosto nello stesso slot.

    Se solo_vuoti=True (modalità "completa bozza"): salta i candidati che
    hanno già data/ora impostate (manualmente o da una bozza precedente) —
    non li tocca — e pre-carica la loro occupazione, cosi' i nuovi
    colloqui generati non si sovrappongono a quelli già fissati.
    """
    from routes.recupero import ANNO_AGO, PERIODO_AGO

    periodo = RecuperoPeriodo.query.filter_by(anno_scol=ANNO, codice=PERIODO_CODICE).first()
    if not periodo:
        return

    giorni = []
    cur = periodo.data_inizio
    while cur <= periodo.data_fine:
        if cur.weekday() < 5:
            giorni.append(cur)
        cur += timedelta(days=1)
    if not giorni:
        return

    def _t(s):
        try:
            h, m = map(int, s.split(':'))
            return h * 60 + m
        except Exception:
            return None

    def _fmt(m):
        return f'{m // 60:02d}:{m % 60:02d}'

    ora_ini_giorno = _t(periodo.ora_inizio) or 8 * 60
    ora_fine_giorno = _t(periodo.ora_fine) or 16 * 60

    candidati_list = RientroCandidato.query.filter_by(anno_scol=ANNO).order_by(
        RientroCandidato.classe, RientroCandidato.cognome).all()

    # Occupazione: per ogni giorno, lista di (ini, fin, set membri)
    occupazione = {g: [] for g in giorni}

    # Pre-popola con gli impegni dei gruppi prova agosto (sola lettura)
    gruppi_agosto = (RecuperoGruppo.query.join(RecGruppoDocente)
                     .filter(RecGruppoDocente.anno_scol == ANNO_AGO,
                             RecuperoGruppo.periodo_codice == PERIODO_AGO).all())
    for g in gruppi_agosto:
        membri_g = set(filter(None, [g.docente.id if g.docente else None, g.id_sorvegliante]))
        for l in g.lezioni:
            if l.data in occupazione:
                occupazione[l.data].append((_t(l.ora_inizio), _t(l.ora_fine), membri_g))

    candidati_da_piazzare = []
    for cand in candidati_list:
        coll = cand.colloquio
        if coll is None:
            coll = RientroColloquio(id_candidato=cand.id)
            db.session.add(coll)
            db.session.commit()

        ha_orario = coll.data and coll.ora_inizio and coll.ora_fine
        if solo_vuoti and ha_orario:
            # Già calendarizzato (a mano o da una bozza precedente): non
            # si tocca, ma il suo slot va comunque registrato come occupato.
            if coll.data in occupazione:
                membri_esistenti = {d.id for d in coll.membri_commissione}
                occupazione[coll.data].append(
                    (_t(coll.ora_inizio), _t(coll.ora_fine), membri_esistenti))
            continue

        candidati_da_piazzare.append(cand)

    def _slot_libero(giorno, ini, fin, membri):
        for oi, of, occ in occupazione[giorno]:
            if oi is not None and oi < fin and ini < of and (membri & occ):
                return False
        return True

    idx_giorno = 0
    minuto_corrente = ora_ini_giorno

    for cand in candidati_da_piazzare:
        coll = cand.colloquio

        membri = {d.id for d in coll.membri_commissione}
        if not membri:
            continue  # commissione non ancora completata: salta

        piazzato = False
        tentativi = 0
        while not piazzato and tentativi < len(giorni) * 20:
            if minuto_corrente + DURATA_MIN > ora_fine_giorno:
                idx_giorno += 1
                minuto_corrente = ora_ini_giorno
                if idx_giorno >= len(giorni):
                    db.session.commit()
                    return  # periodo esaurito
            giorno = giorni[idx_giorno]
            fin = minuto_corrente + DURATA_MIN
            if _slot_libero(giorno, minuto_corrente, fin, membri):
                coll.data = giorno
                coll.ora_inizio = _fmt(minuto_corrente)
                coll.ora_fine = _fmt(fin)
                occupazione[giorno].append((minuto_corrente, fin, membri))
                minuto_corrente = fin
                piazzato = True
            else:
                minuto_corrente += 15
            tentativi += 1

    db.session.commit()


@rientro_bp.route('/rientro/export-xlsx')
def export_xlsx():
    """
    Export del calendario colloqui di rientro: un foglio unico ordinato
    per giornata, poi per classe di provenienza e orario. Per ogni
    colloquio elenca candidato, le 4 materie con il relativo docente, e
    il membro DS/vicario — stesso impianto grafico degli export del
    modulo recupero (giugno/agosto).
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file

    candidati_list = RientroCandidato.query.filter_by(anno_scol=ANNO).order_by(
        RientroCandidato.classe, RientroCandidato.cognome).all()

    GIORNI = ['Lunedì', 'Martedì', 'Mercoledì', 'Giovedì', 'Venerdì', 'Sabato', 'Domenica']

    BLU    = PatternFill('solid', start_color='1e3a5f')
    AZZUR  = PatternFill('solid', start_color='dbeafe')
    GRAY   = PatternFill('solid', start_color='f3f4f6')
    BOLD   = Font(bold=True)
    BOLD_W = Font(bold=True, color='FFFFFF')
    THIN   = Border(left=Side(style='thin', color='d1d5db'),
                     right=Side(style='thin', color='d1d5db'),
                     top=Side(style='thin', color='d1d5db'),
                     bottom=Side(style='thin', color='d1d5db'))
    CENTER = Alignment(horizontal='center', vertical='center')
    WRAP   = Alignment(wrap_text=True, vertical='center')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Calendario'

    from config_istituto import get_dati_istituto
    ws['A1'] = get_dati_istituto()['nome_istituto']
    ws['A1'].font = Font(bold=True, size=13)
    ws['A2'] = f'COLLOQUI DI RIENTRO DALL\'ESTERO — A.S. {ANNO}'
    ws['A2'].font = Font(bold=True, size=11)
    ws.append([])
    ws.append([])

    # Raggruppa per giornata i colloqui già calendarizzati; quelli senza
    # data finiscono in una sezione "Da calendarizzare" in fondo.
    colloqui_per_giorno = {}
    colloqui_senza_data = []
    for cand in candidati_list:
        coll = cand.colloquio
        if not coll:
            continue
        if coll.data and coll.ora_inizio and coll.ora_fine:
            colloqui_per_giorno.setdefault(coll.data, []).append((coll, cand))
        else:
            colloqui_senza_data.append((coll, cand))

    HEADER_COLS = ['Orario', 'Classe', 'Candidato', 'Materia 1', 'Docente 1',
                   'Materia 2', 'Docente 2', 'Materia 3', 'Docente 3',
                   'Materia 4', 'Docente 4', 'DS / Vicario']
    N_COLS = len(HEADER_COLS)

    def _nome_doc(d):
        return f'{d.cognome} {d.nome or ""}'.strip() if d else '—'

    def _scrivi_header_giorno(titolo):
        row = ws.max_row + 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
        cell = ws.cell(row=row, column=1, value=titolo)
        cell.font = BOLD_W
        cell.fill = BLU
        cell.alignment = CENTER
        row += 1
        for col, h in enumerate(HEADER_COLS, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = BOLD
            c.fill = AZZUR
            c.alignment = CENTER
            c.border = THIN
        return row + 1

    def _scrivi_riga_colloquio(row, coll, cand, mostra_orario=True):
        materie_classe = [m.materia for m in RientroMateriaClasse.query.filter_by(
            anno_scol=ANNO, classe=cand.classe).order_by(RientroMateriaClasse.materia).all()]
        docenti = [coll.docente_1, coll.docente_2, coll.docente_3, coll.docente_4]

        orario_str = f'{coll.ora_inizio}–{coll.ora_fine}' if mostra_orario and coll.ora_inizio else '—'
        vals = [orario_str, cand.classe, f'{cand.cognome} {cand.nome}']
        for i in range(4):
            mat = materie_classe[i] if i < len(materie_classe) else ''
            doc = _nome_doc(docenti[i]) if docenti[i] else ('—' if mat else '')
            vals.append(mat)
            vals.append(doc)
        membro_ds = coll.ruolo_istituzionale
        vals.append(f'{membro_ds.ruolo} — {membro_ds.nome_completo}' if membro_ds else '—')

        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = THIN
            c.alignment = CENTER if col in (1, 2) else WRAP
        return row + 1

    for data in sorted(colloqui_per_giorno.keys()):
        coppie = sorted(colloqui_per_giorno[data], key=lambda cc: (cc[1].classe, cc[0].ora_inizio))
        titolo = f'{GIORNI[data.weekday()]} {data.strftime("%d/%m/%Y")}'
        row = _scrivi_header_giorno(titolo)
        for coll, cand in coppie:
            row = _scrivi_riga_colloquio(row, coll, cand)
        ws.append([])

    if colloqui_senza_data:
        row = _scrivi_header_giorno('⚠︎ Da calendarizzare')
        for coll, cand in sorted(colloqui_senza_data, key=lambda cc: (cc[1].classe, cc[1].cognome)):
            row = _scrivi_riga_colloquio(row, coll, cand, mostra_orario=False)
        ws.append([])

    if not colloqui_per_giorno and not colloqui_senza_data:
        ws.append(['Nessun candidato inserito.'])

    # Larghezze colonne
    larghezze = [13, 10, 22, 16, 18, 16, 18, 16, 18, 16, 18, 22]
    for i, w in enumerate(larghezze, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'calendario_rientro_estero_{ANNO}.xlsx',
    )
