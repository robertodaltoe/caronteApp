"""
import_banca_ore.py
Importa docenti e movimenti banca ore dal file Excel nel database SQLite.
Eseguire UNA SOLA VOLTA dalla cartella SupplenzeApp con venv attivo:
    python import_banca_ore.py
"""
import sys, os
# Percorso della cartella radice del progetto (questo script vive in
# scripts/legacy/, due livelli sotto la radice dove sta app.py).
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from app import create_app
from models import db
from models.docente import Docente
from models.movimento_banca_ore import MovimentoBancaOre
from openpyxl import load_workbook
from datetime import date

# ── CONFIGURAZIONE ────────────────────────────────────────────────────
EXCEL_PATH = os.path.join(os.path.dirname(__file__),
             'data', 'Banca_Ore_Docenti_v3.xlsm')

# Colonne nei fogli settimanali (1-indexed)
COL_NOME       = 1
COL_CONTRATTO  = 2
COL_PERMESSI   = 6   # F — assenze orarie / permessi
COL_CIVICA     = 7   # G — libero Ed. Civica
COL_SUPPLENZA  = 9   # I — ore supplenza svolte
COL_PAGAMENTO  = 4   # D nel Riepilogo

RIGA_DATI_START = 3  # prima riga docenti nei fogli settimanali

# Foglio Riepilogo — colonne
RIE_COL_NOME      = 1
RIE_COL_CONTRATTO = 2
RIE_COL_PAGAMENTO = 4

# ── HELPERS ───────────────────────────────────────────────────────────
def nz(v):
    """Restituisce float da cella, 0 se None/vuoto."""
    if v is None or str(v).strip() == '':
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0

def clean_nome(v):
    if not v:
        return ''
    return str(v).strip().upper().replace(chr(160), ' ')

# ── MAIN ──────────────────────────────────────────────────────────────
def run():
    if not os.path.exists(EXCEL_PATH):
        print(f"❌  File non trovato: {EXCEL_PATH}")
        print("    Copia il file Excel in SupplenzeApp/data/ con il nome:")
        print("    Banca_Ore_Docenti_v3.xlsm")
        sys.exit(1)

    print(f"📂  Lettura file: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH, data_only=True)

    app = create_app()
    with app.app_context():

        # ── 1. IMPORTA DOCENTI ─────────────────────────────────────
        print("\n👥  Importazione docenti...")
        ws_doc = wb['Docenti']
        docenti_map = {}   # cognome -> Docente obj
        importati = 0
        saltati   = 0

        for row in ws_doc.iter_rows(min_row=2, values_only=True):
            cognome_raw = row[0]
            if not cognome_raw:
                continue
            cognome = clean_nome(cognome_raw)
            if cognome == 'DOCENTE':
                continue

            ore_contr = int(row[1]) if row[1] and str(row[1]).strip() != '' else 0
            da_sett   = row[2]
            a_sett    = row[3]

            # Docente attivo = nessuna settimana finale o settimana finale > 25
            attivo = (a_sett is None) or (isinstance(a_sett, (int, float)) and a_sett > 25)

            # Controlla se esiste già
            esistente = Docente.query.filter_by(cognome=cognome).first()
            if esistente:
                saltati += 1
                docenti_map[cognome] = esistente
                continue

            d = Docente(
                cognome      = cognome,
                nome         = '',        # non presente nel file
                nome_display = cognome,
                ore_contratto= ore_contr,
                attivo       = attivo,
            )
            db.session.add(d)
            db.session.flush()
            docenti_map[cognome] = d
            importati += 1

        db.session.commit()
        print(f"   ✓ Importati: {importati} | Già presenti: {saltati}")

        # ── 2. IMPORTA MOVIMENTI DAI FOGLI SETTIMANALI ─────────────
        print("\n📊  Importazione movimenti banca ore...")

        # Leggi elenco fogli settimanali dal foglio 'fogli'
        ws_fogli = wb['fogli']
        sett_list = []
        for row in ws_fogli.iter_rows(min_row=2, max_row=40, values_only=True):
            nome_foglio = row[0]
            if not nome_foglio:
                continue
            nome_foglio = str(nome_foglio).strip()
            if nome_foglio in wb.sheetnames:
                # Periodo (col C) come data approssimativa
                periodo = row[2] if len(row) > 2 else None
                sett_list.append((nome_foglio, periodo))

        totale_mov = 0
        for nome_foglio, periodo in sett_list:
            ws = wb[nome_foglio]

            # Estrai numero settimana per costruire una data fittizia
            try:
                num_sett = int(nome_foglio.replace('sett.', ''))
            except ValueError:
                num_sett = 1

            # Date reali dalle settimane dell'anno scolastico 2025/26
            SETT_DATE = {
                1:date(2025,10,20), 2:date(2025,10,27), 3:date(2025,11,3),
                4:date(2025,11,10), 5:date(2025,11,17), 6:date(2025,11,24),
                7:date(2025,12,1),  8:date(2025,12,8),  9:date(2025,12,15),
                10:date(2026,1,7),  11:date(2026,1,12), 12:date(2026,1,19),
                13:date(2026,1,26), 14:date(2026,2,2),  15:date(2026,2,9),
                16:date(2026,2,18), 17:date(2026,2,23), 18:date(2026,3,2),
                19:date(2026,3,9),  20:date(2026,3,16), 21:date(2026,3,23),
                22:date(2026,3,30), 23:date(2026,4,9),  24:date(2026,4,13),
                25:date(2026,4,20), 26:date(2026,4,27), 27:date(2026,5,4),
                28:date(2026,5,11), 29:date(2026,5,18), 30:date(2026,5,25),
            }
            from datetime import timedelta
            data_base = date(2025, 10, 6)
            data_sett = SETT_DATE.get(num_sett, data_base + timedelta(weeks=num_sett-1))

            for row in ws.iter_rows(min_row=RIGA_DATI_START, values_only=True):
                cognome_raw = row[COL_NOME - 1]
                if not cognome_raw:
                    continue
                cognome = clean_nome(cognome_raw)
                if cognome not in docenti_map:
                    continue

                docente = docenti_map[cognome]

                permessi  = nz(row[COL_PERMESSI  - 1])
                civica    = nz(row[COL_CIVICA     - 1])
                supplenza = nz(row[COL_SUPPLENZA  - 1])

                # Permessi orari → debito (negativo)
                if permessi != 0:
                    m = MovimentoBancaOre(
                        id_docente  = docente.id,
                        data        = data_sett,
                        minuti      = int(-abs(permessi) * 60),
                        tipo        = 'permesso',
                        descrizione = f'Permesso/assenza oraria — {nome_foglio}',
                    )
                    db.session.add(m)
                    totale_mov += 1

                # Ed. Civica libero → debito (negativo)
                if civica != 0:
                    m = MovimentoBancaOre(
                        id_docente  = docente.id,
                        data        = data_sett,
                        minuti      = int(-abs(civica) * 60),
                        tipo        = 'civica',
                        descrizione = f'Libero Ed. Civica — {nome_foglio}',
                    )
                    db.session.add(m)
                    totale_mov += 1

                # Supplenza svolta → credito (positivo)
                if supplenza != 0:
                    m = MovimentoBancaOre(
                        id_docente  = docente.id,
                        data        = data_sett,
                        minuti      = int(abs(supplenza) * 60),
                        tipo        = 'supplenza_recupero',
                        descrizione = f'Supplenza svolta — {nome_foglio}',
                    )
                    db.session.add(m)
                    totale_mov += 1

        # ── 3. IMPORTA ORE A PAGAMENTO DAL RIEPILOGO ───────────────
        print("💶  Importazione ore a pagamento dal Riepilogo...")
        ws_rie = wb['Riepilogo']
        pag_mov = 0
        for row in ws_rie.iter_rows(min_row=2, values_only=True):
            cognome_raw = row[RIE_COL_NOME - 1]
            if not cognome_raw:
                continue
            cognome = clean_nome(cognome_raw)
            if cognome not in docenti_map:
                continue
            pagamento = nz(row[RIE_COL_PAGAMENTO - 1])
            if pagamento != 0:
                docente = docenti_map[cognome]
                m = MovimentoBancaOre(
                    id_docente  = docente.id,
                    data        = date(2026, 5, 23),   # data di registrazione
                    minuti      = int(abs(pagamento) * 60),
                    tipo        = 'supplenza_pagamento',
                    descrizione = 'Ore richieste a pagamento (da Riepilogo)',
                )
                db.session.add(m)
                pag_mov += 1

        db.session.commit()
        print(f"   ✓ Movimenti settimanali: {totale_mov}")
        print(f"   ✓ Ore a pagamento: {pag_mov}")
        print(f"\n✅  Importazione completata.")
        print(f"   Docenti nel database: {Docente.query.count()}")
        print(f"   Movimenti nel database: {MovimentoBancaOre.query.count()}")

if __name__ == '__main__':
    run()
