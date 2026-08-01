"""
import_orario.py — versione 3 (struttura con colonna tipo riga)
Struttura foglio '7_ORARIO DEFINITIVO_teachers_ti':
  - Riga r1: titolo
  - Riga r2: giorni (Lunedì, Martedì...)
  - Riga r3: orari (datetime.time)
  - Righe dati: col A = CLASSE | MATERIE | COMPRESENZA
                col B = nome docente (solo su CLASSE)
                col C+ = classi/materie/compresenze nelle colonne orario

Struttura foglio 'Docenti':
  - Riga 1: titolo
  - Riga 2: intestazioni
  - Righe 3+: COGNOME | NOME | MATERIE | ORE_CALCOLATE | TIPO_CONTRATTO | ATTIVO

Compresenze: formato 'COGNOME1 | COGNOME2' nella cella
"""
import sys, os, re, datetime
# Percorso della cartella radice del progetto (questo script vive in
# scripts/legacy/, due livelli sotto la radice dove sta app.py).
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from app import create_app
from models import db
from models.docente import Docente
from models.orario_docente import OrarioDocente
from openpyxl import load_workbook
from collections import Counter

EXCEL_PATH = os.path.join(os.path.dirname(__file__),
             'data', 'ORARIO DEFINITIVO_DOCENTI DA GENNAIO 26.xlsx')

SHEET_ORARIO  = '7_ORARIO DEFINITIVO_teachers_ti'
SHEET_DOCENTI = 'Docenti'

LIBERO = {'---', '-x-', '', 'none'}
GIORNI = ['Lunedì','Martedì','Mercoledì','Giovedì','Venerdì','Sabato']

def clean(v):
    return str(v).strip() if v is not None else ''

def is_libero(v):
    return clean(v).lower() in LIBERO or clean(v) == ''

def is_classe(s):
    return bool(re.match(r'^\d[A-Z]', s.strip()))

def build_col_map(ws):
    """
    Legge riga 2 (giorni) e riga 3 (orari) per costruire
    mappa colonna -> (giorno_num, ora_num).
    """
    col_map = {}
    giorno_corrente = None
    ora_counter = {}   # giorno_num -> contatore ora

    # Riga 2: nomi giorni
    giorno_map = {}
    for c in range(1, ws.max_column + 1):
        v = clean(ws.cell(2, c).value)
        for i, g in enumerate(GIORNI):
            if g.lower() in v.lower():
                giorno_map[c] = i
                break

    # Riga 3: orari — ogni colonna con un orario valido è un'ora
    for c in range(1, ws.max_column + 1):
        v = ws.cell(3, c).value
        if not isinstance(v, datetime.time):
            continue
        # Trova il giorno di questa colonna (cerca la colonna giorno più vicina a sinistra)
        g_num = None
        for gc in sorted(giorno_map.keys(), reverse=True):
            if gc <= c:
                g_num = giorno_map[gc]
                break
        if g_num is None:
            continue
        # Ora numero = progressivo dentro il giorno
        ora_counter[g_num] = ora_counter.get(g_num, 0) + 1
        col_map[c] = (g_num, ora_counter[g_num])

    return col_map

def run():
    if not os.path.exists(EXCEL_PATH):
        print(f"❌  File non trovato: {EXCEL_PATH}")
        print(f"    Atteso in: {EXCEL_PATH}")
        sys.exit(1)

    print(f"📂  Lettura file: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH, data_only=True)

    # ── Celle unite ──────────────────────────────────────────
    ws_or = wb[SHEET_ORARIO]
    merged_masters = {}
    for merge in ws_or.merged_cells.ranges:
        mr, mc = merge.min_row, merge.min_col
        for r in range(merge.min_row, merge.max_row + 1):
            for c in range(merge.min_col, merge.max_col + 1):
                merged_masters[(r, c)] = (mr, mc)

    def val(r, c):
        mr, mc = merged_masters.get((r, c), (r, c))
        return ws_or.cell(mr, mc).value

    # ── Mappa colonne ────────────────────────────────────────
    col_map = build_col_map(ws_or)
    print(f"   Colonne orario mappate: {len(col_map)}")
    # Debug: mostra distribuzione per giorno
    from collections import defaultdict
    per_giorno = defaultdict(list)
    for c, (g, o) in col_map.items():
        per_giorno[g].append(o)
    for g in range(6):
        if per_giorno[g]:
            print(f"   {GIORNI[g]}: {len(per_giorno[g])} ore ({min(per_giorno[g])}ª–{max(per_giorno[g])}ª)")

    app = create_app()
    with app.app_context():

        # ── 1. Importa anagrafica dal foglio Docenti ─────────
        print("\n👥  Importazione anagrafica docenti...")
        ws_doc = wb[SHEET_DOCENTI]
        docenti_db = {d.cognome.upper(): d for d in Docente.query.all()}

        nuovi = 0
        aggiornati_mat = 0
        for r in range(3, ws_doc.max_row + 1):
            cognome_v = ws_doc.cell(r, 1).value
            if not cognome_v:
                continue
            cognome  = clean(cognome_v).upper()
            nome     = clean(ws_doc.cell(r, 2).value)
            materie  = clean(ws_doc.cell(r, 3).value)
            ore_v    = ws_doc.cell(r, 4).value
            attivo_v = clean(ws_doc.cell(r, 6).value).upper()

            ore    = int(ore_v) if ore_v and str(ore_v).strip().isdigit() else 0
            attivo = attivo_v in ('SÌ', 'SI', 'S', '1', 'TRUE', 'YES')

            if cognome in docenti_db:
                d = docenti_db[cognome]
                # Aggiorna materie e ore se non già impostati
                if materie and not d.materia:
                    d.materia = materie
                    aggiornati_mat += 1
                if ore and not d.ore_contratto:
                    d.ore_contratto = ore
            else:
                d = Docente(
                    cognome      = cognome,
                    nome         = nome,
                    nome_display = f"{cognome} {nome[0]}." if nome else cognome,
                    materia      = materie,
                    ore_contratto= ore,
                    attivo       = attivo,
                )
                db.session.add(d)
                db.session.flush()
                docenti_db[cognome] = d
                nuovi += 1

        db.session.commit()
        print(f"   ✓ Nuovi docenti: {nuovi} | Materie aggiornate: {aggiornati_mat}")
        print(f"   ✓ Totale docenti in DB: {Docente.query.count()}")

        # ── 2. Importa orario ────────────────────────────────
        print("\n📅  Importazione orario...")
        OrarioDocente.query.delete()
        db.session.commit()

        slot_totali  = 0
        comp_totali  = 0
        docente_curr = None   # docente corrente mentre scorriamo le righe

        for r in range(4, ws_or.max_row + 1):
            tipo_riga = clean(val(r, 1)).upper()

            if tipo_riga not in ('CLASSE', 'MATERIE', 'COMPRESENZA'):
                continue

            if tipo_riga == 'CLASSE':
                # Nuovo docente
                nome_doc = clean(val(r, 2)).upper()
                if nome_doc:
                    docente_curr = docenti_db.get(nome_doc)
                    if docente_curr is None:
                        # Crea docente mancante
                        docente_curr = Docente(
                            cognome=nome_doc, nome='',
                            nome_display=nome_doc, attivo=True
                        )
                        db.session.add(docente_curr)
                        db.session.flush()
                        docenti_db[nome_doc] = docente_curr
                        print(f"   ➕ Docente creato: {nome_doc}")

                if docente_curr is None:
                    continue

                # Leggi slot orario per questo docente
                # La riga MATERIE è quella successiva
                riga_materie = r + 1  # sarà MATERIE

                for c, (giorno, ora) in col_map.items():
                    classe_v  = val(r, c)
                    materia_v = val(riga_materie, c)

                    classe_s  = clean(classe_v)
                    materia_s = clean(materia_v)

                    if is_libero(classe_s):
                        continue

                    # Tipo ora
                    if is_classe(classe_s):
                        tipo = 'lezione'
                    elif 'POTENZ' in classe_s.upper():
                        tipo = 'potenziamento'
                    elif 'DISPOS' in classe_s.upper():
                        tipo = 'disposizione'
                    else:
                        tipo = 'altro'

                    slot = OrarioDocente(
                        id_docente = docente_curr.id,
                        giorno     = giorno,
                        ora        = ora,
                        classe     = classe_s,
                        materia    = materia_s,
                        tipo_ora   = tipo,
                    )
                    db.session.add(slot)
                    slot_totali += 1

            elif tipo_riga == 'COMPRESENZA':
                # Compresenza: formato 'COGNOME1 | COGNOME2' nella cella
                # La classe reale è nella riga CLASSE sopra (docente_curr già impostato)
                # Dobbiamo trovare la riga CLASSE precedente per la classe
                # In realtà la classe è nella riga CLASSE del docente corrente
                # già processata — qui registriamo il secondo docente

                for c, (giorno, ora) in col_map.items():
                    comp_v = val(r, c)
                    comp_s = clean(comp_v)

                    if is_libero(comp_s) or '|' not in comp_s:
                        continue

                    # Estrai i due cognomi
                    cognomi = [x.strip().upper() for x in comp_s.split('|')]

                    # Trova la classe reale: cerca nella riga CLASSE del docente corrente
                    # Risali fino alla riga CLASSE più recente
                    classe_reale = ''
                    for rr in range(r-1, 3, -1):
                        if clean(val(rr, 1)).upper() == 'CLASSE':
                            classe_reale = clean(val(rr, c))
                            break

                    # Materia: riga MATERIE subito dopo la riga CLASSE
                    materia_comp = ''
                    for rr in range(r-1, 3, -1):
                        if clean(val(rr, 1)).upper() == 'CLASSE':
                            materia_comp = clean(val(rr+1, c))
                            break

                    for cognome_comp in cognomi:
                        doc_comp = docenti_db.get(cognome_comp)
                        if doc_comp is None:
                            doc_comp = Docente(
                                cognome=cognome_comp, nome='',
                                nome_display=cognome_comp, attivo=True
                            )
                            db.session.add(doc_comp)
                            db.session.flush()
                            docenti_db[cognome_comp] = doc_comp
                            print(f"   ➕ Docente compresenza creato: {cognome_comp}")

                        # Evita duplicati (il primo docente è già stato inserito)
                        if doc_comp.id == (docente_curr.id if docente_curr else -1):
                            continue

                        slot = OrarioDocente(
                            id_docente = doc_comp.id,
                            giorno     = giorno,
                            ora        = ora,
                            classe     = classe_reale,
                            materia    = materia_comp,
                            tipo_ora   = 'compresenza',
                        )
                        db.session.add(slot)
                        slot_totali += 1
                        comp_totali += 1

        db.session.commit()

        print(f"   ✓ Slot totali: {OrarioDocente.query.count()}")
        print(f"   ✓ Di cui compresenze: {comp_totali}")
        print(f"\n✅  Importazione completata.")

        # Riepilogo docenti con orario
        doc_con_orario = db.session.query(
            OrarioDocente.id_docente
        ).distinct().count()
        print(f"   Docenti con orario: {doc_con_orario}")

        # Verifica ABRAMINI
        from models.docente import Docente as D
        abr = D.query.filter_by(cognome='ABRAMINI').first()
        if abr:
            slots = OrarioDocente.query.filter_by(id_docente=abr.id).all()
            print(f"\n   Verifica ABRAMINI ({len(slots)} slot):")
            for s in slots[:5]:
                print(f"     {GIORNI[s.giorno]} {s.ora}ª — {s.classe} ({s.materia})")

if __name__ == '__main__':
    run()
